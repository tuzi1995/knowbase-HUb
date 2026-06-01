import argparse
import os
import re
from openpyxl import load_workbook


def extract_section(text, header):
    pattern = re.compile(r"(?mi)^\s*#\s*{}\s*$".format(re.escape(header)))
    next_header = re.compile(r"(?mi)^\s*#\s*.+")
    lines = text.splitlines()
    start_idx = None
    for i, line in enumerate(lines):
        if pattern.match(line):
            start_idx = i + 1
            break
    if start_idx is None:
        return ""
    collected = []
    for j in range(start_idx, len(lines)):
        if next_header.match(lines[j]):
            break
        collected.append(lines[j])
    return "\n".join(collected).rstrip("\n")


def process_excel(path):
    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    wb = load_workbook(path)
    ws = wb.active
    max_row = ws.max_row
    success = 0
    for r in range(2, max_row + 1):
        content = ws.cell(row=r, column=2).value
        if not isinstance(content, str):
            continue
        print("正在处理第{}行".format(r))
        name = extract_section(content, "配件名称")
        models = extract_section(content, "适用机型")
        ws.cell(row=r, column=3).value = name
        ws.cell(row=r, column=4).value = models
        success += 1
    try:
        wb.save(path)
        print("已保存到原文件：{}".format(path))
    except PermissionError:
        base, ext = os.path.splitext(path)
        alt = base + "_已拆分" + ext
        wb.save(alt)
        print("原文件可能被占用，已保存到备用文件：{}".format(alt))
    print("完成填充，共处理 {} 行".format(success))


def main():
    parser = argparse.ArgumentParser(description="从B列内容拆分到C/D列")
    parser.add_argument(
        "excel",
        nargs="?",
        default=r"D:\AI处理\accessory_markdown\KnowledgeBaseTool\Excel_Data\markdown汇总.xlsx",
        help="Excel 文件路径，默认 D:\\AI处理\\accessory_markdown\\KnowledgeBaseTool\\Excel_Data\\markdown汇总.xlsx",
    )
    args = parser.parse_args()
    process_excel(args.excel)


if __name__ == "__main__":
    main()

