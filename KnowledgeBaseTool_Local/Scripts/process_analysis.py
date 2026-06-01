import openpyxl
import os

# 1. 定义设备列表
devices_raw = """P5,T6,T7,T7 Pro,T7S,T7S Plus,T8,T8 Plus,G10,G10 Plus,G10S,G10S Auto,G10S Pro,G10S Pure,G20,G20S,P10,P10 Pro,P10S,P10S Pure,P10S Pro 系列,P10S Pro 系列超薄嵌入式,P20 Pro,G20S Ultra,V20,石头星耀Pro,G30,P20 Plus,P20 Ultra,G30 Space探索版,G30 U,P20,P20极光,P20 Ultra Plus,P20活水版,P20极光 增强版,G30曜石,G30S Pro,P20 Ultra活水版,
A10,A10 Plus,A10 Ultra,A10 UltraE,A20,A20 Air,A20 Pro,U10,A30 Pro,A30,A30 CE,A30 ProCE,A30 Pro Steam,A30 Lite,A30 SE,A30 Muse,A30 U,A30 Combo,A30 Pro Turbo,A30 Pure,A30 Pro Combo,A30 Pro Ultra,拓界 Mist,拓界 Aura,A30 Pro Steam 5合1版,拓界 Pro,A30 CE 畅享版,A30 CE 悦享版,A30 Pro Steam 智享版,A30 Pro CE 悦享版,
H5,H6,H7,H50 Ultra,H50,H50 Pro,
H1,H1 Air,H1 Neo,M1,M1 Pure, 洗烘一体机 Z1, 洗烘一体机 Z1 Pro,Q1,Q1M, 洗烘一体机 Z1 Plus, 洗衣机 Z1,M1S,M1S Pure, 洗衣机 Z1 Max, 洗衣机 Z1 Max 银盾版,洗衣机 Z1 Max 智控版,热泵干衣机 Z1 Max, 分子筛干衣机 Z1 Max, 套装 Z1 Max,Z1 Max Pro,Z1 Max Ultra,Q1 Hello Kitty"""

# 清洗设备列表，生成标准化的集合用于匹配
# 标准化规则：去除所有空格，转小写（防止大小写差异）
device_list = [d.strip() for d in devices_raw.replace('\n', ',').split(',') if d.strip()]
device_set_norm = set(d.replace(' ', '').lower() for d in device_list)

def normalize(text):
    if not isinstance(text, str):
        return ""
    return text.strip().replace(' ', '').lower()

# 需要在 AS 列中移除的后缀/修饰词
CLEAN_AS_TERMS = ["水箱版", "上下水版", "智能上下水版"]

def clean_as(text):
    """移除 AS 中的指定后缀词并去除两端空格"""
    if not isinstance(text, str):
        return ""
    s = text.strip()
    for term in CLEAN_AS_TERMS:
        if term in s:
            s = s.replace(term, "")
    return s.strip()

def is_device(text):
    """判断文本是否为设备"""
    if not text:
        return False
    norm_text = normalize(text)
    # 精确匹配（归一化后）
    if norm_text in device_set_norm:
        return True
    return False

def check_contains(text, keyword):
    """检查 text 是否包含 keyword (归一化包含)"""
    if not text or not keyword:
        return False
    # 这里使用简单的字符串包含，不进行归一化，因为L列是自然语言，I/AS是专有名词
    # 但为了提高匹配率，可以尝试去掉 keyword 的空格
    k_clean = keyword.strip()
    if k_clean and k_clean in text:
        return True
    return False

def process_row(row_idx, ws):
    # 获取列值 (0-based index in list, but openpyxl is 1-based)
    # C=3, I=9, L=12, AS=45
    # cell index = col_num
    
    # 辅助函数：安全获取值
    def get_val(col):
        v = ws.cell(row=row_idx, column=col).value
        return str(v).strip() if v is not None else ""

    val_I = get_val(9)   # 绑定设备
    val_L = get_val(12)  # 用户问题
    val_AS_raw = get_val(45) # 浏览设备/配件（原始）
    val_AS = clean_as(val_AS_raw)  # 清理指定后缀
    # 将清理后的 AS 写回单元格
    ws.cell(row=row_idx, column=45).value = val_AS

    # --- 1. 处理 AT 列 (46) ---
    # 规则：AS为配件直接返原内容；AS为设备则对比I列，一致返「是」，不一致返「否」
    val_AT = ""
    is_as_device = is_device(val_AS)
    
    if is_as_device:
        # AS 是设备，对比 I 列
        # 对比规则：归一化后比较
        if normalize(val_I) == normalize(val_AS):
            val_AT = "是"
        else:
            val_AT = "否"
    else:
        # AS 是配件 (或者无法识别的设备)，返回原内容
        val_AT = val_AS

    # --- 2. 处理 AU 列 (47) ---
    # 规则：提及绑定设备(I)返「绑定设备」，提及浏览设备(AS)返「浏览设备/配件」
    # 均提及优先绑定设备，未明确提及默认返「绑定设备」
    val_AU = "绑定设备" # 默认值
    match_I = False
    match_AS = False
    
    # 判断逻辑：L 是否包含 I 或 AS
    # 注意：如果 I 或 AS 为空，视为未提及
    if val_I and val_I in val_L:
        match_I = True
    if val_AS and val_AS in val_L:
        match_AS = True
        
    if match_I:
        val_AU = "绑定设备"
    elif match_AS:
        val_AU = "浏览设备/配件"
    else:
        val_AU = "绑定设备" # 默认

    # --- 3. 处理 AV 列 (48) ---
    # 规则：引用L列问题片段，结合I/AS列内容，简洁说明AU列判断原因
    val_AV = ""
    
    # 截取 L 列片段 (如果太长)
    l_snippet = val_L[:20] + "..." if len(val_L) > 20 else val_L
    l_snippet = l_snippet.replace('\n', ' ')
    
    if match_I:
        val_AV = f"问题提及'{val_I}'，命中绑定设备"
    elif match_AS:
        val_AV = f"问题提及浏览品'{val_AS}'，未提绑定设备，命中浏览设备"
    else:
        if not val_I and not val_AS:
             val_AV = "无设备信息，默认绑定设备"
        elif not val_I:
             val_AV = "无绑定设备且未提浏览品，默认绑定设备"
        else:
             val_AV = f"问题未明确提及设备，默认绑定设备: {val_I}"

    # 写入单元格
    ws.cell(row=row_idx, column=46).value = val_AT
    ws.cell(row=row_idx, column=47).value = val_AU
    ws.cell(row=row_idx, column=48).value = val_AV

def main():
    file_path = r"D:\AI处理\accessory_markdown\KnowledgeBaseTool\Excel_Data\商品链接-0129_processed.xlsx"
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return

    print(f"正在加载文件: {file_path}")
    # data_only=False 以保留公式（如果有），但在这种处理中通常不需要
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    max_row = ws.max_row
    
    print(f"开始处理，共 {max_row-1} 行数据...")
    
    for r in range(2, max_row + 1):
        process_row(r, ws)
        if (r - 1) % 100 == 0:
            print(f"已处理 {r-1} 行...", end='\r')
            
    print(f"\n处理完成。")
    wb.save(file_path)
    print(f"文件已保存。")

if __name__ == "__main__":
    main()
