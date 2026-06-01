import os
import json
import argparse
from openpyxl import load_workbook


def extract_skuname(text):
    if not isinstance(text, str):
        return ""
    try:
        # 尝试清理可能存在的非 JSON 字符，或者直接解析
        # 有时候 JSON 字符串可能被包裹在其他字符中，或者包含 Markdown 代码块标记
        text = text.strip()
        if text.startswith('`') and text.endswith('`'):
             text = text.strip('`')
        
        # 尝试直接解析 JSON
        data = json.loads(text)
        return data.get("skuName", "")
    except json.JSONDecodeError:
        # 如果解析失败，尝试使用简单的字符串查找作为后备方案
        import re
        pattern = re.compile(r'"skuName"\s*:\s*"([^"]*)"')
        m = pattern.search(text)
        if m:
            return m.group(1)
        return ""
    except Exception as e:
        print(f"解析错误: {e}")
        return ""


def process_excel(path):
    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    print(f"正在打开文件: {path}", flush=True)
    try:
        wb = load_workbook(path)
        ws = wb.active
        max_row = ws.max_row
        print(f"文件加载成功，总行数: {max_row}", flush=True)
    except Exception as e:
        print(f"文件加载失败: {e}", flush=True)
        return

    success = 0
    
    # AR 列是第 44 列
    col_ar_idx = 44
    # AS 列是第 45 列
    col_as_idx = 45

    for r in range(2, max_row + 1):
        val = ws.cell(row=r, column=col_ar_idx).value
        if val is None:
            continue
            
        sku = extract_skuname(str(val))
        
        # 写入 AS 列
        ws.cell(row=r, column=col_as_idx).value = sku
        
        if sku:
            success += 1
            if success % 10 == 0:
                print(f"已处理 {success} 行数据...")

    try:
        wb.save(path)
        print(f"已保存到原文件：{path}")
    except PermissionError:
        base, ext = os.path.splitext(path)
        alt = base + "_已填充" + ext
        wb.save(alt)
        print(f"原文件可能被占用，已保存到备用文件：{alt}")
    print(f"完成填充，共提取 {success} 个 skuName")


def main():
    parser = argparse.ArgumentParser(description="从AR列提取skuName到AS列")
    parser.add_argument(
        "excel",
        nargs="?",
        default=r"D:\AI处理\accessory_markdown\商品链接-0129.xlsx",
        help="Excel 文件路径",
    )
    args = parser.parse_args()
    process_excel(args.excel)


if __name__ == "__main__":
    main()

