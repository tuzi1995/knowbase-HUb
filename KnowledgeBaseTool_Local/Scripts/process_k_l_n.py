import openpyxl
import os

def process_row(row_idx, ws):
    # K列 = 11, L列 = 12, N列 = 14
    
    # 辅助函数：安全获取值
    def get_val(col):
        v = ws.cell(row=row_idx, column=col).value
        # 如果是 None，转为空字符串；如果是其他类型，转为 string
        return str(v).strip() if v is not None else ""

    val_K = get_val(11)  # K列
    val_L = get_val(12)  # L列
    
    val_N = ""

    # 逻辑判断
    # 1. 若是 [], N 列返回 商品设备-无绑定设备
    if val_K == "[]":
        val_N = "商品设备-无绑定设备"
    # 2. 若是同一行 K 列内容和 L 列内容一样，则 N 列返回 同一设备
    # 注意：这里假设 K 和 L 都不为空时才比较，或者空对空也算？
    # 通常 excel 处理中，如果是空对空，可能不算“内容一样”的业务含义，但按字面意思是一样的。
    # 考虑到 K=[] 已经被处理了，这里主要处理非 [] 的情况。
    # 如果 K 和 L 都是空字符串，是否算同一设备？
    # 根据 "无绑定设备" 的逻辑，K=[] 是无绑定。如果 K="" 也是无绑定吗？
    # 假设严格按照用户指令：K 和 L 内容一样。
    elif val_K == val_L:
        val_N = "同一设备"
    else:
        # 其余的不符合要求保持空值
        val_N = ""

    # 写入 N 列 (14)
    ws.cell(row=row_idx, column=14).value = val_N

def main():
    file_path = r"D:\AI处理\accessory_markdown\KnowledgeBaseTool\Excel_Data\商品链接-0129_processed.xlsx"
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        return

    print(f"正在加载文件: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    # 指定 Sheet1
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
    else:
        ws = wb.active
        print(f"未找到 Sheet1，使用当前活动工作表: {ws.title}")

    max_row = ws.max_row
    print(f"开始处理，共 {max_row-1} 行数据...")
    
    for r in range(2, max_row + 1):
        process_row(r, ws)
        if (r - 1) % 100 == 0:
            print(f"已处理 {r-1} 行...", end='\r')
            
    print(f"\n处理完成。")
    wb.save(file_path)
    print(f"文件已保存。")

if __name__ == "__main__":
    main()
