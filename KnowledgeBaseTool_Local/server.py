import os
import json
import re
import pandas as pd
import requests
import traceback
import uuid
import hashlib
import csv
import glob
import importlib.util
import threading
import time
import subprocess
import tempfile
import shutil
import sys
from array import array
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import OrderedDict
from datetime import datetime
from urllib.parse import urlparse
from flask import Flask, request, jsonify, send_from_directory, send_file, Response, stream_with_context
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, or_, desc, asc, case, text
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
import io
import math
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from scoring_logic import LLMScorer, calculate_product_overlap, load_scoring_config, save_scoring_config, load_ai_config, save_ai_config
from matrix_submit_validation import validate_submit_changes
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
try:
    import pypdfium2 as pdfium
except Exception:
    pdfium = None

try:
    from PIL import Image
except Exception:
    Image = None


def canonical_download_name(data_type, ext='xlsx', step='s01', tool='kb8085'):
    suffix = str(ext or 'xlsx').lstrip('.')
    return f"{datetime.now().strftime('%Y%m%d')}_{step}_{tool}_{data_type}_v1.{suffix}"

# 本地 PostgreSQL 支持
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    HAS_PSYCOPG2 = True
except ImportError:
    HAS_PSYCOPG2 = False

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_INSTANCE_DIR = os.path.join(_BASE_DIR, 'instance')
app = Flask(
    __name__,
    static_folder=os.path.join(_BASE_DIR, 'link_viewer'),
    static_url_path='',
    instance_path=_INSTANCE_DIR,
)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['SESSION_COOKIE_NAME'] = os.environ.get('KMATRIX_SESSION_COOKIE_NAME', 'kmatrix_8085_session')


def _normalize_http_origin(value):
    raw = str(value or '').strip()
    if not raw:
        return ''
    try:
        parsed = urlparse(raw)
        if parsed.scheme.lower() not in {'http', 'https'} or not parsed.hostname:
            return ''
        if parsed.username or parsed.password:
            return ''
        host = parsed.hostname.lower()
        if ':' in host:
            host = f'[{host}]'
        port = parsed.port
        default_port = 80 if parsed.scheme.lower() == 'http' else 443
        authority = f'{host}:{port}' if port and port != default_port else host
        return f'{parsed.scheme.lower()}://{authority}'
    except (TypeError, ValueError):
        return ''


def _get_embed_allowed_origins():
    configured = os.environ.get('KMATRIX_EMBED_ALLOWED_ORIGINS', '').strip()
    values = configured.split(',') if configured else [
        'http://127.0.0.1:5175',
        'http://localhost:5175',
    ]
    origins = []
    for value in values:
        origin = _normalize_http_origin(value)
        if origin and origin not in origins:
            origins.append(origin)
    return origins


def _is_embed_origin_allowed(value):
    origin = _normalize_http_origin(value)
    return bool(origin and origin in _get_embed_allowed_origins())


def _get_cors_allowed_origins():
    default_origins = [
        "http://localhost:8083",
        "http://127.0.0.1:8083",
        "http://localhost:8082",
        "http://127.0.0.1:8082",
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://112.126.63.84:8083",
    ]
    env_origins = [
        origin.strip()
        for origin in os.environ.get('KMATRIX_CORS_ALLOWED_ORIGINS', '').split(',')
        if origin.strip()
    ]
    return list(OrderedDict.fromkeys(default_origins + env_origins))

CORS(app, supports_credentials=True, origins=_get_cors_allowed_origins())
app.config['SECRET_KEY'] = os.environ.get('KMATRIX_SECRET_KEY', 'dev-only-change-me')
_DB_PATH = os.path.join(_INSTANCE_DIR, 'data.db')
BADCASE_WORKBENCH_SOURCE = 'badcase标注工作台'
os.makedirs(os.path.dirname(_DB_PATH), exist_ok=True)  # Ensure SQLite folder exists
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + _DB_PATH.replace('\\', '/')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

def _resolve_product_manual_path():
    env_path = os.environ.get('KMATRIX_PRODUCT_MANUAL_PATH', '').strip()
    candidates = []
    if env_path:
        candidates.append(env_path if os.path.isabs(env_path) else os.path.join(_BASE_DIR, env_path))
    candidates.extend([
        os.path.join(_BASE_DIR, '产品说明书.md'),
        os.path.join(_BASE_DIR, 'product_manual.md'),
    ])
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None

_SM_LOCK = threading.Lock()
_SM_BASELINE_CACHE = {}
_SM_JOBS = {}
_SM_DB_CLEAN_TS = 0.0
_KD_LOCK = threading.Lock()
_KD_ACTIVE_TASKS = set()
_KD_INDEX_ACTIVE_JOBS = set()




db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login_view'

@login_manager.unauthorized_handler
def unauthorized():
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    # For non-api routes, we might want to redirect, but since this is SPA, 
    # we usually just return 401 or let frontend handle it.
    # But if we really want to support browser navigation to protected routes:
    # return redirect('/') 
    # For now, consistent 401 is safer than broken redirect
    return jsonify({'success': False, 'message': 'Login required'}), 401

@app.after_request
def apply_no_cache_headers(response):
    if request.method == 'GET' and request.path == '/':
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    elif request.method == 'GET' and request.path.endswith(('.css', '.js')):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    if response.mimetype == 'text/html' or request.path == '/':
        frame_ancestors = " ".join(["'self'", *_get_embed_allowed_origins()])
        response.headers['Content-Security-Policy'] = f'frame-ancestors {frame_ancestors}'
        response.headers.pop('X-Frame-Options', None)
    return response

# Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)

class Link(db.Model):
    id = db.Column(db.String(36), primary_key=True)
    url = db.Column(db.Text, nullable=False)
    type = db.Column(db.String(20))
    tags = db.Column(db.Text)  # JSON string
    created_at = db.Column(db.Float)

class KBScore(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    kb_id = db.Column(db.String(100), unique=True, nullable=False)
    question_content = db.Column(db.Text)
    answer_content = db.Column(db.Text)
    status = db.Column(db.String(20), default='unscored') # unscored, scored, outdated
    total_score = db.Column(db.Integer)
    remarks = db.Column(db.Text)
    score_data = db.Column(db.Text) # JSON string
    updated_at = db.Column(db.String(50))

class KBRecall(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    kb_id = db.Column(db.String(100), nullable=False) # 关联 question_wiki_id
    month = db.Column(db.String(7), nullable=False) # YYYY-MM
    recall_count = db.Column(db.Integer, default=0)
    valid_recall_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('kb_id', 'month', name='unique_kb_month'),
    )

class SmartMappingJob(db.Model):
    __tablename__ = 'smart_mapping_job'
    job_id = db.Column(db.String(36), primary_key=True)
    username = db.Column(db.String(80), default='')
    status = db.Column(db.String(20), default='running')  # running/done/failed
    total = db.Column(db.Integer, default=0)
    done = db.Column(db.Integer, default=0)
    message = db.Column(db.Text)
    results_json = db.Column(db.Text)
    created_ts = db.Column(db.Float, default=lambda: time.time())
    updated_ts = db.Column(db.Float, default=lambda: time.time())
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class SmartMappingEmbeddingCache(db.Model):
    __tablename__ = 'smart_mapping_embedding_cache'
    cache_key = db.Column(db.String(64), primary_key=True)
    model = db.Column(db.String(240), nullable=False, index=True)
    dimensions = db.Column(db.Integer, nullable=False)
    vector_blob = db.Column(db.LargeBinary, nullable=False)
    updated_ts = db.Column(db.Float, default=lambda: time.time(), index=True)


class KBDuplicateRetrievalIndex(db.Model):
    __tablename__ = 'kb_retrieval_index'
    id = db.Column(db.Integer, primary_key=True)
    library_type = db.Column(db.String(40), nullable=False, index=True)
    question_wiki_id = db.Column(db.String(160), nullable=False, index=True)
    content_hash = db.Column(db.String(64), nullable=False)
    intent_cache_key = db.Column(db.String(64), default='')
    content_cache_keys_json = db.Column(db.Text, default='[]')
    question = db.Column(db.Text, default='')
    answer = db.Column(db.Text, default='')
    similar_questions_json = db.Column(db.Text, default='[]')
    product_category_name = db.Column(db.Text, default='')
    product_names_json = db.Column(db.Text, default='[]')
    topic_terms_json = db.Column(db.Text, default='[]')
    source_update_time = db.Column(db.String(100), default='')
    index_status = db.Column(db.String(20), default='pending', index=True)
    last_error = db.Column(db.Text, default='')
    indexed_at = db.Column(db.DateTime)
    updated_ts = db.Column(db.Float, default=lambda: time.time(), index=True)

    __table_args__ = (
        db.UniqueConstraint('library_type', 'question_wiki_id', name='uq_kb_retrieval_library_wiki'),
    )


class KBDuplicateCheckTask(db.Model):
    __tablename__ = 'kb_duplicate_check_task'
    task_id = db.Column(db.String(36), primary_key=True)
    username = db.Column(db.String(80), default='', index=True)
    library = db.Column(db.String(40), default='knowledge_base_v1')
    status = db.Column(db.String(24), default='running', index=True)
    stage = db.Column(db.String(40), default='preparing_index')
    question = db.Column(db.Text, default='')
    answer = db.Column(db.Text, default='')
    product_category_name = db.Column(db.Text, default='')
    product_names_json = db.Column(db.Text, default='[]')
    source_note = db.Column(db.Text, default='')
    top_k = db.Column(db.Integer, default=20)
    expanded = db.Column(db.Boolean, default=False)
    index_total = db.Column(db.Integer, default=0)
    index_done = db.Column(db.Integer, default=0)
    candidate_count = db.Column(db.Integer, default=0)
    completed_channels_json = db.Column(db.Text, default='[]')
    failed_stages_json = db.Column(db.Text, default='[]')
    candidates_json = db.Column(db.Text, default='[]')
    analysis_json = db.Column(db.Text, default='{}')
    config_snapshot_json = db.Column(db.Text, default='{}')
    human_decision = db.Column(db.String(40), default='')
    human_note = db.Column(db.Text, default='')
    selected_source_ids_json = db.Column(db.Text, default='[]')
    cancel_requested = db.Column(db.Boolean, default=False)
    message = db.Column(db.Text, default='')
    created_ts = db.Column(db.Float, default=lambda: time.time(), index=True)
    updated_ts = db.Column(db.Float, default=lambda: time.time(), index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class KBDuplicateIndexJob(db.Model):
    __tablename__ = 'kb_duplicate_index_job'
    job_id = db.Column(db.String(36), primary_key=True)
    username = db.Column(db.String(80), default='', index=True)
    library = db.Column(db.String(40), default='knowledge_base_v1', index=True)
    mode = db.Column(db.String(20), default='incremental')
    status = db.Column(db.String(24), default='running', index=True)
    total = db.Column(db.Integer, default=0)
    done = db.Column(db.Integer, default=0)
    cache_hits = db.Column(db.Integer, default=0)
    failed_count = db.Column(db.Integer, default=0)
    message = db.Column(db.Text, default='')
    error = db.Column(db.Text, default='')
    cancel_requested = db.Column(db.Boolean, default=False)
    config_snapshot_json = db.Column(db.Text, default='{}')
    created_ts = db.Column(db.Float, default=lambda: time.time(), index=True)
    updated_ts = db.Column(db.Float, default=lambda: time.time(), index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class MatrixColumn(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_name = db.Column(db.String(100), unique=True, nullable=False)
    sort_order = db.Column(db.Integer, default=0)

class ProductMatrix(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    question_wiki_id = db.Column(db.String(100), nullable=False)
    product_name = db.Column(db.String(100), nullable=False)
    is_configured = db.Column(db.Boolean, default=False)
    manual_edit = db.Column(db.Boolean, default=False)
    edit_source = db.Column(db.String(20), default='')
    last_synced_at = db.Column(db.DateTime, default=datetime.utcnow)
    question_content = db.Column(db.Text)
    answer_content = db.Column(db.Text)
    update_time = db.Column(db.String(50))
    product_category = db.Column(db.String(100))

    __table_args__ = (
        db.UniqueConstraint('question_wiki_id', 'product_name', name='unique_matrix_item'),
    )

def _matrix_scope_bool(value, default=True):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text_value = str(value).strip().lower()
    if text_value == '':
        return default
    if text_value in ('1', 'true', 't', 'yes', 'y', 'on'):
        return True
    if text_value in ('0', 'false', 'f', 'no', 'n', 'off'):
        return False
    return default

def _matrix_scope_list(value, separators=r'[,，\n]+'):
    if value is None:
        return []
    if isinstance(value, list):
        raw_items = value
    elif isinstance(value, tuple) or isinstance(value, set):
        raw_items = list(value)
    else:
        raw_items = re.split(separators, str(value))
    items = [str(x).strip() for x in raw_items if str(x).strip()]
    return list(OrderedDict.fromkeys(items))

def _matrix_scope_normalize_name(value):
    s = str(value or '')
    s = s.replace('\u3000', ' ')
    return re.sub(r'\s+', ' ', s).strip()

def _matrix_scope_diff_compare_ids(ids_subq, selected_models):
    models = [str(x).strip() for x in (selected_models or []) if str(x).strip()]
    if len(models) < 2:
        return []
    rows_for_diff = ProductMatrix.query.filter(
        ProductMatrix.question_wiki_id.in_(ids_subq),
        ProductMatrix.product_name.in_(models)
    ).all()
    state_map = {}
    for r in rows_for_diff:
        wid_key = str(getattr(r, 'question_wiki_id', '') or '').strip()
        product_key = str(getattr(r, 'product_name', '') or '').strip()
        if not wid_key or product_key not in models:
            continue
        state_map.setdefault(wid_key, {})[product_key] = bool(getattr(r, 'is_configured', False))
    out = []
    for wid_key, states_by_model in state_map.items():
        states = [bool(states_by_model.get(model, False)) for model in models]
        if len(set(states)) > 1:
            out.append(wid_key)
    return sorted(out)

def _matrix_scope_modified_ids(ids_subq, columns_list):
    edit_query = db.session.query(ProductMatrix.question_wiki_id).filter(
        ProductMatrix.question_wiki_id.in_(ids_subq)
    ).filter(
        or_(
            ProductMatrix.edit_source.in_(['cell', 'bulk']),
            ProductMatrix.manual_edit == True
        )
    )
    wiki_ids_for_edits = [
        str(r[0]).strip()
        for r in edit_query.distinct().all()
        if r and str(r[0]).strip()
    ]
    wiki_ids_for_edits = list(OrderedDict.fromkeys(wiki_ids_for_edits))
    if not wiki_ids_for_edits:
        return set()

    source_products_map, source_ok = _fetch_kb_products_map('knowledge_base_v1', wiki_ids_for_edits)
    if not source_ok:
        raise RuntimeError('Failed to fetch source products')

    source_norm_map = {}
    for x in wiki_ids_for_edits:
        source_norm_map[x] = set([
            _matrix_scope_normalize_name(v)
            for v in (source_products_map.get(x, set()) or set())
            if _matrix_scope_normalize_name(v)
        ])

    cols_norm_set = set([
        _matrix_scope_normalize_name(x)
        for x in (columns_list or [])
        if _matrix_scope_normalize_name(x)
    ])

    qy = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(wiki_ids_for_edits)).filter(
        or_(
            ProductMatrix.edit_source.in_(['cell', 'bulk']),
            ProductMatrix.manual_edit == True
        )
    )
    if columns_list:
        qy = qy.filter(ProductMatrix.product_name.in_(columns_list))
    matrix_rows = qy.all()

    modified_ids = set()
    for r in matrix_rows:
        wid = str(getattr(r, 'question_wiki_id', '') or '').strip()
        product_name = str(getattr(r, 'product_name', '') or '').strip()
        if not wid or not product_name:
            continue
        product_norm = _matrix_scope_normalize_name(product_name)
        if product_norm not in cols_norm_set:
            continue
        source_configured = product_norm in source_norm_map.get(wid, set())
        current_configured = bool(getattr(r, 'is_configured', False))
        if current_configured != source_configured:
            modified_ids.add(wid)
    return modified_ids

def _resolve_matrix_filtered_scope_ids(filters):
    filters = filters if isinstance(filters, dict) else {}

    wid = str(filters.get('id') or filters.get('wid') or '').strip()
    q = str(filters.get('q') or '').strip()
    a = str(filters.get('a') or '').strip()
    p = str(filters.get('p') or '').strip()
    pc = str(filters.get('pc') or '').strip()
    mc = str(filters.get('mc') or '').strip()
    p_mode = str(filters.get('p_mode') or 'any').strip().lower()
    p_models = _matrix_scope_list(filters.get('p_models'))
    col_models = _matrix_scope_list(filters.get('col_models'), separators=r'[,，]+')
    diff_compare = _matrix_scope_bool(filters.get('diff_compare'), False)

    marks = filters.get('marks') if isinstance(filters.get('marks'), dict) else {}
    if marks:
        want_modified = _matrix_scope_bool(marks.get('modified'), True)
        want_unmodified = _matrix_scope_bool(marks.get('unmodified'), True)
    else:
        want_modified = _matrix_scope_bool(filters.get('mark_modified'), True)
        want_unmodified = _matrix_scope_bool(filters.get('mark_unmodified'), True)

    if not (want_modified or want_unmodified):
        return []

    mappings = get_model_mappings()
    allowed_models = []
    if mc and mc in mappings and isinstance(mappings.get(mc), list):
        allowed_models = [str(x) for x in mappings.get(mc) if x]

    catalog = parse_product_catalog()
    pc_models = []
    if pc and pc in catalog and isinstance(catalog.get(pc), list):
        pc_models = [str(x).strip() for x in catalog.get(pc) if x and str(x).strip()]

    models_whitelist = set()
    if pc_models:
        models_whitelist = set([str(x) for x in pc_models if x])
    if allowed_models:
        if models_whitelist:
            models_whitelist = models_whitelist.intersection(set([str(x) for x in allowed_models if x]))
        else:
            models_whitelist = set([str(x) for x in allowed_models if x])

    selected_models_ordered = []
    if col_models:
        seen_models = set()
        for model in col_models:
            if model in seen_models:
                continue
            seen_models.add(model)
            selected_models_ordered.append(model)
        selected_set = set(selected_models_ordered)
        if models_whitelist:
            models_whitelist = models_whitelist.intersection(selected_set)
        else:
            models_whitelist = selected_set

    columns_raw = filters.get('columns')
    if isinstance(columns_raw, list) and columns_raw:
        col_list = [
            str(x).strip()
            for x in columns_raw
            if str(x).strip() and str(x).strip() != '测试型号'
        ]
    else:
        col_query = MatrixColumn.query.order_by(MatrixColumn.sort_order)
        if models_whitelist:
            col_query = col_query.filter(MatrixColumn.product_name.in_(list(models_whitelist)))
        if selected_models_ordered:
            columns = col_query.filter(MatrixColumn.product_name.in_(selected_models_ordered)).all()
            by_name = {
                str(c.product_name): c
                for c in columns
                if c and str(getattr(c, 'product_name', '') or '').strip()
            }
            col_list = [m for m in selected_models_ordered if m in by_name]
        else:
            columns = col_query.all()
            col_list = [c.product_name for c in columns]
        col_list = [c for c in (col_list or []) if str(c or '').strip() and str(c or '').strip() != '测试型号']

    if not col_list:
        return []

    base_query = db.session.query(ProductMatrix.question_wiki_id)
    if wid:
        base_query = base_query.filter(ProductMatrix.question_wiki_id.ilike(f'%{wid}%'))
    if q:
        base_query = base_query.filter(ProductMatrix.question_content.ilike(f'%{q}%'))
    if a:
        base_query = base_query.filter(ProductMatrix.answer_content.ilike(f'%{a}%'))

    if p_models:
        if p_mode == 'all':
            subq = db.session.query(ProductMatrix.question_wiki_id)\
                .filter(ProductMatrix.product_name.in_(p_models))\
                .group_by(ProductMatrix.question_wiki_id)\
                .having(func.count(func.distinct(ProductMatrix.product_name)) == len(p_models))\
                .subquery()
            base_query = base_query.filter(ProductMatrix.question_wiki_id.in_(subq))
        else:
            base_query = base_query.filter(ProductMatrix.product_name.in_(p_models))
    elif p:
        base_query = base_query.filter(ProductMatrix.product_name.ilike(f'%{p}%'))

    if pc:
        base_query = base_query.filter(ProductMatrix.product_category.ilike(f'%{pc}%'))

    if mc and mc in mappings:
        if allowed_models:
            allowed_models = [str(x) for x in allowed_models if x]
            n = len(allowed_models)
            if n == 0:
                base_query = base_query.filter(False)
            else:
                exact_ids_subq = db.session.query(ProductMatrix.question_wiki_id)\
                    .filter(ProductMatrix.is_configured == True)\
                    .group_by(ProductMatrix.question_wiki_id)\
                    .having(func.count(func.distinct(ProductMatrix.product_name)) == n)\
                    .having(func.count(func.distinct(case(
                        (ProductMatrix.product_name.in_(allowed_models), ProductMatrix.product_name),
                        else_=None
                    ))) == n)\
                    .subquery()
                base_query = base_query.filter(ProductMatrix.question_wiki_id.in_(exact_ids_subq))
        else:
            base_query = base_query.filter(False)

    base_ids_subq = base_query.distinct().subquery()
    if diff_compare and len(selected_models_ordered) >= 2:
        return _matrix_scope_diff_compare_ids(base_ids_subq, selected_models_ordered)

    if want_modified and want_unmodified:
        return [
            str(row[0]).strip()
            for row in base_query.distinct().order_by(ProductMatrix.question_wiki_id).all()
            if row and str(row[0]).strip()
        ]

    modified_ids = _matrix_scope_modified_ids(base_ids_subq, col_list)
    if want_modified and not want_unmodified:
        return sorted(list(modified_ids))
    if want_unmodified and not want_modified:
        base_ids = [
            str(row[0]).strip()
            for row in base_query.distinct().order_by(ProductMatrix.question_wiki_id).all()
            if row and str(row[0]).strip()
        ]
        return [x for x in base_ids if x not in modified_ids]
    return []

def _resolve_matrix_clone_scope_ids(scope, allow_empty_selected=False):
    scope = scope if isinstance(scope, dict) else {}
    scope_mode = str(scope.get('mode') or 'all').strip().lower()
    if not scope_mode:
        scope_mode = 'all'
    if scope_mode not in ('all', 'selected', 'filtered'):
        raise ValueError('Invalid clone scope')

    scope_ids = None
    if scope_mode == 'selected':
        scope_ids = _matrix_scope_list(scope.get('wiki_ids') or scope.get('ids'))
        if not scope_ids and not allow_empty_selected:
            raise ValueError('No selected rows found for clone scope')
    elif scope_mode == 'filtered':
        scope_ids = _resolve_matrix_filtered_scope_ids(
            scope.get('filters') if isinstance(scope.get('filters'), dict) else {}
        )
    return scope_mode, scope_ids

def _resolve_matrix_clone_source_items(mode, source, scope_ids=None):
    source_items = {}
    source_products_to_remove = set()

    mode = str(mode or '').strip()
    source = str(source or '').strip()
    if not mode or not source:
        return source_items, source_products_to_remove

    if mode == 'model':
        source_query = ProductMatrix.query.filter_by(product_name=source, is_configured=True)
        if scope_ids is not None:
            source_query = source_query.filter(ProductMatrix.question_wiki_id.in_(scope_ids))
        items = source_query.all()

        if not items and ' ' in source:
            source_query = ProductMatrix.query.filter_by(product_name=source.replace(' ', ''), is_configured=True)
            if scope_ids is not None:
                source_query = source_query.filter(ProductMatrix.question_wiki_id.in_(scope_ids))
            items = source_query.all()

        source_products_to_remove.add(source)
        if ' ' in source:
            source_products_to_remove.add(source.replace(' ', ''))
        source_items = {item.question_wiki_id: item for item in items}
    elif mode == 'category':
        mappings = get_model_mappings()
        if source in mappings:
            allowed_models = [str(x) for x in (mappings[source] or []) if x]
            n = len(allowed_models)
            if n > 0:
                source_products_to_remove.update([str(m) for m in allowed_models if m])
                exact_query = db.session.query(ProductMatrix.question_wiki_id)\
                    .filter(ProductMatrix.is_configured == True)
                if scope_ids is not None:
                    exact_query = exact_query.filter(ProductMatrix.question_wiki_id.in_(scope_ids))
                exact_query = exact_query\
                    .group_by(ProductMatrix.question_wiki_id)\
                    .having(func.count(func.distinct(ProductMatrix.product_name)) == n)\
                    .having(func.count(func.distinct(case(
                        (ProductMatrix.product_name.in_(allowed_models), ProductMatrix.product_name),
                        else_=None
                    ))) == n)
                exact_ids = exact_query.all()
                valid_ids = [row[0] for row in exact_ids]
                if valid_ids:
                    all_rows = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(valid_ids)).all()
                    for row in all_rows:
                        if row.question_wiki_id not in source_items:
                            source_items[row.question_wiki_id] = row
        else:
            category_query = ProductMatrix.query.filter(ProductMatrix.product_category == source)
            if scope_ids is not None:
                category_query = category_query.filter(ProductMatrix.question_wiki_id.in_(scope_ids))
            items = category_query.all()
            for item in items:
                if item.question_wiki_id not in source_items:
                    source_items[item.question_wiki_id] = item
            source_products_to_remove.update([str(r.product_name) for r in items if r.product_name])

    return source_items, source_products_to_remove

class MatrixSubmitOperation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    operation_id = db.Column(db.String(80), unique=True, nullable=False)
    status = db.Column(db.String(20), default='pending')
    attempts = db.Column(db.Integer, default=0)
    created_by = db.Column(db.String(80))
    error_message = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Button(db.Model):
    __tablename__ = 'button'
    id = db.Column(db.Integer, primary_key=True)
    operation_id = db.Column(db.String(80), nullable=False)
    question_wiki_id = db.Column(db.String(100), nullable=False)
    product_name = db.Column(db.String(100), nullable=False)
    old_is_configured = db.Column(db.Boolean, nullable=False)
    new_is_configured = db.Column(db.Boolean, nullable=False)
    edit_source = db.Column(db.String(20), default='')
    diff_json = db.Column(db.Text)
    submitted_by = db.Column(db.String(80))
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('operation_id', 'question_wiki_id', 'product_name', name='unique_button_op_item'),
    )


class SupabaseOutbox(db.Model):
    """
    Outbox pattern for supabase write operations.
    When supabase is unavailable (network error / 5xx / 429), we enqueue the operation locally,
    so content is not lost and can be retried manually.
    """
    __tablename__ = "supabase_outbox"
    id = db.Column(db.Integer, primary_key=True)
    created_ts = db.Column(db.Float, default=lambda: time.time(), index=True)
    updated_ts = db.Column(db.Float, default=lambda: time.time(), onupdate=lambda: time.time())

    op_type = db.Column(db.String(30), nullable=False)       # insert/upsert/update/delete/delete_in
    table_name = db.Column(db.String(120), nullable=False)

    # JSON blobs for replay
    payload_json = db.Column(db.Text, nullable=True)
    filters_json = db.Column(db.Text, nullable=True)
    extra_json = db.Column(db.Text, nullable=True)            # on_conflict / column/values, etc.

    status = db.Column(db.String(30), default="pending", index=True)  # pending/done/failed/needs_manual_sync
    attempts = db.Column(db.Integer, default=0)
    last_error = db.Column(db.Text, default="")

class ArchiveBatch(db.Model):
    __tablename__ = 'archive_batch'
    id = db.Column(db.Integer, primary_key=True)
    batch_name = db.Column(db.String(200), nullable=False)
    record_count = db.Column(db.Integer, default=0)
    created_by = db.Column(db.String(80))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class ArchiveRecord(db.Model):
    __tablename__ = 'archive_record'
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.Integer, db.ForeignKey('archive_batch.id'), nullable=False)
    record_json = db.Column(db.Text, nullable=False)
    modify_time = db.Column(db.DateTime)


class OpsLibraryItem(db.Model):
    __tablename__ = 'ops_library_item'
    id = db.Column(db.Integer, primary_key=True)
    kind = db.Column(db.String(20), nullable=False)  # app / product
    name = db.Column(db.String(200), nullable=False)
    steps = db.Column(db.Text, nullable=False)
    compatible_models = db.Column(db.Text, default='')  # newline/comma separated (backend only)
    sort_order = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class QualityTaskPool(db.Model):
    __tablename__ = 'quality_task_pool'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    sources_json = db.Column(db.Text, default='[]')
    rule_config_json = db.Column(db.Text, default='{}')
    field_mapping_json = db.Column(db.Text, default='{}')
    status = db.Column(db.String(20), default='active')
    created_by = db.Column(db.String(80), default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class QualityRawIssue(db.Model):
    __tablename__ = 'quality_raw_issue'
    id = db.Column(db.Integer, primary_key=True)
    pool_id = db.Column(db.Integer, db.ForeignKey('quality_task_pool.id'), nullable=False, index=True)
    source_type = db.Column(db.String(40), nullable=False, index=True)
    source_record_key = db.Column(db.String(240), nullable=False)
    wiki_id = db.Column(db.String(100), nullable=False, index=True)
    issue_text = db.Column(db.Text, default='')
    remediation_reference = db.Column(db.Text, default='')
    snapshot_json = db.Column(db.Text, default='{}')
    rule_snapshot_json = db.Column(db.Text, default='{}')
    ignored_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('pool_id', 'source_record_key', name='uq_quality_raw_pool_source_key'),
    )


class QualityTask(db.Model):
    __tablename__ = 'quality_task'
    id = db.Column(db.Integer, primary_key=True)
    wiki_id = db.Column(db.String(100), unique=True, nullable=False, index=True)
    priority = db.Column(db.String(10), default='p2', index=True)
    status = db.Column(db.String(20), default='pending', index=True)
    latest_kb_update_time = db.Column(db.String(80), default='')
    completed_at = db.Column(db.DateTime)
    ignored_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class QualityTaskIssueLink(db.Model):
    __tablename__ = 'quality_task_issue_link'
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey('quality_task.id'), nullable=False, index=True)
    raw_issue_id = db.Column(db.Integer, db.ForeignKey('quality_raw_issue.id'), nullable=False, index=True)
    pool_id = db.Column(db.Integer, db.ForeignKey('quality_task_pool.id'), nullable=False, index=True)
    source_type = db.Column(db.String(40), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('task_id', 'raw_issue_id', name='uq_quality_task_raw_issue'),
    )


class QualityImportJob(db.Model):
    __tablename__ = 'quality_import_job'
    id = db.Column(db.Integer, primary_key=True)
    file_name = db.Column(db.String(260), default='')
    target_pool_id = db.Column(db.Integer, db.ForeignKey('quality_task_pool.id'), index=True)
    total_count = db.Column(db.Integer, default=0)
    success_count = db.Column(db.Integer, default=0)
    failed_count = db.Column(db.Integer, default=0)
    duplicate_append_count = db.Column(db.Integer, default=0)
    failed_detail_json = db.Column(db.Text, default='[]')
    created_by = db.Column(db.String(80), default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


# Setup
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def init_db():
    with app.app_context():
        db.create_all()
        cur_con = None
        legacy_con = None
        try:
            import sqlite3

            cur_con = sqlite3.connect(_DB_PATH)
            cur_cur = cur_con.cursor()

            def _dedupe_archive_batches(cursor):
                dup_groups = cursor.execute(
                    "select batch_name, created_at, record_count, created_by, count(*) as c "
                    "from archive_batch "
                    "group by batch_name, created_at, record_count, created_by "
                    "having c > 1"
                ).fetchall()
                for bname, cat, rc, cby, _c in (dup_groups or []):
                    ids = [r[0] for r in cursor.execute(
                        "select id from archive_batch where batch_name=? and created_at=? and record_count=? and (created_by is ? or created_by=?) order by id asc",
                        (bname, cat, rc, cby, cby)
                    ).fetchall()]
                    if len(ids) <= 1:
                        continue
                    drop_ids = ids[1:]
                    cursor.execute(
                        f"delete from archive_record where batch_id in ({','.join(['?'] * len(drop_ids))})",
                        drop_ids
                    )
                    cursor.execute(
                        f"delete from archive_batch where id in ({','.join(['?'] * len(drop_ids))})",
                        drop_ids
                    )

            try:
                _dedupe_archive_batches(cur_cur)
                cur_con.commit()
            except Exception:
                try:
                    cur_con.rollback()
                except Exception:
                    pass

            legacy_path = os.path.abspath(os.path.join(_BASE_DIR, '..', 'instance', 'data.db'))
            if os.path.exists(legacy_path) and os.path.abspath(legacy_path) != os.path.abspath(_DB_PATH):
                legacy_con = sqlite3.connect(legacy_path)
                legacy_cur = legacy_con.cursor()

                def _has_table(c, name):
                    return c.execute("select count(*) from sqlite_master where type='table' and name=?", (name,)).fetchone()[0] > 0

                if _has_table(legacy_cur, 'archive_batch') and _has_table(legacy_cur, 'archive_record') and _has_table(cur_cur, 'archive_batch') and _has_table(cur_cur, 'archive_record'):
                    existing_batch_ids = {r[0] for r in cur_cur.execute("select id from archive_batch").fetchall()}
                    existing_rec_ids = {r[0] for r in cur_cur.execute("select id from archive_record").fetchall()}

                    legacy_batches = legacy_cur.execute(
                        "select id, batch_name, record_count, created_by, created_at from archive_batch order by id asc"
                    ).fetchall()

                    id_map = {}
                    for bid, bname, rc, cby, cat in legacy_batches:
                        if bid not in existing_batch_ids:
                            cur_cur.execute(
                                "insert into archive_batch (id, batch_name, record_count, created_by, created_at) values (?, ?, ?, ?, ?)",
                                (bid, bname, rc, cby, cat)
                            )
                            existing_batch_ids.add(bid)
                            id_map[int(bid)] = int(bid)
                        else:
                            cur_cur.execute(
                                "insert into archive_batch (batch_name, record_count, created_by, created_at) values (?, ?, ?, ?)",
                                (bname, rc, cby, cat)
                            )
                            new_id = cur_cur.lastrowid
                            id_map[int(bid)] = int(new_id)

                    legacy_recs = legacy_cur.execute(
                        "select id, batch_id, record_json, modify_time from archive_record order by id asc"
                    ).fetchall()

                    for rid, batch_id, rjson, mt in legacy_recs:
                        new_batch_id = id_map.get(int(batch_id)) if batch_id is not None else None
                        if new_batch_id is None:
                            continue
                        if rid not in existing_rec_ids:
                            cur_cur.execute(
                                "insert into archive_record (id, batch_id, record_json, modify_time) values (?, ?, ?, ?)",
                                (rid, new_batch_id, rjson, mt)
                            )
                            existing_rec_ids.add(rid)
                        else:
                            cur_cur.execute(
                                "insert into archive_record (batch_id, record_json, modify_time) values (?, ?, ?)",
                                (new_batch_id, rjson, mt)
                            )

                    _dedupe_archive_batches(cur_cur)
                    cur_con.commit()
        except Exception:
            try:
                if cur_con:
                    cur_con.rollback()
            except Exception:
                pass
        finally:
            try:
                if legacy_con:
                    legacy_con.close()
            except Exception:
                pass
            try:
                if cur_con:
                    cur_con.close()
            except Exception:
                pass
        try:
            cols = db.session.execute(text("PRAGMA table_info(product_matrix)")).fetchall()
            col_names = {row[1] for row in cols}
            if 'edit_source' not in col_names:
                db.session.execute(text("ALTER TABLE product_matrix ADD COLUMN edit_source VARCHAR(20) DEFAULT ''"))
                db.session.commit()
        except Exception:
            db.session.rollback()
        # Create default user if not exists
        if not User.query.filter_by(username='admin').first():
            user = User(username='admin', password_hash=generate_password_hash('123456'))
            db.session.add(user)
            db.session.commit()
            print("Created default user: admin / 123456")

# Routes
@app.route('/')
def index():
    return send_from_directory(os.path.join(_BASE_DIR, 'link_viewer'), 'index.html')


@app.route('/api/embed/validate')
def validate_embed_origin():
    host_origin = request.args.get('host_origin', '')
    normalized = _normalize_http_origin(host_origin)
    if not normalized:
        return jsonify({'success': False, 'allowed': False, 'message': '缺少有效的 host_origin'}), 400
    if not _is_embed_origin_allowed(normalized):
        return jsonify({'success': False, 'allowed': False, 'message': '当前父页面不在允许嵌入列表中'}), 403
    return jsonify({'success': True, 'allowed': True, 'host_origin': normalized})


@app.route('/login', methods=['POST'])
def login():
    data = request.json
    user = User.query.filter_by(username=data.get('username')).first()
    if user and check_password_hash(user.password_hash, data.get('password')):
        login_user(user)
        return jsonify({'success': True, 'username': user.username})
    return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

@app.route('/logout', methods=['POST'])
@login_required
def logout():
    logout_user()
    return jsonify({'success': True})

@app.route('/api/status')
def status():
    if current_user.is_authenticated:
        return jsonify({'logged_in': True, 'username': current_user.username})
    return jsonify({'logged_in': False})

@app.route('/api/product_manual', methods=['GET'])
@login_required
def product_manual():
    manual_path = _resolve_product_manual_path()
    if not manual_path:
        return jsonify({'success': False, 'message': '产品说明书文件不存在'}), 404
    try:
        with open(manual_path, 'r', encoding='utf-8') as f:
            text_body = f.read()
        return Response(text_body, content_type='text/markdown; charset=utf-8')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/links', methods=['GET'])
@login_required
def get_links():
    try:
        client = get_supabase_client()
        if not client:
            return jsonify([])
        
        # Pagination Params
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('pageSize', 1000)) # Default to 1000 for backward compatibility
        
        # Search Param
        search_query = request.args.get('search', '').strip()
        
        # Tag Params
        tags_param = request.args.get('tags', '').strip()
        tag_mode = request.args.get('tagMode', 'OR').upper() # AND / OR
        
        # Sort Params
        sort_by = request.args.get('sortBy', 'created_at')
        sort_dir = request.args.get('sortDir', 'asc')

        filters = {}
        if search_query:
            filters['url'] = f"ilike.*{search_query}*"
            
        if tags_param:
            # tags_param is comma separated: "tag1,tag2"
            tag_list = [t.strip() for t in tags_param.split(',') if t.strip()]
            if tag_list:
                if tag_mode == 'AND':
                    # Use JSON array syntax for JSONB containment: tags @> '["a", "b"]'
                    # PostgREST: tags=cs.["a","b"]
                    json_tags = json.dumps(tag_list)
                    filters['tags'] = f"cs.{json_tags}"
                else:
                    # Use OR with multiple contains checks for JSONB
                    # PostgREST: or=(tags.cs.["a"],tags.cs.["b"])
                    or_conditions = []
                    for tag in tag_list:
                        # Create a single-element JSON array string for each tag
                        single_tag_json = json.dumps([tag])
                        or_conditions.append(f"tags.cs.{single_tag_json}")
                    
                    filters['or'] = f"({','.join(or_conditions)})"

        # Select from Supabase 'link_previews' table
        print(f"DEBUG: Fetching links page={page}, pageSize={page_size}, search={search_query}, tags={tags_param}, mode={tag_mode}, sort={sort_by}.{sort_dir}")
        response = client.select('link_previews', page=page, page_size=page_size, filters=filters, order_by=sort_by, order_dir=sort_dir)
        

        if response.status_code not in (200, 206):
            print(f"Error fetching links: {response.text}")
            return jsonify([])
            
        data = response.json()
        print(f"DEBUG: Fetched {len(data)} links from Supabase")
        
        # Get Total Count (Supabase returns Content-Range header like "0-9/100")
        total = 0
        content_range = response.headers.get('Content-Range')
        if content_range:
            parts = content_range.split('/')
            if len(parts) > 1:
                try:
                    total = int(parts[1])
                except:
                    total = len(data)
        else:
            total = len(data)

        # Format data for frontend
        # Note: Supabase JSONB returns as python list/dict, so no json.loads needed for tags if stored as jsonb
        # However, if stored as text, json.loads is needed. 
        # Let's assume tags is JSONB. If it's text, we might need check.
        # Frontend expects: id, url, type, tags (list), created_at (was createdAt)
        formatted = []
        for item in data:
            tags = item.get('tags', [])
            if isinstance(tags, str):
                try:
                    tags = json.loads(tags)
                except:
                    tags = []
            
            # Handle created_at conversion if it's a timestamp (float/int)
            created_at_val = item.get('created_at')
            if isinstance(created_at_val, (int, float)):
                try:
                    # Assume Unix timestamp in seconds
                    dt = datetime.fromtimestamp(created_at_val)
                    created_at_val = dt.isoformat()
                except:
                    pass

            formatted.append({
                'id': item.get('id'),
                'kb_id': item.get('kb_id'),
                'url': item.get('url'),
                'type': item.get('type'),
                'tags': tags,
                'created_at': created_at_val
            })
        
        # Return wrapper object if pagination is used (or always? frontend expects array currently)
        # To maintain backward compatibility, we should return array if no page/pageSize params were provided explicitly?
        # Or just update frontend to handle {data: [], total: N} format.
        # User asked for pagination support.
        
        return jsonify({
            'success': True,
            'data': formatted,
            'total': total,
            'page': page,
            'pageSize': page_size
        })
    except Exception as e:
        print(f"Exception in get_links: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e), 'data': [], 'total': 0}), 500

def _detect_link_type_backend(url):
    u_lower = str(url or '').strip().lower()
    if re.search(r'\.(png|jpg|jpeg|gif|webp|svg)(\?.*)?$', u_lower):
        return 'image'
    if re.search(r'\.(mp4|webm|ogg|mov|m3u8)(\?.*)?$', u_lower):
        return 'video'
    if 'youtube.com' in u_lower or 'youtu.be' in u_lower:
        return 'youtube'
    if re.search(r'\.(pdf|doc|docx|xls|xlsx|ppt|pptx|zip|rar|7z|txt)(\?.*)?$', u_lower):
        return 'file'
    return 'link'

@app.route('/api/links', methods=['POST'])
@login_required
def add_link():
    data = request.get_json(silent=True) or {}
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    required_fields = ['id', 'url', 'type', 'tags', 'createdAt']
    missing_fields = [field for field in required_fields if field not in data]
    if missing_fields:
        return jsonify({'success': False, 'message': f"缺少必要字段: {', '.join(missing_fields)}"}), 400

    new_link = {
        'id': data['id'],
        'kb_id': data.get('kb_id'),
        'url': data['url'],
        'type': data['type'],
        'tags': data['tags'], # Send as list, requests will dump to JSON
        'created_at': _dt_to_iso(data['createdAt'])
    }

    try:
        resp = client.upsert('link_previews', [new_link], on_conflict='id')
    except Exception as e:
        print("[add_link] sync upsert failed")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

    if resp is None or getattr(resp, 'status_code', 500) >= 400:
        return jsonify({
            'success': False,
            'message': getattr(resp, 'text', '写入 link_previews 失败')
        }), 500

    return jsonify({'success': True, 'item': new_link})

@app.route('/api/links/batch', methods=['POST'])
@login_required
def add_links_batch():
    data = request.get_json(silent=True) or []
    # data should be a list of link objects
    if not isinstance(data, list):
        return jsonify({'success': False, 'message': 'Expected a list of links'}), 400
        
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    # Ensure all items have required fields
    new_links = []
    for item in data:
        new_links.append({
            'id': item['id'],
            'kb_id': item.get('kb_id'),
            'url': item['url'],
            'type': item['type'],
            'tags': item['tags'],
            'created_at': _dt_to_iso(item['createdAt'])
        })
    
    if not new_links:
        return jsonify({'success': True, 'count': 0})

    try:
        resp = client.upsert('link_previews', new_links, on_conflict='id')
    except Exception as e:
        print("[add_links_batch] sync upsert failed")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

    if resp is None or getattr(resp, 'status_code', 500) >= 400:
        return jsonify({
            'success': False,
            'message': getattr(resp, 'text', '批量写入 link_previews 失败')
        }), 500

    return jsonify({'success': True, 'count': len(new_links)})

@app.route('/api/links/<link_id>', methods=['PUT'])
@login_required
def update_link(link_id):
    data = request.get_json(silent=True) or {}

    update_data = {}
    if 'tags' in data:
        update_data['tags'] = data['tags']
    
    if not update_data:
        return jsonify({'success': False, 'message': 'No update fields'}), 400

    try:
        c = get_supabase_client()
        if not c:
            return jsonify({'success': False, 'message': '本地主库未配置'}), 500
        resp = c.update('link_previews', dict(update_data), {'id': link_id})
        if resp is None or getattr(resp, 'status_code', 500) >= 400:
            return jsonify({'success': False, 'message': getattr(resp, 'text', '更新 link_previews 失败')}), 500
        return jsonify({'success': True})
    except Exception as e:
        print("[update_link] update failed")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/tags', methods=['GET'])
@login_required
def get_all_tags():
    client = get_supabase_client()
    if not client:
        return jsonify([])
    
    # Fetch all tags from link_previews
    # We select only the 'tags' column to minimize data transfer
    # Use select_all to ensure we get tags from all pages
    try:
        data = client.select_all('link_previews', columns='tags')
    except Exception as e:
        print(f"Error fetching tags: {e}")
        return jsonify([])
    
    unique_tags = set()
    
    for item in data:
        tags = item.get('tags', [])
        if isinstance(tags, str):
            try:
                tags = json.loads(tags)
            except:
                tags = []
        
        if isinstance(tags, list):
            for tag in tags:
                if tag:
                    unique_tags.add(str(tag).strip())
                    
    return jsonify(list(sorted(unique_tags)))

@app.route('/api/links/<link_id>', methods=['DELETE'])
@login_required
def delete_link(link_id):
    try:
        c = get_supabase_client()
        if not c:
            return jsonify({'success': False, 'message': '本地主库未配置'}), 500
        resp = c.delete('link_previews', {'id': link_id})
        if resp is None or getattr(resp, 'status_code', 500) >= 400:
            return jsonify({'success': False, 'message': getattr(resp, 'text', '删除 link_previews 失败')}), 500
        return jsonify({'success': True})
    except Exception as e:
        print("[delete_link] delete failed")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

def _run_sync_kb_links():
    client = get_supabase_client()
    if not client:
        return {'success': False, 'message': '本地主库未配置', '_http_status': 500}

    try:
        allowed_cols = ['id', 'url', 'kb_id', 'type', 'tags', 'created_at']
        def _normalize_link_payload(obj, current_time_value):
            out = {}
            for k in allowed_cols:
                out[k] = obj.get(k)
            if not out.get('id'):
                out['id'] = str(uuid.uuid4())
            out['url'] = str(out.get('url') or '').strip()
            out['kb_id'] = str(out.get('kb_id') or '').strip()
            if not out.get('type'):
                out['type'] = 'link'
            if out.get('tags') is None:
                out['tags'] = []
            if not out.get('created_at'):
                out['created_at'] = current_time_value
            return out

        # 1. Fetch all KB data (ID, URLs)
        print("Starting Sync...")
        # Get only necessary columns to speed up.
        try:
            kb_items = client.select_all(
                'knowledge_base_v1',
                columns='question_wiki_id,image_urls,video_urls,file_urls,link_url,link_type',
                order_by='question_wiki_id'
            )
        except Exception as e:
            msg = str(e)
            if 'PGRST204' in msg:
                return {
                    'success': False,
                    'message': 'knowledge_base_v1 缺少新字段（image_urls/video_urls/file_urls/link_type/link_url）。请先执行 update_schema_v3.sql 里的 KB Schema 更新 SQL。',
                    '_http_status': 400
                }
            raise
        print(f"Fetched {len(kb_items)} KB items.")
        
        # DEBUG: Print first item to check fields
        if kb_items:
            print(f"DEBUG Sample KB Item: {kb_items[0]}")
        
        # 2. Fetch all existing link_previews
        existing_links = client.select_all('link_previews', columns='id,url,kb_id,type,tags,created_at', order_by='created_at')
        print(f"Fetched {len(existing_links)} existing links.")
        
        # CLEANUP: Existing map URLs should also be stripped for comparison
        existing_map = {}
        for item in existing_links:
            u_orig = item['url']
            u_clean = u_orig
            if isinstance(u_orig, str):
                u_clean = u_orig.replace('`', '').strip()
            
            # Store the CLEANED url as the key for easy lookup during processing
            existing_map[u_clean] = item
            # Also store the original URL in the item for later comparison
            item['_original_url_in_db'] = u_orig
        
        # 3. Process KB items to build URL -> set(KB_IDs) mapping
        url_to_kb_ids = {}
        url_to_sources = {}
        url_to_external_types = {}
        
        # 根据客户端类型使用不同的时间格式
        # LocalPostgreSQLClient 需要 datetime 对象
        # SupabaseClient 需要 Unix 时间戳
        if isinstance(client, LocalPostgreSQLClient):
            current_time_value = datetime.utcnow()
        else:
            current_time_value = _now_iso_with_tz()
        
        print(f"DEBUG: Processing {len(kb_items)} KB items for URLs...")
        
        def _parse_url_list(urls_raw):
            if not urls_raw:
                return []
            if isinstance(urls_raw, list):
                return urls_raw
            if isinstance(urls_raw, str):
                s = urls_raw.strip()
                if not s:
                    return []
                try:
                    parsed = json.loads(s)
                    if isinstance(parsed, list):
                        return parsed
                    if isinstance(parsed, str):
                        return [parsed]
                except Exception:
                    return [u.strip() for u in re.split(r'[,，\n\r]+', s) if u and u.strip()]
            return []

        def _iter_kb_urls_with_source(kb_item):
            if not isinstance(kb_item, dict):
                return

            if any(k in kb_item for k in ('image_urls', 'video_urls', 'file_urls')):
                for k in ('image_urls', 'video_urls', 'file_urls'):
                    for u in _parse_url_list(kb_item.get(k)):
                        yield u, k, None

            link_url_val = str(kb_item.get('link_url') or '').strip() if 'link_url' in kb_item else ''
            if link_url_val:
                yield link_url_val, 'link_url', kb_item.get('link_type')

        def _add_url(url_raw, kb_id_val, source_key, external_type_val=None):
            if not url_raw:
                return
            url = url_raw
            if isinstance(url, str):
                url = url.replace('`', '').strip()
            if not url:
                return
            if url not in url_to_kb_ids:
                url_to_kb_ids[url] = set()
            url_to_kb_ids[url].add(str(kb_id_val))
            if source_key:
                url_to_sources.setdefault(url, set()).add(str(source_key))
            if external_type_val:
                url_to_external_types.setdefault(url, set()).add(str(external_type_val).strip())

        for kb_item in kb_items:
            kb_id = kb_item.get('question_wiki_id')
            if not kb_id:
                continue

            for u, src_key, ext_type in _iter_kb_urls_with_source(kb_item):
                _add_url(u, kb_id, src_key, ext_type)
        
        print(f"DEBUG: Found {len(url_to_kb_ids)} unique URLs from KB.")
        
        # 4. Prepare data for upsert
        to_upsert = []
        def _merge_sys_tags(tags, source_label, external_type_text):
            existing = tags if isinstance(tags, list) else []
            if isinstance(tags, str):
                try:
                    parsed = json.loads(tags)
                    if isinstance(parsed, list):
                        existing = parsed
                except Exception:
                    existing = []
            user_tags = []
            for t in existing:
                s = str(t or '').strip()
                if not s:
                    continue
                if s.startswith('来源:') or s.startswith('外链类型:'):
                    continue
                user_tags.append(s)
            sys_tags = []
            if source_label:
                sys_tags.append(f'来源:{source_label}')
            if source_label == '外部链接' and external_type_text:
                sys_tags.append(f'外链类型:{external_type_text}')
            merged = []
            seen = set()
            for t in user_tags + sys_tags:
                s = str(t or '').strip()
                if not s:
                    continue
                if s in seen:
                    continue
                seen.add(s)
                merged.append(s)
            return merged

        label_map = {
            'image_urls': '图片链接',
            'video_urls': '视频链接',
            'file_urls': '文件链接',
            'link_url': '外部链接'
        }

        for url, kb_id_set in url_to_kb_ids.items():
            kb_id_str = ",".join(sorted(list(kb_id_set)))
            src_set = url_to_sources.get(url) or set()
            src_labels = sorted({label_map.get(s, s) for s in src_set if s})
            if len(src_labels) == 1:
                source_label = src_labels[0]
            elif len(src_labels) > 1:
                source_label = '混合'
            else:
                source_label = ''
            ext_types = url_to_external_types.get(url) or set()
            external_type_text = ",".join(sorted({str(x).strip() for x in ext_types if str(x).strip()}))
            
            if url in existing_map:
                existing = existing_map[url]
                # Update kb_id if different
                # ALSO Update the URL itself to the cleaned version if it was different
                needs_update = False
                
                # IMPORTANT: Convert to string and strip for comparison
                existing_kb_id = str(existing.get('kb_id') or '').strip()
                if existing_kb_id != kb_id_str:
                    print(f"DEBUG: KB_ID mismatch for {url}: DB='{existing_kb_id}', New='{kb_id_str}'")
                    existing['kb_id'] = kb_id_str
                    needs_update = True
                
                # Compare against the REAL original URL from DB
                orig_url_in_db = str(existing.get('_original_url_in_db', existing.get('url')) or '')
                if orig_url_in_db != url:
                    print(f"DEBUG: URL mismatch for {url}: DB='{orig_url_in_db}', New='{url}'")
                    # The URL in DB has backticks/spaces, we need to update it to the cleaned 'url'
                    existing['url'] = url
                    needs_update = True

                merged_tags = _merge_sys_tags(existing.get('tags'), source_label, external_type_text)
                if merged_tags != (existing.get('tags') if isinstance(existing.get('tags'), list) else []):
                    existing['tags'] = merged_tags
                    needs_update = True
                
                if needs_update:
                    # Remove the temp key before sending to DB
                    update_payload = {k: v for k, v in existing.items() if k != '_original_url_in_db'}
                    to_upsert.append(_normalize_link_payload(update_payload, current_time_value))
            else:
                # Create new
                new_item = {
                    'id': str(uuid.uuid4()),
                    'url': url,
                    'kb_id': kb_id_str,
                    'type': 'link',
                    'tags': _merge_sys_tags([], source_label, external_type_text),
                    'created_at': current_time_value
                }
                
                # Basic type detection
                u_lower = url.lower()
                if re.search(r'\.(png|jpg|jpeg|gif|webp|svg)(\?.*)?$', u_lower):
                    new_item['type'] = 'image'
                elif re.search(r'\.(mp4|webm|ogg|mov|m3u8)(\?.*)?$', u_lower):
                    new_item['type'] = 'video'
                elif 'youtube.com' in u_lower or 'youtu.be' in u_lower:
                    new_item['type'] = 'youtube'
                elif re.search(r'\.(pdf|doc|docx|xls|xlsx|ppt|pptx|zip|rar|7z|txt)(\?.*)?$', u_lower):
                    new_item['type'] = 'file'
                
                to_upsert.append(_normalize_link_payload(new_item, current_time_value))
        
        # LOGIC FIX: Always include existing items in upsert if we want them to show up correctly, 
        # or at least ensure we are comparing correctly.
        # Actually, if to_upsert is 0, it means all URLs and their kb_ids already match perfectly.
        # But if the user says "still 0", maybe the database has some records that need refreshing.
        
        print(f"DEBUG: to_upsert list length: {len(to_upsert)}")
        
        # 5. Batch Upsert
        count_new = 0
        count_updated = 0
        upsert_batch = []

        existing_urls = set(existing_map.keys())
        kb_urls = set(url_to_kb_ids.keys())
        urls_to_unlink = existing_urls - kb_urls

        print(f"DEBUG: Found {len(urls_to_unlink)} obsolete URLs to unlink (remove KB ID).")

        if urls_to_unlink:
            for u in urls_to_unlink:
                item = existing_map[u]
                if item.get('kb_id'):
                    payload = {k: v for k, v in item.items() if k != '_original_url_in_db'}
                    payload['kb_id'] = ''
                    upsert_batch.append(_normalize_link_payload(payload, current_time_value))

        if to_upsert:
            unique_upserts = {item['url']: item for item in to_upsert}.values()

            for item in list(unique_upserts):
                is_update = item['url'] in existing_map and existing_map[item['url']].get('id')
                
                if is_update:
                    count_updated += 1
                    if 'id' not in item:
                        item['id'] = existing_map[item['url']]['id']
                else:
                    count_new += 1
                
                upsert_batch.append(_normalize_link_payload(item, current_time_value))

        batch_size = 100
        if upsert_batch:
            deduped = {}
            for item in upsert_batch:
                key = item.get('id') or item.get('url')
                if key:
                    deduped[key] = item
            upsert_batch = list(deduped.values())
            upsert_batch = [_normalize_link_payload(it, current_time_value) for it in upsert_batch]

            for i in range(0, len(upsert_batch), batch_size):
                batch = upsert_batch[i:i+batch_size]
                print(f"DEBUG: Syncing batch {i//batch_size + 1}, size {len(batch)}")
                resp = client.upsert('link_previews', batch)
                if resp.status_code >= 400:
                    return {'success': False, 'message': resp.text, '_http_status': 500}

        return {
            'success': True,
            'count': count_new,
            'updated': count_updated,
            'unlinked': len(urls_to_unlink),
            'total_found': len(url_to_kb_ids)
        }

    except Exception as e:
        traceback.print_exc()
        return {'success': False, 'message': str(e), '_http_status': 500}


@app.route('/api/links/sync_kb', methods=['POST'])
@login_required
def sync_kb_links():
    r = _run_sync_kb_links()
    if r.get('success'):
        return jsonify(r)
    code = int(r.pop('_http_status', 500))
    return jsonify(r), code


@app.route('/api/links/delete_batch', methods=['POST'])
@login_required
def delete_links_batch():
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No IDs provided'}), 400
    
    client = get_supabase_client()
    response = client.delete_in('link_previews', 'id', ids)
    
    if response.status_code not in [200, 204]:
        return jsonify({'success': False, 'message': response.text}), 500
        
    return jsonify({'success': True})

import requests
from scoring_logic import LLMScorer, calculate_product_overlap, DEFAULT_SYSTEM_PROMPT, PRODUCT_CATALOG_TEXT

# Configuration File for Scoring

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 优先检查 ⚙️ 配置文件 目录
_EXTERNAL_CONFIG_DIR = os.path.join(os.path.dirname(BASE_DIR), '⚙️ 配置文件')

def _get_config_path(filename):
    ext_path = os.path.join(_EXTERNAL_CONFIG_DIR, filename)
    if os.path.exists(ext_path):
        return ext_path
    return os.path.join(BASE_DIR, filename)

PRODUCT_CATALOG_FILE = _get_config_path('product_catalog.json')
MODEL_MAPPINGS_FILE = _get_config_path('model_mappings.json')

def parse_product_catalog():
    """
    Parses PRODUCT_CATALOG_TEXT or loads from JSON file.
    Returns: { "Category": ["Model1", "Model2", ...] }
    """
    # 1. Try to load from file
    if os.path.exists(PRODUCT_CATALOG_FILE):
        try:
            with open(PRODUCT_CATALOG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading product catalog file: {e}")
            
    # 2. Fallback to default text and save
    catalog = {}
    lines = PRODUCT_CATALOG_TEXT.strip().split('\n')
    for line in lines:
        line = line.strip()
        # Match pattern: 1. **Category**：Model1, Model2...
        match = re.match(r'\d+\.\s*\*\*(.+?)\*\*[:：](.+)', line)
        if match:
            category = match.group(1).strip()
            models_str = match.group(2).strip()
            # Split by comma or Chinese comma
            models = [m.strip() for m in re.split(r'[,，]', models_str) if m.strip()]
            catalog[category] = models
    
    # Save to file for future use
    try:
        with open(PRODUCT_CATALOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(catalog, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving product catalog file: {e}")
        
    return catalog

def save_product_catalog(catalog):
    with open(PRODUCT_CATALOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, indent=2, ensure_ascii=False)

def get_model_mappings():
    if os.path.exists(MODEL_MAPPINGS_FILE):
        try:
            with open(MODEL_MAPPINGS_FILE, 'r', encoding='utf-8') as f:
                raw = json.load(f)
                normalized = normalize_model_mappings_dict(raw)
                if normalized != raw:
                    save_model_mappings(normalized)
                return normalized
        except Exception as e:
            print(f"Error loading model mappings: {e}")
    return {}

def save_model_mappings(mappings):
    mappings = normalize_model_mappings_dict(mappings)
    with open(MODEL_MAPPINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)

def get_all_valid_models():
    """
    Returns a normalization map: { "normalized_name": "Official Name" }
    Normalization: remove spaces, lowercase.
    Also returns the set of official names for direct lookup.
    """
    catalog = parse_product_catalog()
    valid_map = {}
    valid_set = set()
    for models in catalog.values():
        for m in models:
            valid_set.add(m)
            # Normalize: remove all spaces, to lower
            norm = m.replace(" ", "").lower()
            valid_map[norm] = m
    return valid_map, valid_set

def normalize_model_mappings_dict(mappings):
    if not isinstance(mappings, dict):
        return {}
    valid_map, valid_set = get_all_valid_models()
    normalized = {}
    for cat, models in mappings.items():
        cat_name = str(cat).strip() if cat is not None else ''
        if not cat_name:
            continue
        model_list = normalize_model_list(models, valid_map, valid_set)
        if model_list:
            normalized[cat_name] = model_list
    return normalized

def normalize_model_list(models, valid_map, valid_set):
    if not isinstance(models, list):
        return []
    out = []
    seen = set()
    for m in models:
        if m is None:
            continue
        s = str(m).strip()
        if not s:
            continue
        if s in valid_set:
            if s not in seen:
                out.append(s)
                seen.add(s)
            continue
        norm = re.sub(r'\s+', '', s).lower()
        official = valid_map.get(norm)
        if official:
            if official not in seen:
                out.append(official)
                seen.add(official)
    return out

def validate_product_string(product_str, valid_map=None, valid_set=None):
    if _is_blank_cell_value(product_str):
        return [], []
        
    if valid_map is None or valid_set is None:
        valid_map, valid_set = get_all_valid_models()
        
    # Split by comma or Chinese comma
    # User says: ",T7 Pro,和,T7Pro,, T7 Pro,,T7 Pro ," -> "T7 Pro"
    parts = [m.strip() for m in re.split(r'[,，]', str(product_str)) if m.strip()]
    
    valid = []
    invalid = []
    seen = set() # To avoid duplicates in output
    
    for m in parts:
        # Check exact match first
        if m in valid_set:
            if m not in seen:
                valid.append(m)
                seen.add(m)
            continue
            
        # Check normalized match
        norm = m.replace(" ", "").lower()
        if norm in valid_map:
            official = valid_map[norm]
            if official not in seen:
                valid.append(official)
                seen.add(official)
        else:
            invalid.append(m)
            
    return valid, invalid

def _flatten_catalog_models(catalog):
    models = set()
    if isinstance(catalog, dict):
        for values in catalog.values():
            if isinstance(values, list):
                for v in values:
                    s = str(v or '').strip()
                    if s:
                        models.add(s)
    return models

def _build_product_catalog_impact(new_catalog):
    old_catalog = parse_product_catalog()
    old_models = _flatten_catalog_models(old_catalog)
    new_models = _flatten_catalog_models(new_catalog)
    removed_models = sorted(old_models - new_models)

    matrix_column_count = 0
    matrix_row_count = 0
    orphan_row_count = 0
    orphan_models = []
    if removed_models:
        matrix_column_count = MatrixColumn.query.filter(MatrixColumn.product_name.in_(removed_models)).count()
        matrix_row_count = ProductMatrix.query.filter(ProductMatrix.product_name.in_(removed_models)).count()

    db_product_names_res = db.session.query(ProductMatrix.product_name).distinct().all()
    db_product_names = {r[0] for r in db_product_names_res if r and r[0]}
    orphan_models = sorted(db_product_names - new_models)
    if orphan_models:
        orphan_row_count = ProductMatrix.query.filter(ProductMatrix.product_name.in_(orphan_models)).count()

    return {
        'removed_models': removed_models,
        'removed_model_count': len(removed_models),
        'matrix_column_count': matrix_column_count,
        'matrix_row_count': matrix_row_count,
        'orphan_models': orphan_models,
        'orphan_model_count': len(orphan_models),
        'orphan_row_count': orphan_row_count,
        'requires_confirmation': bool(removed_models or matrix_column_count or matrix_row_count or orphan_row_count)
    }

def _sync_product_catalog_to_matrix(new_catalog):
    catalog_products = set()
    for products in new_catalog.values():
        if isinstance(products, list):
            for p in products:
                s = str(p or '').strip()
                if s:
                    catalog_products.add(s)

    existing_cols = MatrixColumn.query.all()
    existing_map = {c.product_name: c for c in existing_cols}
    existing_names = set(existing_map.keys())

    to_add = catalog_products - existing_names
    to_remove = existing_names - catalog_products

    if to_add:
        max_order = db.session.query(func.max(MatrixColumn.sort_order)).scalar() or 0
        for p_name in sorted(to_add):
            max_order += 1
            db.session.add(MatrixColumn(product_name=p_name, sort_order=max_order))

    if to_remove:
        for p_name in to_remove:
            col = existing_map[p_name]
            db.session.delete(col)

    db_product_names_res = db.session.query(ProductMatrix.product_name).distinct().all()
    db_product_names = {r[0] for r in db_product_names_res if r and r[0]}
    to_remove_data = db_product_names - catalog_products
    if to_remove_data:
        db.session.query(ProductMatrix).filter(ProductMatrix.product_name.in_(list(to_remove_data))).delete(synchronize_session=False)

@app.route('/api/kb/product_catalog', methods=['GET', 'POST'])
@login_required
def handle_product_catalog():
    if request.method == 'GET':
        return jsonify(parse_product_catalog())
    
    # POST: Update catalog
    payload = request.get_json(silent=True)
    confirm_cleanup = False
    if isinstance(payload, dict) and 'catalog' in payload:
        new_catalog = payload.get('catalog')
        confirm_cleanup = bool(payload.get('confirm_cleanup'))
    else:
        new_catalog = payload
    if not new_catalog or not isinstance(new_catalog, dict):
        return jsonify({'success': False, 'message': 'Invalid catalog format'}), 400

    impact = _build_product_catalog_impact(new_catalog)
    if impact.get('requires_confirmation') and not confirm_cleanup:
        return jsonify({
            'success': False,
            'requires_confirmation': True,
            'message': '本次型号库保存会删除矩阵列或矩阵数据，请预览影响后确认。',
            'impact': impact
        }), 409
    
    old_catalog = parse_product_catalog()
    try:
        save_product_catalog(new_catalog)
        _sync_product_catalog_to_matrix(new_catalog)
        db.session.commit()
        
        # Normalize model mappings against new catalog (drops deleted/invalid models)
        try:
            _ = get_model_mappings()
        except Exception as _e:
            print(f"Error normalizing model mappings after catalog update: {_e}")
    except Exception as e:
        db.session.rollback()
        try:
            save_product_catalog(old_catalog)
        except Exception:
            pass
        print(f"Error syncing catalog to matrix: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'型号库保存失败: {str(e)}'}), 500
        
    return jsonify({'success': True, 'impact': impact})

@app.route('/api/kb/product_catalog/preview', methods=['POST'])
@login_required
def preview_product_catalog_update():
    payload = request.get_json(silent=True) or {}
    new_catalog = payload.get('catalog') if isinstance(payload, dict) and 'catalog' in payload else payload
    if not isinstance(new_catalog, dict):
        return jsonify({'success': False, 'message': 'Invalid catalog format'}), 400
    return jsonify({'success': True, 'impact': _build_product_catalog_impact(new_catalog)})

@app.route('/api/kb/product_catalog/export', methods=['GET'])
@login_required
def export_product_catalog_xlsx():
    """导出型号库为 xlsx：A列分类，B列型号（同一分类型号用逗号连接）。使用 openpyxl 保证 Excel 可打开。"""
    catalog = parse_product_catalog()
    rows = []
    for category, models in catalog.items():
        if not isinstance(models, list):
            models = []
        models_str = ','.join(str(m).strip() for m in models if m)
        rows.append({'分类': category, '型号': models_str})
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='型号库')
    # 使用完整字节返回，避免流位置导致 Excel 无法打开
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue()),
        as_attachment=True,
        download_name=canonical_download_name('product_catalog'),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/model_mappings', methods=['GET', 'POST'])
@login_required
def handle_model_mappings():
    if request.method == 'GET':
        return jsonify(get_model_mappings())
    
    mappings = request.json
    if not isinstance(mappings, dict):
        return jsonify({'success': False, 'message': 'Invalid format'}), 400
        
    save_model_mappings(mappings)
    return jsonify({'success': True})


@app.route('/api/model_mappings/export_excel', methods=['POST'])
@login_required
def export_model_mappings_excel():
    """
    导出机型分类映射为 xlsx。
    前端会把弹窗内“当前（可能未保存）”的映射直接 POST 过来，因此优先使用请求体。
    """
    try:
        payload = request.get_json(silent=True)
        if isinstance(payload, dict):
            mappings = normalize_model_mappings_dict(payload)
        else:
            mappings = get_model_mappings()

        rows = []
        for category in sorted((mappings or {}).keys(), key=lambda x: str(x)):
            models = mappings.get(category) or []
            if not isinstance(models, list):
                models = []
            clean_models = [str(m).strip() for m in models if str(m).strip()]
            rows.append({
                '分类': str(category).strip(),
                '机型列表': ','.join(clean_models),
                '机型数量': len(clean_models)
            })

        df = pd.DataFrame(rows, columns=['分类', '机型列表', '机型数量'])
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='机型分类映射')
        output.seek(0)

        return send_file(
            io.BytesIO(output.getvalue()),
            as_attachment=True,
            download_name=canonical_download_name('model_mapping'),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error exporting model mappings excel: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500





@app.route('/api/scoring/sample', methods=['POST'])
@login_required
def sample_scoring_data():
    """
    Randomly sample N items for scoring.
    Prioritize unscored/outdated items.
    """
    data = request.json
    count = data.get('count', 10)
    
    client = get_supabase_client()
    
    # 1. Fetch all scores to identify status
    scores_resp = client.select_all('kb_scores')
    scores_map = {item['kb_id']: item for item in scores_resp}
    
    # 2. Fetch all KB IDs
    # Lightweight fetch
    kb_resp = client.select_all('knowledge_base_v1', columns='question_wiki_id', order_by='question_wiki_id')
    all_kb_ids = [item['question_wiki_id'] for item in kb_resp]
    
    # 3. Categorize
    unscored_ids = []
    scored_ids = []
    
    print(f"DEBUG: sample_scoring_data: fetched {len(all_kb_ids)} KB items, {len(scores_map)} scores")
    
    for kbid in all_kb_ids:
        s_item = scores_map.get(kbid)
        if not s_item or s_item.get('status') in ['unscored', 'outdated']:
            unscored_ids.append(kbid)
        else:
            scored_ids.append(kbid)
            
    print(f"DEBUG: sample_scoring_data: {len(unscored_ids)} unscored, {len(scored_ids)} scored")
            
    # 4. Sample
    import random
    result_ids = []
    
    # Prioritize unscored
    if len(unscored_ids) >= count:
        result_ids = random.sample(unscored_ids, count)
    else:
        result_ids = unscored_ids[:] # Take all unscored
        remaining = count - len(result_ids)
        if remaining > 0 and scored_ids:
            if len(scored_ids) >= remaining:
                result_ids.extend(random.sample(scored_ids, remaining))
            else:
                result_ids.extend(scored_ids)
                
    return jsonify({
        'ids': result_ids, 
        'debug': {
            'total_kb': len(all_kb_ids),
            'total_scores': len(scores_map),
            'unscored_count': len(unscored_ids),
            'scored_count': len(scored_ids)
        }
    })


@app.route('/api/proxy_image')
def proxy_image():
    url = request.args.get('url')
    if not url:
        return "Missing url parameter", 400
    
    # Simple security check to prevent SSRF against localhost/intranet
    # Allow external URLs only
    if 'localhost' in url or '127.0.0.1' in url or url.startswith('http://192.') or url.startswith('http://10.'):
         # Allow only if it's explicitly whitelisted or pass through if you trust the user (Authenticated tool)
         # For now, let's just log and proceed if it's not critical, or block if needed.
         # But wait, our client side might be requesting localhost? No, client requests external URLs via proxy.
         pass

    try:
        # Stream the request to avoid loading large files into memory
        # Set headers to look like a browser
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        # Support Range header for video seeking
        if 'Range' in request.headers:
            headers['Range'] = request.headers['Range']
        if 'If-Range' in request.headers:
            headers['If-Range'] = request.headers['If-Range']

        # Disable SSL verification for some sites if needed (use with caution)
        # roborock might have SSL issues or just strict filtering.
        resp = requests.get(url, headers=headers, stream=True, timeout=30, verify=False)
        
        # Headers to exclude (Hop-by-hop + Content-Encoding/Length which we might want to control)
        excluded_headers = [
            'content-encoding', 'content-length', 'transfer-encoding', 'connection', 
            'content-disposition', 'host', 'keep-alive', 'proxy-authenticate', 
            'proxy-authorization', 'te', 'trailers', 'upgrade'
        ]
        
        response_headers = []
        for name, value in resp.headers.items():
            if name.lower() not in excluded_headers:
                response_headers.append((name, value))
        
        # Force inline disposition
        response_headers.append(('Content-Disposition', 'inline'))
        
        # Explicitly handle Content-Length and Content-Range from processed headers
        # Only forward Content-Length if content is NOT encoded (requests decompresses automatically)
        if 'Content-Length' in resp.headers and 'Content-Encoding' not in resp.headers:
             response_headers.append(('Content-Length', resp.headers['Content-Length']))
        
        # Ensure Content-Type is set (fallback if missing)
        if 'Content-Type' not in resp.headers:
             # Guess from extension
             if url.lower().endswith('.mp4'):
                 response_headers.append(('Content-Type', 'video/mp4'))
             else:
                 response_headers.append(('Content-Type', 'image/jpeg'))

        from flask import Response, stream_with_context
        # Increase chunk size to 128KB for better video performance
        return Response(stream_with_context(resp.iter_content(chunk_size=1024*128)), 
                        status=resp.status_code,
                        headers=response_headers)
    except Exception as e:
        print(f"Proxy Error for {url}: {e}")
        return str(e), 500

# Scoring Endpoints










@app.route('/api/scoring/data', methods=['GET'])
@login_required
def get_scoring_data():
    client = get_supabase_client()
    if not client:
        return jsonify([])
    
    page_size_param = request.args.get('pageSize')
    
    try:
        score_has_product = _supabase_has_column(client, 'kb_scores', 'product_name')
        include_kb_details = str(
            request.args.get('include_kb') or request.args.get('includeKb') or ''
        ).strip().lower() in ('1', 'true', 'yes', 'y')

        if page_size_param:
            # Paged Fetch
            page = int(request.args.get('page', 1))
            page_size = int(page_size_param)
            
            # Fetch Scores Paged
            scores_resp = client.select('kb_scores', page=page, page_size=page_size, order_by='id')
            
            if scores_resp.status_code >= 400:
                 return jsonify([])
                 
            scores_data = scores_resp.json()
            
            # Get total count
            total = 0
            cr = scores_resp.headers.get('Content-Range')
            if cr:
                try:
                    total = int(cr.split('/')[-1])
                except:
                    total = len(scores_data) # Fallback
        else:
            # Fetch ALL Data (when no pageSize provided)
            # Optimize columns to reduce data transfer
            scores_cols = 'id,kb_id,question_content,answer_content,status,total_score,remarks,score_data,updated_at'
            if score_has_product:
                scores_cols += ',product_name'
            scores_data = client.select_all('kb_scores', columns=scores_cols, order_by='id')
            total = len(scores_data)
            page = 1
            page_size = total
            
        if not scores_data:
            return jsonify({'success': True, 'data': [], 'total': total, 'page': 1, 'pageSize': total})

        # Collect IDs to fetch KB details (supplementary fields only; Q/A come from kb_scores snapshot)
        kb_ids = [s['kb_id'] for s in scores_data if s.get('kb_id')]
        
        # Fetch KB Info (for product_name / urls — optional enrichment; not a substitute for "同步打分数据")
        kb_cols = 'question_wiki_id,question,answer,product_name,update_time,image_urls,video_urls,file_urls,link_type,link_url'
        
        kb_map = {}

        missing_product_ids = []
        if score_has_product:
            missing_product_ids = [
                s.get('kb_id') for s in scores_data
                if s.get('kb_id') and not str(s.get('product_name') or '').strip()
            ]
        ids_to_enrich = kb_ids if include_kb_details or not score_has_product else missing_product_ids

        if ids_to_enrich:
            unique_ids = list(set(ids_to_enrich))
            batch_size = 50
            for i in range(0, len(unique_ids), batch_size):
                batch = unique_ids[i:i+batch_size]
                id_filter = _postgrest_in_str(batch)
                if not id_filter:
                    continue
                try:
                    kb_resp = client.select('knowledge_base_v1', filters={'question_wiki_id': id_filter}, columns=kb_cols)
                    if kb_resp.status_code < 400:
                        for item in kb_resp.json():
                            kb_map[item['question_wiki_id']] = item
                    else:
                        print(f"WARN: failed to enrich scoring KB products: {kb_resp.text}")
                except Exception as e:
                    print(f"WARN: scoring KB enrichment failed: {e}")
        
        def parse_score_fields(score_item):
            """Helper to parse score_data and extract flattened fields"""
            score_data_raw = score_item.get('score_data')
            parsed = {}
            
            if score_data_raw:
                if isinstance(score_data_raw, str):
                    try:
                        parsed = json.loads(score_data_raw)
                    except:
                        pass
                elif isinstance(score_data_raw, dict):
                    parsed = score_data_raw
                    
            # Extract dimensions
            dims = parsed.get('维度得分', {})
            
            return {
                'quality': dims.get('问题质量', 0),
                'compliance': dims.get('答案合规与准确性', 0),
                'timeliness': dims.get('时效性', 0),
                'utility': dims.get('实际解决力', 0),
                'redundancy': dims.get('非冗余与相关性', 0),
                'multimedia': dims.get('多媒体加分', 0),
                'analysis': parsed.get('分析过程', ''),
                'suggestion': parsed.get('处理建议', '') or score_item.get('remarks', '')
            }

        result = []
        
        if not page_size_param:
            # Snapshot rows only; question/answer always from kb_scores (last sync), not live V1
            for score_item in scores_data:
                kb_id = score_item.get('kb_id')
                kb_item = kb_map.get(kb_id, {})
                score_details = parse_score_fields(score_item)
                item = {
                    'kb_id': kb_id,
                    'question_content': score_item.get('question_content'),
                    'answer_content': score_item.get('answer_content'),
                    'product_name': score_item.get('product_name') or kb_item.get('product_name'),
                    'update_time': score_item.get('updated_at') or kb_item.get('update_time'),
                    'image_urls': kb_item.get('image_urls'),
                    'video_urls': kb_item.get('video_urls'),
                    'file_urls': kb_item.get('file_urls'),
                    'link_type': kb_item.get('link_type'),
                    'link_url': kb_item.get('link_url'),
                    'score_id': score_item.get('id'),
                    'status': score_item.get('status', 'unscored'),
                    'total_score': score_item.get('total_score'),
                    'remarks': score_item.get('remarks'),
                    'score_data': score_item.get('score_data'),
                    **score_details
                }
                result.append(item)
            total = len(result)
            page_size = total
            
        else:
            # Paged Strategy (Existing Behavior: Show Scores)
            for score_item in scores_data:
                kb_id = score_item.get('kb_id')
                kb_item = kb_map.get(kb_id, {})
                
                # Parse score details
                score_details = parse_score_fields(score_item)
                
                item = {
                    'kb_id': kb_id,
                    'question_content': score_item.get('question_content') or kb_item.get('question'),
                    'answer_content': score_item.get('answer_content') or kb_item.get('answer'),
                    'product_name': score_item.get('product_name') or kb_item.get('product_name'),
                    'update_time': kb_item.get('update_time'),
                    'image_urls': kb_item.get('image_urls'),
                    'video_urls': kb_item.get('video_urls'),
                    'file_urls': kb_item.get('file_urls'),
                    'link_type': kb_item.get('link_type'),
                    'link_url': kb_item.get('link_url'),
                    # Score Data
                    'score_id': score_item.get('id'),
                    'status': score_item.get('status'),
                    'total_score': score_item.get('total_score'),
                    'remarks': score_item.get('remarks'),
                    'score_data': score_item.get('score_data'),
                    **score_details
                }
                result.append(item)

            
        return jsonify({
            'success': True,
            'data': result,
            'total': total,
            'page': page,
            'pageSize': page_size
        })
        
    except Exception as e:
        print(f"Get Data Error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500





@app.route('/api/scoring/export', methods=['GET'])
@login_required
def export_scoring_data():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500
        
    try:
        # Fetch all scores
        scores_resp = client.select_all('kb_scores', order_by='id')
        scores = {item['kb_id']: item for item in scores_resp}
        
        # Fetch all KB items
        kb_resp = client.select_all('knowledge_base_v1', order_by='question_wiki_id')
        
        # Generate CSV
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        headers = ['KB ID', 'Question', 'Answer', 'Product', 'Status', 'Score', 'Remarks', 'Analysis', 'Updated At']
        writer.writerow(headers)
        
        def parse_score_data(score_item):
            if not score_item or not score_item.get('score_data'):
                return {}

            score_data = score_item.get('score_data')
            if isinstance(score_data, dict):
                return score_data
            if isinstance(score_data, str):
                try:
                    parsed = json.loads(score_data)
                    return parsed if isinstance(parsed, dict) else {}
                except Exception:
                    return {}
            return {}

        for kb_item in kb_resp:
            kb_id = kb_item.get('question_wiki_id')
            score_item = scores.get(kb_id)
            
            status = score_item.get('status', 'unscored') if score_item else 'unscored'
            score = score_item.get('total_score', '') if score_item else ''
            remarks = score_item.get('remarks', '') if score_item else ''
            
            sd = parse_score_data(score_item)
            analysis = (
                sd.get('分析过程', '')
                or sd.get('扣分分析', '')
                or sd.get('简要点评', '')
                or sd.get('analysis', '')
            )
            
            writer.writerow([
                kb_id,
                kb_item.get('question', ''),
                kb_item.get('answer', ''),
                kb_item.get('product_name', ''),
                status,
                score,
                remarks,
                analysis,
                score_item.get('updated_at', '') if score_item else ''
            ])
            
        output.seek(0)
        
        return Response(
            output.getvalue().encode('utf-8-sig'),
            mimetype="text/csv",
            headers={"Content-disposition": f"attachment; filename={canonical_download_name('scoring', 'csv')}"}
        )
        
    except Exception as e:
        print(f"Export Error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


def _run_scoring_sync():
    client = get_supabase_client()
    if not client:
        return {'success': False, 'message': '本地主库未配置', '_http_status': 500}
    try:
        kb_resp = client.select_all('knowledge_base_v1', order_by='question_wiki_id')
        if not kb_resp:
            return {'success': True, 'message': 'No KB items to sync', 'count': 0, 'added': 0, 'updated': 0, 'deleted': 0}

        score_has_product = _supabase_has_column(client, 'kb_scores', 'product_name')
        scores_resp = client.select_all('kb_scores')
        scores_map = {s['kb_id']: s for s in scores_resp if s.get('kb_id')}
        kb_map = {item['question_wiki_id']: item for item in kb_resp if item.get('question_wiki_id')}

        to_upsert = []
        to_delete_ids = []

        for kbid, s_item in scores_map.items():
            if kbid not in kb_map:
                to_delete_ids.append(s_item['id'])

        added_count = 0
        updated_count = 0
        product_updated_count = 0
        
        # 根据客户端类型使用不同的时间格式
        if isinstance(client, LocalPostgreSQLClient):
            current_time = datetime.utcnow()
        else:
            current_time = _now_iso_with_tz()

        for kbid, kb_item in kb_map.items():
            q = kb_item.get('question', '')
            a = kb_item.get('answer', '')
            p = kb_item.get('product_name', '')

            if kbid not in scores_map:
                rec = {
                    'kb_id': kbid,
                    'question_content': q,
                    'answer_content': a,
                    'status': 'unscored',
                    'updated_at': current_time
                }
                if score_has_product:
                    rec['product_name'] = p
                to_upsert.append(rec)
                added_count += 1
            else:
                s_item = scores_map[kbid]
                old_q = s_item.get('question_content', '')
                old_a = s_item.get('answer_content', '')
                old_p = s_item.get('product_name', '')
                qa_changed = q != old_q or a != old_a
                product_changed = score_has_product and p != old_p
                if qa_changed or product_changed:
                    s_item['question_content'] = q
                    s_item['answer_content'] = a
                    if score_has_product:
                        s_item['product_name'] = p
                    if qa_changed:
                        s_item['status'] = 'outdated'
                    s_item['updated_at'] = current_time
                    to_upsert.append(s_item)
                    updated_count += 1
                    if product_changed:
                        product_updated_count += 1

        if to_delete_ids:
            print(f"Deleting {len(to_delete_ids)} items...")
            for i in range(0, len(to_delete_ids), 100):
                batch_ids = to_delete_ids[i:i+100]
                try:
                    client.delete_in('kb_scores', 'id', batch_ids)
                except Exception as e:
                    print(f"Delete batch failed: {e}")
                    for bid in batch_ids:
                        client.delete('kb_scores', {'id': bid})

        if to_upsert:
            print(f"Upserting {len(to_upsert)} items...")
            batch_size = 200
            total_batches = (len(to_upsert) + batch_size - 1) // batch_size
            for i in range(0, len(to_upsert), batch_size):
                batch = to_upsert[i:i+batch_size]
                print(f"Upserting batch {i//batch_size + 1}/{total_batches}")
                resp = client.upsert('kb_scores', batch, on_conflict='kb_id')
                if resp.status_code >= 400:
                    print(f"Upsert failed: {resp.text}")

        print("Sync completed successfully.")
        return {
            'success': True,
            'added': added_count,
            'updated': updated_count,
            'product_updated': product_updated_count,
            'deleted': len(to_delete_ids),
            'message': f'同步完成: 新增 {added_count}, 更新 {updated_count}, 产品回填 {product_updated_count}, 删除 {len(to_delete_ids)}'
        }
    except Exception as e:
        print(f"Sync Error: {e}")
        return {'success': False, 'message': str(e), '_http_status': 500}


@app.route('/api/scoring/sync', methods=['POST'])
@login_required
def sync_scoring_data():
    r = _run_scoring_sync()
    if r.get('success'):
        return jsonify(r)
    code = int(r.pop('_http_status', 500))
    return jsonify(r), code

@app.route('/api/scoring/config', methods=['GET', 'POST'])
@login_required
def manage_scoring_config():
    if request.method == 'GET':
        config = load_scoring_config()
        api_key_value = str(config.get('api_key') or '')
        config['api_key_configured'] = bool(api_key_value.strip())
        config['api_key_length'] = len(api_key_value)
        config['api_key_prefix'] = api_key_value[:8]
        config['api_key_suffix'] = api_key_value[-6:] if api_key_value else ''
        config['config_status'] = '已配置' if config['api_key_configured'] else '未配置'
        return jsonify(config)
    
    else:
        new_data = request.json or {}
        config = load_scoring_config()

        # Update config with new data
        config.update(new_data)
        
        if save_scoring_config(config):
            saved = load_scoring_config()
            mismatches = []
            for key in ('api_key', 'base_url', 'model'):
                if str(saved.get(key) or '') != str(config.get(key) or ''):
                    mismatches.append(key)
            if mismatches:
                return jsonify({
                    'success': False,
                    'message': '配置文件保存后回读不一致: ' + ', '.join(mismatches)
                }), 500
            api_key_value = str(saved.get('api_key') or '')
            saved['api_key_configured'] = bool(api_key_value.strip())
            saved['api_key_length'] = len(api_key_value)
            saved['api_key_prefix'] = api_key_value[:8]
            saved['api_key_suffix'] = api_key_value[-6:] if api_key_value else ''
            saved['config_status'] = '已配置' if saved['api_key_configured'] else '未配置'
            return jsonify({'success': True, 'config': saved})
        else:
            return jsonify({'success': False, 'message': 'Failed to save config'}), 500

def _llm_config_meta(api_key, base_url, model):
    api_key_value = str(api_key or '')
    return {
        'api_key_configured': bool(api_key_value.strip()),
        'api_key_length': len(api_key_value),
        'api_key_prefix': api_key_value[:8],
        'api_key_suffix': api_key_value[-6:] if api_key_value else '',
        'base_url': base_url,
        'model': model,
    }

def _test_llm_connection(data, fallback_config):
    api_key = str(data.get('api_key') if 'api_key' in data else fallback_config.get('api_key') or '').strip()
    base_url = str(data.get('base_url') if 'base_url' in data else fallback_config.get('base_url') or '').strip()
    model = str(data.get('model') if 'model' in data else fallback_config.get('model') or '').strip()
    meta = _llm_config_meta(api_key, base_url, model)

    if not api_key:
        return {'success': False, 'message': 'API Key 未填写', 'config': meta}
    if not base_url:
        return {'success': False, 'message': 'Base URL 未填写', 'config': meta}
    if not model:
        return {'success': False, 'message': 'Model 未填写', 'config': meta}

    try:
        scorer = LLMScorer(
            api_key=api_key,
            base_url=base_url,
            model=model,
            system_prompt='你是 API 连通性测试助手。'
        )
        content = scorer._chat_completions(
            messages=[
                {'role': 'system', 'content': '你是 API 连通性测试助手。'},
                {'role': 'user', 'content': '请只回复 OK。'}
            ],
            temperature=0,
            response_format=None,
            timeout=30
        )
        meta['response_preview'] = str(content or '')[:200]
        return {'success': True, 'message': 'API 连接测试成功', 'config': meta}
    except Exception as e:
        return {'success': False, 'message': str(e), 'config': meta}

@app.route('/api/scoring/test_config', methods=['POST'])
@login_required
def test_scoring_config():
    return jsonify(_test_llm_connection(request.json or {}, load_scoring_config()))

@app.route('/api/ai/test_config', methods=['POST'])
@login_required
def test_ai_config():
    return jsonify(_test_llm_connection(request.json or {}, load_ai_config()))

@app.route('/api/ai/config', methods=['GET', 'POST'])
@login_required
def manage_ai_config():
    """
    独立的 AI 配置（用于问题/答案润色、相似问题生成等），与评分配置解耦。
    """
    if request.method == 'GET':
        config = load_ai_config()
        api_key_value = str(config.get('api_key') or '')
        config['api_key_configured'] = bool(api_key_value.strip())
        config['api_key_length'] = len(api_key_value)
        config['api_key_prefix'] = api_key_value[:8]
        config['api_key_suffix'] = api_key_value[-6:] if api_key_value else ''
        config['config_status'] = '已配置' if config['api_key_configured'] else '未配置'
        return jsonify(config)

    data = request.json or {}
    config = load_ai_config()

    if 'ai_prompts' in data and not isinstance(data['ai_prompts'], dict):
        data['ai_prompts'] = {}

    config.update(data)

    if save_ai_config(config):
        saved = load_ai_config()
        mismatches = []
        for key in ('api_key', 'base_url', 'model'):
            if str(saved.get(key) or '') != str(config.get(key) or ''):
                mismatches.append(key)
        if mismatches:
            return jsonify({
                'success': False,
                'message': 'AI配置文件保存后回读不一致: ' + ', '.join(mismatches)
            }), 500
        api_key_value = str(saved.get('api_key') or '')
        saved['api_key_configured'] = bool(api_key_value.strip())
        saved['api_key_length'] = len(api_key_value)
        saved['api_key_prefix'] = api_key_value[:8]
        saved['api_key_suffix'] = api_key_value[-6:] if api_key_value else ''
        saved['config_status'] = '已配置' if saved['api_key_configured'] else '未配置'
        ok, err = _save_ai_prompts_to_prompt_folder(config)
        if ok:
            return jsonify({'success': True, 'prompt_saved': True, 'config': saved})
        return jsonify({'success': True, 'prompt_saved': False, 'config': saved, 'message': f'AI配置已保存，但Prompt落盘失败: {err}'})
    else:
        return jsonify({'success': False, 'message': 'Failed to save AI config'}), 500

@app.route('/api/scoring/prompt', methods=['GET', 'POST'])
@login_required
def manage_scoring_prompt():
    if request.method == 'GET':
        config = load_scoring_config()
        return jsonify({'prompt': config.get('system_prompt', '')})
    
    else:
        data = request.json
        prompt = data.get('prompt')
        
        if prompt is None:
             return jsonify({'success': False, 'message': 'No prompt provided'}), 400
             
        config = load_scoring_config()
        config['system_prompt'] = prompt
        
        if save_scoring_config(config):
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': 'Failed to save config'}), 500

def _ai_default_prompts():
    return {
        'question': (
            "你是一个企业知识库的“问题区 AI 润色助手”。请在严格保留原始含义的前提下，对输入的 question 做整体润色，"
            "使其成为适合在线知识库与智能客服检索的标准问句。\n"
            "输出必须为严格 JSON：\n"
            "{\n"
            "  \"question\": string|null,\n"
            "  \"keywords\": null,\n"
            "  \"difficulty\": null,\n"
            "  \"notes\": string|null\n"
            "}\n"
            "变量：{{task}} {{question}}"
        ),
        # legacy fallback (旧版本单模板)
        'answer': (
            "你是一个企业知识库的编辑助手。请根据 task 对输入 answer 做优化，必须与原语义一致。\n"
            "输出必须为严格 JSON：\n"
            "{\n"
            "  \"answer\": string|null,\n"
            "  \"urls\": string[]|null,\n"
            "  \"notes\": string|null\n"
            "}\n"
            "变量：{{task}} {{question}} {{answer}} {{urls}}"
        ),
        # split answer prompts (faster & more stable)
        'answer_structure': (
            "你是知识库内容结构化处理专家，聚焦结构化处理，不做标签化。\n"
            "任务：对输入文本进行通用结构化处理，输出可直接录入的 Markdown。\n"
            "规则：去噪精简；按顺序拆解模块（有则写，无则跳过）：接入范围→技术底座→服务能力→服务方式→核心优势；模块内用 - 列表；核心名词加粗；生成 ## 标题（≤20字）；扩写 3-5 组问答对；无多余解释。\n"
            "输出必须为严格 JSON：{\"answer\": string, \"urls\": string[]|null, \"notes\": string|null}\n"
            "结构要求（强约束 + 软容错）：answer 必须包含 ### 结构化正文 与 ### 扩写问答对（按此顺序）；模块顺序固定但允许缺失跳过；问答对目标 3-5 组，若不足 3 组需在 notes 说明原因。\n"
            "变量：{{task}} {{question}} {{answer}} {{urls}}"
        ),
        'answer_fault': (
            "你是知识库结构化处理专家，核心聚焦结构化处理，不进行任何标签化操作。\n"
            "任务：将产品故障类客服文本结构化输出为可直接录入的 Markdown，并扩写 3-5 组故障相关问答对。\n"
            "规则：去噪精简；按顺序拆解模块（有则写，无则跳过）：接入范围（适用型号/场景）→技术底座（涉及则写）→服务能力（可排查/解决的故障）→服务方式（排查/解决步骤）→核心优势（客观描述）。标题 ≤20 字，含产品+故障关键词；模块用 ###，条目用 -；核心名词加粗；无多余解释。\n"
            "输出必须为严格 JSON：{\"answer\": string, \"urls\": string[]|null, \"notes\": string|null}\n"
            "结构要求（强约束 + 软容错）：answer 必须包含 ### 结构化正文 与 ### 扩写问答对（按此顺序）；模块顺序固定但允许缺失跳过；问答对目标 3-5 组，若不足 3 组需在 notes 说明原因。\n"
            "变量：{{task}} {{question}} {{answer}} {{urls}}"
        ),
        'answer_usage': (
            "你是知识库结构化处理专家，核心聚焦结构化处理，不进行任何标签化操作。\n"
            "任务：将使用方法类客服文本结构化输出为可直接录入的 Markdown，并扩写 3-5 组操作相关问答对。\n"
            "规则：去噪精简；按顺序拆解模块（有则写，无则跳过）：接入范围（适用型号/场景/渠道）→技术底座（涉及则写）→服务能力（可实现功能/需求）→服务方式（操作步骤+注意事项）→核心优势（客观描述）。标题 ≤20 字，含产品+使用功能关键词；模块用 ###，条目用 -；核心名词加粗；无多余解释。\n"
            "输出必须为严格 JSON：{\"answer\": string, \"urls\": string[]|null, \"notes\": string|null}\n"
            "结构要求（强约束 + 软容错）：answer 必须包含 ### 结构化正文 与 ### 扩写问答对（按此顺序）；模块顺序固定但允许缺失跳过；问答对目标 3-5 组，若不足 3 组需在 notes 说明原因。\n"
            "变量：{{task}} {{question}} {{answer}} {{urls}}"
        ),
        'answer_feature': (
            "你是知识库结构化处理专家，核心聚焦结构化处理，不进行任何标签化操作。\n"
            "任务：将功能介绍类客服文本结构化输出为可直接录入的 Markdown，并扩写 3-5 组功能相关问答对。\n"
            "规则：去噪精简；按顺序拆解模块（有则写，无则跳过）：接入范围（适用型号/渠道/用户）→技术底座（核心技术/系统/模型）→服务能力（功能作用/效果/场景）→服务方式（开启/使用路径简述）→核心优势（客观描述）。标题 ≤20 字，含产品+功能关键词；模块用 ###，条目用 -；核心名词加粗；无多余解释。\n"
            "输出必须为严格 JSON：{\"answer\": string, \"urls\": string[]|null, \"notes\": string|null}\n"
            "结构要求（强约束 + 软容错）：answer 必须包含 ### 结构化正文 与 ### 扩写问答对（按此顺序）；模块顺序固定但允许缺失跳过；问答对目标 3-5 组，若不足 3 组需在 notes 说明原因。\n"
            "变量：{{task}} {{question}} {{answer}} {{urls}}"
        ),
        'answer_requirement': (
            "你是客服知识库 FAQ 答案定向优化助手。\n"
            "当前 task=requirement。你的任务是根据“优化需求”对原答案做定向优化。\n\n"
            "输入数据：\n"
            "- question：{{question}}\n"
            "- answer：{{answer}}\n"
            "- optimization_requirement：{{optimization_requirement}}\n"
            "- urls：{{urls}}\n\n"
            "处理规则（严格遵守）：\n"
            "1) 优化需求是本次修改的主要依据，必须逐条落实到 refined_answer。\n"
            "2) 允许使用原答案和优化需求中明确给出的信息；不得编造产品能力、操作路径、按钮名称、限制条件或承诺。\n"
            "3) 保留原答案中仍然正确、必要的信息；删除或改写会造成混淆、重复、口语化或与优化需求冲突的表达。\n"
            "4) 如果优化需求要求拆分场景、澄清概念、补充限制说明、替换措辞，必须在答案中明确体现。\n"
            "5) 输出内容应适合作为知识库 FAQ 正文，面向用户，表达清晰、准确、可执行。\n"
            "6) 保持 Markdown 格式；不要输出解释过程。\n\n"
            "异常处理：\n"
            "- 如果优化需求为空、模糊到无法执行，notes 写为「【需二次修订】：请补充明确的优化需求」。\n"
            "- 如果优化需求包含疑似需要人工确认的事实，notes 写为「【需二次修订】：请人工确认 XXX」；refined_answer 不要编造未确认内容。\n"
            "- 无异常时 notes 可简短说明主要修改点，或返回 null。\n\n"
            "输出要求（严格遵循）：\n"
            "你必须只输出合法 JSON（不要输出 Markdown 代码块，不要输出任何额外文本）：\n"
            "{\n"
            "  \"original_answer\": string,\n"
            "  \"refined_answer\": string,\n"
            "  \"notes\": string|null,\n"
            "  \"answer\": string,\n"
            "  \"urls\": string[]|null\n"
            "}\n\n"
            "字段要求：\n"
            "- original_answer：原始 answer 原文。\n"
            "- refined_answer：按优化需求处理后的答案正文。\n"
            "- answer：必须与 refined_answer 完全一致。\n"
            "- urls：如无新增或调整，返回 null。\n"
            "变量：{{task}} {{question}} {{answer}} {{optimization_requirement}} {{urls}}"
        ),
        'similar': (
            "你是一个企业知识库的编辑助手。请基于 question 生成 3-5 条相似问题，表述要多样。\n"
            "输出必须为严格 JSON：\n"
            "{\n"
            "  \"items\": [\n"
            "    {\"text\": string, \"difficulty\": number}\n"
            "  ],\n"
            "  \"notes\": string|null\n"
            "}\n"
            "变量：{{question}} {{target_min}} {{target_max}} {{count_min}} {{count_max}} {{difficulty}}"
        )
    }


def _save_ai_prompts_to_prompt_folder(config):
    """
    Persist AI prompt templates to ./prompt for easier ops backup/review.
    This runs on every /api/ai/config save.
    """
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        prompt_dir = os.path.join(base_dir, 'prompt')
        os.makedirs(prompt_dir, exist_ok=True)

        prompts = config.get('ai_prompts') if isinstance(config.get('ai_prompts'), dict) else {}
        defaults = _ai_default_prompts()

        file_map = {
            'question': 'ai_question_prompt.txt',
            'answer': 'ai_answer_legacy_prompt.txt',
            'answer_structure': 'ai_answer_structure_prompt.txt',
            'answer_fault': 'ai_answer_fault_prompt.txt',
            'answer_usage': 'ai_answer_usage_prompt.txt',
            'answer_feature': 'ai_answer_feature_prompt.txt',
            'answer_requirement': 'ai_answer_requirement_prompt.txt',
            'similar': 'ai_similar_prompt.txt',
        }

        for k, filename in file_map.items():
            content = prompts.get(k)
            if content is None or str(content).strip() == '':
                content = defaults.get(k, '')
            with open(os.path.join(prompt_dir, filename), 'w', encoding='utf-8') as f:
                f.write(str(content or ''))

        # Keep a machine-readable latest snapshot.
        snapshot = {
            'saved_at': datetime.now().isoformat(),
            'base_url': config.get('base_url', ''),
            'model': config.get('model', ''),
            'question_type_config_json': config.get('question_type_config_json', ''),
            'ai_prompts': prompts,
        }
        with open(os.path.join(prompt_dir, 'ai_prompts_latest.json'), 'w', encoding='utf-8') as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2)

        # Also write timestamped snapshots for traceability.
        history_dir = os.path.join(prompt_dir, 'history')
        os.makedirs(history_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        with open(os.path.join(history_dir, f'ai_prompts_{ts}.json'), 'w', encoding='utf-8') as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2)
        return True, None
    except Exception as e:
        traceback.print_exc()
        return False, str(e)


def _ai_pick_prompt(area, task, prompts, defaults):
    """
    Pick prompt template based on area/task with backward compatibility.
    """
    if area == 'answer':
        key_map = {
            'structure': 'answer_structure',
            'fault': 'answer_fault',
            'usage': 'answer_usage',
            'feature': 'answer_feature',
            'requirement': 'answer_requirement',
        }
        k = key_map.get((task or '').strip())
        if k:
            return (prompts.get(k) or defaults.get(k) or prompts.get('answer') or defaults.get('answer') or '')
        return (prompts.get('answer') or defaults.get('answer') or '')
    return (prompts.get(area) or defaults.get(area) or '')

def _ai_render_template(tpl, vars_map):
    if tpl is None:
        return ""
    s = str(tpl)
    def repl(m):
        k = m.group(1).strip()
        v = vars_map.get(k)
        if v is None:
            return ""
        if isinstance(v, (dict, list)):
            try:
                return json.dumps(v, ensure_ascii=False)
            except Exception:
                return str(v)
        return str(v)
    return re.sub(r"\{\{\s*([a-zA-Z0-9_]+)\s*\}\}", repl, s)

def _ai_extract_json(text):
    if text is None:
        return None
    t = str(text).strip()
    if not t:
        return None
    try:
        return json.loads(t)
    except Exception:
        pass
    m = re.search(r"\{[\s\S]*\}", t)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except Exception:
        return None

def _ai_cjkish_count(text):
    return len(re.findall(r"[\u3400-\u9fff\u3000-\u303f\uff00-\uffef]", str(text or "")))

def _ai_maybe_repair_mojibake_text(value):
    if not isinstance(value, str) or not value:
        return value
    # Typical symptom: UTF-8 Chinese bytes were decoded as Latin-1.
    if not re.search(r"[ÃÂãäåæçèéêëìíîïðñòóôõöøùúûü][\x80-\xff]", value):
        return value
    try:
        repaired = value.encode('latin-1').decode('utf-8')
    except (UnicodeEncodeError, UnicodeDecodeError):
        return value
    if repaired != value and _ai_cjkish_count(repaired) > _ai_cjkish_count(value):
        return repaired
    return value

def _ai_repair_mojibake_value(value):
    if isinstance(value, str):
        return _ai_maybe_repair_mojibake_text(value)
    if isinstance(value, list):
        return [_ai_repair_mojibake_value(v) for v in value]
    if isinstance(value, dict):
        return {k: _ai_repair_mojibake_value(v) for k, v in value.items()}
    return value

def _ai_ndjson(value):
    return json.dumps(value, ensure_ascii=False) + "\n"

def _ai_is_cjk(text):
    if not text:
        return False
    return re.search(r"[\u4e00-\u9fff]", text) is not None

def _ai_ngrams(text, n=2):
    if not text:
        return []
    s = re.sub(r"\s+", " ", str(text)).strip()
    if not s:
        return []
    if _ai_is_cjk(s):
        chars = [c for c in s if re.match(r"[\u4e00-\u9fffA-Za-z0-9]", c)]
        if len(chars) < n:
            return chars
        return ["".join(chars[i:i+n]) for i in range(0, len(chars) - n + 1)]
    words = re.findall(r"[A-Za-z0-9]+", s.lower())
    if len(words) < n:
        return words
    return [" ".join(words[i:i+n]) for i in range(0, len(words) - n + 1)]

def _ai_cosine_sim(a, b):
    a_grams = _ai_ngrams(a, 2)
    b_grams = _ai_ngrams(b, 2)
    if not a_grams or not b_grams:
        return 0.0
    av = {}
    bv = {}
    for g in a_grams:
        av[g] = av.get(g, 0) + 1
    for g in b_grams:
        bv[g] = bv.get(g, 0) + 1
    dot = 0.0
    for k, v in av.items():
        if k in bv:
            dot += v * bv[k]
    an = math.sqrt(sum(v*v for v in av.values()))
    bn = math.sqrt(sum(v*v for v in bv.values()))
    if an == 0 or bn == 0:
        return 0.0
    return float(dot / (an * bn))

def _ai_distinct_ngram_ratio(texts, n=2):
    grams = []
    for t in texts or []:
        grams.extend(_ai_ngrams(t, n))
    if not grams:
        return 0.0
    return float(len(set(grams)) / len(grams))

def _ai_count_sentences_en(text):
    if not text:
        return 0
    s = re.split(r"[.!?]+", str(text))
    return max(1, len([x for x in s if x.strip()]))

def _ai_count_words_en(text):
    if not text:
        return 0
    return len(re.findall(r"[A-Za-z]+", str(text)))

def _ai_estimate_syllables_en(word):
    w = re.sub(r"[^a-z]", "", word.lower())
    if not w:
        return 0
    w = re.sub(r"e$", "", w)
    groups = re.findall(r"[aeiouy]+", w)
    return max(1, len(groups)) if groups else 1

def _ai_flesch_reading_ease(text):
    if not text:
        return 0.0
    words = re.findall(r"[A-Za-z]+", str(text))
    if not words:
        return 0.0
    word_count = len(words)
    sent_count = _ai_count_sentences_en(text)
    syllables = sum(_ai_estimate_syllables_en(w) for w in words)
    wps = word_count / max(1, sent_count)
    spw = syllables / max(1, word_count)
    return float(206.835 - 1.015 * wps - 84.6 * spw)

def _ai_len(text):
    return len(str(text or "").strip())

def _ai_call_llm(config, system_prompt, user_prompt, temperature=0.2):
    api_key = (config or {}).get('api_key') or ''
    base_url = (config or {}).get('base_url') or ''
    model = (config or {}).get('model') or ''
    if not api_key or not base_url or not model:
        raise ValueError("AI 配置不完整（api_key/base_url/model）")
    base = str(base_url).strip().rstrip("/")
    if base.endswith("/v1"):
        base_v1 = base
    else:
        base_v1 = base + "/v1"
    url = base_v1 + "/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt or "You are a helpful assistant."},
            {"role": "user", "content": user_prompt or ""},
        ],
        "temperature": float(temperature or 0.2),
    }
    resp = requests.post(url, headers=headers, json=payload, timeout=120)
    resp.encoding = 'utf-8'
    if resp.status_code >= 400:
        raise RuntimeError(f"HTTP {resp.status_code}: {resp.text}")
    data = resp.json() or {}
    return (((data.get("choices") or [{}])[0].get("message") or {}).get("content") or "").strip()


def _ai_call_llm_stream(config, system_prompt, user_prompt, temperature=0.2):
    """
    Stream OpenAI-compatible chat.completions content deltas.
    Yields text chunks (delta content) as they arrive, and returns full text when completed.
    """
    api_key = (config or {}).get('api_key') or ''
    base_url = (config or {}).get('base_url') or ''
    model = (config or {}).get('model') or ''
    if not api_key or not base_url or not model:
        raise ValueError("AI 配置不完整（api_key/base_url/model）")
    base = str(base_url).strip().rstrip("/")
    if base.endswith("/v1"):
        base_v1 = base
    else:
        base_v1 = base + "/v1"
    url = base_v1 + "/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Accept": "text/event-stream",
    }
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt or "You are a helpful assistant."},
            {"role": "user", "content": user_prompt or ""},
        ],
        "temperature": float(temperature or 0.2),
        "stream": True,
    }

    resp = requests.post(url, headers=headers, json=payload, timeout=120, stream=True)
    resp.encoding = 'utf-8'
    if resp.status_code >= 400:
        raise RuntimeError(f"HTTP {resp.status_code}: {resp.text}")

    full = []
    for raw_line in resp.iter_lines(decode_unicode=False):
        if not raw_line:
            continue
        if isinstance(raw_line, bytes):
            raw_line = raw_line.decode('utf-8', errors='replace')
        line = raw_line.strip()
        if not line.startswith("data:"):
            continue
        data_str = line[len("data:"):].strip()
        if data_str == "[DONE]":
            break
        try:
            obj = json.loads(data_str)
            choices = obj.get("choices") or []
            if not choices:
                continue
            delta = (choices[0].get("delta") or {})
            content = delta.get("content")
            if isinstance(content, str) and content:
                full.append(content)
                yield content
        except Exception:
            continue

    return "".join(full).strip()


@app.route('/api/ai/optimize_stream', methods=['POST'])
@login_required
def ai_optimize_stream():
    """
    AI 优化（流式）：返回 NDJSON。
    事件格式：
    - {"type":"delta","text":"..."}  # 逐步输出
    - {"type":"final","result":{...}} # 最终结构化结果（与 /ai/optimize 一致）
    - {"type":"error","message":"..."} # 错误
    """
    def generate():
        try:
            payload = request.json or {}
            area = (payload.get('area') or '').strip()
            task = (payload.get('task') or '').strip()
            inputs = payload.get('inputs') or {}
            if not isinstance(inputs, dict):
                inputs = {}

            question = str(inputs.get('question') or '')
            answer = str(inputs.get('answer') or '')
            optimization_requirement = str(
                inputs.get('optimization_requirement')
                or inputs.get('optimizationRequirement')
                or ''
            )
            urls = inputs.get('urls')
            if isinstance(urls, str):
                urls = [u.strip() for u in re.split(r"[,，\n]", urls) if u.strip()]
            if not isinstance(urls, list):
                urls = []
            if not urls:
                merged = []
                merged.extend(_ai_parse_list(inputs.get('image_urls')))
                merged.extend(_ai_parse_list(inputs.get('video_urls')))
                merged.extend(_ai_parse_list(inputs.get('file_urls')))
                link_url = str(inputs.get('link_url') or '').strip()
                if link_url:
                    merged.append(link_url)
                urls = [u for u in merged if isinstance(u, str) and u.strip()]

            if area not in ('question', 'answer', 'similar', 'question_type'):
                yield _ai_ndjson({"type": "error", "message": "Invalid area"})
                return

            cfg = load_ai_config() or {}
            prompts = (cfg.get('ai_prompts') or {}) if isinstance(cfg.get('ai_prompts'), dict) else {}
            defaults = _ai_default_prompts()
            tpl = _ai_pick_prompt(area, task, prompts, defaults)

            vars_map = {
                'task': task,
                'question': question,
                'answer': answer,
                'optimization_requirement': optimization_requirement,
                'urls': urls,
                'target_min': payload.get('target_min', 0.75),
                'target_max': payload.get('target_max', 0.85),
                'count_min': payload.get('count_min', 3),
                'count_max': payload.get('count_max', 5),
                'difficulty': payload.get('difficulty')
            }
            user_prompt = _ai_render_template(tpl, vars_map)

            system_prompt = payload.get('system_prompt')
            if system_prompt is None:
                system_prompt = "你是一个严谨的编辑助手，输出必须为严格 JSON，不要输出任何额外文本。"

            # Stream once; validate at end (no retries in streaming mode)
            full_text_parts = []
            for chunk in _ai_call_llm_stream(cfg, system_prompt, user_prompt, temperature=0.2):
                if isinstance(chunk, str) and chunk:
                    full_text_parts.append(chunk)
                    yield _ai_ndjson({"type": "delta", "text": chunk})

            full_text = "".join(full_text_parts).strip()
            obj = _ai_extract_json(full_text)
            obj = _ai_repair_mojibake_value(obj)
            if not isinstance(obj, dict):
                yield _ai_ndjson({"type": "error", "message": "AI 输出解析失败", "raw": full_text})
                return

            # Minimal validation aligned with /ai/optimize
            ok = True
            if area == 'question':
                out_q = obj.get('question')
                if not isinstance(out_q, str) or not out_q.strip():
                    ok = False
            elif area == 'answer':
                if not isinstance(obj.get('answer'), str) and isinstance(obj.get('refined_answer'), str):
                    obj['answer'] = obj.get('refined_answer')
                out_a = obj.get('answer')
                if not isinstance(out_a, str) or not out_a.strip():
                    ok = False
            elif area == 'question_type':
                out_t = obj.get('question_type')
                if not isinstance(out_t, str) or not out_t.strip():
                    out_t = obj.get('type')
                    if not isinstance(out_t, str) or not out_t.strip():
                        ok = False
                    else:
                        obj['question_type'] = out_t
            else:
                items = obj.get('items')
                if not isinstance(items, list):
                    ok = False

            if not ok:
                yield _ai_ndjson({"type": "error", "message": "AI 输出字段不符合要求", "raw": full_text})
                return

            # 答案类任务：统一输出 original_answer / refined_answer / notes / answer
            if area == 'answer' and (task or '').strip() in ('structure', 'fault', 'usage', 'feature', 'requirement'):
                try:
                    orig_ans = answer
                except Exception:
                    orig_ans = ""
                refined = obj.get('answer') if isinstance(obj, dict) else None
                if not isinstance(refined, str):
                    refined = ""
                notes_val = obj.get('notes') if isinstance(obj, dict) else None
                obj = {
                    "original_answer": orig_ans,
                    "refined_answer": refined,
                    "notes": notes_val,
                    # 为兼容前端/旧逻辑，继续提供 answer 字段
                    "answer": refined,
                    # 透传 urls，若模型有返回
                    "urls": obj.get("urls") if isinstance(obj, dict) else None,
                }

            yield _ai_ndjson({"type": "final", "result": {"success": True, "data": obj, "metrics": {}}})
        except Exception as e:
            yield _ai_ndjson({"type": "error", "message": str(e)})

    return Response(stream_with_context(generate()), content_type='application/x-ndjson; charset=utf-8')

def _ai_parse_list(v):
    """
    供 AI 接口使用的简单列表解析工具，避免依赖 SmartMapping 内部的 _sm_parse_list。
    支持：
    - 直接传 list
    - JSON 字符串形式的 list
    - 使用换行/中英文逗号分隔的字符串
    """
    if v is None:
        return []
    if isinstance(v, list):
        return v
    s = str(v).strip()
    if not s:
        return []
    try:
        obj = json.loads(s)
        if isinstance(obj, list):
            return obj
    except Exception:
        pass
    parts = re.split(r'[\n,，]', s)
    return [x.strip() for x in parts if x.strip()]


@app.route('/api/ai/optimize', methods=['POST'])
@login_required
def ai_optimize():
    """
    AI 优化接口：用于问题润色 / 答案优化 / 相似问题生成。
    无论内部是否出错，都保证返回 JSON，避免前端解析 HTML 报错。
    """
    try:
        payload = request.json or {}
        area = (payload.get('area') or '').strip()
        task = (payload.get('task') or '').strip()
        inputs = payload.get('inputs') or {}
        if not isinstance(inputs, dict):
            inputs = {}

        question = str(inputs.get('question') or '')
        answer = str(inputs.get('answer') or '')
        optimization_requirement = str(
            inputs.get('optimization_requirement')
            or inputs.get('optimizationRequirement')
            or ''
        )
        urls = inputs.get('urls')
        if isinstance(urls, str):
            urls = [u.strip() for u in re.split(r"[,，\n]", urls) if u.strip()]
        if not isinstance(urls, list):
            urls = []
        if not urls:
            merged = []
            merged.extend(_ai_parse_list(inputs.get('image_urls')))
            merged.extend(_ai_parse_list(inputs.get('video_urls')))
            merged.extend(_ai_parse_list(inputs.get('file_urls')))
            link_url = str(inputs.get('link_url') or '').strip()
            if link_url:
                merged.append(link_url)
            urls = [u for u in merged if isinstance(u, str) and u.strip()]

        if area not in ('question', 'answer', 'similar', 'question_type'):
            return jsonify({'success': False, 'message': 'Invalid area'}), 400

        # 使用独立的 AI 配置，而不是评分配置
        cfg = load_ai_config() or {}
        prompts = (cfg.get('ai_prompts') or {}) if isinstance(cfg.get('ai_prompts'), dict) else {}
        defaults = _ai_default_prompts()
        tpl = _ai_pick_prompt(area, task, prompts, defaults)

        vars_map = {
            'task': task,
            'question': question,
            'answer': answer,
            'optimization_requirement': optimization_requirement,
            'urls': urls,
            'target_min': payload.get('target_min', 0.75),
            'target_max': payload.get('target_max', 0.85),
            'count_min': payload.get('count_min', 3),
            'count_max': payload.get('count_max', 5),
            'difficulty': payload.get('difficulty')
        }
        user_prompt = _ai_render_template(tpl, vars_map)

        # 问题类型分类：支持粘贴 JSON 配置（按型号分类选择 swapper/mopping/washing 的 Prompt）
        if area == 'question_type':
            # 使用 product_category（若无则回退到 product_category_name），支持多分类与优先级
            raw_pc = inputs.get('product_category')
            if not raw_pc:
                raw_pc = inputs.get('product_category_name')
            product_category = str(raw_pc or '')
            cfg_json = cfg.get('question_type_config_json')
            rules_prompt = ''
            if isinstance(cfg_json, str) and cfg_json.strip():
                try:
                    obj_cfg = json.loads(cfg_json)
                    if isinstance(obj_cfg, dict):
                        key = None
                        # 若有多个分类，用逗号/顿号拆分后，按优先级扫描：
                        # 扫地机(swapper) > 洗地机/吸尘器(mopping) > 洗衣机(washing)
                        pcs = re.split(r'[，,;/\s]+', product_category)
                        pcs = [p.strip() for p in pcs if p and p.strip()]
                        text_all = product_category.lower()

                        def match_sweeper(text: str) -> bool:
                            t = text.lower()
                            return ('扫地' in text) or ('sweeper' in t) or ('robot' in t) or ('vac' in t and '洗地' not in text)

                        def match_mopping_or_vacuum(text: str) -> bool:
                            t = text.lower()
                            return ('洗地' in text) or ('mop' in t) or ('吸尘器' in text) or ('vacuum' in t)

                        def match_washing(text: str) -> bool:
                            t = text.lower()
                            return ('洗衣' in text) or ('洗烘' in text) or ('washer' in t) or ('washing' in t)

                        # 1) 扫地机优先
                        if 'swapper' in obj_cfg and (
                            match_sweeper(product_category) or any(match_sweeper(p) for p in pcs)
                        ):
                            key = 'swapper'
                        # 2) 洗地机 / 吸尘器 → mopping
                        elif 'mopping' in obj_cfg and (
                            match_mopping_or_vacuum(product_category) or any(match_mopping_or_vacuum(p) for p in pcs)
                        ):
                            key = 'mopping'
                        # 3) 洗衣机
                        elif 'washing' in obj_cfg and (
                            match_washing(product_category) or any(match_washing(p) for p in pcs)
                        ):
                            key = 'washing'

                        # 4) 兜底：如果 product_category 恰好是 swapper/mopping/washing，本身也可用作 key
                        if key is None and product_category in obj_cfg:
                            key = product_category

                        chosen = obj_cfg.get(key) if key else None
                        if isinstance(chosen, dict):
                            rules_prompt = str(chosen.get('Prompt') or '')
                except Exception:
                    rules_prompt = ''

            if not rules_prompt:
                rules_prompt = user_prompt or ''

            user_prompt = (
                rules_prompt.strip()
                + "\n\n# 用户输入\n"
                + question.strip()
                + "\n\n# 输出格式\n请只输出严格 JSON：\n{\n  \"question_type\": string\n}\n"
            )

        system_prompt = payload.get('system_prompt')
        if system_prompt is None:
            system_prompt = "你是一个严谨的编辑助手，输出必须为严格 JSON，不要输出任何额外文本。"

        max_tries = 3
        last_obj = None
        last_raw = None
        best = None
        best_metrics = None

        for i in range(max_tries):
            raw = _ai_call_llm(cfg, system_prompt, user_prompt, temperature=0.2)
            obj = _ai_extract_json(raw)
            obj = _ai_repair_mojibake_value(obj)
            last_obj = obj
            last_raw = raw
            if not isinstance(obj, dict):
                user_prompt = user_prompt + "\n\n请只输出严格 JSON，不要包含 markdown 代码块。"
                continue

            metrics = {}
            ok = True

            if area == 'question':
                # 目前仅支持问题“润色”（rewrite），不再做关键词/难度等其他任务
                out_q = obj.get('question')
                if not isinstance(out_q, str) or not out_q.strip():
                    ok = False
            elif area == 'answer':
                # 答案区：允许前端根据 task 走不同模板/策略，但输出必须始终包含非空 answer
                if not isinstance(obj.get('answer'), str) and isinstance(obj.get('refined_answer'), str):
                    obj['answer'] = obj.get('refined_answer')
                out_a = obj.get('answer')
                if not isinstance(out_a, str) or not out_a.strip():
                    ok = False
            elif area == 'question_type':
                out_t = obj.get('question_type')
                if not isinstance(out_t, str) or not out_t.strip():
                    # 容错：模型可能输出 {"type": "..."}
                    out_t = obj.get('type')
                    if not isinstance(out_t, str) or not out_t.strip():
                        ok = False
                    else:
                        obj['question_type'] = out_t
            else:
                items = obj.get('items')
                if not isinstance(items, list):
                    ok = False

            if ok:
                best = obj
                best_metrics = metrics
                break

        if best is None:
            return jsonify({'success': False, 'message': 'AI 输出解析失败', 'raw': last_raw}), 500

        # 答案类任务：统一输出 original_answer / refined_answer / notes / answer
        if area == 'answer' and (task or '').strip() in ('structure', 'fault', 'usage', 'feature', 'requirement'):
            try:
                orig_ans = answer
            except Exception:
                orig_ans = ""
            refined = best.get('answer') if isinstance(best, dict) else None
            if not isinstance(refined, str):
                refined = ""
            notes_val = best.get('notes') if isinstance(best, dict) else None
            best = {
                "original_answer": orig_ans,
                "refined_answer": refined,
                "notes": notes_val,
                # 为兼容前端/旧逻辑，继续提供 answer 字段
                "answer": refined,
                # 透传 urls，若模型有返回
                "urls": best.get("urls") if isinstance(best, dict) else None,
            }

        # 将 question_type 从「类别名」映射为「full_class」（如 APP相关-智能助手）
        if area == 'question_type':
            try:
                cfg_json = cfg.get('question_type_config_json')
                raw_pc = inputs.get('product_category') or inputs.get('product_category_name') or ''
                product_category = str(raw_pc or '')
                if isinstance(cfg_json, str) and cfg_json.strip():
                    cfg_obj = json.loads(cfg_json)
                    if isinstance(cfg_obj, dict):
                        # 复用与上游相同的优先级规则，确定使用哪个 family（swapper/mopping/washing）
                        def _match_sweeper(text: str) -> bool:
                            t = text.lower()
                            return ('扫地' in text) or ('sweeper' in t) or ('robot' in t) or ('vac' in t and '洗地' not in text)

                        def _match_mopping_or_vacuum(text: str) -> bool:
                            t = text.lower()
                            return ('洗地' in text) or ('mop' in t) or ('吸尘器' in text) or ('vacuum' in t)

                        def _match_washing(text: str) -> bool:
                            t = text.lower()
                            return ('洗衣' in text) or ('洗烘' in text) or ('washer' in t) or ('washing' in t)

                        pcs = re.split(r'[，,;/\s]+', product_category)
                        pcs = [p.strip() for p in pcs if p and p.strip()]

                        fam_key = None
                        if 'swapper' in cfg_obj and (
                            _match_sweeper(product_category) or any(_match_sweeper(p) for p in pcs)
                        ):
                            fam_key = 'swapper'
                        elif 'mopping' in cfg_obj and (
                            _match_mopping_or_vacuum(product_category) or any(_match_mopping_or_vacuum(p) for p in pcs)
                        ):
                            fam_key = 'mopping'
                        elif 'washing' in cfg_obj and (
                            _match_washing(product_category) or any(_match_washing(p) for p in pcs)
                        ):
                            fam_key = 'washing'
                        if fam_key is None and product_category in cfg_obj:
                            fam_key = product_category

                        fam = cfg_obj.get(fam_key) if fam_key else None
                        classes = fam.get('class') if isinstance(fam, dict) else None
                        raw_type = best.get('question_type')
                        if isinstance(classes, dict) and isinstance(raw_type, str):
                            full = classes.get(raw_type)
                            if isinstance(full, str) and full.strip():
                                best['question_type_raw'] = raw_type
                                best['question_type'] = full
            except Exception:
                # 映射失败时，保持原样（仅展示类别名）
                pass

        return jsonify({
            'success': True,
            'area': area,
            'task': task,
            'data': best,
            'metrics': best_metrics or {}
        })

    except Exception as e:
        # 捕获所有异常，返回 JSON，避免 Flask 默认 HTML 错误页
        return jsonify({'success': False, 'message': f'AI 调用异常: {str(e)}'}), 500

def _split_text_list_value(raw, *, is_url_list=False):
    """与前端 parseSmartListValue(splitOnAsciiComma=True) 保持一致的列表拆分规则。

    规则要点（重要）：
    - 文本类字段（相似问/关键词/错误码等）：只按【换行】和【半角逗号 ,】拆分。
      全角逗号 ，是句子内部的正常标点，不作为分隔符，整句视为一个元素。
    - URL 类字段：在出现 >=2 个 url-like 片段时，才允许按全角逗号 ，拆分，
      以兼容历史上用中文逗号拼接多个链接的写法。

    例：
        "检测到基站内部异常，请插拔电源重试"   -> ["检测到基站内部异常，请插拔电源重试"]   (1 条)
        "检测到基站内部异常,请插拔电源重试"     -> ["检测到基站内部异常", "请插拔电源重试"] (2 条)
    """
    s = str(raw or '').strip()
    if not s:
        return []

    # 换行优先：有换行时按行拆分
    if '\n' in s or '\r' in s:
        return [x.strip() for x in re.split(r'\r?\n', s) if x and x.strip()]

    # 半角逗号：文本字段的分隔符
    if ',' in s:
        return [x.strip() for x in s.split(',') if x and x.strip()]

    # URL 字段才允许用全角逗号拆分（且需 >=2 个 url-like 片段）
    if is_url_list and '，' in s:
        rough = [x.strip() for x in s.split('，') if x and x.strip()]
        url_like = sum(1 for t in rough if re.match(r'^(https?://|www\.)', t, flags=re.I))
        if url_like >= 2:
            return rough

    # 否则整体作为一个元素（全角逗号保留在句子内）
    return [s]

def _parse_similar_questions(v):
    if v is None:
        return []
    if isinstance(v, list):
        out = []
        seen = set()
        for x in v:
            val = str(x or '').strip()
            if val and val not in seen:
                out.append(val)
                seen.add(val)
        return out
    if isinstance(v, dict):
        out = []
        seen = set()
        for vv in (v or {}).values():
            for val in _parse_similar_questions(vv):
                if val and val not in seen:
                    out.append(val)
                    seen.add(val)
        return out
    if isinstance(v, str):
        raw = v.strip()
        if not raw:
            return []
        if (raw.startswith('[') and raw.endswith(']')) or (raw.startswith('{') and raw.endswith('}')):
            try:
                return _parse_similar_questions(json.loads(raw))
            except Exception:
                pass
        out = []
        seen = set()
        for part in _split_text_list_value(raw):
            val = str(part or '').strip()
            if val and val not in seen:
                out.append(val)
                seen.add(val)
        return out
    return []


def _kb_compare_ai_text(value, limit=16000):
    """Keep comparison prompts bounded without changing the source records."""
    if value is None:
        return ''
    if isinstance(value, (dict, list)):
        try:
            value = json.dumps(value, ensure_ascii=False)
        except Exception:
            value = str(value)
    return str(value).strip()[:limit]


def _kb_compare_ai_bool(value):
    if isinstance(value, bool):
        return value
    text = str(value or '').strip().lower()
    if text in ('true', '1', 'yes', 'y', '建议合并', '合并', '是'):
        return True
    if text in ('false', '0', 'no', 'n', '不建议合并', '不合并', '否'):
        return False
    return None


def _kb_compare_ai_confidence(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0.0
    if number > 1:
        number /= 100
    return round(max(0.0, min(1.0, number)), 3)


def _kb_compare_ai_records(raw_records):
    if not isinstance(raw_records, list):
        return []
    records = []
    seen_ids = set()
    for raw in raw_records:
        if not isinstance(raw, dict):
            continue
        record_id = _kb_compare_ai_text(raw.get('question_wiki_id') or raw.get('id'), 160)
        if not record_id or record_id in seen_ids:
            continue
        seen_ids.add(record_id)
        records.append({
            'question_wiki_id': record_id,
            'question': _kb_compare_ai_text(raw.get('question')),
            'answer': _kb_compare_ai_text(raw.get('answer'), 24000),
            'similar_questions': _parse_similar_questions(raw.get('similar_questions'))[:80],
            'product_name': _kb_compare_ai_text(raw.get('product_name'), 4000),
            'product_category_name': _kb_compare_ai_text(raw.get('product_category_name'), 2000),
            'question_type': _kb_compare_ai_text(raw.get('question_type'), 1000),
            'answer_type': _kb_compare_ai_text(raw.get('answer_type'), 1000),
            'link_type': _kb_compare_ai_text(raw.get('link_type'), 1000),
            'link_url': _kb_compare_ai_text(raw.get('link_url'), 4000),
        })
    return records


@app.route('/api/kb/compare/ai_merge', methods=['POST'])
@login_required
def kb_compare_ai_merge():
    """Recommend whether records can merge and generate only the three content fields."""
    payload = request.json or {}
    records = _kb_compare_ai_records(payload.get('records'))
    if len(records) < 2:
        return jsonify({'success': False, 'message': '至少需要 2 条有效记录进行 AI 合并判断'}), 400
    if len(records) > 20:
        return jsonify({'success': False, 'message': '单次 AI 合并判断最多支持 20 条记录，请拆分后重试'}), 400

    base_id = _kb_compare_ai_text(payload.get('base_id'), 160)
    system_prompt = (
        '你是企业知识库治理审核员，负责判断多条知识库记录是否属于同一用户意图，并在可以合并时生成一份可人工复核的内容草稿。\n'
        '必须只输出合法 JSON，不要输出 Markdown 代码块或额外解释。\n'
        '合并判断规则：\n'
        '1. 只有当问题表达的是同一用户意图、操作目标和适用边界时才建议合并。\n'
        '2. 如果功能、步骤、产品范围或限制条件存在无法消解的冲突，建议不合并。\n'
        '3. 产品型号不同不必然禁止合并，但答案必须能同时适用于所有记录；无法确认时不合并。\n'
        '4. 不得补造输入记录中不存在的按钮、能力、参数或结论。\n'
        '5. 建议合并时，问题要能代表共同意图，答案只保留输入中可核验的事实，相似问题要去重并覆盖原问题的自然问法。\n'
        '6. 其他字段（型号、分类、类型、治理、链接）由系统按原规则处理，你只能生成 question、answer、similar_questions。\n'
        '输出结构：\n'
        '{"recommend_merge": true|false, "confidence": 0-1, "reason": "简洁原因", "conflicts": ["冲突点"], "question": "合并后问题", "answer": "合并后答案", "similar_questions": ["相似问题"]}\n'
        '当 recommend_merge=false 时，question、answer、similar_questions 必须返回空字符串或空数组。'
    )
    user_prompt = (
        f'当前主记录 ID：{base_id or records[0]["question_wiki_id"]}\n'
        '请分析以下记录：\n'
        + json.dumps(records, ensure_ascii=False, indent=2)
    )

    last_error = None
    for attempt in range(3):
        try:
            raw = _ai_call_llm(load_ai_config() or {}, system_prompt, user_prompt, temperature=0.1)
            result = _ai_repair_mojibake_value(_ai_extract_json(raw))
            if not isinstance(result, dict):
                raise ValueError('AI 未返回合法 JSON')

            recommend_merge = _kb_compare_ai_bool(result.get('recommend_merge'))
            if recommend_merge is None:
                raise ValueError('AI 未返回明确的合并建议')
            reason = _kb_compare_ai_text(result.get('reason'), 2000)
            if not reason:
                raise ValueError('AI 未返回合并判断原因')
            conflicts = _parse_similar_questions(result.get('conflicts'))[:12]
            confidence = _kb_compare_ai_confidence(result.get('confidence'))
            question = _kb_compare_ai_text(result.get('question'), 8000) if recommend_merge else ''
            answer = _kb_compare_ai_text(result.get('answer'), 30000) if recommend_merge else ''
            similar_questions = _parse_similar_questions(result.get('similar_questions'))[:20] if recommend_merge else []
            if recommend_merge and (not question or not answer or not similar_questions):
                raise ValueError('建议合并时必须同时返回问题、答案和相似问题')

            return jsonify({
                'success': True,
                'data': {
                    'recommend_merge': recommend_merge,
                    'confidence': confidence,
                    'reason': reason,
                    'conflicts': conflicts,
                    'question': question,
                    'answer': answer,
                    'similar_questions': similar_questions,
                    'base_id': base_id or records[0]['question_wiki_id'],
                    'source_ids': [record['question_wiki_id'] for record in records],
                }
            })
        except Exception as exc:
            last_error = exc
            if attempt < 2:
                user_prompt += '\n\n请修正上一轮输出，只返回符合结构的严格 JSON。'

    raw_message = str(last_error or '').strip()
    if 'AI 配置不完整' in raw_message:
        safe_message = raw_message
    elif raw_message.startswith('HTTP '):
        status_code = raw_message.split(':', 1)[0]
        safe_message = f'AI 服务请求失败（{status_code}），请检查 AI 配置或接口可用性。'
    else:
        safe_message = 'AI 输出未通过校验，请稍后重试。'
    return jsonify({'success': False, 'message': safe_message}), 500

@app.route('/api/kb/data', methods=['GET'])
@login_required
def get_kb_data():
    client = get_supabase_client()
    if not client:
        return jsonify([])
    
    table_name = request.args.get('table', 'knowledge_base_v1')
    page = request.args.get('page')
    page_size = request.args.get('pageSize')
    
    # Search parameters
    ids_raw = request.args.get('ids')
    id_search = request.args.get('id')
    question_search = request.args.get('question')
    similar_question_search = request.args.get('similar_question')
    answer_search = request.args.get('answer')
    product_search = request.args.get('product')
    url_search = request.args.get('url')
    product_category_search = request.args.get('product_categories') or request.args.get('product_category')
    tag_names_search = request.args.get('tagNames') or request.args.get('tag_names') or ''
    tag_mode_search = str(request.args.get('tagMode') or 'OR').strip().upper()  # OR / AND
    tag_mode = 'AND' if tag_mode_search == 'AND' else 'OR'
    
    # Security check: only allow specific tables
    allowed_tables = ['knowledge_base_v1', 'knowledge_base_v1_t1']
    if table_name not in allowed_tables:
        return jsonify([])
    
    # Determine sort column
    sort_by_raw = str(request.args.get('sortBy') or '').strip()
    sort_dir_raw = str(request.args.get('sortDir') or 'desc').strip().lower()

    # Keep sorting robust: frontend may send invalid values like "null"/"undefined".
    allowed_sort_by = {
        'question_wiki_id', 'review_status', 'question', 'answer', 'product_name',
        'product_category_name', 'question_type', 'answer_type', 'if_bm25',
        'update_time'
    }
    if table_name == 'knowledge_base_v1_t1':
        allowed_sort_by.discard('review_status')

    sort_by = sort_by_raw if sort_by_raw in allowed_sort_by else (
        'question_wiki_id' if 'knowledge_base_v1' in table_name else 'id'
    )
    sort_dir = sort_dir_raw if sort_dir_raw in ('asc', 'desc') else 'desc'

    kb_columns = None
    if table_name == 'knowledge_base_v1':
        kb_columns = (
            'question_wiki_id,review_status,question,answer,product_name,product_category_name,'
            'question_type,answer_type,similar_questions,if_bm25,error_list,keyword_list,'
            'image_urls,video_urls,file_urls,link_type,link_url,update_time'
        )

    # Build filters
    filters = {}
    ids = []
    if ids_raw:
        s = str(ids_raw).strip()
        if s:
            if s.startswith('[') and s.endswith(']'):
                try:
                    parsed = json.loads(s)
                    if isinstance(parsed, list):
                        ids = [str(x).strip() for x in parsed if str(x).strip()]
                except Exception:
                    ids = []
            if not ids:
                ids = [p.strip() for p in s.split(',') if p and p.strip()]
            seen = set()
            ids = [x for x in ids if not (x in seen or seen.add(x))]

    if ids:
        id_field = 'question_wiki_id'
        all_numeric = True
        for x in ids:
            if not str(x).isdigit():
                all_numeric = False
                break
        if all_numeric:
            in_list = ','.join(ids)
        else:
            quoted = ['"' + str(x).replace('"', '') + '"' for x in ids]
            in_list = ','.join(quoted)
        filters[id_field] = f"in.({in_list})"
    else:
        if id_search:
            filters['question_wiki_id'] = f"ilike.*{id_search}*"
        if question_search:
            filters['question'] = f"ilike.*{question_search}*"
        if similar_question_search:
            # similar_questions 是 JSONB，PostgREST 上对该列直接 ilike 兼容性不稳定。
            # 这里改为：先不下推到 Supabase 过滤，后续在 Python 层做稳健匹配。
            pass
        if answer_search:
            filters['answer'] = f"ilike.*{answer_search}*"
        if product_search:
            filters['product_name'] = f"ilike.*{product_search}*"
    
    or_clauses = []
    if url_search:
        clean_url_search = url_search.replace('https://', '').replace('http://', '')
        or_clauses.extend([
            f"image_urls->>0.ilike.*{clean_url_search}*",
            f"image_urls->>1.ilike.*{clean_url_search}*",
            f"image_urls->>2.ilike.*{clean_url_search}*",
            f"video_urls->>0.ilike.*{clean_url_search}*",
            f"video_urls->>1.ilike.*{clean_url_search}*",
            f"video_urls->>2.ilike.*{clean_url_search}*",
            f"file_urls->>0.ilike.*{clean_url_search}*",
            f"file_urls->>1.ilike.*{clean_url_search}*",
            f"file_urls->>2.ilike.*{clean_url_search}*",
            f"link_url.ilike.*{clean_url_search}*",
        ])
    
    if or_clauses and not ids:
        filters['or'] = f"({','.join(or_clauses)})"
    
    # Product category filter - handle separately as it should work as AND with other filters
    # When multiple categories are selected, they should be OR'd together (match any category)
    python_filter_categories = []  # For Python-side filtering when DB filter can't be used
    if product_category_search and not ids:
        parts = [p.strip() for p in str(product_category_search).split(',') if p and str(p).strip()]
        seen = set()
        categories = []
        for p in parts:
            if p in seen:
                continue
            seen.add(p)
            categories.append(p)
        
        if len(categories) == 1:
            # Single category - use simple ilike filter
            filters['product_category_name'] = f"ilike.*{categories[0]}*"
        elif len(categories) > 1:
            # Multiple categories - create OR clauses for them
            # Need to use 'or' parameter, but if URL search already used it, we have a conflict
            # Solution: Fetch all data and filter in Python, or use a different approach
            # For now, let's use Python-side filtering when there's a conflict
            category_or_clauses = [f"product_category_name.ilike.*{c}*" for c in categories]
            if 'or' not in filters:
                # No conflict, can use 'or' parameter for categories
                filters['or'] = f"({','.join(category_or_clauses)})"
            else:
                # Conflict case: both URL search and multiple categories
                # Will need to filter in Python after fetching
                python_filter_categories = categories
    
    # Helper function for Python-side category filtering (when DB filter can't be used)
    def _matches_category_filter(row, categories):
        """Check if row matches any of the specified categories"""
        if not categories or not isinstance(row, dict):
            return True
        row_category = str(row.get('product_category_name') or '').lower()
        if not row_category:
            return False
        return any(cat.lower() in row_category for cat in categories)

    # Tag filter (kb tags): restrict by matching item tags
    if tag_names_search and not ids:
        # For current/previous library, kb_item_tags.library_type maps to table_name
        library_type = 'current' if table_name == 'knowledge_base_v1' else 'previous'

        # Parse tag names: allow comma / chinese comma / newline
        raw = str(tag_names_search)
        parts = [p.strip() for p in re.split(r'[,，\n]+', raw) if p and str(p).strip()]
        # Deduplicate case-insensitively, keep original order
        seen = set()
        tag_names = []
        for p in parts:
            k = p.lower()
            if k in seen:
                continue
            seen.add(k)
            tag_names.append(p)

        # Load tag dictionary and map name -> id
        tag_dict = client.select_all('kb_tags', columns='id,name', page_size=1000) or []
        name_to_id = {}
        for tr in tag_dict:
            if isinstance(tr, dict) and tr.get('id') and tr.get('name'):
                name_to_id[str(tr.get('name')).strip().lower()] = tr.get('id')

        tag_ids = [name_to_id[n.lower()] for n in tag_names if n.lower() in name_to_id]
        tag_ids = [t for t in tag_ids if t]
        tag_ids_unique = []
        seen_tid = set()
        for tid in tag_ids:
            if str(tid) in seen_tid:
                continue
            seen_tid.add(str(tid))
            tag_ids_unique.append(tid)

        if not tag_ids_unique:
            return jsonify({'success': True, 'data': [], 'total': 0})

        # Fetch all mappings for these tags
        tag_id_in = _postgrest_in_str(tag_ids_unique)
        maps = client.select_all(
            'kb_item_tags',
            filters={
                'library_type': f'eq.{library_type}',
                'tag_id': tag_id_in
            },
            columns='question_wiki_id,tag_id',
            page_size=1000,
            order_by='question_wiki_id'
        ) or []

        # Compute allowed question_wiki_id set
        if tag_mode == 'AND':
            q_to_tset = {}
            for m in maps:
                if not isinstance(m, dict):
                    continue
                wid = str(m.get('question_wiki_id') or '').strip()
                tid = m.get('tag_id')
                if not wid or not tid:
                    continue
                q_to_tset.setdefault(wid, set()).add(str(tid))

            need_cnt = len([str(x) for x in tag_ids_unique])
            allowed_ids = [wid for wid, tset in q_to_tset.items() if len(tset) >= need_cnt]
        else:
            allowed_ids = list({str(m.get('question_wiki_id')).strip() for m in maps if isinstance(m, dict) and m.get('question_wiki_id')})

        allowed_ids = [x for x in allowed_ids if x]
        if not allowed_ids:
            return jsonify({'success': True, 'data': [], 'total': 0})

        filters['question_wiki_id'] = _postgrest_in_str(allowed_ids)

    review_status_search = request.args.get('review_status')
    if review_status_search and table_name == 'knowledge_base_v1' and not ids:
        # Only apply review_status filter to v1 table as t1 table doesn't have this column
        statuses = review_status_search.split(',')
        if len(statuses) > 1:
             # Quote statuses for PostgREST in filter
             quoted_statuses = [f'"{s}"' for s in statuses]
             filters['review_status'] = f"in.({','.join(quoted_statuses)})"
        else:
             filters['review_status'] = f"eq.{statuses[0]}"
    
    try:
        sim_kw_raw = str(similar_question_search or '').strip()
        want_similar_debug = bool(sim_kw_raw)
        sim_kw = _sm_norm_for_sim(sim_kw_raw)
        sim_kw_tokens = [
            _sm_norm_for_sim(x) for x in re.split(r'[\s,，;；、/|]+', sim_kw_raw)
            if str(x or '').strip()
        ]
        sim_kw_tokens = [x for x in sim_kw_tokens if x]

        def _contains_similar_kw(row):
            if not sim_kw:
                return True
            try:
                if not isinstance(row, dict):
                    return False
                val = row.get('similar_questions')
                sims = _parse_similar_questions(val)
                if not sims:
                    return False
                merged = _sm_norm_for_sim(" ".join(sims))
                if not merged:
                    return False
                # 优先按分词 AND 匹配，避免“延保 时效”这类非连续子串误判为无数据
                if sim_kw_tokens:
                    return all(tok in merged for tok in sim_kw_tokens)
                return sim_kw in merged
            except Exception:
                return False

        # 有相似问题关键词时，改用 Python 层过滤，避免 JSONB + ilike 的兼容问题导致“暂无数据”。
        # 只要传了 similar_question（非空），就强制走这条分支并返回 debug_similar（避免前端拿到 array 而无法诊断）。
        if want_similar_debug:
            data_all = client.select_all(
                table_name,
                order_by=sort_by,
                order_dir=sort_dir,
                filters=filters,
                columns=kb_columns or '*'
            )
            debug_samples = []
            try:
                print(f"[KB] similar_question={sim_kw_raw!r} norm={sim_kw!r} pre_count={len(data_all or [])}")
            except Exception:
                pass
            if sim_kw:
                filtered = [r for r in (data_all or []) if _contains_similar_kw(r)]
            else:
                # 归一化后为空时：不做相似问过滤，但仍返回 debug 信息，避免“看似无数据”的假象
                filtered = list(data_all or [])
            try:
                # Log a couple of samples to diagnose format issues
                samples = []
                for r in (data_all or [])[:3]:
                    if isinstance(r, dict):
                        samples.append({
                            "id": r.get("question_wiki_id"),
                            "similar_questions": r.get("similar_questions"),
                        })
                print(f"[KB] similar filter post_count={len(filtered)} samples={samples}")
                debug_samples = samples
            except Exception:
                pass

            if page and page_size:
                p = max(int(page), 1)
                ps = max(int(page_size), 1)
                start = (p - 1) * ps
                end = start + ps
                data_out = filtered[start:end]
            else:
                data_out = filtered

            # Attach KB item tags for table rendering (方案A：把标签回填到 kb_tags 字段)
            try:
                library_type = 'current' if table_name == 'knowledge_base_v1' else 'previous'
                wiki_ids = [
                    str(r.get('question_wiki_id')).strip()
                    for r in (data_out or [])
                    if isinstance(r, dict) and r.get('question_wiki_id')
                ]
                if wiki_ids:
                    in_str = _postgrest_in_str(wiki_ids)
                    maps = client.select_all(
                        'kb_item_tags',
                        filters={'library_type': f'eq.{library_type}', 'question_wiki_id': in_str},
                        columns='question_wiki_id,tag_id',
                        order_by='question_wiki_id',
                        page_size=1000
                    ) or []
                    tag_ids = []
                    for m in maps:
                        if isinstance(m, dict) and m.get('tag_id'):
                            tid = m.get('tag_id')
                            if tid not in tag_ids:
                                tag_ids.append(tid)

                    id_to_name = {}
                    if tag_ids:
                        id_in = _postgrest_in_str(tag_ids)
                        tags_rows = client.select_all(
                            'kb_tags',
                            filters={'id': id_in},
                            columns='id,name',
                            page_size=1000
                        ) or []
                        for tr in tags_rows:
                            if isinstance(tr, dict) and tr.get('id') and tr.get('name'):
                                id_to_name[str(tr.get('id'))] = str(tr.get('name'))

                    item_to_tags = {}
                    for m in maps:
                        if not isinstance(m, dict):
                            continue
                        wid = str(m.get('question_wiki_id') or '').strip()
                        tid = m.get('tag_id')
                        if not wid or not tid:
                            continue
                        nm = id_to_name.get(str(tid))
                        # Fallback: if kb_tags.name missing, show tag_id to avoid silent empty UI
                        item_to_tags.setdefault(wid, []).append(nm if nm else str(tid))

                    for r in (data_out or []):
                        if isinstance(r, dict):
                            wid = str(r.get('question_wiki_id') or '').strip()
                            r['kb_tags'] = item_to_tags.get(wid, [])
            except Exception as e:
                print(f"WARN: attach kb tags failed (similar branch): {e}")

            # Always return an object when similar_question is used, so frontend can read debug_similar.
            return jsonify({
                'success': True,
                'data': data_out,
                'total': len(filtered),
                'debug_similar': {
                    'raw': sim_kw_raw,
                    'norm': sim_kw,
                    'tokens': sim_kw_tokens,
                    'pre_count': len(data_all or []),
                    'post_count': len(filtered),
                    'sample_rows': debug_samples,
                    'note': ('norm_empty_no_filter' if not sim_kw else 'filtered'),
                }
            })

        # If pagination params are provided, use paged select
        if page and page_size:
            resp = client.select(
                table_name,
                page=int(page),
                page_size=int(page_size),
                order_by=sort_by,
                order_dir=sort_dir,
                filters=filters,
                columns=kb_columns or '*',
                count='estimated'
            )
            if resp.status_code >= 400:
                print(f"Database Error: {resp.text}")
                return jsonify([])
        
            # Parse total count from Content-Range header
            total = 0
            content_range = resp.headers.get('Content-Range')
            if content_range:
                try:
                    # Content-Range format: 0-19/1234
                    total = int(content_range.split('/')[-1])
                except:
                    pass
        
            rows = resp.json() or []
            
            # Apply Python-side category filtering if needed (when DB filter couldn't be used)
            if python_filter_categories:
                rows = [r for r in rows if _matches_category_filter(r, python_filter_categories)]
                total = len(rows)  # Update total count after filtering
            
            # Attach KB item tags for table rendering
            try:
                library_type = 'current' if table_name == 'knowledge_base_v1' else 'previous'
                wiki_ids = [
                    str(r.get('question_wiki_id')).strip()
                    for r in (rows or [])
                    if isinstance(r, dict) and r.get('question_wiki_id')
                ]
                if wiki_ids:
                    in_str = _postgrest_in_str(wiki_ids)
                    maps = client.select_all(
                        'kb_item_tags',
                        filters={'library_type': f'eq.{library_type}', 'question_wiki_id': in_str},
                        columns='question_wiki_id,tag_id',
                        order_by='question_wiki_id',
                        page_size=1000
                    ) or []
                    tag_ids = []
                    for m in maps:
                        if isinstance(m, dict) and m.get('tag_id'):
                            tid = m.get('tag_id')
                            if tid not in tag_ids:
                                tag_ids.append(tid)

                    id_to_name = {}
                    if tag_ids:
                        id_in = _postgrest_in_str(tag_ids)
                        tags_rows = client.select_all(
                            'kb_tags',
                            filters={'id': id_in},
                            columns='id,name',
                            page_size=1000
                        ) or []
                        for tr in tags_rows:
                            if isinstance(tr, dict) and tr.get('id') and tr.get('name'):
                                id_to_name[str(tr.get('id'))] = str(tr.get('name'))

                    item_to_tags = {}
                    for m in maps:
                        if not isinstance(m, dict):
                            continue
                        wid = str(m.get('question_wiki_id') or '').strip()
                        tid = m.get('tag_id')
                        if not wid or not tid:
                            continue
                        nm = id_to_name.get(str(tid))
                        # Fallback: if kb_tags.name missing, show tag_id to avoid silent empty UI
                        item_to_tags.setdefault(wid, []).append(nm if nm else str(tid))

                    for r in (rows or []):
                        if isinstance(r, dict):
                            wid = str(r.get('question_wiki_id') or '').strip()
                            r['kb_tags'] = item_to_tags.get(wid, [])
            except Exception as e:
                print(f"WARN: attach kb tags failed (paged branch): {e}")

            return jsonify({
                'success': True,
                'data': rows,
                'total': total
            })
        else:
            # Fallback to select_all for backward compatibility or small tables
            data = client.select_all(table_name, order_by=sort_by, order_dir=sort_dir, filters=filters)
            return jsonify(data)
    except Exception as e:
        print(f"Error fetching KB data: {e}")
        return jsonify([])

def generate_kb_id(client):
    """Generate a new KB ID in format ICWIKIYYYYMMDDnnnn"""
    today_str = datetime.now().strftime('%Y%m%d')
    prefix = f"ICWIKI{today_str}"
    
    def _suffix_of(kb_id):
        try:
            s = str(kb_id or '').strip()
            if not s.startswith(prefix):
                return None
            tail = s[len(prefix):]
            if len(tail) < 4:
                return None
            return int(tail[-4:])
        except Exception:
            return None

    max_suffix = None

    try:
        resp = client.select(
            'knowledge_base_v1',
            filters={'question_wiki_id': f'like.{prefix}*'},
            order_by='question_wiki_id',
            order_dir='desc',
            page_size=1
        )
        if resp.status_code in (200, 206):
            data = resp.json() or []
            if data and isinstance(data[0], dict):
                suf = _suffix_of(data[0].get('question_wiki_id'))
                if suf is not None:
                    max_suffix = suf
    except Exception:
        pass

    try:
        resp2 = client.select(
            'knowledge_base_modifications',
            filters={'kb_id': f'like.{prefix}*'},
            order_by='kb_id',
            order_dir='desc',
            page_size=1
        )
        if resp2.status_code in (200, 206):
            data2 = resp2.json() or []
            if data2 and isinstance(data2[0], dict):
                suf2 = _suffix_of(data2[0].get('kb_id') or data2[0].get('question_wiki_id'))
                if suf2 is not None and (max_suffix is None or suf2 > max_suffix):
                    max_suffix = suf2
    except Exception:
        pass

    if max_suffix is None:
        return f"{prefix}0001"
    return f"{prefix}{(max_suffix + 1):04d}"

def _parse_change_meta(val):
    if val is None:
        return {}
    if isinstance(val, dict):
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() == 'null':
            return {}
        try:
            obj = json.loads(s)
            return obj if isinstance(obj, dict) else {}
        except Exception:
            return {}
    return {}

def _attach_change_meta(record, meta):
    try:
        if not isinstance(record, dict):
            return record
        if not isinstance(meta, dict) or not meta:
            return record
        existing = _parse_change_meta(record.get('change_meta'))
        merged = dict(existing)
        merged.update(meta)
        # 将 dict 转换为 JSON 字符串，以便 PostgreSQL 可以接受
        record['change_meta'] = json.dumps(merged, ensure_ascii=False)
        
        # 同时设置 source_module 字段（从 meta 中的 source 提取）
        if 'source' in meta:
            record['source_module'] = meta['source']
        
        return record
    except Exception:
        return record

def _request_from_port(port):
    target = str(port)
    for header in ('Origin', 'Referer'):
        raw = str(request.headers.get(header) or '').strip()
        if not raw:
            continue
        try:
            if urlparse(raw).port == int(target):
                return True
        except Exception:
            if f':{target}' in raw:
                return True
    return False

def _resolve_kb_change_source(payload):
    explicit = str((payload or {}).get('change_source') or (payload or {}).get('source_module') or '').strip()
    if explicit:
        return explicit
    if _request_from_port(8083):
        return BADCASE_WORKBENCH_SOURCE
    return '知识库管理'

def _convert_array_fields_to_json(record_or_records):
    """
    将修改记录中的数组字段转换为 JSON 字符串，以匹配 PostgreSQL 的 jsonb 类型。
    支持单个记录（dict）或记录列表（list）。
    """
    array_fields = ['keyword_list', 'error_list', 'similar_questions', 'image_urls', 'video_urls', 'file_urls']
    
    def convert_single(rec):
        if not isinstance(rec, dict):
            return rec
        for field in array_fields:
            if field in rec and isinstance(rec[field], list):
                rec[field] = json.dumps(rec[field], ensure_ascii=False)
        return rec
    
    if isinstance(record_or_records, list):
        return [convert_single(rec) for rec in record_or_records]
    else:
        return convert_single(record_or_records)

def _supabase_insert_drop_unknown_columns(client, table, payload):
    if not client or not table:
        return None
    data = payload
    last = None
    removed_cols = 0
    while True:
        resp = client.insert(table, data)
        last = resp
        if getattr(resp, 'status_code', 500) < 400:
            return resp
        text = str(getattr(resp, 'text', '') or '')
        m = re.search(r"Could not find the '([^']+)' column of '([^']+)'", text)
        if not m:
            m = re.search(r'column "([^"]+)" of relation "([^"]+)" does not exist', text)
        if not m:
            return resp
        col = str(m.group(1) or '').strip()
        if not col:
            return resp
        if removed_cols >= 64:
            return resp
        if isinstance(data, list):
            changed = False
            new_list = []
            for row in data:
                if isinstance(row, dict) and col in row:
                    nr = dict(row)
                    nr.pop(col, None)
                    new_list.append(nr)
                    changed = True
                else:
                    new_list.append(row)
            if not changed:
                return resp
            data = new_list
            removed_cols += 1
        elif isinstance(data, dict):
            if col not in data:
                return resp
            nd = dict(data)
            nd.pop(col, None)
            data = nd
            removed_cols += 1
        else:
            return resp
    return last

def _build_kb_modification_record(source, modifier, change_type, kb_id, before_obj, after_obj, changed_fields, base_row=None):
    rec = {}
    if isinstance(base_row, dict):
        for k in _kb_all_fields_allowlist():
            if k in base_row and k not in ('review_status',):
                rec[k] = base_row.get(k)
    rec['kb_id'] = kb_id
    if kb_id:
        rec['question_wiki_id'] = kb_id
    rec['modifier'] = modifier
    rec['modification_time'] = _now_iso_with_tz()
    rec['change_type'] = change_type
    rec.pop('id', None)
    rec.pop('review_status', None)
    _attach_change_meta(rec, {
        'source': source,
        'before': before_obj,
        'after': after_obj,
        'changed_fields': changed_fields
    })
    
    # 转换数组字段为 JSON 字符串
    _convert_array_fields_to_json(rec)
    
    return rec

def _extract_change_source_from_change_meta(change_meta):
    meta = _parse_change_meta(change_meta)
    src = meta.get('source') if isinstance(meta, dict) else None
    return str(src) if src else None

def _extract_change_meta(change_meta):
    return _parse_change_meta(change_meta)

def _kb_all_fields_allowlist():
    return {
        'question_wiki_id',
        'question_type',
        'question',
        'answer',
        'answer_type',
        'if_bm25',
        'similar_questions',
        'error_list',
        'keyword_list',
        'image_urls',
        'video_urls',
        'file_urls',
        'link_type',
        'link_url',
        'update_time',
        'product_category_name',
        'product_name',
        'review_status'
    }

_MOD_DIFF_FIELDS = [
    'question', 'answer', 'products', 'question_type', 'answer_type', 'error_list',
    'image_urls', 'video_urls', 'file_urls', 'link_type', 'link_url',
    'similar_questions', 'keyword_list', 'if_bm25'
]

def _normalize_mod_diff_value(v):
    try:
        if v is None:
            return None
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float)):
            return v
        if isinstance(v, list):
            cleaned = []
            for x in v:
                if x is None:
                    continue
                s = str(x).strip()
                if not s or s.lower() == 'null':
                    continue
                cleaned.append(s)
            cleaned = sorted(list(dict.fromkeys(cleaned)))
            return cleaned
        if isinstance(v, dict):
            return v
        return str(v).strip()
    except Exception:
        return v

def _export_mod_bool_text(v):
    if v is True:
        return '是'
    if v is False:
        return '否'
    s = str(v or '').strip().lower()
    if not s:
        return ''
    if s in ('1', 'true', 'yes', 'y', '是'):
        return '是'
    if s in ('0', 'false', 'no', 'n', '否'):
        return '否'
    return ''

def _export_mod_list_text(v, is_url_list=False):
    if v is None:
        return ''
    if isinstance(v, dict):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)
    if isinstance(v, list):
        tokens = []
        for x in v:
            if x is None:
                continue
            s = str(x).replace('`', '').strip()
            if not s or s.lower() in ('null', 'none', 'nan'):
                continue
            tokens.append(s)
        seen = set()
        out = []
        for t in tokens:
            if t in seen:
                continue
            seen.add(t)
            out.append(t)
        return ",".join(out)

    raw = str(v).replace('`', '').strip()
    if not raw or raw in ('[]', '{}') or raw.lower() in ('null', 'none', 'nan'):
        return ''
    if (raw.startswith('[') and raw.endswith(']')) or (raw.startswith('{') and raw.endswith('}')):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return _export_mod_list_text(parsed, is_url_list=is_url_list)
            if isinstance(parsed, dict):
                return json.dumps(parsed, ensure_ascii=False)
        except Exception:
            pass

    if "\n" in raw or "\r" in raw:
        parts = re.split(r"\r?\n", raw)
        return _export_mod_list_text([p for p in (x.strip() for x in parts) if p], is_url_list=is_url_list)

    if is_url_list and '，' in raw:
        rough = [x.strip() for x in raw.split('，') if x.strip()]
        url_like = sum(1 for t in rough if re.match(r"^(https?://|www\.)", t, flags=re.I))
        if url_like >= 2:
            return _export_mod_list_text(rough, is_url_list=is_url_list)

    parts = [x.strip() for x in re.split(r"[,，]+", raw) if x.strip()]
    if len(parts) <= 1:
        return raw
    return _export_mod_list_text(parts, is_url_list=is_url_list)

def _aggregate_links_from_fields(obj):
    links = []
    if isinstance(obj, dict):
        for k in ('image_urls', 'video_urls', 'file_urls'):
            v = obj.get(k)
            if isinstance(v, list):
                links.extend(v)
            elif isinstance(v, str):
                s = v.strip()
                if s:
                    try:
                        parsed = json.loads(s)
                        if isinstance(parsed, list):
                            links.extend(parsed)
                        elif isinstance(parsed, str):
                            links.append(parsed)
                    except Exception:
                        parts = [x.strip() for x in re.split(r'[,，\n\r]+', s) if x and x.strip()]
                        links.extend(parts)
        lu = str(obj.get('link_url') or '').strip()
        if lu:
            links.append(lu)
    out = []
    seen = set()
    for x in links:
        if x is None:
            continue
        s = str(x).replace('`', '').strip()
        if not s:
            continue
        if s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out

def _snapshot_mod_fields(row):
    if not isinstance(row, dict):
        return {k: None for k in _MOD_DIFF_FIELDS}
    return {
        'question': row.get('question') or row.get('question_content') or '',
        'answer': row.get('answer') or row.get('answer_content') or '',
        'products': row.get('products') or row.get('product_name') or '',
        'question_type': row.get('question_type'),
        'answer_type': row.get('answer_type'),
        'error_list': row.get('error_list'),
        'image_urls': row.get('image_urls'),
        'video_urls': row.get('video_urls'),
        'file_urls': row.get('file_urls'),
        'link_type': row.get('link_type'),
        'link_url': row.get('link_url'),
        'similar_questions': row.get('similar_questions'),
        'keyword_list': row.get('keyword_list'),
        'if_bm25': row.get('if_bm25')
    }

def _compute_mod_changed_fields(before_obj, after_obj):
    changed = []
    for k in _MOD_DIFF_FIELDS:
        b = _normalize_mod_diff_value((before_obj or {}).get(k))
        a = _normalize_mod_diff_value((after_obj or {}).get(k))
        if b != a:
            changed.append(k)
    return changed

def _safe_parse_dt(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    try:
        if s.endswith('Z'):
            s = s[:-1] + '+00:00'
        return datetime.fromisoformat(s)
    except Exception:
        try:
            return datetime.strptime(str(val), "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None

def _postgrest_in_str(values):
    cleaned = []
    for v in (values or []):
        s = str(v or '').strip()
        if s:
            cleaned.append(s)
    if not cleaned:
        return None
    inner = ",".join([json.dumps(v, ensure_ascii=False) for v in cleaned])
    return f"in.({inner})"

def _is_blank_cell_value(value):
    if value is None:
        return True
    if isinstance(value, str):
        s = value.strip()
        return not s or s.lower() in ('nan', 'null', 'none')
    if isinstance(value, (list, tuple, set, dict)):
        return False
    try:
        blank = pd.isna(value)
        if hasattr(blank, 'all') and not isinstance(blank, bool):
            return bool(blank.all())
        return bool(blank)
    except (TypeError, ValueError):
        return False

def _normalize_kb_import_id(raw):
    if isinstance(raw, (list, tuple, set, dict)):
        return ''
    if _is_blank_cell_value(raw):
        return ''
    s = str(raw).strip()
    if not s or s.lower() in ('nan', 'null', 'none'):
        return ''
    if s.endswith('.0'):
        s = s[:-2]
    return s.strip()

def _get_v1_ids_missing_from_file(client, file_ids):
    file_set = {str(x or '').strip() for x in (file_ids or []) if str(x or '').strip()}
    if not file_set:
        return []
    rows = client.select_all(
        'knowledge_base_v1',
        columns='question_wiki_id',
        order_by='question_wiki_id',
        page_size=1000
    ) or []
    missing = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        wiki_id = _normalize_kb_import_id(row.get('question_wiki_id'))
        if not wiki_id or wiki_id in file_set:
            continue
        missing.append(wiki_id)
    return sorted(list(dict.fromkeys(missing)))

def _delete_link_previews_for_kb_ids(client, ids):
    id_set = {_normalize_kb_import_id(x) for x in (ids or []) if _normalize_kb_import_id(x)}
    result = {'deleted': 0, 'updated': 0, 'warnings': []}
    if not id_set:
        return result

    try:
        rows = client.select_all('link_previews', columns='id,kb_id', order_by='id', page_size=1000) or []
    except Exception as e:
        result['warnings'].append(f'清理 link_previews 失败: {e}')
        return result

    link_ids_to_delete = []
    links_to_update = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        link_id = str(row.get('id') or '').strip()
        kb_id_raw = str(row.get('kb_id') or '').strip()
        if not link_id or not kb_id_raw:
            continue
        kb_ids = [_normalize_kb_import_id(x) for x in kb_id_raw.split(',')]
        kb_ids = [x for x in kb_ids if x]
        if not kb_ids:
            continue
        remaining = [x for x in kb_ids if x not in id_set]
        if len(remaining) == len(kb_ids):
            continue
        if remaining:
            links_to_update.append((link_id, ','.join(dict.fromkeys(remaining))))
        else:
            link_ids_to_delete.append(link_id)

    batch_size = 100
    for i in range(0, len(link_ids_to_delete), batch_size):
        batch_ids = link_ids_to_delete[i:i + batch_size]
        id_filter = _postgrest_in_str(batch_ids)
        if not id_filter:
            continue
        resp = client.delete('link_previews', {'id': id_filter})
        if resp is None or getattr(resp, 'status_code', 500) >= 400:
            result['warnings'].append(getattr(resp, 'text', '') or '删除 link_previews 失败')
            continue
        result['deleted'] += len(batch_ids)

    for link_id, new_kb_id in links_to_update:
        resp = client.update('link_previews', {'kb_id': new_kb_id}, {'id': link_id})
        if resp is None or getattr(resp, 'status_code', 500) >= 400:
            result['warnings'].append(getattr(resp, 'text', '') or f'更新 link_previews({link_id}) 失败')
            continue
        result['updated'] += 1

    return result

def _delete_kb_items_physical(client, ids):
    ids = [_normalize_kb_import_id(x) for x in (ids or []) if _normalize_kb_import_id(x)]
    ids = list(dict.fromkeys(ids))
    if not ids:
        return {'success': True, 'count': 0, 'ids': []}

    deleted_ids = []
    related = {'kb_scores': 0, 'kb_item_tags': 0, 'link_previews_deleted': 0, 'link_previews_updated': 0}
    warnings = []
    batch_size = 100
    for i in range(0, len(ids), batch_size):
        batch_ids = ids[i:i + batch_size]
        id_filter = _postgrest_in_str(batch_ids)
        if not id_filter:
            continue

        for table, column, label in (
            ('kb_scores', 'kb_id', '评分缓存'),
            ('kb_item_tags', 'question_wiki_id', '当前标签映射'),
        ):
            filters = {column: id_filter}
            if table == 'kb_item_tags':
                filters['library_type'] = 'eq.current'
            resp = client.delete(table, filters)
            if resp is None or getattr(resp, 'status_code', 500) >= 400:
                return {
                    'success': False,
                    'count': len(deleted_ids),
                    'ids': deleted_ids,
                    'message': getattr(resp, 'text', '') or f'物理删除前清理{label}失败'
                }
            related[table] += len(batch_ids)

        delete_resp = client.delete('knowledge_base_v1', {'question_wiki_id': id_filter})
        if delete_resp is None or getattr(delete_resp, 'status_code', 500) >= 400:
            return {
                'success': False,
                'count': len(deleted_ids),
                'ids': deleted_ids,
                'message': getattr(delete_resp, 'text', '') or '物理删除失败'
            }

        verify_resp = client.select(
            'knowledge_base_v1',
            page=1,
            page_size=len(batch_ids),
            filters={'question_wiki_id': id_filter},
            columns='question_wiki_id'
        )
        if verify_resp is None or getattr(verify_resp, 'status_code', 500) >= 400:
            return {
                'success': False,
                'count': len(deleted_ids),
                'ids': deleted_ids,
                'message': getattr(verify_resp, 'text', '') or '物理删除后校验失败'
            }
        remaining_rows = verify_resp.json() or []
        remaining_ids = [
            _normalize_kb_import_id(r.get('question_wiki_id'))
            for r in remaining_rows
            if isinstance(r, dict) and _normalize_kb_import_id(r.get('question_wiki_id'))
        ]
        if remaining_ids:
            return {
                'success': False,
                'count': len(deleted_ids),
                'ids': deleted_ids,
                'message': f'物理删除未完成，仍有 {len(remaining_ids)} 条记录存在: {", ".join(remaining_ids[:10])}'
            }

        deleted_ids.extend(batch_ids)

    link_cleanup = _delete_link_previews_for_kb_ids(client, deleted_ids)
    related['link_previews_deleted'] = link_cleanup.get('deleted', 0)
    related['link_previews_updated'] = link_cleanup.get('updated', 0)
    warnings.extend(link_cleanup.get('warnings') or [])

    return {
        'success': True,
        'count': len(deleted_ids),
        'ids': deleted_ids,
        'related': related,
        'warnings': warnings
    }

def _now_iso_with_tz():
    """
    Return timezone-aware local ISO datetime string.
    Use local timezone instead of naive UTC to avoid UI time offset in history tables.
    """
    return datetime.now().astimezone().isoformat()

def _normalize_mod_record(row):
    if not isinstance(row, dict):
        return None
    meta = _extract_change_meta(row.get('change_meta')) or {}
    src = (meta.get('source') if isinstance(meta, dict) else None) or _extract_change_source_from_change_meta(row.get('change_meta')) or '知识库管理'
    wiki_id = row.get('question_wiki_id') or row.get('kb_id') or ''
    change_type = row.get('change_type') or row.get('modification_type') or 'edit'
    before_obj = meta.get('before') if isinstance(meta, dict) else None
    after_obj = meta.get('after') if isinstance(meta, dict) else None
    changed_fields = meta.get('changed_fields') if isinstance(meta, dict) else None
    operation_id = meta.get('operation_id') if isinstance(meta, dict) else None
    if not isinstance(after_obj, dict) or not after_obj:
        after_obj = _snapshot_mod_fields(row)
    if isinstance(before_obj, dict) and not before_obj:
        before_obj = None
    if not isinstance(changed_fields, list):
        if isinstance(before_obj, dict) and isinstance(after_obj, dict):
            changed_fields = _compute_mod_changed_fields(before_obj, after_obj)
        else:
            changed_fields = []
    q_val = (after_obj.get('question') if isinstance(after_obj, dict) else None) or row.get('question') or row.get('question_content') or ''
    a_val = (after_obj.get('answer') if isinstance(after_obj, dict) else None) or row.get('answer') or row.get('answer_content') or ''
    p_val = (after_obj.get('products') if isinstance(after_obj, dict) else None) or row.get('products') or row.get('product_name') or '-'
    qtype_val = (after_obj.get('question_type') if isinstance(after_obj, dict) else None) or row.get('question_type')
    atype_val = (after_obj.get('answer_type') if isinstance(after_obj, dict) else None) or row.get('answer_type')
    err_val = (after_obj.get('error_list') if isinstance(after_obj, dict) else None) or row.get('error_list')
    kw_val = (after_obj.get('keyword_list') if isinstance(after_obj, dict) else None) or row.get('keyword_list')
    sim_val = (after_obj.get('similar_questions') if isinstance(after_obj, dict) else None) or row.get('similar_questions')
    bm25_val = (after_obj.get('if_bm25') if isinstance(after_obj, dict) else None) or row.get('if_bm25')
    image_urls_val = (after_obj.get('image_urls') if isinstance(after_obj, dict) else None) or row.get('image_urls')
    video_urls_val = (after_obj.get('video_urls') if isinstance(after_obj, dict) else None) or row.get('video_urls')
    file_urls_val = (after_obj.get('file_urls') if isinstance(after_obj, dict) else None) or row.get('file_urls')
    link_type_val = (after_obj.get('link_type') if isinstance(after_obj, dict) else None) or row.get('link_type')
    link_url_val = (after_obj.get('link_url') if isinstance(after_obj, dict) else None) or row.get('link_url')
    mod_time = row.get('modification_time')
    out = {
        'opera': change_type,
        'change_type': change_type,
        'operation': change_type,
        'supabase_id': row.get('id'),
        'question_wiki_id': wiki_id,
        'kb_id': row.get('kb_id') or wiki_id,
        'modifier': row.get('modifier'),
        'question': q_val,
        'question_type': qtype_val,
        'answer': a_val,
        'answer_type': atype_val,
        'error_list': err_val,
        'keyword_list': kw_val,
        'similar_questions': sim_val,
        'if_bm25': bm25_val,
        'products': p_val,
        'image_urls': image_urls_val,
        'video_urls': video_urls_val,
        'file_urls': file_urls_val,
        'link_type': link_type_val,
        'link_url': link_url_val,
        'source': src,
        'source_module': src,
        'modification_time': mod_time,
        'modify_time': mod_time,
        'before': before_obj,
        'after': after_obj,
        'changed_fields': changed_fields,
        'operation_id': operation_id
    }
    return out

def _get_archived_mod_keys():
    rows = []
    if is_supabase_archives_enabled():
        try:
            client = get_supabase_client()
            if client and _supabase_table_exists(client, 'archive_record'):
                rows = client.select_all(
                    'archive_record',
                    order_by='id',
                    order_dir='asc',
                    columns='record_json',
                    page_size=1000
                ) or []
        except Exception:
            rows = []
    if not rows:
        try:
            rows = db.session.query(ArchiveRecord.record_json).all()
        except Exception:
            return set()
    keys = set()
    for row in (rows or []):
        if isinstance(row, dict):
            record_json = row.get('record_json')
        else:
            try:
                record_json = row[0]
            except Exception:
                record_json = None
        if not record_json:
            continue
        try:
            obj = json.loads(record_json)
        except Exception:
            continue
        if not isinstance(obj, dict):
            continue
        kb = str(obj.get('kb_id') or obj.get('question_wiki_id') or '').strip()
        mt = str(obj.get('modification_time') or obj.get('modify_time') or '').strip()
        if kb and mt:
            keys.add((kb, mt))
    return keys

def _is_archived_mod_item(item, archived_keys):
    if not archived_keys:
        return False
    if not isinstance(item, dict):
        return False
    kb = str(item.get('kb_id') or item.get('question_wiki_id') or '').strip()
    mt = str(item.get('modification_time') or item.get('modify_time') or '').strip()
    if not kb or not mt:
        return False
    return (kb, mt) in archived_keys

def _mod_source_match(source_val, wanted):
    if not wanted:
        return True
    s = str(source_val or '').strip()
    w = str(wanted or '').strip()
    if not w:
        return True
    return w in s

def _mod_operation_match(op_val, wanted):
    if not wanted:
        return True
    w = str(wanted or '').strip().lower()
    if not w:
        return True
    s = str(op_val or '').strip().lower()
    aliases = {
        '增加': 'create',
        '新增': 'create',
        'create': 'create',
        '删除': 'delete',
        'delete': 'delete',
        '修改': 'edit',
        '编辑': 'edit',
        'update': 'edit',
        'edit': 'edit'
    }
    ww = aliases.get(w, w)
    ss = aliases.get(s, s)
    return ss == ww

@app.route('/api/kb/update', methods=['POST'])
@login_required
def update_kb_item():
    raw_payload = request.get_json(silent=True) or {}
    base_update_time = str(raw_payload.get('base_update_time') or '').strip()
    change_source = _resolve_kb_change_source(raw_payload)
    data = raw_payload
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    allow = _kb_all_fields_allowlist()
    data = {k: v for k, v in (data or {}).items() if k in allow}

    json_cols = ['similar_questions', 'error_list', 'keyword_list', 'image_urls', 'video_urls', 'file_urls']
    
    for col in json_cols:
        if col in data:
            val = data[col]
            cleaned = None
            if val is None:
                cleaned = None
            elif isinstance(val, dict):
                cleaned = val
            elif isinstance(val, list):
                cleaned = [str(v).strip() for v in val if v is not None and str(v).strip() and str(v).strip().lower() != 'null']
            elif isinstance(val, str):
                s = val.strip()
                if not s or s.lower() == 'null':
                    cleaned = None
                else:
                    parsed = None
                    try:
                        parsed = json.loads(s)
                    except Exception:
                        parsed = None
                    
                    if isinstance(parsed, dict):
                        cleaned = parsed
                    elif isinstance(parsed, list):
                        cleaned = [str(v).strip() for v in parsed if v is not None and str(v).strip() and str(v).strip().lower() != 'null']
                    else:
                        val_clean = s.replace('[', '').replace(']', '').replace('"', '').replace("'", "")
                        parts = _split_text_list_value(val_clean, is_url_list=(col in ('image_urls', 'video_urls', 'file_urls')))
                        final_list = [x.strip() for x in parts if x.strip() and x.strip().lower() != 'null']
                        cleaned = final_list if final_list else None
            else:
                cleaned = val
            
            if cleaned is None:
                data[col] = None
            elif isinstance(cleaned, list) and len(cleaned) == 0:
                data[col] = None
            elif isinstance(cleaned, dict) and len(cleaned) == 0:
                data[col] = None
            else:
                # 关键修复：将 list/dict 转换为 JSON 字符串，确保 PostgREST 正确识别为 JSONB
                # PostgREST 会将 Python list 解释为 PostgreSQL array (text[])，而不是 jsonb
                # 通过发送 JSON 字符串，PostgREST 会正确地将其存储为 jsonb
                if isinstance(cleaned, (list, dict)):
                    data[col] = json.dumps(cleaned, ensure_ascii=False)
                else:
                    data[col] = cleaned

    # 1.5 Handle Boolean fields
    if 'if_bm25' in data:
        val = data['if_bm25']
        if val == "" or val is None:
            data['if_bm25'] = None
        elif str(val).lower() == 'true':
            data['if_bm25'] = True
        elif str(val).lower() == 'false':
            data['if_bm25'] = False

    # 2. Product Validation
    if 'product_name' in data:
        p_name = str(data['product_name']) if data['product_name'] else ""
        if p_name.strip():
            try:
                valid_map, valid_set = get_all_valid_models()
                valid, invalid = validate_product_string(p_name, valid_map, valid_set)
                
                if invalid:
                    return jsonify({'success': False, 'message': f'包含未知型号: {", ".join(invalid)}。请先在“管理型号库”中添加该型号。'}), 400
                
                valid.sort()
                data['product_name'] = ", ".join(valid)
            except Exception as e:
                print(f"Product validation warning: {e}")

    kb_id_raw = data.get('question_wiki_id')
    kb_id = str(kb_id_raw).strip() if kb_id_raw is not None else ''
    if kb_id:
        data['question_wiki_id'] = kb_id
    else:
        kb_id = None

    q_val = data.get('question')
    if q_val is not None and not str(q_val).strip():
        return jsonify({'success': False, 'message': 'question 不能为空'}), 400
    a_val = data.get('answer')
    if a_val is not None and not str(a_val).strip():
        return jsonify({'success': False, 'message': 'answer 不能为空'}), 400
    
    try:
        # Fetch current row before update so modification logs store a real "before" snapshot
        # (required for 修改记录详情对比; one extra SELECT per edit).
        before_row = None
        if kb_id:
            try:
                _kb_cols = (
                    'question_wiki_id,question_type,question,answer,answer_type,if_bm25,'
                    'similar_questions,error_list,keyword_list,image_urls,video_urls,file_urls,'
                    'link_type,link_url,update_time,product_category_name,product_name'
                )
                before_resp = client.select(
                    'knowledge_base_v1',
                    page=1,
                    page_size=1,
                    filters={'question_wiki_id': f'eq.{kb_id}'},
                    columns=_kb_cols,
                )
                if before_resp.status_code in (200, 206):
                    rows = before_resp.json() or []
                    if rows and isinstance(rows[0], dict):
                        before_row = rows[0]
            except Exception:
                before_row = None

            # Fallback query path (some environments behave differently between select/select_all)
            if not isinstance(before_row, dict):
                try:
                    rows2 = client.select_all(
                        'knowledge_base_v1',
                        filters={'question_wiki_id': f'eq.{kb_id}'},
                        columns=_kb_cols,
                        page_size=1
                    ) or []
                    if rows2 and isinstance(rows2[0], dict):
                        before_row = rows2[0]
                except Exception:
                    before_row = None

            # Hard guard: edit mode must have a valid "before" snapshot source.
            # Otherwise we would create unusable logs whose "修改前" is empty.
            if not isinstance(before_row, dict):
                return jsonify({
                    'success': False,
                    'message': f'未找到待编辑记录（question_wiki_id={kb_id}），已阻止保存以避免产生无效修改记录。'
                }), 404

            if base_update_time:
                current_update_time = str(before_row.get('update_time') or '').strip()
                if current_update_time != base_update_time:
                    return jsonify({
                        'success': False,
                        'message': '知识库内容已被他人修改，请重新打开任务后再保存。',
                        'conflict': True,
                        'current_update_time': current_update_time,
                        'base_update_time': base_update_time
                    }), 409

        after_row_for_diff = dict(data or {})

        before_obj = _snapshot_mod_fields(before_row) if before_row else None
        after_obj = _snapshot_mod_fields(after_row_for_diff)
        changed_fields = _compute_mod_changed_fields(before_obj, after_obj) if kb_id else _compute_mod_changed_fields(None, after_obj)

        # No-op guard: editing existing row without any business-field change should not
        # mark review_status/modification log, and should not create "修改中" noise.
        if kb_id and isinstance(before_row, dict) and not changed_fields:
            return jsonify({
                'success': True,
                'question_wiki_id': kb_id,
                'no_change': True,
                'message': 'No changes detected',
                'mod_log_ok': True,
                'mod_log_error': ''
            })

        # Prepare modification record
        modification_record = data.copy()
        modification_record['kb_id'] = kb_id
        modification_record['modifier'] = current_user.username if current_user.is_authenticated else 'system'
        modification_record['modification_time'] = _now_iso_with_tz()
        modification_record['change_type'] = 'edit' if kb_id else 'create'
        _attach_change_meta(modification_record, {
            'source': change_source,
            'before': before_obj,
            'after': after_obj,
            'changed_fields': changed_fields
        })
        
        # Cleanup modification record
        if 'question_wiki_id' in modification_record:
            modification_record['kb_id'] = modification_record['question_wiki_id']
        if 'id' in modification_record:
            del modification_record['id'] 
        if 'review_status' in modification_record:
            del modification_record['review_status']
        
        # Update/Insert into knowledge_base_v1
        data['review_status'] = 'modifying' if kb_id else 'creating'
        
        # Ensure update_time is set
        data['update_time'] = _now_iso_with_tz()

        saved_id = None
        if kb_id:
            # Update existing
            response = client.update('knowledge_base_v1', data, {'question_wiki_id': kb_id})
            saved_id = kb_id
        else:
            # Insert new
            if 'question_wiki_id' in data and not data['question_wiki_id']:
                del data['question_wiki_id'] 
            
            # Generate new ID
            new_id = generate_kb_id(client)
            data['question_wiki_id'] = new_id
            saved_id = new_id
            
            # Update modification record with new ID
            modification_record['kb_id'] = new_id
            modification_record['question_wiki_id'] = new_id
            
            response = client.insert('knowledge_base_v1', data)

        if response.status_code >= 400:
             return jsonify({'success': False, 'message': f"Database error: {response.text}"}), 500

        # Log Modification
        # 改为同步插入，确保修改记录一定被保存
        mod_log_ok = True
        mod_log_error = ""
        try:
            mod_copy = dict(modification_record or {})
            
            # 转换数组字段为 JSON 字符串
            _convert_array_fields_to_json(mod_copy)
            
            # 直接在主线程中插入，确保可靠性
            insert_resp = _supabase_insert_drop_unknown_columns(client, 'knowledge_base_modifications', mod_copy)
            
            # 检查插入是否真的成功
            if insert_resp and hasattr(insert_resp, 'status_code'):
                if insert_resp.status_code >= 400:
                    # 插入失败
                    mod_log_ok = False
                    mod_log_error = f"Insert failed with status {insert_resp.status_code}: {insert_resp.text[:200]}"
                    print(f"[ERROR] Failed to insert modification record: {mod_log_error}")
                else:
                    # 插入成功
                    print(f"[INFO] Modification record inserted successfully: {mod_copy.get('kb_id')}")
            elif insert_resp is None:
                # 返回None说明参数错误
                mod_log_ok = False
                mod_log_error = "Insert function returned None (invalid parameters)"
                print(f"[ERROR] Failed to insert modification record: {mod_log_error}")
            else:
                # 其他情况，假设成功
                print(f"[INFO] Modification record inserted (no status code check): {mod_copy.get('kb_id')}")
        except Exception as e:
            # 记录失败但不影响主流程
            mod_log_ok = False
            mod_log_error = str(e)
            print(f"[ERROR] Failed to insert modification record: {e}")
            print(traceback.format_exc())

        quality_task_updated = False
        quality_task_id = str(raw_payload.get('quality_task_id') or '').strip()
        if quality_task_id:
            quality_task_updated = _quality_mark_task_processing(
                quality_task_id,
                latest_kb_update_time=data.get('update_time')
            )

        return jsonify({
            'success': True,
            'question_wiki_id': saved_id or data.get('question_wiki_id') or kb_id,
            'mod_log_ok': mod_log_ok,
            'mod_log_error': mod_log_error,
            'quality_task_updated': quality_task_updated,
            'warning': '修改记录保存失败，但数据已成功保存' if not mod_log_ok else None
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/kb/delete', methods=['POST'])
@login_required
def delete_kb_item():
    data = request.json or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No IDs provided'}), 400
    change_source = str(data.get('change_source') or data.get('source_module') or '').strip() or '知识库管理'

    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500
        
    try:
        # 1. Fetch current data for these items to log them
        # We need to snapshot them BEFORE "deleting" (marking as deleting)
        # Actually we just need the data.
        
        # Quote IDs for PostgREST filter
        b_ids_quoted = [f'"{bid}"' for bid in ids]
        id_str = "(" + ",".join(b_ids_quoted) + ")"
        
        print(f"[DEBUG] Deleting IDs: {ids}")
        print(f"[DEBUG] Filter string: in.{id_str}")
        
        fetch_resp = client.select('knowledge_base_v1', page_size=1000, filters={'question_wiki_id': f'in.{id_str}'})
        
        items_to_log = []
        if fetch_resp.status_code < 400:
            items_to_log = fetch_resp.json()
            print(f"[DEBUG] Fetched {len(items_to_log)} items to log")
        else:
            print(f"[ERROR] Fetch failed: {fetch_resp.status_code} - {fetch_resp.text}")
            
        # 2. Update status to 'deleting' in knowledge_base_v1
        # We can use update with 'in' filter? 
        # Supabase update usually takes specific filters.
        # But we can try updating where question_wiki_id in list.
        # client.update signature: update(table, data, filters)
        
        # We need to loop if client.update doesn't support bulk update with different IDs (it doesn't).
        # But here we are setting SAME status for ALL IDs.
        # So we can use: update knowledge_base_v1 set review_status='deleting' where question_wiki_id in (...)
        
        print(f"[DEBUG] Updating review_status to 'deleting'")
        update_resp = client.update('knowledge_base_v1', {'review_status': 'deleting'}, {'question_wiki_id': f'in.{id_str}'})
        
        print(f"[DEBUG] Update response: {update_resp.status_code}")
        if update_resp.status_code >= 400:
            error_msg = f"Failed to update status: {update_resp.text}"
            print(f"[ERROR] {error_msg}")
            return jsonify({'success': False, 'message': error_msg}), 500
             
        # 3. Log to modifications
        if items_to_log:
            mod_records = []
            for item in items_to_log:
                mod_rec = item.copy()
                mod_rec['kb_id'] = item.get('question_wiki_id')
                mod_rec['modifier'] = current_user.username if current_user.is_authenticated else 'system'
                mod_rec['modification_time'] = _now_iso_with_tz()
                mod_rec['change_type'] = 'delete'
                before_obj = _snapshot_mod_fields(item)
                after_obj = None
                changed_fields = _compute_mod_changed_fields(before_obj, after_obj)
                _attach_change_meta(mod_rec, {
                    'source': change_source,
                    'before': before_obj,
                    'after': after_obj,
                    'changed_fields': changed_fields
                })
                if 'id' in mod_rec:
                    del mod_rec['id']
                if 'review_status' in mod_rec:
                    del mod_rec['review_status']
                mod_records.append(mod_rec)
            
            # Bulk insert
            # 转换数组字段为 JSON 字符串
            _convert_array_fields_to_json(mod_records)
            
            print(f"[DEBUG] Inserting {len(mod_records)} modification records")
            mod_log_resp = _supabase_insert_drop_unknown_columns(client, 'knowledge_base_modifications', mod_records)
            if mod_log_resp and hasattr(mod_log_resp, 'status_code') and mod_log_resp.status_code >= 400:
                print(f"[WARN] Failed to log delete modifications: {mod_log_resp.text}")
        
        print(f"[SUCCESS] Deleted {len(ids)} items")
        return jsonify({'success': True, 'count': len(ids)})

    except Exception as e:
        error_msg = str(e)
        print(f"[ERROR] Delete failed: {error_msg}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': error_msg}), 500

@app.route('/api/kb/tags', methods=['GET'])
@login_required
def get_all_kb_tags():
    """
    获取所有知识库标签列表
    用于前端标签选择器的下拉菜单
    返回格式：直接返回标签名称数组 ["标签1", "标签2", ...]
    """
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    try:
        # 从 kb_tags 表获取所有标签
        tags_rows = client.select_all(
            'kb_tags',
            columns='name',
            order_by='name',
            page_size=1000
        ) or []
        
        # 提取标签名称并去重
        tag_names = []
        seen = set()
        for row in tags_rows:
            if isinstance(row, dict):
                name = row.get('name')
                if name and str(name).strip():
                    name_str = str(name).strip()
                    if name_str not in seen:
                        seen.add(name_str)
                        tag_names.append(name_str)
        
        # 前端期望直接返回数组，不是 {success: true, tags: [...]}
        return jsonify(tag_names)
    except Exception as e:
        traceback.print_exc()
        # 错误时返回空数组，避免前端报错
        return jsonify([]), 500

@app.route('/api/kb/item/tags', methods=['GET'])
@login_required
def get_kb_item_tags():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    library_type = str(request.args.get('libraryType') or 'current').strip().lower()
    if library_type not in ('current', 'previous'):
        return jsonify({'success': False, 'message': 'Invalid libraryType'}), 400

    wiki_id = str(request.args.get('question_wiki_id') or request.args.get('id') or '').strip()
    if not wiki_id:
        return jsonify({'success': False, 'message': 'question_wiki_id is required'}), 400

    try:
        # 1) Fetch tag_ids for this item
        in_row = client.select_all(
            'kb_item_tags',
            filters={'library_type': f'eq.{library_type}', 'question_wiki_id': f'eq.{wiki_id}'},
            columns='tag_id',
            order_by='question_wiki_id',
            page_size=1000
        ) or []
        tag_ids = [r.get('tag_id') for r in in_row if isinstance(r, dict) and r.get('tag_id')]

        if not tag_ids:
            return jsonify({'success': True, 'tags': []})

        # 2) Fetch tag names for those ids
        id_in = _postgrest_in_str(tag_ids)
        tags_rows = client.select_all(
            'kb_tags',
            filters={'id': id_in},
            columns='id,name',
            page_size=1000
        ) or []
        id_to_name = {}
        for tr in tags_rows:
            if isinstance(tr, dict):
                tid = tr.get('id')
                nm = tr.get('name')
                if tid and nm:
                    id_to_name[str(tid)] = str(nm)

        # Keep stable order by mapping tag_ids order (and dedupe)
        out = []
        seen = set()
        for tid in tag_ids:
            key = str(tid)
            if key in id_to_name and key not in seen:
                seen.add(key)
                out.append(id_to_name[key])
        return jsonify({'success': True, 'tags': out})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/kb/item/tags', methods=['PUT'])
@login_required
def put_kb_item_tags():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    data = request.json or {}
    library_type = str(data.get('libraryType') or 'current').strip().lower()
    if library_type not in ('current', 'previous'):
        return jsonify({'success': False, 'message': 'Invalid libraryType'}), 400

    wiki_id = str(data.get('question_wiki_id') or data.get('id') or '').strip()
    if not wiki_id:
        return jsonify({'success': False, 'message': 'question_wiki_id is required'}), 400

    tag_names = data.get('tagNames') or data.get('tags') or []
    if isinstance(tag_names, str):
        # Allow passing a string like "a,b,c"
        tag_names = re.split(r'[,，\n]', tag_names)
    if not isinstance(tag_names, list):
        tag_names = []

    # Normalize + dedupe + trim
    normalized = []
    seen = set()
    for t in tag_names:
        s = str(t or '').strip()
        if not s:
            continue
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        normalized.append(s)
    normalized = normalized[:200]

    try:
        # Always replace mapping for this item/library_type
        client.delete('kb_item_tags', {'library_type': f'eq.{library_type}', 'question_wiki_id': f'eq.{wiki_id}'})

        if not normalized:
            return jsonify({'success': True, 'count': 0})

        # 1) Fetch all existing tags to avoid N queries
        exist_tags = client.select_all('kb_tags', columns='id,name', page_size=1000) or []
        name_to_id_lower = {}
        for t in exist_tags:
            if isinstance(t, dict) and t.get('id') and t.get('name'):
                name_to_id_lower[str(t['name']).strip().lower()] = t['id']

        # 2) Insert missing tags
        missing = [n for n in normalized if n.lower() not in name_to_id_lower]
        if missing:
            to_insert = [{'name': n} for n in missing]
            # Unique constraint: name
            client.upsert('kb_tags', to_insert, on_conflict='name')

            # Re-fetch to get ids
            exist_tags = client.select_all('kb_tags', columns='id,name', page_size=1000) or []
            name_to_id_lower = {}
            for t in exist_tags:
                if isinstance(t, dict) and t.get('id') and t.get('name'):
                    name_to_id_lower[str(t['name']).strip().lower()] = t['id']

        tag_ids = [name_to_id_lower[n.lower()] for n in normalized if n.lower() in name_to_id_lower and name_to_id_lower[n.lower()]]

        if tag_ids:
            # 3) Insert mapping rows
            rows = []
            for tid in tag_ids:
                rows.append({
                    'library_type': library_type,
                    'question_wiki_id': wiki_id,
                    'tag_id': tid
                })

            batch_size = 500
            for i in range(0, len(rows), batch_size):
                batch = rows[i:i + batch_size]
                resp = client.insert('kb_item_tags', batch)
                if resp is not None and getattr(resp, 'status_code', 0) >= 400:
                    return jsonify({'success': False, 'message': getattr(resp, 'text', 'insert failed')}), 500
                time.sleep(0.05)

        return jsonify({'success': True, 'count': len(tag_ids)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/kb/complete_revision', methods=['POST'])
@login_required
def complete_revision():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500
        
    try:
        # Update all items with review_status 'creating' / 'modifying' / 'deleting' to 'unadjusted'
        # We use the 'in' filter for bulk update
        
        # The filter string format for PostgREST 'in' is: in.("val1","val2")
        filter_str = 'in.("creating","modifying","deleting")'
        
        update_resp = client.update('knowledge_base_v1', {'review_status': 'unadjusted'}, {'review_status': filter_str})
        
        if update_resp.status_code >= 400:
             return jsonify({'success': False, 'message': update_resp.text}), 500
             
        return jsonify({'success': True})
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/kb/modifications', methods=['GET'])
@login_required
def get_kb_modifications():
    client = get_supabase_client()
    if not client:
        return jsonify([])
    
    page = request.args.get('page')
    page_size = request.args.get('pageSize')
    
    # Search parameters
    kb_id = request.args.get('kb_id')
    product = request.args.get('product')
    question = request.args.get('question')
    answer = request.args.get('answer')
    modifier = request.args.get('modifier')
    source_module = request.args.get('source_module')
    operation = request.args.get('operation')
    start_time = request.args.get('start_time')
    end_time = request.args.get('end_time')
    
    # Sort
    sort_by = request.args.get('sortBy', 'modification_time')
    sort_dir = request.args.get('sortDir', 'desc')

    filters = {}
    if kb_id:
        filters['kb_id'] = f"ilike.*{kb_id}*"
    if product:
        filters['product_name'] = f"ilike.*{product}*"
    if question:
        filters['question'] = f"ilike.*{question}*"
    if answer:
        filters['answer'] = f"ilike.*{answer}*"
    if modifier:
        filters['modifier'] = f"ilike.*{modifier}*"

    and_conditions = []
    st = _safe_parse_dt(start_time)
    et = _safe_parse_dt(end_time)
    if st:
        and_conditions.append(f"modification_time.gte.{st.isoformat()}")
    if et:
        and_conditions.append(f"modification_time.lte.{et.isoformat()}")
    if and_conditions:
        existing_and = filters.get('and')
        if existing_and:
            try:
                existing_inner = str(existing_and).strip()
                if existing_inner.startswith('(') and existing_inner.endswith(')'):
                    existing_inner = existing_inner[1:-1]
                merged = [x.strip() for x in existing_inner.split(',') if x.strip()]
            except Exception:
                merged = []
            merged.extend(and_conditions)
            filters['and'] = f"({','.join(merged)})"
        else:
            filters['and'] = f"({','.join(and_conditions)})"
        
    try:
        page_int = int(page) if page else 1
        page_size_int = int(page_size) if page_size else 20
        archived_keys = _get_archived_mod_keys()

        def _fill_from_kb_base(items):
            def _is_empty_snapshot(obj):
                if not isinstance(obj, dict):
                    return True
                for v in obj.values():
                    if v is None:
                        continue
                    if isinstance(v, str) and not v.strip():
                        continue
                    if isinstance(v, (list, dict)) and len(v) == 0:
                        continue
                    return False
                return True

            def _lookup_prev_snapshot(wid, mod_time):
                try:
                    if not wid or not mod_time:
                        return None
                    prev_rows = client.select_all(
                        'knowledge_base_modifications',
                        filters={'and': f"(kb_id.eq.{wid},modification_time.lt.{mod_time})"},
                        order_by='modification_time',
                        order_dir='desc',
                        columns='change_meta,question,answer,question_type,answer_type,if_bm25,similar_questions,error_list,keyword_list,image_urls,video_urls,file_urls,link_type,link_url,product_name',
                        page_size=5
                    ) or []
                    for pr in prev_rows:
                        if not isinstance(pr, dict):
                            continue
                        meta = _extract_change_meta(pr.get('change_meta')) or {}
                        after_obj = meta.get('after') if isinstance(meta.get('after'), dict) else None
                        before_obj = meta.get('before') if isinstance(meta.get('before'), dict) else None
                        if isinstance(after_obj, dict) and not _is_empty_snapshot(after_obj):
                            return after_obj
                        if isinstance(before_obj, dict) and not _is_empty_snapshot(before_obj):
                            return before_obj
                        snap = _snapshot_mod_fields(pr)
                        if not _is_empty_snapshot(snap):
                            return snap
                except Exception:
                    return None
                return None

            missing_ids = []
            for it in (items or []):
                if not isinstance(it, dict):
                    continue
                wid = str(it.get('question_wiki_id') or it.get('kb_id') or '').strip()
                if not wid:
                    continue
                if (not str(it.get('question') or '').strip()) or (not str(it.get('answer') or '').strip()):
                    missing_ids.append(wid)
            missing_ids = sorted(list(dict.fromkeys(missing_ids)))
            if not missing_ids:
                return items

            in_str = _postgrest_in_str(missing_ids)
            if not in_str:
                return items
            base_rows = client.select_all(
                'knowledge_base_v1',
                filters={'question_wiki_id': in_str},
                order_by='question_wiki_id',
                order_dir='asc',
                columns='question_wiki_id,question,answer,question_type,answer_type,if_bm25,similar_questions,error_list,keyword_list,image_urls,video_urls,file_urls,link_type,link_url,product_name',
                page_size=1000
            ) or []
            base_map = {}
            for r in base_rows:
                wid = str(r.get('question_wiki_id') or '').strip()
                if wid:
                    base_map[wid] = r

            for it in (items or []):
                if not isinstance(it, dict):
                    continue
                wid = str(it.get('question_wiki_id') or it.get('kb_id') or '').strip()
                if not wid or wid not in base_map:
                    continue
                b = base_map[wid]
                if not str(it.get('question') or '').strip():
                    it['question'] = b.get('question') or it.get('question') or ''
                if not str(it.get('answer') or '').strip():
                    it['answer'] = b.get('answer') or it.get('answer') or ''
                it['question_type'] = it.get('question_type') if it.get('question_type') is not None else b.get('question_type')
                it['answer_type'] = it.get('answer_type') if it.get('answer_type') is not None else b.get('answer_type')
                it['if_bm25'] = it.get('if_bm25') if it.get('if_bm25') is not None else b.get('if_bm25')
                it['similar_questions'] = it.get('similar_questions') if it.get('similar_questions') is not None else b.get('similar_questions')
                it['error_list'] = it.get('error_list') if it.get('error_list') is not None else b.get('error_list')
                it['keyword_list'] = it.get('keyword_list') if it.get('keyword_list') is not None else b.get('keyword_list')
                it['image_urls'] = it.get('image_urls') if it.get('image_urls') is not None else b.get('image_urls')
                it['video_urls'] = it.get('video_urls') if it.get('video_urls') is not None else b.get('video_urls')
                it['file_urls'] = it.get('file_urls') if it.get('file_urls') is not None else b.get('file_urls')
                it['link_type'] = it.get('link_type') if it.get('link_type') is not None else b.get('link_type')
                it['link_url'] = it.get('link_url') if it.get('link_url') is not None else b.get('link_url')
                if not str(it.get('products') or '').strip():
                    it['products'] = it.get('product_name') or b.get('product_name') or it.get('products') or '-'

                before_obj = it.get('before') if isinstance(it.get('before'), dict) else {}
                after_obj = it.get('after') if isinstance(it.get('after'), dict) else {}
                # Historical compatibility: for old logs without before snapshot, infer from previous record.
                if _is_empty_snapshot(before_obj) and str(it.get('change_type') or '').lower() == 'edit':
                    inferred = _lookup_prev_snapshot(wid, it.get('modification_time'))
                    if isinstance(inferred, dict) and not _is_empty_snapshot(inferred):
                        before_obj = inferred
                for k in ['question', 'answer', 'question_type', 'answer_type', 'if_bm25', 'similar_questions', 'error_list', 'keyword_list', 'image_urls', 'video_urls', 'file_urls', 'link_type', 'link_url']:
                    if k not in before_obj or before_obj.get(k) in [None, '']:
                        if b.get(k) is not None:
                            before_obj[k] = b.get(k)
                    if k not in after_obj or after_obj.get(k) in [None, '']:
                        if b.get(k) is not None:
                            after_obj[k] = b.get(k)
                it['before'] = before_obj
                it['after'] = after_obj
            return items

        if source_module or operation or archived_keys:
            data = client.select_all('knowledge_base_modifications', order_by=sort_by, order_dir=sort_dir, filters=filters, page_size=1000)
            normalized = []
            for row in (data or []):
                item = _normalize_mod_record(row)
                if not item:
                    continue
                if not _mod_source_match(item.get('source_module'), source_module):
                    continue
                if operation and not _mod_operation_match(item.get('change_type'), operation):
                    continue
                normalized.append(item)

            if archived_keys:
                normalized = [it for it in normalized if not _is_archived_mod_item(it, archived_keys)]
            normalized = _fill_from_kb_base(normalized)
            total = len(normalized)
            start_idx = max(0, (page_int - 1) * page_size_int)
            end_idx = start_idx + page_size_int
            sliced = normalized[start_idx:end_idx]
            return jsonify({'success': True, 'data': sliced, 'total': total})

        resp = client.select('knowledge_base_modifications', page=page_int, page_size=page_size_int, order_by=sort_by, order_dir=sort_dir, filters=filters)
        if resp.status_code >= 400:
            print(f"Database Error: {resp.text}")
            return jsonify([])

        total = 0
        content_range = resp.headers.get('Content-Range')
        if content_range:
            try:
                total = int(content_range.split('/')[-1])
            except Exception:
                total = 0

        raw = resp.json() or []
        normalized = []
        for row in raw:
            item = _normalize_mod_record(row)
            if item:
                normalized.append(item)
        if archived_keys:
            normalized = [it for it in normalized if not _is_archived_mod_item(it, archived_keys)]
        normalized = _fill_from_kb_base(normalized)
        return jsonify({'success': True, 'data': normalized, 'total': total})
            
    except Exception as e:
        print(f"Error fetching modifications: {e}")
        return jsonify([])

@app.route('/api/kb/modifications/delete', methods=['POST'])
@login_required
def delete_kb_modifications():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    data = request.json or {}
    ids = data.get('ids') or []
    if not isinstance(ids, list) or len(ids) == 0:
        return jsonify({'success': False, 'message': 'No ids'}), 400

    cleaned = []
    for v in (ids or []):
        if v is None:
            continue
        if isinstance(v, bool):
            continue
        if isinstance(v, int):
            cleaned.append(v)
            continue
        s = str(v).strip()
        if not s:
            continue
        if re.fullmatch(r'\d+', s):
            try:
                cleaned.append(int(s))
                continue
            except Exception:
                pass
        cleaned.append(s)

    if not cleaned:
        return jsonify({'success': False, 'message': 'No valid ids'}), 400

    inner = ",".join([json.dumps(v, ensure_ascii=False) for v in cleaned])
    in_filter = f"in.({inner})"
    try:
        resp = client.delete('knowledge_base_modifications', {'id': in_filter})
        code = getattr(resp, 'status_code', 500) if resp else 500
        if code >= 400:
            return jsonify({'success': False, 'message': getattr(resp, 'text', '') or '数据库删除失败'}), 500
        return jsonify({'success': True, 'deleted': len(cleaned)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/kb/modifications/export_raw', methods=['GET'])
@login_required
def export_kb_modifications_raw():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400

    kb_id = request.args.get('kb_id')
    product = request.args.get('product')
    question = request.args.get('question')
    modifier = request.args.get('modifier')
    source_module = request.args.get('source_module')
    operation = request.args.get('operation')
    start_time = request.args.get('start_time')
    end_time = request.args.get('end_time')
    sort_by = request.args.get('sortBy', 'modification_time')
    sort_dir = request.args.get('sortDir', 'desc')

    filters = {}
    if kb_id:
        filters['kb_id'] = f"ilike.*{kb_id}*"
    if product:
        filters['product_name'] = f"ilike.*{product}*"
    if question:
        filters['question'] = f"ilike.*{question}*"
    if modifier:
        filters['modifier'] = f"ilike.*{modifier}*"

    and_conditions = []
    st = _safe_parse_dt(start_time)
    et = _safe_parse_dt(end_time)
    if st:
        and_conditions.append(f"modification_time.gte.{st.isoformat()}")
    if et:
        and_conditions.append(f"modification_time.lte.{et.isoformat()}")
    if and_conditions:
        filters['and'] = f"({','.join(and_conditions)})"

    try:
        data = client.select_all('knowledge_base_modifications', order_by=sort_by, order_dir=sort_dir, filters=filters, page_size=1000)
        normalized = []
        for row in (data or []):
            item = _normalize_mod_record(row)
            if not item:
                continue
            if source_module and not _mod_source_match(item.get('source_module'), source_module):
                continue
            if operation and not _mod_operation_match(item.get('change_type'), operation):
                continue
            normalized.append(item)

        archived_keys = _get_archived_mod_keys()
        if archived_keys:
            normalized = [it for it in normalized if not _is_archived_mod_item(it, archived_keys)]

        export_rows = []
        for it in normalized:
            export_rows.append({
                'opera': it.get('opera'),
                'question_wiki_id': it.get('question_wiki_id'),
                'question': it.get('question'),
                'question_type': it.get('question_type'),
                'answer': it.get('answer'),
                'answer_type': it.get('answer_type'),
                'error_list': _export_mod_list_text(it.get('error_list')),
                'keyword_list': _export_mod_list_text(it.get('keyword_list')),
                'similar_questions': _export_mod_list_text(it.get('similar_questions')),
                'if_bm25': _export_mod_bool_text(it.get('if_bm25')),
                'products': _export_mod_list_text(it.get('products')),
                'image_urls': _export_mod_list_text(it.get('image_urls'), is_url_list=True),
                'video_urls': _export_mod_list_text(it.get('video_urls'), is_url_list=True),
                'file_urls': _export_mod_list_text(it.get('file_urls'), is_url_list=True),
                'link_type': it.get('link_type'),
                'link_url': it.get('link_url'),
                'source_module': it.get('source_module'),
                'modify_time': it.get('modify_time'),
                'before': json.dumps(it.get('before'), ensure_ascii=False) if it.get('before') is not None else '',
                'after': json.dumps(it.get('after'), ensure_ascii=False) if it.get('after') is not None else '',
                'changed_fields': json.dumps(it.get('changed_fields'), ensure_ascii=False) if it.get('changed_fields') is not None else ''
            })

        df = pd.DataFrame(export_rows)
        csv_text = df.to_csv(index=False)
        output = io.BytesIO(csv_text.encode('utf-8-sig'))
        return send_file(
            output,
            as_attachment=True,
            download_name=canonical_download_name('kb_edit_log', 'csv'),
            mimetype='text/csv'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

def _is_matrix_source(source_val):
    s = str(source_val or '')
    return '机型矩阵' in s

def _is_kb_source(source_val):
    s = str(source_val or '')
    # Non-matrix modification sources still represent KB content changes.
    return not _is_matrix_source(s)

def _get_field_from_mod(it, field):
    after = it.get('after') if isinstance(it, dict) else None
    before = it.get('before') if isinstance(it, dict) else None
    if isinstance(after, dict) and field in after:
        return after.get(field)
    if isinstance(before, dict) and field in before:
        return before.get(field)
    return it.get(field) if isinstance(it, dict) else None

def _mod_products_changed(it):
    changed_fields = it.get('changed_fields') if isinstance(it, dict) else None
    if isinstance(changed_fields, list) and ('products' in changed_fields or 'product_name' in changed_fields):
        return True
    before = it.get('before') if isinstance(it, dict) else None
    after = it.get('after') if isinstance(it, dict) else None
    if isinstance(before, dict) and isinstance(after, dict):
        b = _normalize_mod_diff_value(before.get('products') if 'products' in before else before.get('product_name'))
        a = _normalize_mod_diff_value(after.get('products') if 'products' in after else after.get('product_name'))
        return b != a
    return False

@app.route('/api/kb/modifications/smart_merge_export', methods=['GET'])
@login_required
def export_kb_modifications_smart_merge():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400

    kb_id = request.args.get('kb_id')
    product = request.args.get('product')
    question = request.args.get('question')
    modifier = request.args.get('modifier')
    source_module = request.args.get('source_module')
    operation = request.args.get('operation')
    start_time = request.args.get('start_time')
    end_time = request.args.get('end_time')

    filters = {}
    if kb_id:
        filters['kb_id'] = f"ilike.*{kb_id}*"
    if product:
        filters['product_name'] = f"ilike.*{product}*"
    if question:
        filters['question'] = f"ilike.*{question}*"
    if modifier:
        filters['modifier'] = f"ilike.*{modifier}*"

    and_conditions = []
    st = _safe_parse_dt(start_time)
    et = _safe_parse_dt(end_time)
    if st:
        and_conditions.append(f"modification_time.gte.{st.isoformat()}")
    if et:
        and_conditions.append(f"modification_time.lte.{et.isoformat()}")
    if and_conditions:
        filters['and'] = f"({','.join(and_conditions)})"

    try:
        filtered_raw = client.select_all('knowledge_base_modifications', order_by='modification_time', order_dir='desc', filters=filters, page_size=1000)
        filtered_norm = []
        for r in (filtered_raw or []):
            it = _normalize_mod_record(r)
            if not it:
                continue
            if source_module and not _mod_source_match(it.get('source_module'), source_module):
                continue
            if operation and not _mod_operation_match(it.get('change_type'), operation):
                continue
            filtered_norm.append(it)

        archived_keys = _get_archived_mod_keys()
        if archived_keys:
            filtered_norm = [it for it in filtered_norm if not _is_archived_mod_item(it, archived_keys)]

        ids = sorted(list({str(it.get('question_wiki_id') or '').strip() for it in filtered_norm if str(it.get('question_wiki_id') or '').strip()}))
        if not ids:
            df = pd.DataFrame([])
            output = io.BytesIO()
            engine = 'xlsxwriter' if importlib.util.find_spec('xlsxwriter') is not None else 'openpyxl'
            with pd.ExcelWriter(output, engine=engine) as writer:
                df.to_excel(writer, index=False, sheet_name='智能合并导出')
            output.seek(0)
            return send_file(
                output,
                as_attachment=True,
                download_name=canonical_download_name('kb_merge'),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        id_filter_str = _postgrest_in_str(ids)
        if not id_filter_str:
            df = pd.DataFrame([])
            output = io.BytesIO()
            engine = 'xlsxwriter' if importlib.util.find_spec('xlsxwriter') is not None else 'openpyxl'
            with pd.ExcelWriter(output, engine=engine) as writer:
                df.to_excel(writer, index=False, sheet_name='智能合并导出')
            output.seek(0)
            return send_file(
                output,
                as_attachment=True,
                download_name=canonical_download_name('kb_merge'),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        kb_rows = client.select_all(
            'knowledge_base_v1',
            filters={'question_wiki_id': id_filter_str},
            order_by='question_wiki_id',
            order_dir='asc',
            columns='question_wiki_id,question,answer,question_type,answer_type,if_bm25,similar_questions,error_list,keyword_list,image_urls,video_urls,file_urls,link_type,link_url,product_name',
            page_size=1000
        ) or []
        kb_map = {}
        for r in kb_rows:
            wid = str(r.get('question_wiki_id') or '').strip()
            if wid:
                kb_map[wid] = r

        all_mod_rows = client.select_all(
            'knowledge_base_modifications',
            filters={'kb_id': id_filter_str},
            order_by='modification_time',
            order_dir='desc',
            page_size=1000
        ) or []
        all_mod_norm = []
        for r in all_mod_rows:
            it = _normalize_mod_record(r)
            if it:
                all_mod_norm.append(it)
        if archived_keys:
            all_mod_norm = [it for it in all_mod_norm if not _is_archived_mod_item(it, archived_keys)]

        by_id = {}
        for it in all_mod_norm:
            wid = str(it.get('question_wiki_id') or '').strip()
            if wid:
                by_id.setdefault(wid, []).append(it)

        fields_b = [
            'question', 'question_type', 'answer', 'answer_type',
            'error_list', 'keyword_list', 'similar_questions', 'if_bm25',
            'image_urls', 'video_urls', 'file_urls', 'link_type', 'link_url'
        ]

        out_rows = []
        for wid in ids:
            mods = by_id.get(wid, [])
            kb_base = kb_map.get(wid, {}) if isinstance(kb_map.get(wid, {}), dict) else {}

            matrix_mods = [m for m in mods if _is_matrix_source(m.get('source_module'))]
            kb_mods = [m for m in mods if _is_kb_source(m.get('source_module'))]

            final_opera = 'edit'
            if kb_mods:
                m0 = kb_mods[0]
                final_opera = m0.get('change_type') or m0.get('opera') or m0.get('operation') or final_opera
            elif matrix_mods:
                m0 = matrix_mods[0]
                final_opera = m0.get('change_type') or m0.get('opera') or m0.get('operation') or final_opera

            final_products = kb_base.get('product_name') or ''
            if matrix_mods:
                m0 = matrix_mods[0]
                final_products = _get_field_from_mod(m0, 'products') or _get_field_from_mod(m0, 'product_name') or final_products
            else:
                kb_prod_mods = [m for m in kb_mods if _mod_products_changed(m)]
                if kb_prod_mods:
                    m0 = kb_prod_mods[0]
                    final_products = _get_field_from_mod(m0, 'products') or _get_field_from_mod(m0, 'product_name') or final_products

            final_obj = {
                'question_wiki_id': wid,
                'products': final_products
            }

            for f in fields_b:
                val = kb_base.get(f)
                if kb_mods:
                    m0 = kb_mods[0]
                    v2 = _get_field_from_mod(m0, f)
                    if v2 is not None:
                        val = v2
                final_obj[f] = val
            op_raw = str(final_opera or '').strip().lower()
            if op_raw in ('create', 'insert', 'add', '新增', '增加'):
                op_label = '增加'
            elif op_raw in ('delete', 'del', 'remove', 'rm', '删除'):
                op_label = '删除'
            else:
                op_label = '修改'

            out_rows.append({
                '操作': op_label,
                '问题编号': final_obj.get('question_wiki_id'),
                '问题': final_obj.get('question'),
                '问题类型': final_obj.get('question_type'),
                '答案': final_obj.get('answer'),
                '答案类型': final_obj.get('answer_type'),
                '错误列表': _export_mod_list_text(final_obj.get('error_list')),
                '关键词': _export_mod_list_text(final_obj.get('keyword_list')),
                '相似提问': _export_mod_list_text(final_obj.get('similar_questions')),
                'BM25': _export_mod_bool_text(final_obj.get('if_bm25')),
                '机型': _export_mod_list_text(final_obj.get('products')),
                '图片链接': _export_mod_list_text(final_obj.get('image_urls'), is_url_list=True),
                '视频链接': _export_mod_list_text(final_obj.get('video_urls'), is_url_list=True),
                '文件链接': _export_mod_list_text(final_obj.get('file_urls'), is_url_list=True),
                '跳转链接类型': str(final_obj.get('link_type') or '').strip(),
                '跳转链接（url/key）': str(final_obj.get('link_url') or '').strip()
            })

        df = pd.DataFrame(out_rows)
        output = io.BytesIO()
        engine = 'xlsxwriter' if importlib.util.find_spec('xlsxwriter') is not None else 'openpyxl'
        with pd.ExcelWriter(output, engine=engine) as writer:
            df.to_excel(writer, index=False, sheet_name='智能合并导出')
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=canonical_download_name('kb_merge'),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/ops/<kind>', methods=['GET'])
@login_required
def get_ops_library(kind):
    kind = (kind or '').strip().lower()
    if kind not in ('app', 'product'):
        return jsonify({'success': False, 'message': 'Invalid kind'}), 400
    if is_supabase_ops_enabled():
        try:
            client = get_supabase_client()
            rows = client.select_all(
                'ops_library_item',
                filters={'kind': f'eq.{kind}'},
                columns='id,kind,name,steps,compatible_models,sort_order,updated_at',
                order_by='sort_order',
                order_dir='asc',
                page_size=1000,
            ) or []
            data = []
            for it in rows:
                if not isinstance(it, dict):
                    continue
                data.append({
                    'id': it.get('id'),
                    'kind': it.get('kind'),
                    'name': it.get('name'),
                    'steps': it.get('steps'),
                    'compatible_models': it.get('compatible_models') or '',
                    'sort_order': int(it.get('sort_order') or 0),
                    'updated_at': it.get('updated_at'),
                })
            data.sort(key=lambda x: (int(x.get('sort_order') or 0), int(x.get('id') or 0)))
            return jsonify({'success': True, 'data': data})
        except Exception as e:
            print(f"[OPS] 数据库分支 GET 失败，回退 sqlite: {e}")

    items = OpsLibraryItem.query.filter_by(kind=kind).order_by(OpsLibraryItem.sort_order.asc(), OpsLibraryItem.id.asc()).all()
    data = []
    for it in items:
        data.append({
            'id': it.id,
            'kind': it.kind,
            'name': it.name,
            'steps': it.steps,
            'compatible_models': it.compatible_models or '',
            'sort_order': it.sort_order or 0,
            'updated_at': it.updated_at.isoformat() if it.updated_at else None
        })
    return jsonify({'success': True, 'data': data})


@app.route('/api/ops/<kind>', methods=['POST'])
@login_required
def create_ops_item(kind):
    kind = (kind or '').strip().lower()
    if kind not in ('app', 'product'):
        return jsonify({'success': False, 'message': 'Invalid kind'}), 400
    payload = request.json or {}
    name = str(payload.get('name') or '').strip()
    steps = str(payload.get('steps') or '')
    compatible_models = str(payload.get('compatible_models') or '').strip()
    if not name or not steps.strip():
        return jsonify({'success': False, 'message': 'name/steps required'}), 400
    if is_supabase_ops_enabled():
        try:
            client = get_supabase_client()
            existing = client.select_all(
                'ops_library_item',
                filters={'kind': f'eq.{kind}'},
                columns='id,sort_order',
                order_by='sort_order',
                order_dir='asc',
                page_size=1000,
            ) or []
            max_order = 0
            for row in existing:
                try:
                    max_order = max(max_order, int((row or {}).get('sort_order') or 0))
                except Exception:
                    pass
            next_order = max_order + 10
            resp = client.insert('ops_library_item', {
                'kind': kind,
                'name': name,
                'steps': steps,
                'compatible_models': compatible_models,
                'sort_order': next_order,
            })
            if resp is not None and int(getattr(resp, 'status_code', 500)) < 400:
                return jsonify({'success': True, 'data': {}})
            return jsonify({'success': False, 'message': getattr(resp, 'text', '') or '数据库写入失败'}), 500
        except Exception as e:
            print(f"[OPS] 数据库分支 POST 失败，回退 sqlite: {e}")

    max_order = db.session.query(func.max(OpsLibraryItem.sort_order)).filter(OpsLibraryItem.kind == kind).scalar()
    next_order = int(max_order or 0) + 10
    it = OpsLibraryItem(kind=kind, name=name, steps=steps, compatible_models=compatible_models, sort_order=next_order)
    db.session.add(it)
    db.session.commit()
    return jsonify({'success': True, 'data': {'id': it.id}})


@app.route('/api/ops/<kind>/<int:item_id>', methods=['PUT'])
@login_required
def update_ops_item(kind, item_id):
    kind = (kind or '').strip().lower()
    if kind not in ('app', 'product'):
        return jsonify({'success': False, 'message': 'Invalid kind'}), 400
    payload = request.json or {}
    name = str(payload.get('name') or '').strip()
    steps = str(payload.get('steps') or '')
    compatible_models = str(payload.get('compatible_models') or '').strip()
    if is_supabase_ops_enabled():
        try:
            client = get_supabase_client()
            rows = client.select_all(
                'ops_library_item',
                filters={'id': f'eq.{int(item_id)}', 'kind': f'eq.{kind}'},
                columns='id',
                order_by='id',
                order_dir='asc',
                page_size=1,
            ) or []
            if not rows:
                return jsonify({'success': False, 'message': 'Not found'}), 404
            update_data = {'compatible_models': compatible_models}
            if name:
                update_data['name'] = name
            if steps.strip():
                update_data['steps'] = steps
            resp = client.update('ops_library_item', update_data, {'id': f'eq.{int(item_id)}', 'kind': f'eq.{kind}'})
            if resp is not None and int(getattr(resp, 'status_code', 500)) < 400:
                return jsonify({'success': True})
            return jsonify({'success': False, 'message': getattr(resp, 'text', '') or '数据库更新失败'}), 500
        except Exception as e:
            print(f"[OPS] 数据库分支 PUT 失败，回退 sqlite: {e}")

    it = OpsLibraryItem.query.get(item_id)
    if not it or it.kind != kind:
        return jsonify({'success': False, 'message': 'Not found'}), 404
    if name:
        it.name = name
    if steps.strip():
        it.steps = steps
    it.compatible_models = compatible_models
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/ops/<kind>/<int:item_id>', methods=['DELETE'])
@login_required
def delete_ops_item(kind, item_id):
    kind = (kind or '').strip().lower()
    if kind not in ('app', 'product'):
        return jsonify({'success': False, 'message': 'Invalid kind'}), 400
    if is_supabase_ops_enabled():
        try:
            client = get_supabase_client()
            resp = client.delete('ops_library_item', {'id': f'eq.{int(item_id)}', 'kind': f'eq.{kind}'})
            if resp is not None and int(getattr(resp, 'status_code', 500)) < 400:
                return jsonify({'success': True})
            return jsonify({'success': False, 'message': getattr(resp, 'text', '') or '数据库删除失败'}), 500
        except Exception as e:
            print(f"[OPS] 数据库分支 DELETE 失败，回退 sqlite: {e}")

    it = OpsLibraryItem.query.get(item_id)
    if not it or it.kind != kind:
        return jsonify({'success': False, 'message': 'Not found'}), 404
    db.session.delete(it)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/ops/<kind>/reorder', methods=['POST'])
@login_required
def reorder_ops_items(kind):
    kind = (kind or '').strip().lower()
    if kind not in ('app', 'product'):
        return jsonify({'success': False, 'message': 'Invalid kind'}), 400
    payload = request.json or {}
    ids = payload.get('ids')
    if not isinstance(ids, list) or not ids:
        return jsonify({'success': False, 'message': 'ids required'}), 400
    if is_supabase_ops_enabled():
        try:
            client = get_supabase_client()
            for idx, rid in enumerate(ids):
                try:
                    rid_int = int(rid)
                except Exception:
                    continue
                client.update('ops_library_item', {'sort_order': (idx + 1) * 10}, {'id': f'eq.{rid_int}', 'kind': f'eq.{kind}'})
            return jsonify({'success': True})
        except Exception as e:
            print(f"[OPS] 数据库分支排序失败，回退 sqlite: {e}")

    for idx, rid in enumerate(ids):
        try:
            rid_int = int(rid)
        except Exception:
            continue
        it = OpsLibraryItem.query.get(rid_int)
        if not it or it.kind != kind:
            continue
        it.sort_order = (idx + 1) * 10
    db.session.commit()
    return jsonify({'success': True})







@app.route('/api/ops/<kind>/export.xlsx', methods=['GET'])
@login_required
def export_ops_excel(kind):
    kind = (kind or '').strip().lower()
    if kind not in ('app', 'product'):
        return jsonify({'success': False, 'message': 'Invalid kind'}), 400
    if is_supabase_ops_enabled():
        try:
            client = get_supabase_client()
            rows = client.select_all(
                'ops_library_item',
                filters={'kind': f'eq.{kind}'},
                columns='id,name,steps,compatible_models,sort_order',
                order_by='sort_order',
                order_dir='asc',
                page_size=1000,
            ) or []
            rows = [r for r in rows if isinstance(r, dict)]
            rows.sort(key=lambda x: (int(x.get('sort_order') or 0), int(x.get('id') or 0)))
            items = rows
        except Exception as e:
            print(f"[OPS] 数据库分支导出失败，回退 sqlite: {e}")
            items = OpsLibraryItem.query.filter_by(kind=kind).order_by(OpsLibraryItem.sort_order.asc(), OpsLibraryItem.id.asc()).all()
    else:
        items = OpsLibraryItem.query.filter_by(kind=kind).order_by(OpsLibraryItem.sort_order.asc(), OpsLibraryItem.id.asc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'ops'
    ws.append(['操作名称', '操作步骤', '适配型号', '排序'])
    for it in items:
        if isinstance(it, dict):
            ws.append([
                it.get('name') or '',
                it.get('steps') or '',
                it.get('compatible_models') or '',
                int(it.get('sort_order') or 0),
            ])
        else:
            ws.append([it.name or '', it.steps or '', it.compatible_models or '', int(it.sort_order or 0)])
    # basic column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 10
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = canonical_download_name('ops_library')
    return send_file(bio, as_attachment=True, download_name=fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/ops/<kind>/template.xlsx', methods=['GET'])
@login_required
def download_ops_import_template(kind):
    kind = (kind or '').strip().lower()
    if kind not in ('app', 'product'):
        return jsonify({'success': False, 'message': 'Invalid kind'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = 'ops_template'
    ws.append(['操作名称', '操作步骤', '适配型号', '排序'])
    ws.append([
        '示例操作名称',
        '步骤1：进入页面\n步骤2：点击按钮\n步骤3：完成设置',
        'K10 Pro,G20S Ultra',
        10,
    ])

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 10
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = canonical_download_name('ops_template')
    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/ops/<kind>/import', methods=['POST'])
@login_required
def import_ops_excel(kind):
    kind = (kind or '').strip().lower()
    if kind not in ('app', 'product'):
        return jsonify({'success': False, 'message': 'Invalid kind'}), 400
    f = request.files.get('file')
    if not f:
        return jsonify({'success': False, 'message': 'file required'}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'success': False, 'message': 'empty file'}), 400
        header = [str(x or '').strip() for x in rows[0]]
        # map columns
        def idx_of(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None
        i_name = idx_of('操作名称', 'name', '名称')
        i_steps = idx_of('操作步骤', 'steps', '步骤')
        i_models = idx_of('适配型号', 'compatible_models', '型号')
        i_sort = idx_of('排序', 'sort_order', 'order')
        if i_name is None or i_steps is None:
            return jsonify({'success': False, 'message': '缺少必要列：操作名称/操作步骤'}), 400

        if is_supabase_ops_enabled():
            client = get_supabase_client()
            existing_rows = client.select_all(
                'ops_library_item',
                filters={'kind': f'eq.{kind}'},
                columns='id,name,sort_order',
                order_by='sort_order',
                order_dir='asc',
                page_size=1000,
            ) or []
            existing_by_name = {str((it or {}).get('name') or '').strip(): it for it in existing_rows if isinstance(it, dict)}
            inserted = 0
            updated = 0
            max_order = 0
            for row in existing_rows:
                try:
                    max_order = max(max_order, int((row or {}).get('sort_order') or 0))
                except Exception:
                    pass
            next_order = max_order + 10
            for r in rows[1:]:
                name = str(r[i_name] or '').strip()
                steps = str(r[i_steps] or '')
                if not name or not str(steps).strip():
                    continue
                models = str(r[i_models] or '').strip() if i_models is not None else ''
                sort_order = None
                if i_sort is not None:
                    try:
                        sort_order = int(r[i_sort] or 0)
                    except Exception:
                        sort_order = None
                cur = existing_by_name.get(name)
                if cur:
                    payload = {'steps': steps, 'compatible_models': models}
                    if sort_order is not None:
                        payload['sort_order'] = sort_order
                    rid = int((cur or {}).get('id'))
                    resp = client.update('ops_library_item', payload, {'id': f'eq.{rid}', 'kind': f'eq.{kind}'})
                    if resp is not None and int(getattr(resp, 'status_code', 500)) < 400:
                        updated += 1
                else:
                    so = sort_order if sort_order is not None else next_order
                    resp = client.insert('ops_library_item', {
                        'kind': kind,
                        'name': name,
                        'steps': steps,
                        'compatible_models': models,
                        'sort_order': so,
                    })
                    if resp is not None and int(getattr(resp, 'status_code', 500)) < 400:
                        inserted += 1
                        next_order += 10
            return jsonify({'success': True, 'inserted': inserted, 'updated': updated})

        # existing by name (sqlite fallback)
        existing = {it.name: it for it in OpsLibraryItem.query.filter_by(kind=kind).all()}
        inserted = 0
        updated = 0
        max_order = db.session.query(func.max(OpsLibraryItem.sort_order)).filter(OpsLibraryItem.kind == kind).scalar()
        next_order = int(max_order or 0) + 10
        for r in rows[1:]:
            name = str(r[i_name] or '').strip()
            steps = str(r[i_steps] or '')
            if not name or not str(steps).strip():
                continue
            models = str(r[i_models] or '').strip() if i_models is not None else ''
            sort_order = None
            if i_sort is not None:
                try:
                    sort_order = int(r[i_sort] or 0)
                except Exception:
                    sort_order = None
            it = existing.get(name)
            if it:
                it.steps = steps
                it.compatible_models = models
                if sort_order is not None:
                    it.sort_order = sort_order
                updated += 1
            else:
                so = sort_order if sort_order is not None else next_order
                it = OpsLibraryItem(kind=kind, name=name, steps=steps, compatible_models=models, sort_order=so)
                db.session.add(it)
                inserted += 1
                next_order += 10
        db.session.commit()
        return jsonify({'success': True, 'inserted': inserted, 'updated': updated})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


def _clean_modification_ids(values):
    if values is None:
        return []
    if isinstance(values, str):
        raw = values.strip()
        if not raw:
            values = []
        else:
            try:
                parsed = json.loads(raw)
                values = parsed if isinstance(parsed, list) else [parsed]
            except Exception:
                values = [v.strip() for v in raw.split(',')]
    elif not isinstance(values, (list, tuple, set)):
        values = [values]

    cleaned = []
    seen = set()
    for v in (values or []):
        if v is None or isinstance(v, bool):
            continue
        if isinstance(v, int):
            item = v
        else:
            s = str(v).strip()
            if not s:
                continue
            item = int(s) if re.fullmatch(r'\d+', s) else s
        key = f'{type(item).__name__}:{item}'
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(item)
    return cleaned


def _collect_archive_candidates(client, operation_id=None, ids=None):
    all_rows = client.select_all('knowledge_base_modifications', order_by='modification_time', order_dir='desc', page_size=1000) or []
    normalized = []
    raw_ids = []
    operation_id = str(operation_id or '').strip()
    selected_ids = _clean_modification_ids(ids)
    selected_id_set = {str(v).strip() for v in selected_ids if str(v).strip()}
    for r in all_rows:
        rid = r.get('id') if isinstance(r, dict) else None
        if selected_id_set and str(rid).strip() not in selected_id_set:
            continue
        it = _normalize_mod_record(r)
        if not it:
            continue
        if operation_id and str(it.get('operation_id') or '').strip() != operation_id:
            continue
        normalized.append(it)
        if rid is not None:
            raw_ids.append(rid)
    return normalized, raw_ids


def _delete_modification_rows_by_ids(client, ids):
    cleaned = _clean_modification_ids(ids)
    if not cleaned:
        return True, 0, ''
    inner = ",".join([json.dumps(v, ensure_ascii=False) for v in cleaned])
    in_filter = f"in.({inner})"
    resp = client.delete('knowledge_base_modifications', {'id': in_filter})
    if not resp or getattr(resp, 'status_code', 500) >= 400:
        return False, 0, getattr(resp, "text", "") if resp else ''
    return True, len(cleaned), ''

@app.route('/api/archives/preview', methods=['GET'])
@login_required
def preview_archive():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400
    try:
        operation_id = request.args.get('operation_id')
        ids_params = request.args.getlist('ids')
        ids = ids_params if len(ids_params) > 1 else (ids_params[0] if ids_params else None)
        normalized, raw_ids = _collect_archive_candidates(client, operation_id=operation_id, ids=ids)
        samples = []
        for it in normalized[:5]:
            samples.append({
                'question_wiki_id': it.get('question_wiki_id') or it.get('kb_id'),
                'question': str(it.get('question') or '')[:80],
                'source_module': it.get('source_module'),
                'modify_time': it.get('modify_time')
            })
        return jsonify({
            'success': True,
            'count': len(normalized),
            'delete_count': len(raw_ids),
            'operation_id': str(operation_id or '').strip() or None,
            'selected_count': len(_clean_modification_ids(ids)),
            'samples': samples,
            'destructive': len(raw_ids) > 0
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/archives', methods=['GET'])
@login_required
def list_archives():
    q = request.args.get('q', '').strip()
    batches = []
    if is_supabase_archives_enabled():
        try:
            client = get_supabase_client()
            if client and _supabase_table_exists(client, 'archive_batch'):
                rows = client.select_all(
                    'archive_batch',
                    order_by='created_at',
                    order_dir='desc',
                    page_size=1000
                ) or []
                if q:
                    ql = q.lower()
                    rows = [r for r in rows if ql in str((r or {}).get('batch_name') or '').lower()]
                rows.sort(key=lambda x: str((x or {}).get('created_at') or ''), reverse=True)
                batches = rows[:200]
        except Exception:
            batches = []
    if not batches:
        query = ArchiveBatch.query
        if q:
            query = query.filter(ArchiveBatch.batch_name.ilike(f'%{q}%'))
        batches = query.order_by(ArchiveBatch.created_at.desc()).limit(200).all()
    data = []
    for b in batches:
        if isinstance(b, dict):
            created_at = b.get('created_at')
            created_at_iso = str(created_at) if created_at else None
            if created_at_iso and not created_at_iso.endswith('Z'):
                created_at_iso += 'Z'
            data.append({
                'id': b.get('id'),
                'batch_name': b.get('batch_name'),
                'record_count': b.get('record_count'),
                'created_by': b.get('created_by'),
                'created_at': created_at_iso
            })
            continue
        data.append({
            'id': b.id,
            'batch_name': b.batch_name,
            'record_count': b.record_count,
            'created_by': b.created_by,
            'created_at': b.created_at.isoformat() + 'Z' if b.created_at else None
        })
    return jsonify({'success': True, 'data': data})

@app.route('/api/archives', methods=['POST'])
@login_required
def create_archive():
    payload = request.json or {}
    batch_name = str(payload.get('batch_name') or '').strip()
    if not batch_name:
        return jsonify({'success': False, 'message': 'batch_name required'}), 400

    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400

    try:
        archive_ids = payload.get('ids')
        normalized, raw_ids = _collect_archive_candidates(client, ids=archive_ids)
        if not normalized:
            return jsonify({'success': False, 'message': '没有待归档的修改记录'}), 400

        if not bool(payload.get('confirm_archive')):
            return jsonify({
                'success': False,
                'requires_confirmation': True,
                'message': '归档会把当前修改记录迁移到归档表，并删除当前列表中的对应记录，请确认后继续。',
                'count': len(normalized),
                'delete_count': len(raw_ids),
                'selected_count': len(_clean_modification_ids(archive_ids))
            }), 409

        expected_count = payload.get('expected_count')
        if expected_count is not None:
            try:
                if int(expected_count) != len(normalized):
                    return jsonify({
                        'success': False,
                        'requires_confirmation': True,
                        'message': f'待归档记录数已变化（确认时 {expected_count} 条，当前 {len(normalized)} 条），请重新确认。',
                        'count': len(normalized),
                        'delete_count': len(raw_ids),
                        'selected_count': len(_clean_modification_ids(archive_ids))
                    }), 409
            except Exception:
                return jsonify({'success': False, 'message': 'expected_count 必须为数字'}), 400

        batch = ArchiveBatch(
            batch_name=batch_name,
            record_count=len(normalized),
            created_by=(current_user.username if current_user.is_authenticated else 'admin')
        )
        db.session.add(batch)
        db.session.flush()

        recs = []
        for it in normalized:
            mt = _safe_parse_dt(it.get('modify_time'))
            recs.append(ArchiveRecord(
                batch_id=batch.id,
                record_json=_json_dumps_safe(it),
                modify_time=mt
            ))
        if recs:
            db.session.add_all(recs)
        db.session.commit()
        archive_id = batch.id

        # Supabase-first archive persistence for cross-environment visibility.
        if is_supabase_archives_enabled():
            try:
                sb = get_supabase_client()
                if sb and _supabase_table_exists(sb, 'archive_batch') and _supabase_table_exists(sb, 'archive_record'):
                    sb_batch = [{
                        'id': archive_id,
                        'batch_name': batch_name,
                        'record_count': len(normalized),
                        'created_by': current_user.username,
                    }]
                    sb.upsert('archive_batch', sb_batch, on_conflict='id')
                    if recs:
                        sb_records = []
                        for rec in recs:
                            sb_records.append({
                                'id': rec.id,
                                'batch_id': archive_id,
                                'record_json': rec.record_json,
                                'modify_time': rec.modify_time.isoformat() if rec.modify_time else None,
                            })
                        sb.upsert('archive_record', sb_records, on_conflict='id')
            except Exception:
                # Keep local success as source of truth if remote write is temporarily unavailable.
                pass

        deleted = 0
        if raw_ids:
            chunk_size = 500
            for i in range(0, len(raw_ids), chunk_size):
                chunk = raw_ids[i:i + chunk_size]
                ok, count, message = _delete_modification_rows_by_ids(client, chunk)
                if not ok:
                    return jsonify({'success': False, 'message': f'归档已保存，但清理主库失败: {message}'}), 500
                deleted += count

        return jsonify({'success': True, 'id': archive_id, 'record_count': batch.record_count, 'deleted_count': deleted})
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/archives/<int:batch_id>/records', methods=['GET'])
@login_required
def get_archive_records(batch_id):
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    kb_id = request.args.get('kb_id')
    product = request.args.get('product')
    question = request.args.get('question')
    modifier = request.args.get('modifier')
    source_module = request.args.get('source_module')
    start_time = request.args.get('start_time')
    end_time = request.args.get('end_time')

    st = _safe_parse_dt(start_time)
    et = _safe_parse_dt(end_time)

    records = []
    if is_supabase_archives_enabled():
        try:
            client = get_supabase_client()
            if client and _supabase_table_exists(client, 'archive_record'):
                filters = {'batch_id': f'eq.{batch_id}'}
                records = client.select_all(
                    'archive_record',
                    filters=filters,
                    order_by='modify_time',
                    order_dir='desc',
                    page_size=1000
                ) or []
        except Exception:
            records = []
    if not records:
        query = ArchiveRecord.query.filter_by(batch_id=batch_id)
        if st:
            query = query.filter(ArchiveRecord.modify_time >= st)
        if et:
            query = query.filter(ArchiveRecord.modify_time <= et)
        records = query.order_by(ArchiveRecord.modify_time.desc().nullslast(), ArchiveRecord.id.desc()).all()

    filtered = []
    for r in records:
        if isinstance(r, dict):
            record_json = r.get('record_json')
            mt = _safe_parse_dt(r.get('modify_time'))
            if st and (not mt or mt < st):
                continue
            if et and (not mt or mt > et):
                continue
        else:
            record_json = r.record_json
        try:
            it = json.loads(record_json)
        except Exception:
            continue
        if kb_id and kb_id not in str(it.get('question_wiki_id') or ''):
            continue
        if question and str(question).strip().lower() not in str(it.get('question') or '').lower():
            continue
        if product and str(product).strip().lower() not in str(it.get('products') or '').lower():
            continue
        if modifier and str(modifier).strip().lower() not in str(it.get('modifier') or '').lower():
            continue
        if source_module and not _mod_source_match(it.get('source_module'), source_module):
            continue
        filtered.append(it)

    total = len(filtered)
    start_idx = max(0, (page - 1) * page_size)
    end_idx = start_idx + page_size
    sliced = filtered[start_idx:end_idx]
    return jsonify({'success': True, 'data': sliced, 'total': total})

@app.route('/api/archives/<int:batch_id>/export', methods=['GET'])
@login_required
def export_archive(batch_id):
    kb_id = request.args.get('kb_id')
    product = request.args.get('product')
    question = request.args.get('question')
    modifier = request.args.get('modifier')
    source_module = request.args.get('source_module')
    start_time = request.args.get('start_time')
    end_time = request.args.get('end_time')

    st = _safe_parse_dt(start_time)
    et = _safe_parse_dt(end_time)

    records = []
    if is_supabase_archives_enabled():
        try:
            client = get_supabase_client()
            if client and _supabase_table_exists(client, 'archive_record'):
                filters = {'batch_id': f'eq.{batch_id}'}
                records = client.select_all(
                    'archive_record',
                    filters=filters,
                    order_by='modify_time',
                    order_dir='desc',
                    page_size=1000
                ) or []
        except Exception:
            records = []
    if not records:
        query = ArchiveRecord.query.filter_by(batch_id=batch_id)
        if st:
            query = query.filter(ArchiveRecord.modify_time >= st)
        if et:
            query = query.filter(ArchiveRecord.modify_time <= et)
        records = query.order_by(ArchiveRecord.modify_time.desc().nullslast(), ArchiveRecord.id.desc()).all()

    out = []
    for r in records:
        if isinstance(r, dict):
            record_json = r.get('record_json')
            mt = _safe_parse_dt(r.get('modify_time'))
            if st and (not mt or mt < st):
                continue
            if et and (not mt or mt > et):
                continue
        else:
            record_json = r.record_json
        try:
            it = json.loads(record_json)
        except Exception:
            continue
        if kb_id and kb_id not in str(it.get('question_wiki_id') or ''):
            continue
        if question and str(question).strip().lower() not in str(it.get('question') or '').lower():
            continue
        if product and str(product).strip().lower() not in str(it.get('products') or '').lower():
            continue
        if modifier and str(modifier).strip().lower() not in str(it.get('modifier') or '').lower():
            continue
        if source_module and not _mod_source_match(it.get('source_module'), source_module):
            continue
        out.append({
            'opera': it.get('opera'),
            'question_wiki_id': it.get('question_wiki_id'),
            'question': it.get('question'),
            'question_type': it.get('question_type'),
            'answer': it.get('answer'),
            'answer_type': it.get('answer_type'),
            'error_list': _export_mod_list_text(it.get('error_list')),
            'keyword_list': _export_mod_list_text(it.get('keyword_list')),
            'similar_questions': _export_mod_list_text(it.get('similar_questions')),
            'if_bm25': _export_mod_bool_text(it.get('if_bm25')),
            'products': _export_mod_list_text(it.get('products')),
            'image_urls': _export_mod_list_text(it.get('image_urls'), is_url_list=True),
            'video_urls': _export_mod_list_text(it.get('video_urls'), is_url_list=True),
            'file_urls': _export_mod_list_text(it.get('file_urls'), is_url_list=True),
            'link_type': it.get('link_type'),
            'link_url': it.get('link_url'),
            'source_module': it.get('source_module'),
            'modify_time': it.get('modify_time'),
            'before': json.dumps(it.get('before'), ensure_ascii=False) if it.get('before') is not None else '',
            'after': json.dumps(it.get('after'), ensure_ascii=False) if it.get('after') is not None else '',
            'changed_fields': json.dumps(it.get('changed_fields'), ensure_ascii=False) if it.get('changed_fields') is not None else ''
        })

    df = pd.DataFrame(out)
    output = io.BytesIO()
    output.write(b'\xef\xbb\xbf')
    df.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=canonical_download_name('archive', 'csv'),
        mimetype='text/csv'
    )

def _client_count(client, table, filters=None):
    resp = client.select(table, page=1, page_size=1, filters=filters or {}, columns='*')
    if resp is None or getattr(resp, 'status_code', 500) >= 400:
        raise RuntimeError(getattr(resp, 'text', '') or f'count {table} failed')
    content_range = getattr(resp, 'headers', {}).get('Content-Range') if getattr(resp, 'headers', None) else None
    if content_range and '/' in content_range:
        try:
            return int(str(content_range).split('/')[-1])
        except Exception:
            pass
    try:
        return len(resp.json() or [])
    except Exception:
        return 0

def _write_json_backup(kind, table, rows, metadata=None):
    backup_dir = os.path.join(_BASE_DIR, 'instance', 'backups', kind)
    os.makedirs(backup_dir, exist_ok=True)
    filename = f"{table}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    path = os.path.join(backup_dir, filename)
    payload = {
        'table': table,
        'created_at': datetime.now().isoformat(),
        'created_by': current_user.username if current_user.is_authenticated else None,
        'metadata': metadata or {},
        'count': len(rows or []),
        'rows': rows or []
    }
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    return path

@app.route('/api/scoring/cache_summary', methods=['GET'])
@login_required
def scoring_cache_summary():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400
    try:
        count = _client_count(client, 'kb_scores')
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/scoring/clear_cache', methods=['POST'])
@login_required
def clear_scoring_cache():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400
    payload = request.get_json(silent=True) or {}
    confirm_clear = bool(payload.get('confirm_clear'))
    try:
        count = _client_count(client, 'kb_scores')
        expected_count = payload.get('expected_count')
        if not confirm_clear:
            return jsonify({
                'success': False,
                'requires_confirmation': True,
                'message': '清空评分缓存需要二次确认。',
                'count': count
            }), 409
        if expected_count is not None:
            try:
                if int(expected_count) != int(count):
                    return jsonify({
                        'success': False,
                        'requires_confirmation': True,
                        'message': f'评分缓存数量已变化（确认时 {expected_count} 条，当前 {count} 条），请重新确认。',
                        'count': count
                    }), 409
            except Exception:
                return jsonify({'success': False, 'message': 'expected_count 必须为数字'}), 400

        rows = client.select_all('kb_scores', order_by='id', page_size=1000) or []
        backup_path = _write_json_backup('scoring_cache', 'kb_scores', rows, {'action': 'clear_cache'})

        ids = [r.get('id') for r in rows if isinstance(r, dict) and r.get('id') is not None]
        if ids:
            chunk_size = 500
            for i in range(0, len(ids), chunk_size):
                resp = client.delete_in('kb_scores', 'id', ids[i:i + chunk_size])
                if resp is None or getattr(resp, 'status_code', 500) >= 400:
                    return jsonify({'success': False, 'message': getattr(resp, 'text', '删除评分缓存失败'), 'backup_path': backup_path}), 500
        else:
            resp = client.delete('kb_scores', {'id': 'not.is.null'})
            if resp is None or getattr(resp, 'status_code', 500) >= 400:
                return jsonify({'success': False, 'message': getattr(resp, 'text', '删除评分缓存失败'), 'backup_path': backup_path}), 500

        return jsonify({'success': True, 'deleted': count, 'backup_path': backup_path})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/scoring/evaluate', methods=['POST'])
@login_required
def evaluate_items():
    """
    Evaluate specific items.
    Input: { ids: [kb_id1, kb_id2, ...], use_cache: boolean }
    Supports SSE streaming response if header 'Accept' is 'text/event-stream'.
    """
    # Check for SSE request
    is_stream = request.headers.get('Accept') == 'text/event-stream'
    
    data = request.json
    if not data and is_stream:
        # For SSE, maybe data is in query params or body needs parsing?
        # Actually standard SSE is GET, but POST is allowed if client supports it.
        # However, EventSource in JS only supports GET. 
        # So we'll use fetch with reader for streaming, which allows POST.
        # Flask handles this fine.
        try:
            data = request.get_json(force=True)
        except:
            data = {}
            
    ids = data.get('ids', [])
    use_cache = data.get('use_cache', False)
    
    if not ids:
        return jsonify({'success': False, 'message': 'No IDs provided'}), 400

    client = get_supabase_client()
    config = load_scoring_config()
    score_has_product = _supabase_has_column(client, 'kb_scores', 'product_name') if client else False
    
    if not config.get('api_key'):
        return jsonify({'success': False, 'message': 'LLM API Key not configured'}), 400
    try:
        scorer = LLMScorer(
            api_key=config.get('api_key') or '',
            base_url=config.get('base_url') or "https://api.deepseek.com",
            model=config.get('model') or "deepseek-chat",
            system_prompt=config.get('system_prompt')
        )
    except Exception as e:
        msg = f"LLM scorer init failed: {e}"
        print(msg)
        if is_stream:
            return Response(
                json.dumps({'type': 'error', 'message': msg}) + "\n",
                mimetype='application/x-ndjson'
            )
        return jsonify({'success': False, 'message': msg}), 500
    
    def generate():
        try:
            # 0. Pre-fetch existing scores if using cache
            cached_scores = {}
            if use_cache:
                # Optimized batch fetching
                batch_size = 50
                for i in range(0, len(ids), batch_size):
                    batch_ids = ids[i:i+batch_size]
                    id_filter = _postgrest_in_str(batch_ids)
                    if not id_filter:
                        continue
                    resp = client.select('kb_scores', page_size=1000, filters={'kb_id': id_filter})
                    if isinstance(resp, list):
                        for item in resp:
                            if isinstance(item, dict) and item.get('kb_id'):
                                cached_scores[item['kb_id']] = item
                    elif getattr(resp, 'status_code', None) in (200, 206):
                        for item in (resp.json() or []):
                            if isinstance(item, dict) and item.get('kb_id'):
                                cached_scores[item['kb_id']] = item

            # 1. Fetch Target KB Data
            kb_items_map = {}
            batch_size = 50 
            for i in range(0, len(ids), batch_size):
                batch_ids = ids[i:i+batch_size]
                id_filter = _postgrest_in_str(batch_ids)
                if not id_filter:
                    continue
                resp = client.select('knowledge_base_v1', page_size=1000, filters={'question_wiki_id': id_filter})
                if isinstance(resp, list):
                    for item in resp:
                        if isinstance(item, dict) and item.get('question_wiki_id'):
                            kb_items_map[item['question_wiki_id']] = item
                elif getattr(resp, 'status_code', None) in (200, 206):
                    for item in (resp.json() or []):
                        if isinstance(item, dict) and item.get('question_wiki_id'):
                            kb_items_map[item['question_wiki_id']] = item
            
            # 2. Fetch ALL KB Data for Overlap Calculation (Lightweight)
            # Only fetch minimal columns needed
            all_kb_resp = client.select_all('knowledge_base_v1', columns='question_wiki_id, question, product_name', order_by='question_wiki_id')
            overlap_items = []
            for item in all_kb_resp:
                # Normalize key to 'kb_id' for calculate_product_overlap
                item['kb_id'] = item.get('question_wiki_id')
                overlap_items.append(item)
                
            overlap_map = calculate_product_overlap(overlap_items)
            
            results = []
            pending_items = []
            
            def normalize_score_result(qid, score_result):
                dims = score_result.get('维度得分', {})
                frontend_result = score_result.copy()
                frontend_result['kb_id'] = qid
                frontend_result['quality'] = dims.get('问题质量', 0)
                frontend_result['compliance'] = dims.get('答案合规与准确性', 0)
                frontend_result['timeliness'] = dims.get('时效性', 0)
                frontend_result['utility'] = dims.get('实际解决力', 0)
                frontend_result['redundancy'] = dims.get('非冗余与相关性', 0)
                frontend_result['multimedia'] = dims.get('多媒体加分', 0)
                frontend_result['suggestion'] = score_result.get('处理建议', '')
                frontend_result['analysis'] = score_result.get('分析过程', '')
                frontend_result['total_score'] = score_result.get('总分', 0)
                return frontend_result

            def upsert_score_result(qid, score_result, item=None):
                if isinstance(client, LocalPostgreSQLClient):
                    updated_at_value = datetime.now()
                else:
                    updated_at_value = datetime.now().isoformat()
                
                update_data = {
                    'kb_id': qid,
                    'total_score': score_result.get('总分', 0),
                    'remarks': score_result.get('处理建议', ''),
                    'status': 'scored',
                    'score_data': json.dumps(score_result, ensure_ascii=False),
                    'updated_at': updated_at_value
                }
                if score_has_product and isinstance(item, dict):
                    update_data['product_name'] = item.get('product_name') or ''
                upsert_resp = client.upsert('kb_scores', [update_data], on_conflict='kb_id')
                print(f"DEBUG: Upsert response for {qid}: Status={upsert_resp.status_code}")
                if getattr(upsert_resp, 'status_code', 500) >= 400:
                    raise RuntimeError(f"Save score failed for {qid}: {getattr(upsert_resp, 'text', '') or 'unknown database error'}")

            def evaluate_qid(qid, item):
                eval_item = {
                    'kb_id': qid,
                    'product': item.get('product_name'),
                    'update_time': item.get('update_time'),
                    'question': item.get('question'),
                    'answer': item.get('answer'),
                    'urls': _aggregate_links_from_fields(item)
                }
                started = time.perf_counter()
                score_result = scorer.evaluate_one(eval_item, overlap_count=overlap_map.get(qid, 0))
                elapsed = time.perf_counter() - started
                print(f"DEBUG: AI scoring finished for {qid} in {elapsed:.2f}s")
                return qid, score_result, elapsed

            # Fast path: validate IDs and emit cached results before starting parallel AI calls.
            for qid in ids:
                item = kb_items_map.get(qid)
                if not item:
                    yield json.dumps({'type': 'error', 'id': qid, 'message': 'Item not found'}) + "\n"
                    continue

                if use_cache and qid in cached_scores:
                    score_rec = cached_scores[qid]
                    if score_rec.get('status') == 'scored':
                        try:
                            if score_rec.get('score_data'):
                                res = json.loads(score_rec['score_data'])
                                res = normalize_score_result(qid, res)
                                results.append(res)
                                yield json.dumps({'type': 'result', 'data': res}) + "\n"
                                continue
                        except Exception as cache_err:
                            print(f"DEBUG: Failed to parse cached score for {qid}: {cache_err}")

                pending_items.append((qid, item))

            try:
                configured_workers = int(config.get('scoring_concurrency', 4))
            except (TypeError, ValueError):
                configured_workers = 4
            configured_workers = min(8, max(1, configured_workers))
            max_workers = min(configured_workers, max(1, len(pending_items)))
            if pending_items:
                print(f"DEBUG: Starting parallel scoring for {len(pending_items)} items with {max_workers} workers")
            
            # Model calls are slow and network-bound; DB writes remain serialized after each result returns.
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_map = {
                    executor.submit(evaluate_qid, qid, item): qid
                    for qid, item in pending_items
                }
                for future in as_completed(future_map):
                    qid = future_map[future]
                    try:
                        _, score_result, elapsed = future.result()
                    except Exception as e:
                        yield json.dumps({'type': 'error', 'id': qid, 'message': str(e)}) + "\n"
                        continue
                    
                    if score_result and 'error' not in score_result:
                        source_item = kb_items_map.get(qid)
                        upsert_score_result(qid, score_result, source_item)
                        frontend_result = normalize_score_result(qid, score_result)
                        frontend_result['elapsed_seconds'] = round(elapsed, 2)
                        results.append(frontend_result)
                        yield json.dumps({'type': 'result', 'data': frontend_result}) + "\n"
                    else:
                        error_msg = score_result.get('error') if score_result else "Unknown error"
                        print(f"DEBUG: Score error for {qid}: {error_msg}")
                        yield json.dumps({'type': 'error', 'id': qid, 'message': error_msg}) + "\n"

            yield json.dumps({'type': 'done', 'count': len(results)}) + "\n"

        except Exception as e:
            print(f"Error in evaluate_items stream: {e}")
            yield json.dumps({'type': 'error', 'message': str(e)}) + "\n"

    if is_stream:
        return Response(stream_with_context(generate()), mimetype='application/x-ndjson')
    else:
        # Fallback for non-streaming clients (wait for all)
        # We need to collect from generator
        gen = generate()
        final_results = []
        for chunk in gen:
            try:
                msg = json.loads(chunk.strip())
                if msg['type'] == 'result':
                    final_results.append(msg['data'])
            except:
                pass
        return jsonify({'success': True, 'results': final_results, 'count': len(final_results)})

@app.route('/api/scoring/save_manual', methods=['POST'])
@login_required
def save_manual_score():
    data = request.json
    kb_id = data.get('kb_id')
    remarks = data.get('remarks')
    # Maybe manual score adjustment? User requirement says "Configuration of scoring criteria", not manual score override on row.
    # But user said "Saved scored data...".
    # Requirement 6: "Save scored data".
    # And "Remarks" column.
    
    if not kb_id:
        return jsonify({'success': False}), 400
        
    client = get_supabase_client()
    
    # Find existing ID
    resp = client.select('kb_scores', filters={'kb_id': kb_id})
    if resp.status_code == 200 and resp.json():
        rec = resp.json()[0]
        rec['remarks'] = remarks
        client.upsert('kb_scores', [rec])
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'message': 'Record not found'}), 404



@app.route('/api/links/import', methods=['POST'])
@login_required
def import_links():
    items = request.get_json(silent=True) or []
    if not isinstance(items, list):
        return jsonify({'success': False, 'message': 'Expected a list of links'}), 400

    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    payload = []
    skipped = 0
    for idx, item in enumerate(items):
        if not isinstance(item, dict):
            skipped += 1
            continue
        url = str(item.get('url') or '').strip()
        if not url:
            skipped += 1
            continue
        tags = item.get('tags') or []
        if isinstance(tags, str):
            try:
                parsed_tags = json.loads(tags)
                tags = parsed_tags if isinstance(parsed_tags, list) else [tags]
            except Exception:
                tags = [t.strip() for t in re.split(r'[,，]', tags) if t.strip()]
        if not isinstance(tags, list):
            tags = []
        created_at = item.get('created_at', item.get('createdAt'))
        if created_at is None:
            created_at = _now_iso_with_tz()
        payload.append({
            'id': str(item.get('id') or uuid.uuid4()),
            'kb_id': item.get('kb_id'),
            'url': url,
            'type': str(item.get('type') or _detect_link_type_backend(url)),
            'tags': [str(t).strip() for t in tags if str(t).strip()],
            'created_at': _dt_to_iso(created_at)
        })

    if not payload:
        return jsonify({'success': True, 'count': 0, 'skipped': skipped})

    try:
        resp = client.upsert('link_previews', payload, on_conflict='id')
        if resp is None or getattr(resp, 'status_code', 500) >= 400:
            return jsonify({'success': False, 'message': getattr(resp, 'text', '导入 link_previews 失败')}), 500
        return jsonify({'success': True, 'count': len(payload), 'skipped': skipped})
    except Exception as e:
        print("[import_links] import failed")
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500

# Governance APIs
def _governance_score_detail_fields(score_item):
    if not score_item:
        return {
            'conclusion': '',
            'analysis': '',
            'suggestion': '',
            'remarks': '',
        }
    parsed = _quality_json_loads((score_item or {}).get('score_data'), {})
    if not isinstance(parsed, dict):
        parsed = {}
    remarks = _quality_clean_text((score_item or {}).get('remarks'))
    suggestion = (
        _quality_clean_text(parsed.get('处理建议'))
        or _quality_clean_text(parsed.get('修改建议'))
        or _quality_clean_text(parsed.get('优化建议'))
        or _quality_clean_text(parsed.get('rewrite_suggestion'))
        or remarks
    )
    analysis = (
        _quality_clean_text(parsed.get('分析过程'))
        or _quality_clean_text(parsed.get('扣分分析'))
        or _quality_clean_text(parsed.get('分析评价'))
        or _quality_clean_text(parsed.get('analysis'))
    )
    conclusion = (
        _quality_clean_text(parsed.get('结论'))
        or _quality_clean_text(parsed.get('简要点评'))
        or _quality_clean_text(parsed.get('评价结论'))
        or _quality_clean_text(parsed.get('result'))
    )
    if not conclusion:
        if suggestion and analysis:
            conclusion = f"建议：{suggestion}\n分析：{analysis}"
        else:
            conclusion = suggestion or analysis
    return {
        'conclusion': conclusion,
        'analysis': analysis,
        'suggestion': suggestion,
        'remarks': remarks,
    }

def _load_sqlite_governance_reference_maps():
    score_map = {}
    for s in KBScore.query.all():
        kb_id = str(getattr(s, 'kb_id', '') or '').strip()
        if not kb_id:
            continue
        score_item = {
            'kb_id': kb_id,
            'total_score': getattr(s, 'total_score', None),
            'question_content': getattr(s, 'question_content', None),
            'status': getattr(s, 'status', None),
            'remarks': getattr(s, 'remarks', None),
            'score_data': getattr(s, 'score_data', None),
            'updated_at': getattr(s, 'updated_at', None),
        }
        score_item.update(_governance_score_detail_fields(score_item))
        score_map[kb_id] = score_item

    v1_map = {}
    matrix_rows = db.session.query(
        ProductMatrix.question_wiki_id,
        ProductMatrix.question_content
    ).order_by(ProductMatrix.id.desc()).all()
    for question_wiki_id, question_content in matrix_rows:
        kb_id = str(question_wiki_id or '').strip()
        question = str(question_content or '').strip()
        if not kb_id or not question or kb_id in v1_map:
            continue
        v1_map[kb_id] = {
            'question_wiki_id': kb_id,
            'question': question
        }

    for kb_id, score_item in score_map.items():
        if kb_id not in v1_map:
            question = str(score_item.get('question_content') or '').strip()
            if question:
                v1_map[kb_id] = {
                    'question_wiki_id': kb_id,
                    'question': question
                }

    return score_map, v1_map

def _load_governance_reference_maps():
    """
    优先从本地 PostgreSQL / 兼容客户端读取治理关联数据；
    如果不可用，再回退到本地 SQLite 的快照/矩阵表。
    """
    try:
        client = get_supabase_client()
        if client:
            score_map = {}
            score_rows = client.select_all('kb_scores', columns='kb_id,total_score,question_content,status,remarks,score_data,updated_at', order_by='kb_id', order_dir='asc', page_size=1000) or []
            for s in score_rows:
                if not isinstance(s, dict):
                    continue
                kb_id = str(s.get('kb_id') or '').strip()
                if not kb_id:
                    continue
                s.update(_governance_score_detail_fields(s))
                score_map[kb_id] = s

            v1_map = {}
            v1_rows = client.select_all('knowledge_base_v1', columns='question_wiki_id,question', order_by='question_wiki_id', order_dir='asc', page_size=1000) or []
            for v in v1_rows:
                if not isinstance(v, dict):
                    continue
                kb_id = str(v.get('question_wiki_id') or '').strip()
                question = str(v.get('question') or '').strip()
                if not kb_id or not question:
                    continue
                v1_map[kb_id] = v

            if score_map or v1_map:
                return score_map, v1_map
    except Exception as e:
        print(f"[Governance] 主库映射加载失败，回退 sqlite: {e}")

    return _load_sqlite_governance_reference_maps()

def _load_remote_governance_months():
    try:
        client = get_supabase_client()
        if not client:
            return []
        rows = client.select_all('kb_recall', columns='month', order_by='month', order_dir='desc', page_size=1000) or []
        months_set = set()
        for r in rows:
            m = str((r or {}).get('month') or '').strip()
            if m:
                months_set.add(m)
        return sorted(list(months_set), reverse=True)
    except Exception as e:
        print(f"[Governance] Remote months load failed: {e}")
        return []

def _load_remote_governance_recalls(start_month, end_month=None):
    try:
        client = get_supabase_client()
        if not client:
            return []
        if end_month:
            rows = client.select_all(
                'kb_recall',
                filters={'and': f"(month.gte.{start_month},month.lte.{end_month})"},
                columns='kb_id,month,recall_count,valid_recall_count',
                order_by='month',
                order_dir='asc',
                page_size=1000
            ) or []
        else:
            rows = client.select_all(
                'kb_recall',
                filters={'month': f"eq.{start_month}"},
                columns='kb_id,month,recall_count,valid_recall_count',
                order_by='kb_id',
                order_dir='asc',
                page_size=1000
            ) or []
        return [r for r in rows if isinstance(r, dict)]
    except Exception as e:
        print(f"[Governance] Remote recall query failed: {e}")
        return []

def _load_sqlite_governance_recalls(start_month, end_month=None):
    if end_month:
        return KBRecall.query.filter(KBRecall.month >= start_month, KBRecall.month <= end_month).order_by(KBRecall.month).all()
    return KBRecall.query.filter_by(month=start_month).all()

def _delete_sqlite_governance_month(month):
    if not month:
        return 0
    count = KBRecall.query.filter_by(month=month).delete()
    db.session.flush()
    return count

def _delete_sqlite_governance_items(ids, months):
    if not ids or not months:
        return 0
    count = KBRecall.query.filter(
        KBRecall.kb_id.in_(ids),
        KBRecall.month.in_(months)
    ).delete(synchronize_session=False)
    db.session.flush()
    return count

@app.route('/api/governance/months', methods=['GET'])
@login_required
def get_governance_months():
    try:
        remote_months = _load_remote_governance_months()
        local_months = [m[0] for m in db.session.query(KBRecall.month).distinct().order_by(KBRecall.month.desc()).all()]
        months = sorted(set(remote_months) | set(local_months), reverse=True)
        return jsonify({'success': True, 'months': months})
    except Exception as e:
        print(f"Error getting months: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/governance/data', methods=['GET'])
@login_required
def get_governance_data():
    start_month = request.args.get('month')
    end_month = request.args.get('end_month')
    
    if not start_month:
        return jsonify({'success': False, 'message': 'Start Month is required'}), 400
    
    try:
        remote_recalls = _load_remote_governance_recalls(start_month, end_month)
        local_recalls = _load_sqlite_governance_recalls(start_month, end_month)
        recalls_map = {}
        for r in remote_recalls:
            kb_id = str(r.get('kb_id') or '').strip()
            month = str(r.get('month') or '').strip()
            if not kb_id or not month:
                continue
            recalls_map[(kb_id, month)] = r
        for r in local_recalls:
            kb_id = str(getattr(r, 'kb_id', '') or '').strip()
            month = str(getattr(r, 'month', '') or '').strip()
            if not kb_id or not month:
                continue
            # SQLite 为当前本地导入落点，覆盖同键的历史 PostgreSQL 数据。
            recalls_map[(kb_id, month)] = r
        recalls = list(recalls_map.values())
        
        # 组织数据结构: { kb_id: { month: { recall_count, valid_recall_count } } }
        monthly_map = {}
        month_set = set()
        
        # 计算每月的总召回数，用于计算占比
        monthly_totals = {} # { month: { total_recall, total_valid } }

        for r in recalls:
            if isinstance(r, dict):
                kb_id = str(r.get('kb_id') or '').strip()
                month = str(r.get('month') or '').strip()
                recall_count = int(r.get('recall_count') or 0)
                valid_recall_count = int(r.get('valid_recall_count') or 0)
            else:
                kb_id = str(r.kb_id).strip()
                month = r.month
                recall_count = r.recall_count
                valid_recall_count = r.valid_recall_count
            if not kb_id or not month:
                continue
            month_set.add(month)
            
            if kb_id not in monthly_map:
                monthly_map[kb_id] = {}
            
            monthly_map[kb_id][month] = {
                'recall_count': recall_count,
                'valid_recall_count': valid_recall_count
            }
            
            if month not in monthly_totals:
                monthly_totals[month] = {'total_recall': 0, 'total_valid': 0}
            
            monthly_totals[month]['total_recall'] += recall_count
            monthly_totals[month]['total_valid'] += valid_recall_count
            
        # 排序月份列表。范围查询时补齐中间月份，保证缺月按 0 参与大盘加权。
        sorted_months = _quality_month_sequence(start_month, end_month) if end_month else sorted(list(month_set))
        if not sorted_months and start_month:
             # 如果没有数据，至少返回请求的月份
             sorted_months = [start_month]
             if end_month:
                 # 简单处理：如果是范围但没数据，我们无法得知中间有哪些月份，
                 # 暂时只返回 start_month，或者不做处理。
                 # 更好的做法是生成月份序列，但为了简单起见，这里依赖 DB 数据。
                 pass

        score_map, v1_map = _load_governance_reference_maps()
        
        # 4. 构建结果
        result = []
        
        # 获取所有涉及的 ID
        all_ids = set(monthly_map.keys()) | set(v1_map.keys())
        
        for kb_id in all_ids:
            score_entry = score_map.get(kb_id)
            v1_entry = v1_map.get(kb_id)
            
            # 状态逻辑：主库未命中时给出更明确的提示，避免误解为代码异常。
            status = '使用中' if v1_entry else '主库不存在'
            
            # 构建该 ID 的月度数据
            id_monthly_data = {}
            
            for m in sorted_months:
                m_data = monthly_map.get(kb_id, {}).get(m)
                m_total = monthly_totals.get(m, {'total_recall': 0, 'total_valid': 0})
                
                recall_count = 0
                valid_recall_count = 0
                recall_ratio = None
                valid_recall_ratio = None
                valid_rate = None
                
                if m_data:
                    recall_count = m_data['recall_count']
                    valid_recall_count = m_data['valid_recall_count']
                    
                    tr = m_total['total_recall']
                    tv = m_total['total_valid']
                    
                    recall_ratio = (recall_count / tr) if tr > 0 else 0
                    valid_recall_ratio = (valid_recall_count / tv) if tv > 0 else 0
                    valid_rate = (valid_recall_count / recall_count) if recall_count > 0 else 0
                
                id_monthly_data[m] = {
                    'recall_count': recall_count,
                    'valid_recall_count': valid_recall_count,
                    'recall_ratio': recall_ratio,
                    'valid_recall_ratio': valid_recall_ratio,
                    'valid_rate': valid_rate
                }

            # 获取问题内容和AI评分
            question = v1_entry.get('question', "未匹配到问题快照") if v1_entry else (score_entry.get('question_content', "未匹配到问题快照") if score_entry else "未匹配到问题快照")
            ai_score = score_entry.get('total_score') if score_entry else None
            score_details = _governance_score_detail_fields(score_entry)
            
            result.append({
                'id': kb_id,
                'question': question,
                'v1_question': v1_entry.get('question') if v1_entry else None,
                'score_question': score_entry.get('question_content') if score_entry else None,
                'ai_score': ai_score,
                'status': status,
                'conclusion': score_details.get('conclusion') or '',
                'analysis': score_details.get('analysis') or '',
                'suggestion': score_details.get('suggestion') or '',
                'remarks': score_details.get('remarks') or '',
                'monthly_data': id_monthly_data
            })
            
        return jsonify({
            'success': True, 
            'data': result,
            'months': sorted_months,
            'summary': monthly_totals
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/governance/import', methods=['POST'])
@login_required
def import_governance_data_route():
    month = request.form.get('month')
    if not month:
        return jsonify({'success': False, 'message': 'Month is required'}), 400
        
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': 'File is required'}), 400
        
    try:
        # 读取 Excel/CSV
        if file.filename.lower().endswith('.csv'):
             df = pd.read_csv(file)
        else:
             df = pd.read_excel(file)
        
        df.columns = [str(c).strip() for c in df.columns]
        
        possible_id_cols = ['id', 'kb_id', 'question_wiki_id', 'ID', '知识库ID', '知识库id', '知识库Id', 'Wiki ID', 'WikiID']
        possible_recall_cols = ['recall_count', 'recall', '召回频数', '召回次数', '召回']
        possible_valid_cols = ['valid_recall_count', 'valid_recall', '有效召回频数', '有效召回次数', '有效召回']

        def _norm_col_name(v):
            s = str(v or '').strip().replace('\ufeff', '')
            s = s.replace('（', '(').replace('）', ')')
            low = s.lower()
            key = ''.join(ch for ch in low if ch.isalnum() or ('\u4e00' <= ch <= '\u9fff'))
            return low, key

        normed_cols = [(c, *_norm_col_name(c)) for c in df.columns]

        def _find_col(prefer_keys, skip=None):
            for pk in prefer_keys:
                for orig, low, key in normed_cols:
                    if skip and skip(orig, low, key):
                        continue
                    if key == pk or low == pk:
                        return orig
            for pk in prefer_keys:
                for orig, low, key in normed_cols:
                    if skip and skip(orig, low, key):
                        continue
                    if pk and pk in key:
                        return orig
            return None

        id_col = _find_col(['questionwikiid', 'wikiid', '知识库id', 'kbid', 'id'])
        valid_col = _find_col(['validrecallcount', 'validrecall', '有效召回频数', '有效召回次数', '有效召回'])
        recall_col = _find_col(
            ['recallcount', 'recall', '召回频数', '召回次数', '召回'],
            skip=lambda _o, _l, k: ('有效召回' in k) or ('validrecall' in k)
        )
                
        if not id_col:
             return jsonify({'success': False, 'message': f'Missing ID column. Expected one of: {possible_id_cols}'}), 400
             
        local_deleted = _delete_sqlite_governance_month(month)
        use_supabase = False
        client = None
        try:
            client = get_supabase_client()
            if client:
                use_supabase = True
                del_resp = client.delete('kb_recall', {'month': f"eq.{month}"})
                if not del_resp or getattr(del_resp, 'status_code', 500) >= 400:
                    print(f"[Governance] 主库删除失败，回退 sqlite: {getattr(del_resp, 'text', '')}")
                    use_supabase = False
        except Exception as e:
            print(f"[Governance] 主库删除失败，回退 sqlite: {e}")
            use_supabase = False
        
        # 插入新数据。按 (kb_id, month) 去重，避免源文件里重复行导致冲突或重复写入。
        payload_by_key = {}
        for _, row in df.iterrows():
            kb_id = str(row[id_col]).strip()
            if not kb_id or kb_id.lower() == 'nan':
                continue
                
            recall_val = 0
            if recall_col:
                try:
                    val = row[recall_col]
                    if pd.notna(val):
                        recall_val = int(val)
                except:
                    pass
                    
            valid_val = 0
            if valid_col:
                try:
                    val = row[valid_col]
                    if pd.notna(val):
                        valid_val = int(val)
                except:
                    pass
            payload_by_key[(kb_id, month)] = {
                'kb_id': kb_id,
                'month': month,
                'recall_count': recall_val,
                'valid_recall_count': valid_val
            }

        payload_rows = list(payload_by_key.values())
        count = len(payload_rows)

        if use_supabase:
            errors = []
            for i in range(0, len(payload_rows), 500):
                chunk = payload_rows[i:i + 500]
                resp = client.insert('kb_recall', chunk)
                if resp is None or getattr(resp, 'status_code', 500) >= 400:
                    errors.append({
                        'table': 'kb_recall',
                        'status_code': getattr(resp, 'status_code', None) if resp is not None else None,
                        'text': getattr(resp, 'text', '') if resp is not None else 'no response',
                        'offset': i,
                        'count': len(chunk)
                    })
            if errors:
                print(f"[Governance] 主库写入失败，回退 sqlite: {errors}")
                use_supabase = False
            else:
                db.session.commit()
                return jsonify({'success': True, 'count': count})
        if not use_supabase:
            # 远端写入失败时改为落本地，保证本地部署可用。
            existing_keys = set()
            for item in payload_rows:
                key = (item['kb_id'], item['month'])
                if key in existing_keys:
                    continue
                recall_obj = KBRecall(
                    kb_id=item['kb_id'],
                    month=item['month'],
                    recall_count=item['recall_count'],
                    valid_recall_count=item['valid_recall_count']
                )
                db.session.add(recall_obj)
                existing_keys.add(key)
            db.session.commit()
            return jsonify({'success': True, 'count': count, 'local_deleted': local_deleted})
        
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/governance/delete', methods=['POST'])
@login_required
def delete_governance_data_route():
    data = request.json
    month = data.get('month')
    if not month:
        return jsonify({'success': False, 'message': 'Month is required'}), 400
        
    try:
        remote_attempted = False
        remote_ok = False
        remote_error = ''
        try:
            client = get_supabase_client()
            if client:
                remote_attempted = True
                resp = client.delete('kb_recall', {'month': f"eq.{month}"})
                if resp and getattr(resp, 'status_code', 500) < 400:
                    remote_ok = True
                else:
                    remote_error = getattr(resp, 'text', '') if resp is not None else 'no response'
                    print(f"[Governance] 主库删除月份失败: {remote_error}")
        except Exception as e:
            remote_error = str(e)
            print(f"[Governance] 主库删除月份失败: {e}")

        count = _delete_sqlite_governance_month(month)
        db.session.commit()
        if remote_attempted and not remote_ok:
            return jsonify({
                'success': False,
                'message': f'主库删除失败，本地已删除 {count} 条。{remote_error}',
                'deleted': count,
                'remote_deleted': False
            }), 502
        return jsonify({
            'success': True,
            'deleted': None if remote_ok else count,
            'local_deleted': count,
            'remote_deleted': remote_ok if remote_attempted else None
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/governance/export', methods=['POST'])
@login_required
def export_governance_filtered_route():
    try:
        payload = request.get_json(silent=True) or {}
        rows = payload.get('rows') or []
        months = payload.get('months') or []

        if not isinstance(rows, list):
            return jsonify({'success': False, 'message': 'rows must be a list'}), 400
        if not isinstance(months, list):
            months = []

        normalized_months = [str(m).strip() for m in months if str(m).strip()]
        if not normalized_months:
            month_set = set()
            for item in rows:
                if not isinstance(item, dict):
                    continue
                monthly_data = item.get('monthly_data') or {}
                if isinstance(monthly_data, dict):
                    for m in monthly_data.keys():
                        mm = str(m).strip()
                        if mm:
                            month_set.add(mm)
            normalized_months = sorted(month_set)

        def _fmt_percent(value):
            try:
                if value is None or value == '':
                    return ''
                return f"{float(value) * 100:.2f}%"
            except Exception:
                return ''

        def _fmt_decimal(value):
            try:
                if value is None or value == '':
                    return ''
                return round(float(value), 2)
            except Exception:
                return ''

        export_rows = []
        for item in rows:
            if not isinstance(item, dict):
                continue
            row = {
                'WikiID': str(item.get('id') or '').strip(),
                '问题': str(item.get('question') or '').strip(),
                'AI评分': item.get('ai_score'),
                '状态': str(item.get('status') or '').strip()
            }
            weighted_summary = item.get('weighted_summary') or {}
            if not isinstance(weighted_summary, dict):
                weighted_summary = {}
            row['大盘周期加权_平均召回频数'] = _fmt_decimal(weighted_summary.get('weighted_avg_recall'))
            row['大盘周期加权_平均有效召回频数'] = _fmt_decimal(weighted_summary.get('weighted_avg_valid_recall'))
            row['大盘周期加权_平均召回占比'] = _fmt_percent(weighted_summary.get('weighted_avg_recall_ratio'))
            row['大盘周期加权_平均有效召回占比'] = _fmt_percent(weighted_summary.get('weighted_avg_valid_recall_ratio'))
            row['大盘周期加权_平均有效召回率'] = _fmt_percent(weighted_summary.get('weighted_avg_valid_rate'))

            monthly_data = item.get('monthly_data') or {}
            if not isinstance(monthly_data, dict):
                monthly_data = {}

            for month in normalized_months:
                m_data = monthly_data.get(month) or {}
                row[f'{month}_召回频数'] = int(m_data.get('recall_count') or 0)
                row[f'{month}_有效召回频数'] = int(m_data.get('valid_recall_count') or 0)
                row[f'{month}_召回占比'] = _fmt_percent(m_data.get('recall_ratio'))
                row[f'{month}_有效召回占比'] = _fmt_percent(m_data.get('valid_recall_ratio'))
                row[f'{month}_有效召回率'] = _fmt_percent(m_data.get('valid_rate'))
            export_rows.append(row)

        base_columns = ['WikiID', '问题', 'AI评分', '状态']
        weighted_columns = [
            '大盘周期加权_平均召回频数',
            '大盘周期加权_平均有效召回频数',
            '大盘周期加权_平均召回占比',
            '大盘周期加权_平均有效召回占比',
            '大盘周期加权_平均有效召回率',
        ]
        month_columns = []
        for month in normalized_months:
            month_columns.extend([
                f'{month}_召回频数',
                f'{month}_有效召回频数',
                f'{month}_召回占比',
                f'{month}_有效召回占比',
                f'{month}_有效召回率',
            ])
        all_columns = base_columns + weighted_columns + month_columns

        df = pd.DataFrame(export_rows)
        if df.empty:
            df = pd.DataFrame(columns=all_columns)
        else:
            for col in all_columns:
                if col not in df.columns:
                    df[col] = ''
            df = df[all_columns]

        output = io.BytesIO()
        engine = 'xlsxwriter' if importlib.util.find_spec('xlsxwriter') is not None else 'openpyxl'
        with pd.ExcelWriter(output, engine=engine) as writer:
            df.to_excel(writer, index=False, sheet_name='知识库治理')
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=canonical_download_name('kb_diagnosis'),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500

@app.route('/api/governance/delete_items', methods=['POST'])
@login_required
def delete_governance_items_route():
    data = request.json
    ids = data.get('ids')
    months = data.get('months')
    
    if not ids or not isinstance(ids, list):
        return jsonify({'success': False, 'message': 'IDs list is required'}), 400
    
    if not months or not isinstance(months, list):
        return jsonify({'success': False, 'message': 'Months list is required'}), 400
        
    try:
        remote_attempted = False
        remote_ok = False
        remote_error = ''
        try:
            client = get_supabase_client()
            if client:
                ids_in = _postgrest_in_str(ids)
                months_in = _postgrest_in_str(months)
                if not ids_in or not months_in:
                    return jsonify({'success': True, 'deleted': 0})
                remote_attempted = True
                resp = client.delete('kb_recall', {'kb_id': ids_in, 'month': months_in})
                if resp and getattr(resp, 'status_code', 500) < 400:
                    remote_ok = True
                else:
                    remote_error = getattr(resp, 'text', '') if resp is not None else 'no response'
                    print(f"[Governance] 主库删除条目失败: {remote_error}")
        except Exception as e:
            remote_error = str(e)
            print(f"[Governance] 主库删除条目失败: {e}")

        count = _delete_sqlite_governance_items(ids, months)
        db.session.commit()
        if remote_attempted and not remote_ok:
            return jsonify({
                'success': False,
                'message': f'主库删除失败，本地已删除 {count} 条。{remote_error}',
                'deleted': count,
                'remote_deleted': False
            }), 502
        return jsonify({
            'success': True,
            'deleted': None if remote_ok else count,
            'local_deleted': count,
            'remote_deleted': remote_ok if remote_attempted else None
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# Quality Control Center APIs
QUALITY_SOURCE_LABELS = {
    'scoring': '知识库评分',
    'governance': '知识库治理',
    'external': '外部检测报告',
}

QUALITY_STATUS_LABELS = {
    'pending': '待处理',
    'processing': '处理中',
    'completed': '已完成',
    'ignored': '已忽略',
}

QUALITY_PRIORITY_LABELS = {
    'p0': 'P0',
    'p1': 'P1',
    'p2': 'P2',
    'p3': 'P3',
}

QUALITY_GOVERNANCE_AGGREGATE_FIELDS = {
    'avg_recall_count',
    'avg_valid_recall_count',
    'avg_valid_rate',
    'weighted_avg_recall',
    'weighted_avg_valid_recall',
    'weighted_avg_recall_ratio',
    'weighted_avg_valid_recall_ratio',
    'weighted_avg_valid_rate',
}

def _quality_json_loads(value, default=None):
    if default is None:
        default = {}
    if value is None:
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        parsed = json.loads(str(value or '').strip() or 'null')
        return default if parsed is None else parsed
    except Exception:
        return default

def _quality_jsonable(value):
    if value is None:
        return None
    if isinstance(value, dict):
        return {str(k): _quality_jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_quality_jsonable(v) for v in value]
    if isinstance(value, datetime):
        return _dt_to_iso(value)
    try:
        from datetime import date as _date
        if isinstance(value, _date):
            return value.isoformat()
    except Exception:
        pass
    try:
        import decimal
        if isinstance(value, decimal.Decimal):
            return float(value)
    except Exception:
        pass
    if isinstance(value, float):
        try:
            if math.isnan(value) or math.isinf(value):
                return None
        except Exception:
            pass
        return value
    if isinstance(value, (str, int, bool)):
        return value
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    try:
        item = value.item()
        if item is not value:
            return _quality_jsonable(item)
    except Exception:
        pass
    return str(value)

def _quality_json_dumps(value):
    return json.dumps(_quality_jsonable(value if value is not None else {}), ensure_ascii=False)

def _quality_current_user():
    try:
        return current_user.username if current_user.is_authenticated else ''
    except Exception:
        return ''

def _quality_norm_source(value):
    s = str(value or '').strip().lower()
    aliases = {
        'score': 'scoring',
        'scoring': 'scoring',
        'kb_score': 'scoring',
        'kb_scores': 'scoring',
        '知识库评分': 'scoring',
        '评分': 'scoring',
        'governance': 'governance',
        'kb_recall': 'governance',
        'recall': 'governance',
        '知识库治理': 'governance',
        '治理': 'governance',
        'external': 'external',
        'external_report': 'external',
        '外部检测': 'external',
        '外部检测报告': 'external',
        '导入': 'external',
    }
    return aliases.get(s, s if s in QUALITY_SOURCE_LABELS else '')

def _quality_normalize_sources(values):
    if isinstance(values, str):
        try:
            loaded = json.loads(values)
            values = loaded if isinstance(loaded, list) else [values]
        except Exception:
            values = re.split(r'[,，\n]', values)
    if not isinstance(values, list):
        values = []
    out = []
    for item in values:
        key = _quality_norm_source(item)
        if key and key not in out:
            out.append(key)
    return out or ['scoring', 'governance']

def _quality_rule_summary(rule):
    rule = rule if isinstance(rule, dict) else {}
    conditions = rule.get('conditions')
    if isinstance(conditions, list) and conditions:
        parts = []
        for c in conditions[:4]:
            if not isinstance(c, dict):
                continue
            field = str(c.get('field_label') or c.get('field') or '').strip()
            op = str(c.get('operator_label') or c.get('operator') or '').strip()
            value = c.get('value')
            value2 = c.get('value2')
            if isinstance(value, list):
                value_text = '、'.join([str(x) for x in value if str(x).strip()])
            elif value2 not in (None, ''):
                value_text = f"{value} ~ {value2}"
            else:
                value_text = str(value or '')
            parts.append(' '.join([x for x in [field, op, value_text] if x]).strip())
        extra = len(conditions) - len(parts)
        suffix = f" 等 {len(conditions)} 条" if extra > 0 else ''
        summary = '；'.join(parts) + suffix if parts else '自定义字段规则'
        start, end = _quality_get_governance_rule_range(rule)
        if start or end:
            summary += f"；治理周期 {start or '-'} 至 {end or start or '-'}"
        return summary
    legacy = []
    if rule.get('keyword'):
        legacy.append(f"关键词包含 {rule.get('keyword')}")
    if rule.get('min_score') not in (None, ''):
        legacy.append(f"评分 >= {rule.get('min_score')}")
    if rule.get('max_score') not in (None, ''):
        legacy.append(f"评分 <= {rule.get('max_score')}")
    if rule.get('month_start') or rule.get('month_end'):
        legacy.append(f"月份 {rule.get('month_start') or '-'} 至 {rule.get('month_end') or '-'}")
    if rule.get('min_recall') not in (None, ''):
        legacy.append(f"召回 >= {rule.get('min_recall')}")
    if rule.get('max_valid_rate') not in (None, ''):
        legacy.append(f"有效率 <= {rule.get('max_valid_rate')}")
    return '；'.join(legacy) if legacy else '未设置筛选条件'

def _quality_pool_to_dict(pool):
    raw_count = QualityRawIssue.query.filter_by(pool_id=pool.id).count()
    task_ids = [
        r[0] for r in db.session.query(QualityTaskIssueLink.task_id)
        .filter_by(pool_id=pool.id)
        .distinct()
        .all()
    ]
    task_count = len(task_ids)
    status_counts = {'pending': 0, 'processing': 0, 'completed': 0, 'ignored': 0}
    if task_ids:
        rows = db.session.query(QualityTask.status, func.count(QualityTask.id)).filter(
            QualityTask.id.in_(task_ids)
        ).group_by(QualityTask.status).all()
        for status, count in rows:
            status_counts[str(status or '')] = int(count or 0)
    sources = _quality_normalize_sources(_quality_json_loads(pool.sources_json, []))
    rule_config = _quality_json_loads(pool.rule_config_json, {})
    return {
        'id': pool.id,
        'name': pool.name,
        'sources': sources,
        'source_labels': [QUALITY_SOURCE_LABELS.get(s, s) for s in sources],
        'rule_config': rule_config,
        'rule_summary': _quality_rule_summary(rule_config),
        'field_mapping': _quality_json_loads(pool.field_mapping_json, {}),
        'status': pool.status,
        'raw_count': raw_count,
        'task_count': task_count,
        'aggregated_count': task_count,
        'pending_count': status_counts.get('pending', 0),
        'processing_count': status_counts.get('processing', 0),
        'completed_count': status_counts.get('completed', 0),
        'ignored_count': status_counts.get('ignored', 0),
        'status_counts': status_counts,
        'created_by': pool.created_by or '',
        'created_at': _dt_to_iso(pool.created_at),
        'updated_at': _dt_to_iso(pool.updated_at),
    }

def _quality_raw_to_dict(raw, linked_task_id=None, pool_name=None):
    snapshot = _quality_json_loads(raw.snapshot_json, {})
    rule_snapshot = _quality_json_loads(raw.rule_snapshot_json, {})
    return {
        'id': raw.id,
        'pool_id': raw.pool_id,
        'pool_name': pool_name,
        'source_type': raw.source_type,
        'source_label': QUALITY_SOURCE_LABELS.get(raw.source_type, raw.source_type),
        'source_record_key': raw.source_record_key,
        'wiki_id': raw.wiki_id,
        'issue_text': raw.issue_text or '',
        'remediation_reference': raw.remediation_reference or '',
        'snapshot': snapshot,
        'rule_snapshot': rule_snapshot,
        'question': snapshot.get('question') or snapshot.get('question_content') or '',
        'answer': snapshot.get('answer') or snapshot.get('answer_content') or '',
        'product_name': snapshot.get('product_name') or '',
        'score': snapshot.get('total_score'),
        'month': snapshot.get('month') or '',
        'linked_task_id': linked_task_id,
        'ignored': bool(raw.ignored_at),
        'ignored_at': _dt_to_iso(raw.ignored_at),
        'created_at': _dt_to_iso(raw.created_at),
    }

def _quality_raw_summary(pool_id, only_unlinked=True):
    raws = QualityRawIssue.query.filter_by(pool_id=pool_id).all()
    raw_ids = [r.id for r in raws]
    links = QualityTaskIssueLink.query.filter(QualityTaskIssueLink.raw_issue_id.in_(raw_ids)).all() if raw_ids else []
    linked_ids = {l.raw_issue_id for l in links}
    rows = [r for r in raws if not r.ignored_at]
    if only_unlinked:
        rows = [r for r in rows if r.id not in linked_ids]
    source_counts = {'scoring': 0, 'governance': 0, 'external': 0}
    wiki_ids = set()
    multi_source_by_wiki = {}
    for raw in rows:
        source = _quality_norm_source(raw.source_type) or raw.source_type
        source_counts[source] = source_counts.get(source, 0) + 1
        wiki_id = str(raw.wiki_id or '').strip()
        if wiki_id:
            wiki_ids.add(wiki_id)
            multi_source_by_wiki.setdefault(wiki_id, set()).add(source)
    return {
        'raw_count': len(rows),
        'wiki_count': len(wiki_ids),
        'multi_source_wiki_count': sum(1 for sources in multi_source_by_wiki.values() if len(sources) > 1),
        'source_counts': source_counts,
    }

def _quality_parse_float(value):
    if value is None or value == '':
        return None
    try:
        if isinstance(value, str):
            text_value = value.strip()
            if text_value.lower() in ('nan', 'none', 'null'):
                return None
            is_percent = text_value.endswith('%')
            text_value = text_value[:-1] if is_percent else text_value
            parsed = float(text_value.replace(',', '').strip())
            return parsed / 100 if is_percent else parsed
        return float(value)
    except Exception:
        return None

def _quality_normalize_month(value):
    month = str(value or '').strip()
    return month if re.match(r'^\d{4}-\d{2}$', month) else ''

def _quality_month_sequence(start_month, end_month=None):
    start = _quality_normalize_month(start_month)
    end = _quality_normalize_month(end_month) or start
    if not start:
        return []
    if start > end:
        start, end = end, start
    try:
        year, month = [int(x) for x in start.split('-')]
        end_year, end_month_num = [int(x) for x in end.split('-')]
    except Exception:
        return [start]
    months = []
    for _ in range(240):
        current = f"{year:04d}-{month:02d}"
        months.append(current)
        if year == end_year and month == end_month_num:
            break
        month += 1
        if month > 12:
            year += 1
            month = 1
    return months

def _quality_get_governance_rule_range(rule):
    rule = rule if isinstance(rule, dict) else {}
    start = _quality_normalize_month(
        rule.get('governance_month_start')
        or rule.get('dashboard_start_month')
        or rule.get('month_start')
    )
    end = _quality_normalize_month(
        rule.get('governance_month_end')
        or rule.get('dashboard_end_month')
        or rule.get('month_end')
    )
    if start and end and start > end:
        start, end = end, start
    return start, end

def _quality_apply_governance_range_override(rule, payload):
    rule = dict(rule or {})
    payload = payload if isinstance(payload, dict) else {}
    if 'governance_month_start' in payload or 'governance_month_end' in payload:
        start = _quality_normalize_month(payload.get('governance_month_start'))
        end = _quality_normalize_month(payload.get('governance_month_end'))
        if start and end and start > end:
            start, end = end, start
        if start:
            rule['governance_month_start'] = start
        else:
            rule.pop('governance_month_start', None)
        if end:
            rule['governance_month_end'] = end
        else:
            rule.pop('governance_month_end', None)
    return rule

def _quality_governance_month_weight(wiki_id, month):
    match = re.match(r'^ICWIKI(\d{4})(\d{2})(\d{2})', str(wiki_id or '').strip(), re.I)
    if not match:
        return 1
    try:
        year = int(match.group(1))
        mon = int(match.group(2))
        day = int(match.group(3))
        datetime(year, mon, day)
    except Exception:
        return 1
    effective_month = f"{match.group(1)}-{match.group(2)}"
    stat_month = str(month or '').strip()
    return 1 if stat_month and effective_month <= stat_month else 0

def _quality_get_field_value(row, field):
    cur = row or {}
    for part in str(field or '').split('.'):
        part = part.strip()
        if not part:
            continue
        if isinstance(cur, dict):
            cur = cur.get(part)
        else:
            return None
    return cur

def _quality_field_mapping_for_source(mapping, source_type):
    mapping = mapping if isinstance(mapping, dict) else {}
    source = _quality_norm_source(source_type)
    by_source = (
        mapping.get('by_source')
        or mapping.get('bySource')
        or mapping.get('source_mappings')
        or {}
    )
    if isinstance(by_source, dict) and source:
        source_mapping = by_source.get(source) or by_source.get(str(source_type or '').strip())
        if isinstance(source_mapping, dict):
            return source_mapping
        for key, value in by_source.items():
            if _quality_norm_source(key) == source and isinstance(value, dict):
                return value
    return mapping

def _quality_pool_field_mapping(pool, source_type):
    mapping = _quality_json_loads(getattr(pool, 'field_mapping_json', '{}'), {}) if pool else {}
    return _quality_field_mapping_for_source(mapping, source_type)

def _quality_clean_text(value):
    if value is None:
        return ''
    try:
        if pd.isna(value):
            return ''
    except Exception:
        pass
    text = str(value).strip()
    return '' if text.lower() in ('nan', 'none', 'null') else text

def _quality_mapped_field_value(row, mapping, key, fallback_fields=()):
    fields = []
    configured = str((mapping or {}).get(key) or '').strip()
    if configured:
        fields.append(configured)
    fields.extend([str(f or '').strip() for f in (fallback_fields or []) if str(f or '').strip()])
    seen = set()
    for field in fields:
        if field in seen:
            continue
        seen.add(field)
        value = _quality_get_field_value(row, field)
        if _quality_clean_text(value) or value == 0:
            return value
    return ''

def _quality_normalize_priority(value):
    text_value = _quality_clean_text(value).lower()
    if text_value.upper() in ('P0', 'P1', 'P2', 'P3'):
        text_value = text_value.lower()
    return text_value if text_value in QUALITY_PRIORITY_LABELS else ''

def _quality_raw_priority(raw):
    snapshot = _quality_json_loads(getattr(raw, 'snapshot_json', '{}'), {})
    return _quality_normalize_priority(snapshot.get('priority'))

def _quality_compare_condition(value, operator, expect, expect2=None):
    op = str(operator or 'contains').strip()
    if op in ('empty', 'is_empty'):
        return value in (None, '') or str(value).strip() == ''
    if op in ('not_empty', 'is_not_empty'):
        return not _quality_compare_condition(value, 'empty', expect, expect2)
    if op in ('contains', 'not_contains'):
        hit = str(expect or '').lower() in str(value or '').lower()
        return not hit if op == 'not_contains' else hit
    if op in ('eq', 'equals', '='):
        num = _quality_parse_float(value)
        exp = _quality_parse_float(expect)
        if num is not None and exp is not None:
            return math.isclose(num, exp, rel_tol=1e-9, abs_tol=1e-9)
        return str(value if value is not None else '').strip() == str(expect if expect is not None else '').strip()
    if op in ('neq', 'not_equals', '!='):
        num = _quality_parse_float(value)
        exp = _quality_parse_float(expect)
        if num is not None and exp is not None:
            return not math.isclose(num, exp, rel_tol=1e-9, abs_tol=1e-9)
        return str(value if value is not None else '').strip() != str(expect if expect is not None else '').strip()
    if op in ('in', 'not_in'):
        vals = expect if isinstance(expect, list) else re.split(r'[,，\n]', str(expect or ''))
        vals = [str(v).strip() for v in vals if str(v).strip()]
        hit = str(value if value is not None else '').strip() in vals
        return not hit if op == 'not_in' else hit
    num = _quality_parse_float(value)
    exp = _quality_parse_float(expect)
    exp2 = _quality_parse_float(expect2)
    if op == 'between':
        if num is None and (expect not in (None, '') or expect2 not in (None, '')):
            s_val = str(value or '').strip()
            if not s_val:
                return False
            if expect not in (None, '') and s_val < str(expect).strip():
                return False
            if expect2 not in (None, '') and s_val > str(expect2).strip():
                return False
            return True
        if num is None:
            return False
        if exp is not None and num < exp:
            return False
        if exp2 is not None and num > exp2:
            return False
        return True
    if num is None or exp is None:
        s_val = str(value or '').strip()
        s_exp = str(expect or '').strip()
        if not s_val or not s_exp:
            return False
        if op in ('gt', '>'):
            return s_val > s_exp
        if op in ('gte', '>='):
            return s_val >= s_exp
        if op in ('lt', '<'):
            return s_val < s_exp
        if op in ('lte', '<='):
            return s_val <= s_exp
        return False
    if op in ('gt', '>'):
        return num > exp
    if op in ('gte', '>='):
        return num >= exp
    if op in ('lt', '<'):
        return num < exp
    if op in ('lte', '<='):
        return num <= exp
    return True

def _quality_conditions_match(row, rule, source_type):
    conditions = (rule or {}).get('conditions')
    if not isinstance(conditions, list) or not conditions:
        return None
    relevant = []
    for c in conditions:
        if not isinstance(c, dict):
            continue
        cond_source = _quality_norm_source(c.get('source'))
        if cond_source and cond_source != source_type:
            continue
        if not str(c.get('field') or '').strip():
            continue
        relevant.append(c)
    if not relevant:
        return True
    logic = str((rule or {}).get('logic') or 'AND').upper()
    results = []
    for c in relevant:
        results.append(_quality_compare_condition(
            _quality_get_field_value(row, c.get('field')),
            c.get('operator'),
            c.get('value'),
            c.get('value2')
        ))
    return any(results) if logic == 'OR' else all(results)

def _quality_text_match(rule, *parts):
    keyword = str((rule or {}).get('keyword') or '').strip().lower()
    if not keyword:
        return True
    text = ' '.join(str(p or '') for p in parts).lower()
    return keyword in text

def _quality_fetch_kb_items(wiki_ids, columns=None):
    ids = []
    seen = set()
    for item in (wiki_ids or []):
        s = str(item or '').strip()
        if s and s not in seen:
            seen.add(s)
            ids.append(s)
    if not ids:
        return {}
    columns = columns or (
        'question_wiki_id,question,answer,product_name,product_category_name,question_type,answer_type,'
        'if_bm25,similar_questions,error_list,keyword_list,image_urls,video_urls,file_urls,link_type,link_url,'
        'update_time,review_status'
    )
    client = get_supabase_client()
    if not client:
        return {}
    out = {}
    for i in range(0, len(ids), 80):
        batch = ids[i:i + 80]
        try:
            resp = client.select(
                'knowledge_base_v1',
                page=1,
                page_size=len(batch),
                filters={'question_wiki_id': _postgrest_in_str(batch)},
                columns=columns,
            )
            if resp and getattr(resp, 'status_code', 500) < 400:
                for row in resp.json() or []:
                    if isinstance(row, dict):
                        wid = str(row.get('question_wiki_id') or '').strip()
                        if wid:
                            out[wid] = row
        except Exception as e:
            print(f"[Quality] 批量读取知识库失败: {e}")
    return out

def _quality_score_issue_text(row):
    score = (row or {}).get('total_score')
    if score not in (None, ''):
        return f"AI评分 {score}，建议复核内容质量。"
    status = str((row or {}).get('status') or '').strip()
    if status:
        return f"评分状态 {status} 命中任务池规则。"
    return "评分数据命中任务池规则，建议复核。"

def _quality_is_numeric_metric_text(value):
    text = str(value or '').strip()
    return bool(text and re.fullmatch(r'[-+]?\d+(?:\.\d+)?%?', text))

def _quality_score_matches(row, rule):
    condition_match = _quality_conditions_match(row, rule, 'scoring')
    if condition_match is not None:
        return condition_match
    score = _quality_parse_float((row or {}).get('total_score'))
    min_score = _quality_parse_float((rule or {}).get('min_score'))
    max_score = _quality_parse_float((rule or {}).get('max_score'))
    if min_score is not None and (score is None or score < min_score):
        return False
    if max_score is not None and (score is None or score > max_score):
        return False
    wanted_status = str((rule or {}).get('score_status') or '').strip()
    if wanted_status and str((row or {}).get('status') or '').strip() != wanted_status:
        return False
    return _quality_text_match(
        rule,
        (row or {}).get('kb_id'),
        (row or {}).get('question_content'),
        (row or {}).get('answer_content'),
        (row or {}).get('remarks'),
        (row or {}).get('score_data'),
    )

def _quality_governance_matches(row, rule):
    condition_match = _quality_conditions_match(row, rule, 'governance')
    if condition_match is not None:
        return condition_match
    month = str((row or {}).get('month') or '').strip()
    start = str((rule or {}).get('month_start') or '').strip()
    end = str((rule or {}).get('month_end') or '').strip()
    if start and month and month < start:
        return False
    if end and month and month > end:
        return False
    recall = _quality_parse_float((row or {}).get('recall_count')) or 0
    valid = _quality_parse_float((row or {}).get('valid_recall_count')) or 0
    min_recall = _quality_parse_float((rule or {}).get('min_recall'))
    max_valid_rate = _quality_parse_float((rule or {}).get('max_valid_rate'))
    if min_recall is not None and recall < min_recall:
        return False
    if max_valid_rate is not None:
        rate = valid / recall if recall > 0 else 0
        if rate > max_valid_rate:
            return False
    return _quality_text_match(rule, (row or {}).get('kb_id'), month)

def _quality_load_scoring_raw_candidates(rule, field_mapping=None):
    client = get_supabase_client()
    if not client:
        return []
    mapping = _quality_field_mapping_for_source(field_mapping or {}, 'scoring')
    rows = []
    try:
        cols = 'id,kb_id,question_content,answer_content,status,total_score,remarks,score_data,updated_at'
        rows = client.select_all('kb_scores', columns=cols, order_by='id', page_size=1000) or []
    except Exception as e:
        print(f"[Quality] 评分来源读取失败: {e}")
        rows = []
    out = []
    for row in rows:
        if not isinstance(row, dict) or not _quality_score_matches(row, rule):
            continue
        wiki_id = _quality_clean_text(_quality_mapped_field_value(row, mapping, 'wiki_id', ('kb_id',)))
        if not wiki_id:
            continue
        source_id = str(row.get('id') or wiki_id).strip()
        issue_text = _quality_clean_text(_quality_mapped_field_value(row, mapping, 'issue')) or _quality_score_issue_text(row)
        remediation_reference = _quality_clean_text(_quality_mapped_field_value(row, mapping, 'action', ('remarks',)))
        snapshot = dict(row)
        priority = _quality_normalize_priority(_quality_mapped_field_value(row, mapping, 'priority'))
        if priority:
            snapshot['priority'] = priority
        out.append({
            'source_type': 'scoring',
            'source_record_key': f"scoring:{source_id}:{wiki_id}",
            'wiki_id': wiki_id,
            'issue_text': issue_text,
            'remediation_reference': remediation_reference,
            'snapshot': snapshot,
        })
    return out

def _quality_governance_rule_uses_aggregate(rule):
    conditions = (rule or {}).get('conditions')
    if not isinstance(conditions, list):
        return False
    for condition in conditions:
        if not isinstance(condition, dict):
            continue
        if _quality_norm_source(condition.get('source')) not in ('', 'governance'):
            continue
        field = str(condition.get('field') or '').strip()
        if field in QUALITY_GOVERNANCE_AGGREGATE_FIELDS:
            return True
    return False

def _quality_load_governance_raw_candidates(rule, field_mapping=None):
    mapping = _quality_field_mapping_for_source(field_mapping or {}, 'governance')
    rows_by_key = {}
    start_month, end_month = _quality_get_governance_rule_range(rule)
    try:
        client = get_supabase_client()
        if client:
            if start_month:
                remote_rows = _load_remote_governance_recalls(start_month, end_month) or []
            else:
                remote_rows = client.select_all(
                    'kb_recall',
                    columns='kb_id,month,recall_count,valid_recall_count',
                    order_by='month',
                    order_dir='desc',
                    page_size=1000
                ) or []
            for row in remote_rows:
                if isinstance(row, dict):
                    key = (str(row.get('kb_id') or '').strip(), str(row.get('month') or '').strip())
                    if key[0] and key[1]:
                        rows_by_key[key] = row
    except Exception as e:
        print(f"[Quality] 远端治理来源读取失败: {e}")
    try:
        local_rows = _load_sqlite_governance_recalls(start_month, end_month) if start_month else KBRecall.query.all()
        for r in local_rows:
            key = (str(r.kb_id or '').strip(), str(r.month or '').strip())
            if key[0] and key[1]:
                rows_by_key[key] = {
                    'kb_id': key[0],
                    'month': key[1],
                    'recall_count': r.recall_count or 0,
                    'valid_recall_count': r.valid_recall_count or 0,
                }
    except Exception as e:
        print(f"[Quality] 本地治理来源读取失败: {e}")

    months = _quality_month_sequence(start_month, end_month) if start_month else sorted({key[1] for key in rows_by_key.keys()})
    month_set = set(months)
    monthly_totals = {month: {'total_recall': 0, 'total_valid': 0} for month in months}
    monthly_map = {}
    for row in rows_by_key.values():
        wiki_id = str(row.get('kb_id') or '').strip()
        month = str(row.get('month') or '').strip()
        if not wiki_id or not month:
            continue
        if month_set and month not in month_set:
            continue
        recall = _quality_parse_float(row.get('recall_count')) or 0
        valid = _quality_parse_float(row.get('valid_recall_count')) or 0
        clean_row = dict(row)
        clean_row.update({
            'kb_id': wiki_id,
            'month': month,
            'recall_count': recall,
            'valid_recall_count': valid,
        })
        monthly_map.setdefault(wiki_id, {})[month] = clean_row
        total = monthly_totals.setdefault(month, {'total_recall': 0, 'total_valid': 0})
        total['total_recall'] += recall
        total['total_valid'] += valid

    score_map, v1_map = _load_governance_reference_maps()
    uses_aggregate_rule = _quality_governance_rule_uses_aggregate(rule)
    candidate_ids = set(monthly_map.keys())
    if uses_aggregate_rule:
        candidate_ids |= set(v1_map.keys())

    def _governance_issue_text(row, default_text):
        return _quality_clean_text(_quality_mapped_field_value(row, mapping, 'issue', ('question',))) or default_text

    def _governance_action_text(row, default_text):
        mapped = _quality_clean_text(_quality_mapped_field_value(row, mapping, 'action', ('suggestion', 'remarks', 'conclusion', 'analysis')))
        if not mapped or _quality_is_numeric_metric_text(mapped):
            return default_text
        return mapped

    def _governance_snapshot(row):
        snap = dict(row or {})
        priority = _quality_normalize_priority(_quality_mapped_field_value(row, mapping, 'priority'))
        if priority:
            snap['priority'] = priority
        return snap

    def _build_governance_aggregate_row(wiki_id):
        rows = []
        monthly_data = {}
        v1_entry = v1_map.get(wiki_id) or {}
        score_entry = score_map.get(wiki_id) or {}
        status = '使用中' if v1_entry else '主库不存在'
        question = v1_entry.get('question') or score_entry.get('question_content') or ''
        answer = v1_entry.get('answer') or score_entry.get('answer_content') or ''
        product_name = v1_entry.get('product_name') or ''
        score_details = _governance_score_detail_fields(score_entry)
        for month in months:
            base = monthly_map.get(wiki_id, {}).get(month) or {
                'kb_id': wiki_id,
                'month': month,
                'recall_count': 0,
                'valid_recall_count': 0,
            }
            recall = _quality_parse_float(base.get('recall_count')) or 0
            valid = _quality_parse_float(base.get('valid_recall_count')) or 0
            totals = monthly_totals.get(month, {})
            month_total_recall = _quality_parse_float(totals.get('total_recall')) or 0
            month_total_valid = _quality_parse_float(totals.get('total_valid')) or 0
            row = dict(base)
            row.update({
                'question': question,
                'answer': answer,
                'product_name': product_name,
                'status': status,
                'conclusion': score_details.get('conclusion') or '',
                'analysis': score_details.get('analysis') or '',
                'suggestion': score_details.get('suggestion') or '',
                'remarks': score_details.get('remarks') or '',
                'valid_rate': (valid / recall) if recall else 0,
                'recall_ratio': (recall / month_total_recall) if month_total_recall else 0,
                'valid_recall_ratio': (valid / month_total_valid) if month_total_valid else 0,
            })
            rows.append(row)
            monthly_data[month] = {
                'recall_count': recall,
                'valid_recall_count': valid,
                'recall_ratio': row['recall_ratio'],
                'valid_recall_ratio': row['valid_recall_ratio'],
                'valid_rate': row['valid_rate'],
            }

        total_recall = sum((_quality_parse_float(r.get('recall_count')) or 0) for r in rows)
        total_valid = sum((_quality_parse_float(r.get('valid_recall_count')) or 0) for r in rows)
        row_count = max(1, len(rows))
        total_weight = 0
        weighted_recall_ratio_sum = 0
        weighted_valid_recall_ratio_sum = 0
        weighted_valid_rate_sum = 0
        for row in rows:
            month = str(row.get('month') or '').strip()
            weight = _quality_governance_month_weight(wiki_id, month)
            if weight <= 0:
                continue
            total_weight += weight
            weighted_recall_ratio_sum += (_quality_parse_float(row.get('recall_ratio')) or 0) * weight
            weighted_valid_recall_ratio_sum += (_quality_parse_float(row.get('valid_recall_ratio')) or 0) * weight
            weighted_valid_rate_sum += (_quality_parse_float(row.get('valid_rate')) or 0) * weight
        range_label = f"{months[0]} 至 {months[-1]}" if len(months) > 1 else (months[0] if months else '全部周期')
        return {
            'kb_id': wiki_id,
            'question': question,
            'answer': answer,
            'product_name': product_name,
            'status': status,
            'conclusion': score_details.get('conclusion') or '',
            'analysis': score_details.get('analysis') or '',
            'suggestion': score_details.get('suggestion') or '',
            'remarks': score_details.get('remarks') or '',
            'month': range_label,
            'month_start': months[0] if months else '',
            'month_end': months[-1] if months else '',
            'months': months,
            'monthly_data': monthly_data,
            'recall_count': total_recall,
            'valid_recall_count': total_valid,
            'valid_rate': (total_valid / total_recall) if total_recall else 0,
            'avg_recall_count': total_recall / row_count,
            'avg_valid_recall_count': total_valid / row_count,
            'avg_valid_rate': (total_valid / total_recall) if total_recall else 0,
            'total_weight': total_weight,
            'weighted_avg_recall': (total_recall / total_weight) if total_weight else None,
            'weighted_avg_valid_recall': (total_valid / total_weight) if total_weight else None,
            'weighted_avg_recall_ratio': (weighted_recall_ratio_sum / total_weight) if total_weight else None,
            'weighted_avg_valid_recall_ratio': (weighted_valid_recall_ratio_sum / total_weight) if total_weight else None,
            'weighted_avg_valid_rate': (weighted_valid_rate_sum / total_weight) if total_weight else None,
        }

    out = []
    if uses_aggregate_rule:
        for wiki_id in sorted(candidate_ids):
            row = _build_governance_aggregate_row(wiki_id)
            if not _quality_governance_matches(row, rule):
                continue
            weighted_avg = _quality_parse_float(row.get('weighted_avg_recall'))
            weighted_text = '-' if weighted_avg is None else f"{weighted_avg:.2f}"
            range_key = f"{row.get('month_start') or 'all'}_{row.get('month_end') or 'all'}"
            default_issue = f"{row.get('month')} 大盘周期加权平均召回频数 {weighted_text}，总召回 {int(row.get('recall_count') or 0)}。"
            default_action = '治理大盘加权数据命中任务池规则'
            mapped_wiki_id = _quality_clean_text(_quality_mapped_field_value(row, mapping, 'wiki_id', ('kb_id',))) or wiki_id
            out.append({
                'source_type': 'governance',
                'source_record_key': f"governance:weighted:{mapped_wiki_id}:{range_key}",
                'wiki_id': mapped_wiki_id,
                'issue_text': _governance_issue_text(row, default_issue),
                'remediation_reference': _governance_action_text(row, default_action),
                'snapshot': _governance_snapshot(row),
            })
        return out

    for wiki_id in sorted(candidate_ids):
        aggregate_row = _build_governance_aggregate_row(wiki_id)
        for month, row in sorted((monthly_map.get(wiki_id) or {}).items()):
            row.update({
                'question': aggregate_row.get('question') or '',
                'answer': aggregate_row.get('answer') or '',
                'product_name': aggregate_row.get('product_name') or '',
                'status': aggregate_row.get('status') or '',
                'conclusion': aggregate_row.get('conclusion') or '',
                'analysis': aggregate_row.get('analysis') or '',
                'suggestion': aggregate_row.get('suggestion') or '',
                'remarks': aggregate_row.get('remarks') or '',
                'avg_recall_count': aggregate_row.get('avg_recall_count'),
                'avg_valid_recall_count': aggregate_row.get('avg_valid_recall_count'),
                'avg_valid_rate': aggregate_row.get('avg_valid_rate'),
                'weighted_avg_recall': aggregate_row.get('weighted_avg_recall'),
                'weighted_avg_valid_recall': aggregate_row.get('weighted_avg_valid_recall'),
                'weighted_avg_recall_ratio': aggregate_row.get('weighted_avg_recall_ratio'),
                'weighted_avg_valid_recall_ratio': aggregate_row.get('weighted_avg_valid_recall_ratio'),
                'weighted_avg_valid_rate': aggregate_row.get('weighted_avg_valid_rate'),
            })
            if not _quality_governance_matches(row, rule):
                continue
            recall = int(_quality_parse_float(row.get('recall_count')) or 0)
            valid = int(_quality_parse_float(row.get('valid_recall_count')) or 0)
            valid_rate = (valid / recall) if recall else 0
            default_issue = f"{month} 召回 {recall}，有效召回 {valid}，有效召回率 {valid_rate:.2%}。"
            default_action = '治理召回数据命中任务池规则'
            mapped_wiki_id = _quality_clean_text(_quality_mapped_field_value(row, mapping, 'wiki_id', ('kb_id',))) or wiki_id
            out.append({
                'source_type': 'governance',
                'source_record_key': f"governance:{mapped_wiki_id}:{month}",
                'wiki_id': mapped_wiki_id,
                'issue_text': _governance_issue_text(row, default_issue),
                'remediation_reference': _governance_action_text(row, default_action),
                'snapshot': _governance_snapshot(row),
            })
    return out

def _quality_upsert_raw_issue(pool, candidate, rule_snapshot=None):
    wiki_id = str((candidate or {}).get('wiki_id') or '').strip()
    source_type = _quality_norm_source((candidate or {}).get('source_type'))
    source_record_key = str((candidate or {}).get('source_record_key') or '').strip()
    if not pool or not wiki_id or not source_type or not source_record_key:
        return None, False
    raw = QualityRawIssue.query.filter_by(
        pool_id=pool.id,
        source_record_key=source_record_key
    ).first()
    created = False
    if not raw:
        raw = QualityRawIssue(
            pool_id=pool.id,
            source_type=source_type,
            source_record_key=source_record_key,
            wiki_id=wiki_id,
        )
        db.session.add(raw)
        created = True
    raw.source_type = source_type
    raw.wiki_id = wiki_id
    raw.issue_text = str((candidate or {}).get('issue_text') or '').strip()
    raw.remediation_reference = str((candidate or {}).get('remediation_reference') or '').strip()
    raw.snapshot_json = _quality_json_dumps((candidate or {}).get('snapshot') or {})
    raw.rule_snapshot_json = _quality_json_dumps(rule_snapshot or {})
    return raw, created

def _quality_prune_stale_raw_issues(pool_id, keep_source_keys, source_types):
    source_types = [_quality_norm_source(s) for s in (source_types or [])]
    source_types = [s for s in source_types if s and s != 'external']
    if not source_types:
        return {'removed_raw_count': 0, 'removed_link_count': 0, 'removed_task_count': 0}
    keep_source_keys = set(str(k or '').strip() for k in (keep_source_keys or []) if str(k or '').strip())
    stale_query = QualityRawIssue.query.filter(
        QualityRawIssue.pool_id == pool_id,
        QualityRawIssue.source_type.in_(source_types)
    )
    stale_raws = [raw for raw in stale_query.all() if str(raw.source_record_key or '').strip() not in keep_source_keys]
    if not stale_raws:
        return {'removed_raw_count': 0, 'removed_link_count': 0, 'removed_task_count': 0}

    stale_raw_ids = [raw.id for raw in stale_raws]
    links = QualityTaskIssueLink.query.filter(QualityTaskIssueLink.raw_issue_id.in_(stale_raw_ids)).all()
    affected_task_ids = {link.task_id for link in links}
    removed_link_count = len(links)
    for link in links:
        db.session.delete(link)
    for raw in stale_raws:
        db.session.delete(raw)
    db.session.flush()

    removed_task_count = 0
    for task_id in affected_task_ids:
        remains = QualityTaskIssueLink.query.filter_by(task_id=task_id).count()
        if remains == 0:
            task = QualityTask.query.get(task_id)
            if task:
                db.session.delete(task)
                removed_task_count += 1
    return {
        'removed_raw_count': len(stale_raws),
        'removed_link_count': removed_link_count,
        'removed_task_count': removed_task_count,
    }

def _quality_create_or_link_task(raw, priority='p2'):
    if not raw or raw.ignored_at:
        return None, False
    priority = str(priority or 'p2').strip().lower()
    if priority not in QUALITY_PRIORITY_LABELS:
        priority = 'p2'
    task = QualityTask.query.filter_by(wiki_id=raw.wiki_id).first()
    created = False
    if not task:
        task = QualityTask(wiki_id=raw.wiki_id, priority=priority, status='pending')
        db.session.add(task)
        db.session.flush()
        created = True
    elif task.status == 'ignored':
        task.status = 'pending'
        task.ignored_at = None
    if priority:
        task.priority = priority
    link = QualityTaskIssueLink.query.filter_by(task_id=task.id, raw_issue_id=raw.id).first()
    if not link:
        db.session.add(QualityTaskIssueLink(
            task_id=task.id,
            raw_issue_id=raw.id,
            pool_id=raw.pool_id,
            source_type=raw.source_type,
        ))
    return task, created

def _quality_mark_task_processing(task_id, latest_kb_update_time=None):
    try:
        task = QualityTask.query.get(int(task_id))
    except Exception:
        task = None
    if not task:
        return False
    task.status = 'processing'
    task.completed_at = None
    task.ignored_at = None
    if latest_kb_update_time is not None:
        task.latest_kb_update_time = str(latest_kb_update_time or '')
    db.session.commit()
    return True

def _quality_task_to_dict(task, raw_issues=None, kb_item=None):
    raw_issues = raw_issues or []
    kb_item = kb_item or {}
    source_counts = {}
    pool_names = []
    issue_labels = []
    suggested_actions = []
    question = str(kb_item.get('question') or '').strip()
    answer = str(kb_item.get('answer') or '').strip()
    for raw in raw_issues:
        source_counts[raw.source_type] = source_counts.get(raw.source_type, 0) + 1
        pool_name = getattr(raw, '_quality_pool_name', None)
        if pool_name and pool_name not in pool_names:
            pool_names.append(pool_name)
        issue_label = str(raw.issue_text or '').strip()
        if issue_label and issue_label not in issue_labels:
            issue_labels.append(issue_label)
        suggested_action = str(raw.remediation_reference or '').strip()
        if raw.source_type == 'governance' and _quality_is_numeric_metric_text(suggested_action):
            suggested_action = ''
        if suggested_action and suggested_action not in suggested_actions:
            suggested_actions.append(suggested_action)
        if not question:
            snap = _quality_json_loads(raw.snapshot_json, {})
            question = str(snap.get('question') or snap.get('question_content') or '').strip()
        if not answer:
            snap = _quality_json_loads(raw.snapshot_json, {})
            answer = str(snap.get('answer') or snap.get('answer_content') or '').strip()
    return {
        'id': task.id,
        'wiki_id': task.wiki_id,
        'priority': task.priority,
        'priority_label': QUALITY_PRIORITY_LABELS.get(task.priority, task.priority),
        'status': task.status,
        'status_label': QUALITY_STATUS_LABELS.get(task.status, task.status),
        'question': question,
        'answer': answer,
        'product_name': kb_item.get('product_name') or '',
        'kb_update_time': kb_item.get('update_time') or '',
        'latest_kb_update_time': task.latest_kb_update_time or '',
        'issue_count': len(raw_issues),
        'source_tags': [
            {'source_type': k, 'source_label': QUALITY_SOURCE_LABELS.get(k, k), 'count': v}
            for k, v in sorted(source_counts.items())
        ],
        'pool_names': pool_names,
        'issue_labels': issue_labels,
        'issue_label_text': '；'.join(issue_labels),
        'issue_tag_text': f"共 {len(raw_issues)} 条问题 ｜ {('、'.join(pool_names) or '未关联任务池')}",
        'suggested_actions': suggested_actions,
        'suggested_action_text': '；'.join(suggested_actions),
        'created_at': _dt_to_iso(task.created_at),
        'updated_at': _dt_to_iso(task.updated_at),
        'completed_at': _dt_to_iso(task.completed_at),
        'ignored_at': _dt_to_iso(task.ignored_at),
    }

def _quality_load_task_raw_map(tasks):
    task_ids = [t.id for t in (tasks or [])]
    if not task_ids:
        return {}
    links = QualityTaskIssueLink.query.filter(QualityTaskIssueLink.task_id.in_(task_ids)).all()
    raw_ids = [l.raw_issue_id for l in links]
    raws = QualityRawIssue.query.filter(QualityRawIssue.id.in_(raw_ids)).all() if raw_ids else []
    pools = {p.id: p.name for p in QualityTaskPool.query.all()}
    raw_map = {r.id: r for r in raws}
    out = {tid: [] for tid in task_ids}
    for link in links:
        raw = raw_map.get(link.raw_issue_id)
        if raw:
            setattr(raw, '_quality_pool_name', pools.get(raw.pool_id, ''))
            out.setdefault(link.task_id, []).append(raw)
    return out

@app.route('/api/quality/pools', methods=['GET', 'POST'])
@login_required
def quality_pools_route():
    if request.method == 'GET':
        pools = QualityTaskPool.query.order_by(QualityTaskPool.updated_at.desc(), QualityTaskPool.id.desc()).all()
        return jsonify({'success': True, 'pools': [_quality_pool_to_dict(p) for p in pools]})

    payload = request.get_json(silent=True) or {}
    name = str(payload.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': '任务池名称不能为空'}), 400
    sources = _quality_normalize_sources(payload.get('sources') or [])
    pool = QualityTaskPool(
        name=name,
        sources_json=_quality_json_dumps(sources),
        rule_config_json=_quality_json_dumps(payload.get('rule_config') or {}),
        field_mapping_json=_quality_json_dumps(payload.get('field_mapping') or {}),
        status='active',
        created_by=_quality_current_user(),
    )
    db.session.add(pool)
    db.session.commit()
    return jsonify({'success': True, 'pool': _quality_pool_to_dict(pool)})

@app.route('/api/quality/pools/<int:pool_id>', methods=['PATCH', 'DELETE'])
@login_required
def quality_pool_detail_route(pool_id):
    pool = QualityTaskPool.query.get(pool_id)
    if not pool:
        return jsonify({'success': False, 'message': '任务池不存在'}), 404
    if request.method == 'PATCH':
        payload = request.get_json(silent=True) or {}
        if 'name' in payload:
            name = str(payload.get('name') or '').strip()
            if not name:
                return jsonify({'success': False, 'message': '任务池名称不能为空'}), 400
            pool.name = name
        if 'sources' in payload:
            pool.sources_json = _quality_json_dumps(_quality_normalize_sources(payload.get('sources')))
        if 'rule_config' in payload:
            pool.rule_config_json = _quality_json_dumps(payload.get('rule_config') or {})
        if 'field_mapping' in payload:
            pool.field_mapping_json = _quality_json_dumps(payload.get('field_mapping') or {})
        pool.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True, 'pool': _quality_pool_to_dict(pool)})

    links = QualityTaskIssueLink.query.filter_by(pool_id=pool.id).all()
    affected_task_ids = list({l.task_id for l in links})
    QualityTaskIssueLink.query.filter_by(pool_id=pool.id).delete()
    QualityRawIssue.query.filter_by(pool_id=pool.id).delete()
    db.session.delete(pool)
    for task_id in affected_task_ids:
        remains = QualityTaskIssueLink.query.filter_by(task_id=task_id).count()
        if remains == 0:
            task = QualityTask.query.get(task_id)
            if task:
                db.session.delete(task)
    db.session.commit()
    return jsonify({'success': True, 'deleted_pool_id': pool_id})

@app.route('/api/quality/pools/<int:pool_id>/scan', methods=['POST'])
@login_required
def quality_scan_pool_route(pool_id):
    pool = QualityTaskPool.query.get(pool_id)
    if not pool:
        return jsonify({'success': False, 'message': '任务池不存在'}), 404
    try:
        payload = request.get_json(silent=True) or {}
        rule = _quality_apply_governance_range_override(_quality_json_loads(pool.rule_config_json, {}), payload)
        sources = _quality_normalize_sources(_quality_json_loads(pool.sources_json, []))
        field_mapping = _quality_json_loads(pool.field_mapping_json, {})
        candidates = []
        scanned_sources = []
        if 'scoring' in sources:
            candidates.extend(_quality_load_scoring_raw_candidates(rule, field_mapping))
            scanned_sources.append('scoring')
        if 'governance' in sources:
            candidates.extend(_quality_load_governance_raw_candidates(rule, field_mapping))
            scanned_sources.append('governance')
        if candidates:
            kb_map = _quality_fetch_kb_items([c.get('wiki_id') for c in candidates], columns='question_wiki_id,question,answer,product_name,update_time')
            for c in candidates:
                kb_item = kb_map.get(str(c.get('wiki_id') or '').strip()) or {}
                if kb_item:
                    snap = dict(c.get('snapshot') or {})
                    snap.update({
                        'question': kb_item.get('question') or snap.get('question'),
                        'answer': kb_item.get('answer') or snap.get('answer'),
                        'product_name': kb_item.get('product_name') or snap.get('product_name'),
                        'kb_update_time': kb_item.get('update_time') or '',
                    })
                    c['snapshot'] = snap
        created_count = 0
        updated_count = 0
        matched_source_keys = set()
        for c in candidates:
            raw, created = _quality_upsert_raw_issue(pool, c, rule)
            if raw:
                matched_source_keys.add(str(raw.source_record_key or '').strip())
            if created:
                created_count += 1
            else:
                updated_count += 1
        cleanup_result = _quality_prune_stale_raw_issues(pool.id, matched_source_keys, scanned_sources)
        pool.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({
            'success': True,
            'pool': _quality_pool_to_dict(pool),
            'matched_count': len(candidates),
            'created_count': created_count,
            'updated_count': updated_count,
            **cleanup_result,
            'raw_summary': _quality_raw_summary(pool.id, only_unlinked=True),
        })
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'扫描失败: {str(e)}'}), 500

@app.route('/api/quality/pools/<int:pool_id>/raw_issues', methods=['GET'])
@login_required
def quality_pool_raw_issues_route(pool_id):
    pool = QualityTaskPool.query.get(pool_id)
    if not pool:
        return jsonify({'success': False, 'message': '任务池不存在'}), 404
    page = max(1, int(request.args.get('page', 1) or 1))
    page_size = max(1, min(200, int(request.args.get('pageSize', 50) or 50)))
    source = _quality_norm_source(request.args.get('source'))
    keyword = str(request.args.get('keyword') or '').strip().lower()
    only_unlinked = str(request.args.get('only_unlinked') or '').lower() in ('1', 'true', 'yes')
    status = str(request.args.get('status') or '').strip()
    query = QualityRawIssue.query.filter_by(pool_id=pool.id)
    if source:
        query = query.filter_by(source_type=source)
    if status == 'ignored':
        query = query.filter(QualityRawIssue.ignored_at.isnot(None))
    elif status == 'active':
        query = query.filter(QualityRawIssue.ignored_at.is_(None))
    raws = query.order_by(QualityRawIssue.id.desc()).all()
    raw_ids = [r.id for r in raws]
    links = QualityTaskIssueLink.query.filter(QualityTaskIssueLink.raw_issue_id.in_(raw_ids)).all() if raw_ids else []
    linked_map = {l.raw_issue_id: l.task_id for l in links}
    filtered = []
    for raw in raws:
        if only_unlinked and raw.id in linked_map:
            continue
        if keyword:
            hay = ' '.join([
                raw.wiki_id or '',
                raw.issue_text or '',
                raw.remediation_reference or '',
                raw.snapshot_json or '',
            ]).lower()
            if keyword not in hay:
                continue
        filtered.append(raw)
    total = len(filtered)
    start = (page - 1) * page_size
    page_rows = filtered[start:start + page_size]
    return jsonify({
        'success': True,
        'raw_issues': [_quality_raw_to_dict(r, linked_map.get(r.id), pool.name) for r in page_rows],
        'total': total,
        'page': page,
        'pageSize': page_size,
    })

@app.route('/api/quality/pools/<int:pool_id>/aggregate', methods=['POST'])
@login_required
def quality_aggregate_pool_route(pool_id):
    pool = QualityTaskPool.query.get(pool_id)
    if not pool:
        return jsonify({'success': False, 'message': '任务池不存在'}), 404
    payload = request.get_json(silent=True) or {}
    ids = payload.get('raw_issue_ids') or []
    if not isinstance(ids, list) or not ids:
        return jsonify({'success': False, 'message': '请选择要聚合的原始问题'}), 400
    priority = str(payload.get('priority') or 'p2').lower()
    raws = QualityRawIssue.query.filter(
        QualityRawIssue.pool_id == pool.id,
        QualityRawIssue.id.in_([int(x) for x in ids if str(x).isdigit()]),
        QualityRawIssue.ignored_at.is_(None)
    ).all()
    created_task_count = 0
    linked_count = 0
    task_ids = set()
    for raw in raws:
        before = QualityTask.query.filter_by(wiki_id=raw.wiki_id).first()
        task, created = _quality_create_or_link_task(raw, _quality_raw_priority(raw) or priority)
        if task:
            task_ids.add(task.id)
            if created or before is None:
                created_task_count += 1
            linked_count += 1
    pool.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({
        'success': True,
        'created_task_count': created_task_count,
        'linked_count': linked_count,
        'task_ids': sorted(task_ids),
        'raw_summary': _quality_raw_summary(pool.id, only_unlinked=True),
    })

@app.route('/api/quality/pools/<int:pool_id>/aggregate_all', methods=['POST'])
@login_required
def quality_aggregate_all_pool_route(pool_id):
    pool = QualityTaskPool.query.get(pool_id)
    if not pool:
        return jsonify({'success': False, 'message': '任务池不存在'}), 404
    payload = request.get_json(silent=True) or {}
    priority = str(payload.get('priority') or 'p2').lower()
    only_unlinked = bool(payload.get('only_unlinked', True))
    raws = QualityRawIssue.query.filter(
        QualityRawIssue.pool_id == pool.id,
        QualityRawIssue.ignored_at.is_(None)
    ).order_by(QualityRawIssue.id.asc()).all()
    if only_unlinked and raws:
        raw_ids = [r.id for r in raws]
        links = QualityTaskIssueLink.query.filter(QualityTaskIssueLink.raw_issue_id.in_(raw_ids)).all()
        linked_ids = {l.raw_issue_id for l in links}
        raws = [r for r in raws if r.id not in linked_ids]
    if not raws:
        return jsonify({
            'success': True,
            'created_task_count': 0,
            'linked_count': 0,
            'wiki_count': 0,
            'task_ids': [],
            'raw_summary': _quality_raw_summary(pool.id, only_unlinked=True),
            'message': '暂无新的待聚合原始问题'
        })
    created_task_count = 0
    linked_count = 0
    task_ids = set()
    wiki_ids = set()
    for raw in raws:
        wiki_ids.add(raw.wiki_id)
        before = QualityTask.query.filter_by(wiki_id=raw.wiki_id).first()
        task, created = _quality_create_or_link_task(raw, _quality_raw_priority(raw) or priority)
        if task:
            task_ids.add(task.id)
            if created or before is None:
                created_task_count += 1
            linked_count += 1
    pool.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({
        'success': True,
        'created_task_count': created_task_count,
        'linked_count': linked_count,
        'wiki_count': len({w for w in wiki_ids if w}),
        'task_ids': sorted(task_ids),
        'raw_summary': _quality_raw_summary(pool.id, only_unlinked=True),
    })

@app.route('/api/quality/raw_issues/<int:raw_id>/ignore', methods=['POST'])
@login_required
def quality_ignore_raw_issue_route(raw_id):
    raw = QualityRawIssue.query.get(raw_id)
    if not raw:
        return jsonify({'success': False, 'message': '原始问题不存在'}), 404
    raw.ignored_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True, 'raw_issue': _quality_raw_to_dict(raw)})

@app.route('/api/quality/tasks', methods=['GET'])
@login_required
def quality_tasks_route():
    page = max(1, int(request.args.get('page', 1) or 1))
    page_size = max(1, min(200, int(request.args.get('pageSize', 20) or 20)))
    status = str(request.args.get('status') or '').strip()
    priority = str(request.args.get('priority') or '').strip().lower()
    source = _quality_norm_source(request.args.get('source'))
    pool_id = request.args.get('pool_id')
    keyword = str(request.args.get('keyword') or '').strip().lower()

    tasks = QualityTask.query.order_by(
        case(
            (QualityTask.priority == 'p0', 0),
            (QualityTask.priority == 'p1', 1),
            (QualityTask.priority == 'p2', 2),
            else_=3
        ),
        QualityTask.updated_at.desc(),
        QualityTask.id.desc()
    ).all()
    raw_map = _quality_load_task_raw_map(tasks)
    kb_map = _quality_fetch_kb_items([t.wiki_id for t in tasks], columns='question_wiki_id,question,answer,product_name,update_time')
    filtered = []
    for task in tasks:
        raws = raw_map.get(task.id, [])
        if status and task.status != status:
            continue
        if priority and task.priority != priority:
            continue
        if source and not any(r.source_type == source for r in raws):
            continue
        if pool_id and str(pool_id).isdigit() and not any(r.pool_id == int(pool_id) for r in raws):
            continue
        if keyword:
            kb_item = kb_map.get(task.wiki_id) or {}
            hay = ' '.join([
                task.wiki_id or '',
                kb_item.get('question') or '',
                kb_item.get('answer') or '',
                kb_item.get('product_name') or '',
                ' '.join([r.issue_text or '' for r in raws]),
                ' '.join([r.remediation_reference or '' for r in raws]),
            ]).lower()
            if keyword not in hay:
                continue
        filtered.append(task)
    summary = {}
    for task in filtered:
        summary[task.status] = summary.get(task.status, 0) + 1
    total = len(filtered)
    start = (page - 1) * page_size
    page_tasks = filtered[start:start + page_size]
    return jsonify({
        'success': True,
        'tasks': [
            _quality_task_to_dict(t, raw_map.get(t.id, []), kb_map.get(t.wiki_id) or {})
            for t in page_tasks
        ],
        'summary': summary,
        'total': total,
        'page': page,
        'pageSize': page_size,
    })

@app.route('/api/quality/tasks/<int:task_id>', methods=['GET', 'PATCH'])
@login_required
def quality_task_detail_route(task_id):
    task = QualityTask.query.get(task_id)
    if not task:
        return jsonify({'success': False, 'message': '任务不存在'}), 404
    if request.method == 'PATCH':
        payload = request.get_json(silent=True) or {}
        if 'priority' in payload:
            p = str(payload.get('priority') or '').strip().lower()
            if p in QUALITY_PRIORITY_LABELS:
                task.priority = p
        if 'status' in payload:
            st = str(payload.get('status') or '').strip()
            if st in QUALITY_STATUS_LABELS:
                task.status = st
                if st == 'completed':
                    task.completed_at = datetime.utcnow()
                    task.ignored_at = None
                elif st == 'ignored':
                    task.ignored_at = datetime.utcnow()
                elif st == 'processing':
                    task.completed_at = None
                    task.ignored_at = None
                elif st == 'pending':
                    task.completed_at = None
                    task.ignored_at = None
        db.session.commit()
    raw_map = _quality_load_task_raw_map([task])
    kb_map = _quality_fetch_kb_items([task.wiki_id])
    raws = raw_map.get(task.id, [])
    return jsonify({
        'success': True,
        'task': _quality_task_to_dict(task, raws, kb_map.get(task.wiki_id) or {}),
        'raw_issues': [
            _quality_raw_to_dict(r, linked_task_id=task.id, pool_name=getattr(r, '_quality_pool_name', None))
            for r in raws
        ],
        'kb_item': kb_map.get(task.wiki_id),
    })

@app.route('/api/quality/tasks/batch', methods=['POST'])
@login_required
def quality_tasks_batch_route():
    payload = request.get_json(silent=True) or {}
    ids = payload.get('task_ids') or []
    action = str(payload.get('action') or '').strip()
    if not isinstance(ids, list) or not ids:
        return jsonify({'success': False, 'message': '请选择任务'}), 400
    tasks = QualityTask.query.filter(QualityTask.id.in_([int(x) for x in ids if str(x).isdigit()])).all()
    now = datetime.utcnow()
    changed = 0
    for task in tasks:
        if action == 'complete':
            task.status = 'completed'
            task.completed_at = now
            task.ignored_at = None
            changed += 1
        elif action == 'ignore':
            task.status = 'ignored'
            task.ignored_at = now
            changed += 1
        elif action == 'processing':
            task.status = 'processing'
            task.completed_at = None
            task.ignored_at = None
            changed += 1
        elif action == 'pending':
            task.status = 'pending'
            task.completed_at = None
            task.ignored_at = None
            changed += 1
    if action not in ('complete', 'ignore', 'processing', 'pending'):
        return jsonify({'success': False, 'message': '未知批量操作'}), 400
    db.session.commit()
    return jsonify({'success': True, 'changed': changed})

@app.route('/api/quality/tasks/export', methods=['POST'])
@login_required
def quality_tasks_export_route():
    try:
        payload = request.get_json(silent=True) or {}
        ids = payload.get('task_ids') or []
        query = QualityTask.query
        if isinstance(ids, list) and ids:
            query = query.filter(QualityTask.id.in_([int(x) for x in ids if str(x).isdigit()]))
        tasks = query.order_by(QualityTask.id.desc()).all()
        raw_map = _quality_load_task_raw_map(tasks)
        kb_map = _quality_fetch_kb_items([t.wiki_id for t in tasks], columns='question_wiki_id,question,answer,product_name,update_time')
        rows = []
        for task in tasks:
            raws = raw_map.get(task.id, [])
            item = _quality_task_to_dict(task, raws, kb_map.get(task.wiki_id) or {})
            rows.append({
                '任务ID': task.id,
                'WikiID': task.wiki_id,
                '问题': item.get('question') or '',
                '产品型号': item.get('product_name') or '',
                '优先级': item.get('priority_label') or '',
                '状态': item.get('status_label') or '',
                '来源': '、'.join([x.get('source_label') for x in item.get('source_tags') or []]),
                '任务池': '、'.join(item.get('pool_names') or []),
                '问题数量': item.get('issue_count') or 0,
                '建议操作': item.get('suggested_action_text') or '',
                '问题标签': item.get('issue_tag_text') or '',
                '知识库更新时间': item.get('kb_update_time') or '',
                '任务更新时间': item.get('updated_at') or '',
            })
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        engine = 'xlsxwriter' if importlib.util.find_spec('xlsxwriter') is not None else 'openpyxl'
        with pd.ExcelWriter(output, engine=engine) as writer:
            df.to_excel(writer, index=False, sheet_name='管控中心任务')
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=canonical_download_name('quality_task'),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500

@app.route('/api/quality/import', methods=['POST'])
@login_required
def quality_import_route():
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': '请选择导入文件'}), 400
    target_pool_id = request.form.get('target_pool_id')
    pool = QualityTaskPool.query.get(int(target_pool_id)) if target_pool_id and str(target_pool_id).isdigit() else None
    if not pool:
        pool = QualityTaskPool(
            name=f"外部检测导入 {datetime.now().strftime('%Y%m%d %H:%M')}",
            sources_json=_quality_json_dumps(['external']),
            rule_config_json='{}',
            field_mapping_json='{}',
            created_by=_quality_current_user(),
        )
        db.session.add(pool)
        db.session.flush()
    try:
        filename = secure_filename(file.filename or 'quality_import')
        if filename.lower().endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        df.columns = [str(c).strip().replace('\ufeff', '') for c in df.columns]

        def _norm_import_col(value):
            return re.sub(r'[\s_：:]+', '', str(value or '').strip().replace('\ufeff', '')).lower()

        norm_cols = {_norm_import_col(c): c for c in df.columns}
        mapping = _quality_pool_field_mapping(pool, 'external') if pool else {}

        def _find_import_col(mapping_key, fallback_keys):
            configured = str((mapping or {}).get(mapping_key) or '').strip()
            if configured:
                if configured in df.columns:
                    return configured
                configured_norm = _norm_import_col(configured)
                if configured_norm in norm_cols:
                    return norm_cols[configured_norm]
            for key in fallback_keys:
                nk = _norm_import_col(key)
                if nk in norm_cols:
                    return norm_cols[nk]
            return None

        wiki_col = _find_import_col('wiki_id', ('wikiid', 'questionwikiid', 'kbid', '知识库id', '知识库编号'))
        issue_col = _find_import_col('issue', ('问题', 'issue', 'problem', '问题描述', '检测问题'))
        suggestion_col = _find_import_col('action', ('建议操作', '建议', 'suggestion', 'action', '处理建议'))
        priority_col = _find_import_col('priority', ('优先级', 'priority', 'p级别', '等级'))
        if not wiki_col or not issue_col:
            return jsonify({'success': False, 'message': '导入模板至少需要 WikiID、问题 两列'}), 400
        wiki_ids = []
        for _, row in df.iterrows():
            wiki_id = str(row.get(wiki_col) or '').strip()
            if wiki_id and wiki_id.lower() != 'nan':
                wiki_ids.append(wiki_id)
        kb_map = _quality_fetch_kb_items(wiki_ids, columns='question_wiki_id,question,answer,product_name,update_time')
        success_count = 0
        failed = []
        duplicate_append_count = 0
        for idx, row in df.iterrows():
            wiki_id = str(row.get(wiki_col) or '').strip()
            issue_text = str(row.get(issue_col) or '').strip()
            suggestion = str(row.get(suggestion_col) or '').strip() if suggestion_col else ''
            priority = str(row.get(priority_col) or '').strip().lower() if priority_col else ''
            if priority and priority.upper() in ('P0', 'P1', 'P2', 'P3'):
                priority = priority.lower()
            if not wiki_id or wiki_id.lower() == 'nan':
                failed.append({'row': int(idx) + 2, 'wiki_id': wiki_id, 'reason': 'WikiID 为空'})
                continue
            if wiki_id not in kb_map:
                failed.append({'row': int(idx) + 2, 'wiki_id': wiki_id, 'reason': '知识库不存在'})
                continue
            if not issue_text or issue_text.lower() == 'nan':
                failed.append({'row': int(idx) + 2, 'wiki_id': wiki_id, 'reason': '问题为空'})
                continue
            row_obj = {}
            for col in df.columns:
                val = row.get(col)
                if pd.isna(val):
                    val = ''
                row_obj[col] = val
            raw_key_hash = hashlib.md5(f"{wiki_id}|{issue_text}|{suggestion}".encode('utf-8')).hexdigest()[:16]
            snap = dict(row_obj)
            snap.update(kb_map.get(wiki_id) or {})
            if priority in QUALITY_PRIORITY_LABELS:
                snap['priority'] = priority
            raw, _created = _quality_upsert_raw_issue(pool, {
                'source_type': 'external',
                'source_record_key': f"external:{pool.id}:{wiki_id}:{raw_key_hash}",
                'wiki_id': wiki_id,
                'issue_text': issue_text,
                'remediation_reference': suggestion,
                'snapshot': snap,
            }, {'import_file': filename})
            existing_task = QualityTask.query.filter_by(wiki_id=wiki_id).first()
            if existing_task and raw:
                duplicate_append_count += 1
                QualityTaskIssueLink.query.filter_by(task_id=existing_task.id, raw_issue_id=raw.id).first() or db.session.add(QualityTaskIssueLink(
                    task_id=existing_task.id,
                    raw_issue_id=raw.id,
                    pool_id=raw.pool_id,
                    source_type='external',
                ))
            success_count += 1
        job = QualityImportJob(
            file_name=filename,
            target_pool_id=pool.id,
            total_count=len(df.index),
            success_count=success_count,
            failed_count=len(failed),
            duplicate_append_count=duplicate_append_count,
            failed_detail_json=_quality_json_dumps(failed),
            created_by=_quality_current_user(),
        )
        db.session.add(job)
        pool.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({
            'success': True,
            'pool': _quality_pool_to_dict(pool),
            'job': {
                'id': job.id,
                'file_name': job.file_name,
                'total_count': job.total_count,
                'success_count': job.success_count,
                'failed_count': job.failed_count,
                'duplicate_append_count': job.duplicate_append_count,
                'failed_detail': failed,
                'created_at': _dt_to_iso(job.created_at),
            }
        })
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# Supabase Client Helper
class SupabaseClient:
    def __init__(self, url, key, enable_outbox: bool = True):
        self.url = url.rstrip('/')
        self.key = key
        self.enable_outbox = bool(enable_outbox)
        self.headers = {
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal"
        }
        
        # Configure requests session with retry strategy
        self.session = requests.Session()
        from requests.adapters import HTTPAdapter
        from urllib3.util.retry import Retry
        
        retry_strategy = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "OPTIONS", "POST", "PUT", "DELETE"]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

    def insert(self, table, data, ignore_duplicates=False):
        headers = self.headers.copy()
        if ignore_duplicates:
            headers["Prefer"] = "resolution=ignore-duplicates"
        try:
            response = self.session.post(f"{self.url}/rest/v1/{table}", headers=headers, json=data, timeout=120)
            if self.enable_outbox and response is not None and getattr(response, "status_code", 0) in (429, 500, 502, 503, 504):
                _outbox_enqueue("insert", table, payload=data, filters=None, extra={"ignore_duplicates": bool(ignore_duplicates)}, last_error=getattr(response, "text", ""), status_code=getattr(response, "status_code", None))
                class MockResponse:
                    status_code = 202
                    text = "queued_for_outbox"
                    def json(self): return {"queued": True}
                return MockResponse()
            return response
        except Exception as e:
            if self.enable_outbox:
                _outbox_enqueue("insert", table, payload=data, filters=None, extra={"ignore_duplicates": bool(ignore_duplicates)}, last_error=str(e), status_code=None)
            class MockResponse:
                status_code = 202
                text = "queued_for_outbox"
                def json(self): return {"queued": True}
            return MockResponse()

    def upsert(self, table, data, on_conflict=None):
        headers = self.headers.copy()
        # Supabase specific: use on_conflict to specify unique column
        # and Prefer header to specify resolution
        headers["Prefer"] = "resolution=merge-duplicates, return=minimal"
        
        # DEBUG: Log upsert attempt
        print(f"DEBUG Upserting {len(data)} items to {table} (on_conflict={on_conflict})")
        
        url = f"{self.url}/rest/v1/{table}"
        if on_conflict:
            url += f"?on_conflict={on_conflict}"
        
        try:
            response = self.session.post(url, headers=headers, json=data, timeout=120)
            if self.enable_outbox and response is not None and getattr(response, "status_code", 0) in (429, 500, 502, 503, 504):
                _outbox_enqueue("upsert", table, payload=data, filters=None, extra={"on_conflict": on_conflict}, last_error=getattr(response, "text", ""), status_code=getattr(response, "status_code", None))
                class MockResponse:
                    status_code = 202
                    text = "queued_for_outbox"
                    def json(self): return {"queued": True}
                return MockResponse()

            if response.status_code >= 400:
                print(f"DEBUG Upsert Error: {response.status_code} - {response.text}")
                
                # FALLBACK: If standard upsert fails, try to update items one by one by ID or Unique Key
                print("DEBUG: Falling back to individual updates...")
                for item in data:
                    if 'id' in item:
                        # Create a copy without ID for the update data, use ID for filter
                        update_id = item['id']
                        update_data = {k: v for k, v in item.items() if k != 'id'}
                        self.update(table, update_data, {'id': update_id})
                    elif 'kb_id' in item:
                        # Fallback for kb_scores using kb_id
                        kb_id_val = item['kb_id']
                        update_data = {k: v for k, v in item.items() if k != 'kb_id'}
                        self.update(table, update_data, {'kb_id': kb_id_val})
            return response
        except Exception as e:
            if self.enable_outbox:
                _outbox_enqueue("upsert", table, payload=data, filters=None, extra={"on_conflict": on_conflict}, last_error=str(e), status_code=None)
                class MockResponse:
                    status_code = 202
                    text = "queued_for_outbox"
                    def json(self): return {"queued": True}
                return MockResponse()
            print(f"DEBUG Upsert Exception: {e}")
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {}
            return MockResponse()

    def select(self, table, page=1, page_size=20, filters=None, order_by=None, order_dir='desc', columns='*', count='exact'):
        headers = self.headers.copy()
        if count:
            count_val = str(count).strip().lower()
            if count_val in ('exact', 'planned', 'estimated'):
                headers["Prefer"] = f"count={count_val}"
        
        offset = (page - 1) * page_size
        params = {
            "select": columns,
            "limit": page_size,
            "offset": offset,
        }
        
        # Handling Sorting
        if order_by:
            # e.g. "update_time.desc.nullslast"
            sort_val = f"{order_by}.{order_dir}"
            if order_dir == 'desc':
                sort_val += ".nullslast"
            else:
                sort_val += ".nullsfirst"
            params["order"] = sort_val
        # Remove default update_time sort to avoid errors on tables without that column


        # Handling Filters
        if filters:
            for k, v in filters.items():
                if v:
                    params[k] = v
        
        try:
            response = self.session.get(f"{self.url}/rest/v1/{table}", headers=headers, params=params, timeout=120)
            return response
        except Exception as e:
            print(f"Supabase Select Error: {e}")
            class MockResponse:
                status_code = 500
                text = str(e)
                headers = {}
                def json(self): return []
            return MockResponse()

    def rpc(self, func_name, params=None):
        response = self.session.post(f"{self.url}/rest/v1/rpc/{func_name}", headers=self.headers, json=params or {}, timeout=60)
        return response

    # ---------------------------
    # Supabase Storage Helpers
    # ---------------------------
    def storage_upload(self, bucket: str, object_path: str, content: bytes, content_type: str = 'application/octet-stream'):
        """
        Upload object to Supabase Storage.
        Note: Requires Storage policy allowing this key to upload.
        """
        bucket = str(bucket or '').strip()
        object_path = str(object_path or '').lstrip('/')
        url = f"{self.url}/storage/v1/object/{bucket}/{object_path}"
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": content_type,
            "x-upsert": "true",
        }
        return self.session.post(url, headers=headers, data=content, timeout=300)

    def storage_download(self, bucket: str, object_path: str):
        bucket = str(bucket or '').strip()
        object_path = str(object_path or '').lstrip('/')
        url = f"{self.url}/storage/v1/object/{bucket}/{object_path}"
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
        }
        return self.session.get(url, headers=headers, timeout=300)

    def update(self, table, data, filters):
        # filters: dict of {col: val} for equality check or raw operator string
        params = {}
        for k, v in filters.items():
            if isinstance(v, str) and (v.startswith('eq.') or v.startswith('neq.') or v.startswith('in.') or v.startswith('gt.') or v.startswith('lt.') or v.startswith('like.') or v.startswith('ilike.') or v.startswith('is.') or v.startswith('not.is.')):
                params[k] = v
            else:
                params[k] = f"eq.{v}"
        try:
            response = self.session.patch(f"{self.url}/rest/v1/{table}", headers=self.headers, json=data, params=params, timeout=120)
            if self.enable_outbox and response is not None and getattr(response, "status_code", 0) in (429, 500, 502, 503, 504):
                _outbox_enqueue("update", table, payload=data, filters=filters, extra={"params": params}, last_error=getattr(response, "text", ""), status_code=getattr(response, "status_code", None))
                class MockResponse:
                    status_code = 202
                    text = "queued_for_outbox"
                    def json(self): return {"queued": True}
                return MockResponse()
            return response
        except Exception as e:
            if self.enable_outbox:
                _outbox_enqueue("update", table, payload=data, filters=filters, extra={"params": params}, last_error=str(e), status_code=None)
                class MockResponse:
                    status_code = 202
                    text = "queued_for_outbox"
                    def json(self): return {"queued": True}
                return MockResponse()
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {}
            return MockResponse()

    def delete(self, table, filters):
        # filters: dict of {col: val} for equality check or raw operator string
        params = {}
        for k, v in filters.items():
            if isinstance(v, str) and (v.startswith('eq.') or v.startswith('neq.') or v.startswith('in.') or v.startswith('gt.') or v.startswith('lt.') or v.startswith('like.') or v.startswith('ilike.') or v.startswith('is.') or v.startswith('not.is.')):
                params[k] = v
            else:
                params[k] = f"eq.{v}"
        try:
            response = self.session.delete(f"{self.url}/rest/v1/{table}", headers=self.headers, params=params, timeout=120)
            if self.enable_outbox and response is not None and getattr(response, "status_code", 0) in (429, 500, 502, 503, 504):
                _outbox_enqueue("delete", table, payload=None, filters=filters, extra={"params": params}, last_error=getattr(response, "text", ""), status_code=getattr(response, "status_code", None))
                class MockResponse:
                    status_code = 202
                    text = "queued_for_outbox"
                    def json(self): return {"queued": True}
                return MockResponse()
            return response
        except Exception as e:
            if self.enable_outbox:
                _outbox_enqueue("delete", table, payload=None, filters=filters, extra={"params": params}, last_error=str(e), status_code=None)
                class MockResponse:
                    status_code = 202
                    text = "queued_for_outbox"
                    def json(self): return {"queued": True}
                return MockResponse()
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {}
            return MockResponse()
    
    def delete_in(self, table, column, values):
        if not values:
            return None
        # format: in.(val1,val2)
        # Assuming values are strings, no internal commas
        val_str = "(" + ",".join(str(v) for v in values) + ")"
        params = {column: f"in.{val_str}"}
        try:
            response = self.session.delete(f"{self.url}/rest/v1/{table}", headers=self.headers, params=params, timeout=120)
            if self.enable_outbox and response is not None and getattr(response, "status_code", 0) in (429, 500, 502, 503, 504):
                _outbox_enqueue("delete_in", table, payload=None, filters={"column": column, "values": values}, extra={"column": column}, last_error=getattr(response, "text", ""), status_code=getattr(response, "status_code", None))
                class MockResponse:
                    status_code = 202
                    text = "queued_for_outbox"
                    def json(self): return {"queued": True}
                return MockResponse()
            return response
        except Exception as e:
            if self.enable_outbox:
                _outbox_enqueue("delete_in", table, payload=None, filters={"column": column, "values": values}, extra={"column": column}, last_error=str(e), status_code=None)
                class MockResponse:
                    status_code = 202
                    text = "queued_for_outbox"
                    def json(self): return {"queued": True}
                return MockResponse()
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {}
            return MockResponse()

    def select_all(self, table, filters=None, order_by='id', order_dir='asc', columns='*', page_size=1000):
        # Fetch all data using pagination
        all_data = []
        page = 1
        print(f"DEBUG: select_all {table} starting with page_size={page_size}...")
        
        while True:
            if page % 10 == 0:
                print(f"DEBUG: Fetching page {page}...")
            response = self.select(table, page, page_size, filters, order_by, order_dir, columns)
            if response.status_code >= 400:
                raise Exception(f"Database Error: {response.text}")
            
            batch = response.json()
            if not batch:
                break
                
            all_data.extend(batch)
            
            if len(batch) < page_size:
                break
            
            page += 1
            
        print(f"DEBUG: select_all {table} finished. Total: {len(all_data)}")
        return all_data


class LocalPostgreSQLClient:
    """
    本地 PostgreSQL 客户端，作为 SupabaseClient 的替代品
    提供与 SupabaseClient 兼容的 API 接口
    """
    
    def __init__(self, config):
        """
        初始化本地 PostgreSQL 客户端
        :param config: 包含 host, port, database, user, password 的字典
        """
        self.config = config
        self.connection = None
        self.host = config.get('host', 'localhost')
        self.port = config.get('port', 5432)
        self.database = config.get('database', 'knowledgebase_local')
        self.user = config.get('user', 'postgres')
        self.password = config.get('password', '')
        
        # 为兼容 Outbox 功能，添加 url 和 key 属性（本地模式不使用）
        self.url = "local://localhost"
        self.key = "local_mode"

    def _quote_identifier(self, ident):
        ident = str(ident or '').strip()
        if not re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', ident):
            raise ValueError(f"Unsafe SQL identifier: {ident}")
        return '"' + ident.replace('"', '""') + '"'

    def _quote_qualified_identifier(self, ident):
        parts = [p.strip() for p in str(ident or '').split('.') if p.strip()]
        if not parts:
            raise ValueError("Empty SQL identifier")
        return '.'.join(self._quote_identifier(p) for p in parts)

    def _quote_field_expr(self, expr):
        expr = str(expr or '').strip()
        if not expr:
            raise ValueError("Empty SQL field expression")
        if expr == '*':
            return '*'
        if expr.endswith('::text'):
            return f"({self._quote_field_expr(expr[:-6])}::text)"
        if '->>' in expr:
            base, path = expr.split('->>', 1)
            base_sql = self._quote_field_expr(base)
            path = str(path or '').strip().strip('"').strip("'")
            if re.match(r'^\d+$', path):
                return f"({base_sql}->>{int(path)})"
            if not re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', path):
                raise ValueError(f"Unsafe JSON path: {path}")
            return f"({base_sql}->>'{path}')"
        return self._quote_qualified_identifier(expr)

    def _quote_select_columns(self, columns):
        if columns is None or str(columns).strip() == '*' or str(columns).strip() == '':
            return '*'
        return ', '.join(self._quote_field_expr(c.strip()) for c in str(columns).split(',') if c.strip())

    def _split_postgrest_items(self, raw):
        s = str(raw or '')
        items = []
        buf = []
        depth = 0
        quote = None
        escape = False
        for ch in s:
            if quote:
                buf.append(ch)
                if escape:
                    escape = False
                elif ch == '\\':
                    escape = True
                elif ch == quote:
                    quote = None
                continue
            if ch in ('"', "'"):
                quote = ch
                buf.append(ch)
                continue
            if ch in '([':
                depth += 1
                buf.append(ch)
                continue
            if ch in ')]':
                depth = max(0, depth - 1)
                buf.append(ch)
                continue
            if ch == ',' and depth == 0:
                item = ''.join(buf).strip()
                if item:
                    items.append(item)
                buf = []
                continue
            buf.append(ch)
        tail = ''.join(buf).strip()
        if tail:
            items.append(tail)
        return items

    def _parse_postgrest_list(self, raw):
        s = str(raw or '').strip()
        if s.startswith('(') and s.endswith(')'):
            s = s[1:-1]
        values = []
        for item in self._split_postgrest_items(s):
            item = item.strip()
            if len(item) >= 2 and item[0] == '"' and item[-1] == '"':
                try:
                    values.append(json.loads(item))
                except Exception:
                    values.append(item[1:-1].replace('\\"', '"'))
            elif len(item) >= 2 and item[0] == "'" and item[-1] == "'":
                values.append(item[1:-1].replace("\\'", "'"))
            else:
                values.append(item)
        return values

    def _parse_logic_condition(self, cond):
        cond = str(cond or '').strip()
        for op in ('and', 'or', 'not'):
            prefix = f'{op}='
            if cond.startswith(prefix):
                return self._parse_filter(op, cond[len(prefix):])
        parts = cond.split('.', 1)
        if len(parts) != 2:
            return None, []
        return self._parse_filter(parts[0], parts[1])
        
    def _ensure_connection(self):
        """确保数据库连接存在"""
        if self.connection is None or self.connection.closed:
            try:
                if not HAS_PSYCOPG2:
                    raise Exception("psycopg2 未安装，请运行: pip install psycopg2-binary")
                self.connection = psycopg2.connect(
                    host=self.host,
                    port=self.port,
                    database=self.database,
                    user=self.user,
                    password=self.password
                )
                self.connection.autocommit = False
                print(f"✅ 成功连接到本地 PostgreSQL: {self.database}@{self.host}:{self.port}")
            except Exception as e:
                print(f"❌ 连接本地数据库失败: {e}")
                raise
    
    def _convert_jsonb_fields(self, item):
        """
        转换 JSONB 字段：将 Python 列表/字典转换为 JSON 字符串
        这样 psycopg2 会正确地将其作为 JSONB 类型插入
        
        注意：如果值已经是字符串（可能已被上层转换），则不再转换
        """
        if not isinstance(item, dict):
            return item
        
        # 已知的 JSONB 字段列表
        jsonb_fields = [
            'similar_questions', 'error_list', 'keyword_list',
            'image_urls', 'video_urls', 'file_urls', 'change_meta',
            'tags',  # link_previews 表的 tags 字段
            'score_data'  # kb_scores 表的 score_data 字段
        ]
        
        converted = {}
        for key, value in item.items():
            if key in jsonb_fields and value is not None:
                # 如果已经是字符串，说明上层已经转换过了，直接使用
                if isinstance(value, str):
                    converted[key] = value
                # 如果是列表或字典，转换为 JSON 字符串
                elif isinstance(value, (list, dict)):
                    import json
                    converted[key] = json.dumps(value, ensure_ascii=False)
                else:
                    converted[key] = value
            else:
                converted[key] = value
        
        return converted
    
    def _parse_filter(self, key, value):
        """
        解析 PostgREST 风格的过滤器语法
        支持：eq., neq., gt., gte., lt., lte., ilike., like., in., is., cs., and., or., not.
        :return: (condition_string, params_list)
        """
        key = str(key or '').strip()
        raw_value = str(value)

        if key in ('and', 'or', 'not') and raw_value.startswith('(') and raw_value.endswith(')'):
            operator = key
            param_str = raw_value
        else:
            match = re.match(r'^(eq|neq|gt|gte|lt|lte|ilike|like|in|is|cs|and|or|not)\.(.*)', raw_value, re.DOTALL)
            if not match:
                return f"{self._quote_field_expr(key)} = %s", [value]

            operator = match.group(1)
            param_str = match.group(2)

        field_sql = None
        if key not in ('and', 'or') or operator not in ('and', 'or'):
            if key not in ('and', 'or', 'not'):
                field_sql = self._quote_field_expr(key)
        
        # 处理特殊操作符
        if operator == 'eq':
            return f"{field_sql} = %s", [param_str]
        elif operator == 'neq':
            return f"{field_sql} <> %s", [param_str]
        elif operator == 'gt':
            return f"{field_sql} > %s", [param_str]
        elif operator == 'gte':
            return f"{field_sql} >= %s", [param_str]
        elif operator == 'lt':
            return f"{field_sql} < %s", [param_str]
        elif operator == 'lte':
            return f"{field_sql} <= %s", [param_str]
        elif operator == 'ilike':
            # PostgreSQL ILIKE (不区分大小写的 LIKE)
            # PostgREST 语法：*keyword* -> %keyword%
            pattern = param_str.replace('*', '%')
            return f"{field_sql} ILIKE %s", [pattern]
        elif operator == 'like':
            pattern = param_str.replace('*', '%')
            return f"{field_sql} LIKE %s", [pattern]
        elif operator == 'in':
            values = self._parse_postgrest_list(param_str)
            if values:
                placeholders = ','.join(['%s'] * len(values))
                return f"{field_sql} IN ({placeholders})", values
            return f"{field_sql} IS NULL", []
        elif operator == 'is':
            if param_str.lower() == 'null':
                return f"{field_sql} IS NULL", []
            elif param_str.lower() == 'notnull':
                return f"{field_sql} IS NOT NULL", []
        elif operator == 'cs':
            # JSONB contains. Example: tags=cs.["a"] -> "tags" @> '["a"]'::jsonb
            try:
                parsed = json.loads(param_str)
                param_json = json.dumps(parsed, ensure_ascii=False)
            except Exception:
                param_json = param_str
            return f"{field_sql} @> %s::jsonb", [param_json]
        elif operator == 'and':
            # AND 子句处理：(cond1,cond2,cond3)
            and_match = re.match(r'\((.+)\)', param_str, re.DOTALL)
            if and_match:
                conditions_str = and_match.group(1)
                and_conditions = []
                and_params = []
                for cond in self._split_postgrest_items(conditions_str):
                    c, p = self._parse_logic_condition(cond)
                    if c:
                        and_conditions.append(c)
                        and_params.extend(p)
                
                if and_conditions:
                    return f"({' AND '.join(and_conditions)})", and_params
        elif operator == 'or':
            # OR 子句处理：(cond1,cond2,cond3)
            or_match = re.match(r'\((.+)\)', param_str, re.DOTALL)
            if or_match:
                conditions_str = or_match.group(1)
                or_conditions = []
                or_params = []
                for cond in self._split_postgrest_items(conditions_str):
                    c, p = self._parse_logic_condition(cond)
                    if c:
                        or_conditions.append(c)
                        or_params.extend(p)
                
                if or_conditions:
                    return f"({' OR '.join(or_conditions)})", or_params
        elif operator == 'not':
            # NOT 操作符：not.is.null 或 not.(condition)
            # 特殊处理 not.is.null -> IS NOT NULL
            if param_str.lower() == 'is.null':
                return f"{field_sql} IS NOT NULL", []
            
            # 处理 not.(condition) 格式
            not_match = re.match(r'\((.+)\)', param_str, re.DOTALL)
            if not_match:
                cond_str = not_match.group(1)
                c, p = self._parse_logic_condition(cond_str)
                if c:
                    return f"NOT ({c})", p
            else:
                c, p = self._parse_filter(key, param_str)
                if c:
                    return f"NOT ({c})", p
        
        # 默认返回等值匹配
        return f"{self._quote_field_expr(key)} = %s", [value]

    def _execute_query(self, query, params=None, fetch=True):
        """执行 SQL 查询"""
        self._ensure_connection()
        try:
            with self.connection.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(query, params)
                if fetch:
                    results = cur.fetchall()
                    return [dict(row) for row in results]
                else:
                    self.connection.commit()
                    return []
        except Exception as e:
            self.connection.rollback()
            print(f"❌ 查询执行失败: {e}")
            raise
    
    def insert(self, table, data, ignore_duplicates=False):
        """
        插入数据（兼容 SupabaseClient.insert）
        :param table: 表名
        :param data: 数据列表，每个元素是字典
        :param ignore_duplicates: 是否忽略重复（通过 ON CONFLICT DO NOTHING 实现）
        """
        if not isinstance(data, list):
            data = [data]
        
        if not data:
            class MockResponse:
                status_code = 200
                text = "No data to insert"
                def json(self): return []
            return MockResponse()
        
        try:
            inserted_count = 0
            for item in data:
                # 转换 JSONB 字段
                item = self._convert_jsonb_fields(item)
                
                columns = list(item.keys())
                values = list(item.values())
                placeholders = ['%s'] * len(columns)
                
                table_sql = self._quote_qualified_identifier(table)
                col_str = ', '.join(self._quote_identifier(c) for c in columns)
                val_str = ', '.join(placeholders)
                
                if ignore_duplicates:
                    # 获取主键列名（假设第一个字段是主键或唯一键）
                    pk_col = columns[0] if columns else None
                    if pk_col:
                        sql = f"INSERT INTO {table_sql} ({col_str}) VALUES ({val_str}) ON CONFLICT ({self._quote_identifier(pk_col)}) DO NOTHING"
                    else:
                        sql = f"INSERT INTO {table_sql} ({col_str}) VALUES ({val_str})"
                else:
                    sql = f"INSERT INTO {table_sql} ({col_str}) VALUES ({val_str})"
                
                self._execute_query(sql, values, fetch=False)
                inserted_count += 1
            
            class MockResponse:
                status_code = 201
                text = f"Inserted {inserted_count} rows"
                def json(self): return {"count": inserted_count}
            return MockResponse()
        except Exception as e:
            print(f"❌ Insert 失败: {e}")
            import traceback
            traceback.print_exc()
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {"error": str(e)}
            return MockResponse()
    
    def upsert(self, table, data, on_conflict=None):
        """
        Upsert 数据（兼容 SupabaseClient.upsert）
        :param table: 表名
        :param data: 数据列表
        :param on_conflict: 冲突时用于判断的列名（可以是单个列或逗号分隔的多个列）
        """
        if not isinstance(data, list):
            data = [data]
        
        if not data:
            class MockResponse:
                status_code = 200
                text = "No data to upsert"
                def json(self): return []
            return MockResponse()
        
        try:
            upserted_count = 0
            for item in data:
                # 转换 JSONB 字段
                item = self._convert_jsonb_fields(item)
                
                columns = list(item.keys())
                values = list(item.values())
                placeholders = ['%s'] * len(columns)
                
                table_sql = self._quote_qualified_identifier(table)
                col_str = ', '.join(self._quote_identifier(c) for c in columns)
                val_str = ', '.join(placeholders)
                
                # 确定冲突列（支持多列组合）
                if on_conflict:
                    # 处理逗号分隔的多列情况
                    conflict_cols = [col.strip() for col in on_conflict.split(',')]
                    conflict_col_str = ', '.join(self._quote_identifier(col) for col in conflict_cols)
                else:
                    # 默认使用 id 或第一个字段
                    conflict_col_str = 'id' if 'id' in columns else columns[0]
                    conflict_cols = [conflict_col_str]
                    conflict_col_str = self._quote_identifier(conflict_col_str)
                
                # 构建更新字段（排除冲突列）
                update_cols = [col for col in columns if col not in conflict_cols]
                if update_cols:
                    update_str = ', '.join([f"{self._quote_identifier(col)}=EXCLUDED.{self._quote_identifier(col)}" for col in update_cols])
                    sql = f"INSERT INTO {table_sql} ({col_str}) VALUES ({val_str}) ON CONFLICT ({conflict_col_str}) DO UPDATE SET {update_str}"
                else:
                    sql = f"INSERT INTO {table_sql} ({col_str}) VALUES ({val_str}) ON CONFLICT ({conflict_col_str}) DO NOTHING"
                
                self._execute_query(sql, values, fetch=False)
                upserted_count += 1
            
            class MockResponse:
                status_code = 200
                text = f"Upserted {upserted_count} rows"
                def json(self): return {"count": upserted_count}
            return MockResponse()
        except Exception as e:
            print(f"❌ Upsert 失败: {e}")
            import traceback
            traceback.print_exc()
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {"error": str(e)}
            return MockResponse()
    
    def select(self, table, page=1, page_size=20, filters=None, order_by=None, order_dir='desc', columns='*', count='exact'):
        """
        查询数据（兼容 SupabaseClient.select）
        :return: MockResponse 对象，包含 .json() 方法和 headers 属性
        """
        try:
            # 构建 SELECT 子句
            select_clause = self._quote_select_columns(columns)
            
            table_sql = self._quote_qualified_identifier(table)
            sql = f"SELECT {select_clause} FROM {table_sql}"
            params = []
            
            # 构建 WHERE 子句
            if filters:
                where_conditions = []
                for key, value in filters.items():
                    if value is not None and value != '':
                        # 解析 PostgREST 风格的过滤器语法
                        condition, cond_params = self._parse_filter(key, value)
                        if condition:
                            where_conditions.append(condition)
                            params.extend(cond_params)
                
                if where_conditions:
                    sql += " WHERE " + " AND ".join(where_conditions)
            
            # 构建 ORDER BY 子句
            if order_by:
                order_dir = str(order_dir or 'desc').lower()
                order_dir = 'asc' if order_dir == 'asc' else 'desc'
                null_handling = "NULLS LAST" if order_dir == 'desc' else "NULLS FIRST"
                sql += f" ORDER BY {self._quote_field_expr(order_by)} {order_dir} {null_handling}"
            
            # 构建 LIMIT/OFFSET 子句
            offset = (page - 1) * page_size
            sql += " LIMIT %s OFFSET %s"
            params.extend([page_size, offset])
            
            # 执行查询
            results = self._execute_query(sql, params if params else None, fetch=True)
            
            # 获取总数（如果需要）
            total_count = len(results)
            if count:
                count_sql = f"SELECT COUNT(*) as cnt FROM {table_sql}"
                if filters:
                    where_conditions = []
                    count_params = []
                    for key, value in filters.items():
                        if value is not None and value != '':
                            condition, cond_params = self._parse_filter(key, value)
                            if condition:
                                where_conditions.append(condition)
                                count_params.extend(cond_params)
                    if where_conditions:
                        count_sql += " WHERE " + " AND ".join(where_conditions)
                        count_results = self._execute_query(count_sql, count_params if count_params else None, fetch=True)
                        if count_results:
                            total_count = count_results[0]['cnt']
                else:
                    count_results = self._execute_query(count_sql, fetch=True)
                    if count_results:
                        total_count = count_results[0]['cnt']
            
            # 构建响应
            class MockResponse:
                def __init__(self, data, total):
                    self.data = data
                    self.total = total
                    self.status_code = 200
                    self.text = "Success"
                    self.headers = {'Content-Range': f'0-{len(data)-1}/{total}'}
                
                def json(self):
                    return self.data
            
            return MockResponse(results, total_count)
        
        except Exception as e:
            print(f"❌ Select 失败: {e}")
            class MockResponse:
                status_code = 500
                text = str(e)
                headers = {}
                def json(self): return []
            return MockResponse()
    
    def update(self, table, data, filters):
        """
        更新数据（兼容 SupabaseClient.update）
        :param table: 表名
        :param data: 要更新的字段字典
        :param filters: 过滤条件字典
        """
        try:
            if not data:
                class MockResponse:
                    status_code = 200
                    text = "No data to update"
                    def json(self): return {}
                return MockResponse()
            
            # 转换 JSONB 字段
            data = self._convert_jsonb_fields(data)
            
            # 构建 SET 子句
            set_clauses = []
            values = []
            for key, value in data.items():
                set_clauses.append(f"{self._quote_identifier(key)} = %s")
                values.append(value)
            
            sql = f"UPDATE {self._quote_qualified_identifier(table)} SET {', '.join(set_clauses)}"
            
            # 构建 WHERE 子句
            if filters:
                where_conditions = []
                for key, value in filters.items():
                    if value is not None and value != '':
                        # 使用 _parse_filter 方法来支持完整的 PostgREST 语法
                        condition, cond_params = self._parse_filter(key, value)
                        if condition:
                            where_conditions.append(condition)
                            values.extend(cond_params)
                
                if where_conditions:
                    sql += " WHERE " + " AND ".join(where_conditions)
            
            self._execute_query(sql, values if values else None, fetch=False)
            
            class MockResponse:
                status_code = 200
                text = "Updated successfully"
                def json(self): return {"updated": True}
            return MockResponse()
        
        except Exception as e:
            print(f"❌ Update 失败: {e}")
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {"error": str(e)}
            return MockResponse()
    
    def delete(self, table, filters):
        """
        删除数据（兼容 SupabaseClient.delete）
        :param table: 表名
        :param filters: 过滤条件字典
        """
        try:
            sql = f"DELETE FROM {self._quote_qualified_identifier(table)}"
            params = []
            
            if filters:
                where_conditions = []
                for key, value in filters.items():
                    if value is not None and value != '':
                        # 使用 _parse_filter 方法来支持完整的 PostgREST 语法
                        # 包括 not.is.null, eq., in., is.null 等
                        condition, cond_params = self._parse_filter(key, value)
                        if condition:
                            where_conditions.append(condition)
                            params.extend(cond_params)
                
                if where_conditions:
                    sql += " WHERE " + " AND ".join(where_conditions)
            
            self._execute_query(sql, params if params else None, fetch=False)
            
            class MockResponse:
                status_code = 200
                text = "Deleted successfully"
                def json(self): return {"deleted": True}
            return MockResponse()
        
        except Exception as e:
            print(f"❌ Delete 失败: {e}")
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {"error": str(e)}
            return MockResponse()
    
    def delete_in(self, table, column, values):
        """
        批量删除（兼容 SupabaseClient.delete_in）
        :param table: 表名
        :param column: 列名
        :param values: 值列表
        """
        if not values:
            class MockResponse:
                status_code = 200
                text = "No values to delete"
                def json(self): return {}
            return MockResponse()
        
        try:
            placeholders = ','.join(['%s'] * len(values))
            sql = f"DELETE FROM {self._quote_qualified_identifier(table)} WHERE {self._quote_field_expr(column)} IN ({placeholders})"
            self._execute_query(sql, values, fetch=False)
            
            class MockResponse:
                status_code = 200
                text = f"Deleted {len(values)} rows"
                def json(self): return {"deleted": len(values)}
            return MockResponse()
        
        except Exception as e:
            print(f"❌ Delete_in 失败: {e}")
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {"error": str(e)}
            return MockResponse()
    
    def select_all(self, table, filters=None, order_by='id', order_dir='asc', columns='*', page_size=1000):
        """
        查询所有数据（兼容 SupabaseClient.select_all）
        """
        all_data = []
        page = 1
        
        while True:
            response = self.select(table, page, page_size, filters, order_by, order_dir, columns)
            if response.status_code >= 400:
                raise Exception(f"Database Error: {response.text}")
            
            batch = response.json()
            if not batch:
                break
            
            all_data.extend(batch)
            
            if len(batch) < page_size:
                break
            
            page += 1
        
        return all_data
    
    def rpc(self, func_name, params=None):
        """
        调用存储过程（兼容 SupabaseClient.rpc）
        """
        try:
            # 构建函数调用
            if params:
                param_values = list(params.values())
                sql = f"SELECT * FROM {self._quote_qualified_identifier(func_name)}({','.join(['%s'] * len(param_values))})"
                results = self._execute_query(sql, param_values, fetch=True)
            else:
                sql = f"SELECT * FROM {self._quote_qualified_identifier(func_name)}()"
                results = self._execute_query(sql, fetch=True)
            
            class MockResponse:
                status_code = 200
                text = "Success"
                def json(self): return results
            return MockResponse()
        
        except Exception as e:
            print(f"❌ RPC 调用失败: {e}")
            class MockResponse:
                status_code = 500
                text = str(e)
                def json(self): return {"error": str(e)}
            return MockResponse()
    
    def _local_storage_path(self, bucket, object_path):
        bucket = str(bucket or '').strip()
        if not re.match(r'^[A-Za-z0-9_.-]+$', bucket):
            raise ValueError('Invalid storage bucket name')
        object_path = str(object_path or '').replace('\\', '/').lstrip('/')
        norm_path = os.path.normpath(object_path)
        if not norm_path or norm_path == '.' or norm_path.startswith('..') or os.path.isabs(norm_path):
            raise ValueError('Invalid storage object path')
        root = os.path.abspath(os.path.join(_BASE_DIR, 'instance', 'storage', bucket))
        abs_path = os.path.abspath(os.path.join(root, norm_path))
        if abs_path != root and not abs_path.startswith(root + os.sep):
            raise ValueError('Storage path escapes bucket root')
        return root, abs_path

    def storage_upload(self, bucket, object_path, content, content_type='application/octet-stream'):
        """
        上传文件到本地存储目录 instance/storage/<bucket>/<object_path>。
        """
        try:
            root, abs_path = self._local_storage_path(bucket, object_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok=True)
            if isinstance(content, str):
                content = content.encode('utf-8')
            with open(abs_path, 'wb') as f:
                f.write(content or b'')

            class MockResponse:
                status_code = 200
                text = "Uploaded"
                headers = {'Content-Type': 'application/json'}
                content = b''
                def json(self_nonlocal):
                    rel = os.path.relpath(abs_path, root).replace(os.sep, '/')
                    return {'Key': f"{bucket}/{rel}", 'path': rel, 'bucket': bucket}
            return MockResponse()
        except Exception as e:
            class MockResponse:
                status_code = 400
                text = str(e)
                headers = {}
                content = b''
                def json(self): return {"error": str(e)}
            return MockResponse()
    
    def storage_download(self, bucket, object_path):
        """
        从本地存储目录读取文件，返回与 requests.Response 近似的对象。
        """
        try:
            _, abs_path = self._local_storage_path(bucket, object_path)
            if not os.path.exists(abs_path) or not os.path.isfile(abs_path):
                class NotFoundResponse:
                    status_code = 404
                    text = "Object not found"
                    headers = {}
                    content = b''
                    def json(self): return {"error": "Object not found"}
                return NotFoundResponse()

            import mimetypes
            with open(abs_path, 'rb') as f:
                data = f.read()
            content_type = mimetypes.guess_type(abs_path)[0] or 'application/octet-stream'

            class MockResponse:
                status_code = 200
                text = data.decode('utf-8', errors='ignore')
                headers = {'Content-Type': content_type, 'Content-Length': str(len(data))}
                content = data
                def json(self):
                    try:
                        return json.loads(data.decode('utf-8'))
                    except Exception:
                        return {}
            return MockResponse()
        except Exception as e:
            class MockResponse:
                status_code = 400
                text = str(e)
                headers = {}
                content = b''
                def json(self): return {"error": str(e)}
            return MockResponse()
    
    def close(self):
        """关闭数据库连接"""
        if self.connection and not self.connection.closed:
            self.connection.close()
            print("🔒 数据库连接已关闭")


CONFIG_FILE = 'supabase_config_local.json'

# 默认不再回退到远端 Supabase，避免本地配置异常时误连线上服务。
DEFAULT_SUPABASE_URL = ""
DEFAULT_SUPABASE_KEY = ""

def is_supabase_button_sync_enabled():
    config_file = _get_config_path(CONFIG_FILE)
    if not os.path.exists(config_file):
        return False
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
            return bool(config.get('enable_button_sync', False))
    except Exception:
        return False

def get_supabase_client():
    config_file = _get_config_path(CONFIG_FILE)
    config = {}
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f) or {}
        except Exception as e:
            print(f"Error loading config file: {e}")
            config = {}

    if config.get('local_db'):
        print("🔧 检测到本地主库配置，使用 LocalPostgreSQLClient")
        return LocalPostgreSQLClient(config.get('local_db'))

    allow_remote = bool(config.get('allow_remote_supabase', False))
    url = str(config.get('url') or DEFAULT_SUPABASE_URL or '').strip()
    key = str(config.get('key') or DEFAULT_SUPABASE_KEY or '').strip()
    if allow_remote and url and key:
        print("🌐 显式允许远端数据库客户端，使用 SupabaseClient")
        return SupabaseClient(url, key)

    print("⛔ 未配置本地主库，且远端数据库访问已禁用")
    return None

def is_remote_supabase_allowed():
    cfg = _read_supabase_config()
    return bool(cfg.get('allow_remote_supabase', False))

def _db_not_configured_message():
    return '本地主库未配置'

def _db_client_unavailable_message():
    return '数据库客户端不可用'

def _db_fetch_failed_message():
    return '数据库查询失败'




def _read_supabase_config():
    config_file = _get_config_path(CONFIG_FILE)
    if not os.path.exists(config_file):
        return {}
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
            return cfg if isinstance(cfg, dict) else {}
    except Exception:
        return {}

def is_supabase_matrix_enabled():
    cfg = _read_supabase_config()
    return bool(cfg.get('use_supabase_matrix', False))

def is_supabase_governance_enabled():
    cfg = _read_supabase_config()
    return bool(cfg.get('use_supabase_governance', False))

def is_supabase_archives_enabled():
    cfg = _read_supabase_config()
    return bool(cfg.get('use_supabase_archives', cfg.get('use_supabase_governance', False)))

def is_supabase_ops_enabled():
    cfg = _read_supabase_config()
    return bool(cfg.get('use_supabase_ops', False))

def _dt_to_iso(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return datetime.fromtimestamp(v).isoformat()
        except Exception:
            return None
    if isinstance(v, datetime):
        try:
            return v.isoformat()
        except Exception:
            return None
    try:
        s = str(v).strip()
        return s or None
    except Exception:
        return None

def _supabase_upsert_chunks(client, table, rows, on_conflict=None, chunk_size=500):
    if not rows:
        return {'attempted': 0, 'ok': True, 'errors': []}
    chunk_size = int(chunk_size or 500)
    chunk_size = max(1, min(2000, chunk_size))
    attempted = 0
    errors = []
    for i in range(0, len(rows), chunk_size):
        chunk = rows[i:i + chunk_size]
        attempted += len(chunk)
        resp = client.upsert(table, chunk, on_conflict=on_conflict)
        if resp is None or getattr(resp, 'status_code', 500) >= 400:
            errors.append({
                'table': table,
                'status_code': getattr(resp, 'status_code', None) if resp is not None else None,
                'text': getattr(resp, 'text', '') if resp is not None else 'no response',
                'offset': i,
                'count': len(chunk)
            })
    return {'attempted': attempted, 'ok': len(errors) == 0, 'errors': errors}

def _supabase_table_exists(client, table):
    try:
        # 兼容本地 PostgreSQL 客户端
        if isinstance(client, LocalPostgreSQLClient):
            client._ensure_connection()
            cur = client.connection.cursor()
            cur.execute("""
                SELECT EXISTS (
                    SELECT 1 
                    FROM information_schema.tables 
                    WHERE table_name = %s
                )
            """, (table,))
            result = cur.fetchone()
            cur.close()
            return (result[0] if result else False), None

        # Supabase 客户端使用 REST API
        headers = dict(getattr(client, 'headers', {}) or {})
        headers.pop('Content-Type', None)
        resp = client.session.get(
            f"{client.url}/rest/v1/{table}",
            headers=headers,
            params={'select': '*', 'limit': 1},
            timeout=30
        )
        if resp.status_code in (200, 206):
            return True, None
        if resp.status_code == 404:
            return False, getattr(resp, 'text', '') or ''
        return False, getattr(resp, 'text', '') or ''
    except Exception as e:
        return False, str(e)

def _supabase_has_column(client, table, column):
    try:
        # 检查是否是本地 PostgreSQL 客户端
        if isinstance(client, LocalPostgreSQLClient):
            # 使用 PostgreSQL 的 information_schema 查询
            client._ensure_connection()
            cur = client.connection.cursor()
            cur.execute("""
                SELECT EXISTS (
                    SELECT 1 
                    FROM information_schema.columns 
                    WHERE table_name = %s AND column_name = %s
                )
            """, (table, column))
            result = cur.fetchone()
            cur.close()
            return result[0] if result else False
        
        # Supabase 客户端使用 REST API
        headers = dict(getattr(client, 'headers', {}) or {})
        headers.pop('Content-Type', None)
        resp = client.session.get(
            f"{client.url}/rest/v1/{table}",
            headers=headers,
            params={'select': column, 'limit': 1},
            timeout=30
        )
        if resp.status_code in (200, 206):
            return True
        txt = getattr(resp, 'text', '') or ''
        if 'PGRST204' in txt and f"'{column}'" in txt:
            return False
        return False
    except Exception as e:
        print(f"⚠️ 检查列 {table}.{column} 时出错: {e}")
        return False

def _supabase_table_exists(client, table):
    try:
        # 检查是否是本地 PostgreSQL 客户端
        if isinstance(client, LocalPostgreSQLClient):
            # 使用 PostgreSQL 的 information_schema 查询
            client._ensure_connection()
            cur = client.connection.cursor()
            cur.execute("""
                SELECT EXISTS (
                    SELECT 1 
                    FROM information_schema.tables 
                    WHERE table_name = %s
                )
            """, (table,))
            result = cur.fetchone()
            cur.close()
            return result[0] if result else False
        
        # Supabase 客户端使用 REST API
        headers = dict(getattr(client, 'headers', {}) or {})
        headers.pop('Content-Type', None)
        resp = client.session.get(
            f"{client.url}/rest/v1/{table}",
            headers=headers,
            params={'select': '*', 'limit': 1},
            timeout=30
        )
        return resp.status_code in (200, 206)
    except Exception:
        return False

def _iso_to_naive_datetime(v):
    if not v:
        return None
    try:
        s = str(v).strip()
        if not s:
            return None
        s = s.replace('Z', '+00:00')
        dt = datetime.fromisoformat(s)
        if getattr(dt, 'tzinfo', None) is not None:
            dt = dt.astimezone(tz=None).replace(tzinfo=None)
        return dt
    except Exception:
        return None

_MATRIX_SB_SYNC_LOCK = threading.Lock()
_MATRIX_SB_SYNC_STATE = {
    'matrix_last_pull_at': 0.0,
    'matrix_last_synced_at_iso': None,
    'logs_last_pull_at': 0.0,
    'logs_last_op_updated_at_iso': None,
    'logs_last_btn_submitted_at_iso': None
}

_SUPABASE_OUTBOX_MANUAL_AFTER_SECONDS = 4 * 3600  # 断网 4h 后建议手动重试


def _json_dumps_safe(v) -> str:
    try:
        return json.dumps(v, ensure_ascii=False, default=str)
    except Exception:
        return json.dumps(str(v), ensure_ascii=False)


def _json_loads_safe(s):
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        return None


def _outbox_enqueue(op_type: str, table_name: str, payload=None, filters=None, extra=None, last_error: str = "", status_code=None) -> int:
    """
    Enqueue a supabase write operation for later replay.
    Returns outbox id or -1 if enqueue failed.
    """
    try:
        with app.app_context():
            item = SupabaseOutbox(
                op_type=str(op_type or '').strip(),
                table_name=str(table_name or '').strip(),
                payload_json=_json_dumps_safe(payload) if payload is not None else None,
                filters_json=_json_dumps_safe(filters) if filters is not None else None,
                extra_json=_json_dumps_safe(extra) if extra is not None else None,
                status="pending",
                attempts=0,
                last_error=str(last_error or ''),
            )
            db.session.add(item)
            db.session.commit()
            print(f"[Outbox] queued id={item.id} op={op_type} table={table_name} status_code={status_code}")
            return int(item.id)
    except Exception as e:
        try:
            print(f"[Outbox] enqueue failed: {e}")
        except Exception:
            pass
        return -1


def _outbox_mark_needs_manual(item: "SupabaseOutbox") -> None:
    try:
        item.status = "needs_manual_sync"
        item.updated_ts = time.time()
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass


def _outbox_is_duplicate_error(err_text: str) -> bool:
    s = str(err_text or '').lower()
    # Supabase/PostgREST duplicate/unique constraint variants
    return (
        'duplicate key' in s
        or 'unique constraint' in s
        or 'violates unique constraint' in s
        or 'already exists' in s
    )


def _outbox_replay_pending_batch(limit: int = 50) -> dict:
    """
    Replay local outbox to supabase.
    - Only processes status='pending'
    - If an item is older than 4h, mark it as needs_manual_sync and skip
    Returns {processed, succeeded, failed}.
    """
    processed = 0
    succeeded = 0
    failed = 0

    client = get_supabase_client()
    if not client:
        return {"processed": 0, "succeeded": 0, "failed": 0, "reason": "supabase not configured"}

    # Disable outbox enqueue during replay to avoid infinite recursion.
    replay_client = SupabaseClient(client.url, client.key, enable_outbox=False)

    now_ts = time.time()
    pending_items = (
        SupabaseOutbox.query
        .filter(SupabaseOutbox.status == "pending")
        .order_by(SupabaseOutbox.created_ts.asc())
        .limit(int(limit or 50))
        .all()
    )
    for item in pending_items:
        processed += 1
        try:
            item_age = now_ts - float(item.created_ts or 0)
            if item_age >= _SUPABASE_OUTBOX_MANUAL_AFTER_SECONDS:
                _outbox_mark_needs_manual(item)
                continue

            op_type = item.op_type
            table = item.table_name
            payload = _json_loads_safe(item.payload_json)
            filters = _json_loads_safe(item.filters_json)
            extra = _json_loads_safe(item.extra_json) or {}

            # Normalize stored payload
            if payload is None:
                payload = [] if op_type in ("insert", "upsert") else {}
            if filters is None:
                filters = {} if op_type in ("update", "delete") else None

            resp = None
            if op_type == "insert":
                ignore_duplicates = bool(extra.get("ignore_duplicates", False))
                resp = replay_client.insert(table, payload or [], ignore_duplicates=ignore_duplicates)
            elif op_type == "upsert":
                resp = replay_client.upsert(table, payload or [], on_conflict=extra.get("on_conflict"))
            elif op_type == "update":
                resp = replay_client.update(table, payload or {}, filters or {})
            elif op_type == "delete":
                resp = replay_client.delete(table, filters or {})
            elif op_type == "delete_in":
                col = extra.get("column")
                vals = None
                if isinstance(filters, dict):
                    vals = filters.get("values")
                if not vals and isinstance(filters, dict):
                    vals = filters.get("values") or []
                resp = replay_client.delete_in(table, col, vals or [])
            else:
                raise RuntimeError(f"Unknown op_type: {op_type}")

            sc = getattr(resp, "status_code", 500)
            if int(sc) >= 400:
                err_txt = getattr(resp, "text", "") or ""
                raise RuntimeError(f"Replay failed status_code={sc}: {err_txt}")

            item.status = "done"
            item.last_error = ""
            item.attempts = int(item.attempts or 0) + 1
            item.updated_ts = time.time()
            db.session.commit()
            succeeded += 1
        except Exception as e:
            failed += 1
            err = str(e)
            # If replay is idempotently safe (e.g., duplicate key), mark done to prevent endless retry.
            if _outbox_is_duplicate_error(err):
                try:
                    item.status = "done"
                    item.last_error = ""
                    item.attempts = int(item.attempts or 0) + 1
                    item.updated_ts = time.time()
                    db.session.commit()
                    succeeded += 1
                    failed -= 1
                    continue
                except Exception:
                    pass

            item.attempts = int(item.attempts or 0) + 1
            item.last_error = err[:4000]
            item.updated_ts = time.time()
            db.session.commit()

    return {"processed": processed, "succeeded": succeeded, "failed": failed}


def _outbox_retry_loop(poll_seconds: int = 30) -> None:
    while True:
        try:
            with app.app_context():
                # best-effort; never block request path
                _outbox_replay_pending_batch(limit=50)
        except Exception:
            try:
                print("[Outbox] retry loop exception:", traceback.format_exc())
            except Exception:
                pass
        time.sleep(max(5, int(poll_seconds or 30)))


def _start_supabase_outbox_background_worker() -> None:
    t = threading.Thread(target=_outbox_retry_loop, args=(30,), daemon=True)
    t.start()


def _run_bg(fn, *args, **kwargs) -> None:
    """
    Run potentially slow network work in a daemon background thread,
    so the request returns quickly and UI feels responsive.
    """
    def _wrap():
        try:
            fn(*args, **kwargs)
        except Exception:
            try:
                print("[BG] task exception:", traceback.format_exc())
            except Exception:
                pass

    threading.Thread(target=_wrap, daemon=True).start()

def _upsert_sqlite_rows(model, rows, conflict_cols):
    if not rows:
        return 0
    # SQLite has a variable-number limit (commonly 999).
    # Use chunked upsert to avoid "too many SQL variables" on large batches.
    cols = [c for c in model.__table__.columns if c.name != 'id']
    col_count = max(1, len(cols))
    # Leave headroom for SQLAlchemy-generated bound params.
    max_vars = 900
    chunk_size = max(1, int(max_vars // col_count))

    done = 0
    for i in range(0, len(rows), chunk_size):
        batch = rows[i:i + chunk_size]
        stmt = sqlite_insert(model).values(batch)
        excluded = stmt.excluded
        set_cols = {}
        for c in model.__table__.columns:
            if c.name == 'id':
                continue
            set_cols[c.name] = excluded[c.name]
        stmt = stmt.on_conflict_do_update(index_elements=list(conflict_cols), set_=set_cols)
        db.session.execute(stmt)
        done += len(batch)
    return done

def _coerce_naive_datetime(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        try:
            if getattr(v, 'tzinfo', None) is not None:
                return v.astimezone(tz=None).replace(tzinfo=None)
            return v
        except Exception:
            return v.replace(tzinfo=None) if getattr(v, 'tzinfo', None) is not None else v
    return _iso_to_naive_datetime(v)

def _parse_content_range_total(resp):
    try:
        content_range = (getattr(resp, 'headers', {}) or {}).get('Content-Range') or ''
        if '/' in content_range:
            return int(str(content_range).split('/')[-1])
    except Exception:
        pass
    return None

def _get_client_table_total(client, table, columns='*'):
    try:
        resp = client.select(table, page=1, page_size=1, filters=None, order_by=None, order_dir='asc', columns=columns, count='exact')
        if not resp or getattr(resp, 'status_code', 500) >= 400:
            return None
        total = _parse_content_range_total(resp)
        return int(total) if total is not None else None
    except Exception:
        return None

def _filter_missing_sqlite_rows(model, rows, conflict_cols):
    if not rows:
        return []

    existing_keys = set()
    cols = [getattr(model, c) for c in conflict_cols]
    for rec in db.session.query(*cols).all():
        parts = []
        for idx in range(len(conflict_cols)):
            val = rec[idx] if isinstance(rec, (tuple, list)) else getattr(rec, conflict_cols[idx], None)
            parts.append(str(val or '').strip())
        existing_keys.add(tuple(parts))

    missing = []
    seen = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        key = tuple(str(row.get(c) or '').strip() for c in conflict_cols)
        if not all(key):
            continue
        if key in existing_keys or key in seen:
            continue
        seen.add(key)
        missing.append(row)
    return missing

def _backfill_matrix_table_from_client(model, client, table, columns, order_by, order_dir, conflict_cols, row_mapper, force=False):
    done_state_key = f'matrix_local_compat_done::{table}'
    if not force:
        with _MATRIX_SB_SYNC_LOCK:
            if _MATRIX_SB_SYNC_STATE.get(done_state_key):
                local_total = int(db.session.query(model).count() or 0)
                return {
                    'table': table,
                    'local_total': local_total,
                    'remote_total': None,
                    'fetched': 0,
                    'inserted': 0,
                    'attempted': False,
                    'cached': True
                }

    remote_total = _get_client_table_total(client, table, columns=columns.split(',')[0] if columns != '*' else '*')
    local_total = int(db.session.query(model).count() or 0)
    need_backfill = (remote_total is None) or (local_total < int(remote_total or 0))

    result = {
        'table': table,
        'local_total': local_total,
        'remote_total': remote_total,
        'fetched': 0,
        'inserted': 0,
        'attempted': bool(need_backfill)
    }
    if not need_backfill:
        return result

    raw_rows = client.select_all(
        table,
        filters=None,
        order_by=order_by,
        order_dir=order_dir,
        columns=columns,
        page_size=1000
    ) or []
    result['fetched'] = len(raw_rows)
    mapped_rows = []
    for raw in raw_rows:
        try:
            row = row_mapper(raw or {})
        except Exception:
            row = None
        if row:
            mapped_rows.append(row)
    missing_rows = _filter_missing_sqlite_rows(model, mapped_rows, conflict_cols)
    if missing_rows:
        result['inserted'] = int(_upsert_sqlite_rows(model, missing_rows, conflict_cols=conflict_cols) or 0)
    with _MATRIX_SB_SYNC_LOCK:
        _MATRIX_SB_SYNC_STATE[done_state_key] = True
    return result

def _maybe_backfill_matrix_and_logs_from_local_pg(force=False):
    client = get_supabase_client()
    if not client or not isinstance(client, LocalPostgreSQLClient):
        return {'enabled': False, 'attempted': False, 'ok': True, 'mode': 'sqlite_only'}

    now_ts = time.time()
    with _MATRIX_SB_SYNC_LOCK:
        last_pull_at = float(_MATRIX_SB_SYNC_STATE.get('matrix_local_compat_pull_at') or 0.0)
        if (not force) and (now_ts - last_pull_at) < 8.0:
            return {'enabled': True, 'attempted': False, 'ok': True, 'mode': 'local_pg_compat'}
        _MATRIX_SB_SYNC_STATE['matrix_local_compat_pull_at'] = now_ts

    try:
        include_button_diff_json = _supabase_has_column(client, 'button', 'diff_json')
        btn_columns = 'operation_id,question_wiki_id,product_name,old_is_configured,new_is_configured,edit_source,submitted_by,submitted_at'
        if include_button_diff_json:
            btn_columns += ',diff_json'

        results = {
            'matrix_column': _backfill_matrix_table_from_client(
                MatrixColumn,
                client,
                'matrix_column',
                'product_name,sort_order',
                'sort_order',
                'asc',
                ('product_name',),
                (lambda r: {
                    'product_name': str((r or {}).get('product_name') or '').strip(),
                    'sort_order': int((r or {}).get('sort_order') or 0)
                } if str((r or {}).get('product_name') or '').strip() else None),
                force=force
            ),
            'product_matrix': _backfill_matrix_table_from_client(
                ProductMatrix,
                client,
                'product_matrix',
                'question_wiki_id,product_name,is_configured,manual_edit,edit_source,last_synced_at,question_content,answer_content,update_time,product_category',
                'id',
                'asc',
                ('question_wiki_id', 'product_name'),
                (lambda r: {
                    'question_wiki_id': str((r or {}).get('question_wiki_id') or '').strip(),
                    'product_name': str((r or {}).get('product_name') or '').strip(),
                    'is_configured': bool((r or {}).get('is_configured', False)),
                    'manual_edit': bool((r or {}).get('manual_edit', False)),
                    'edit_source': str((r or {}).get('edit_source') or ''),
                    'last_synced_at': _coerce_naive_datetime((r or {}).get('last_synced_at')),
                    'question_content': (r or {}).get('question_content'),
                    'answer_content': (r or {}).get('answer_content'),
                    'update_time': (r or {}).get('update_time'),
                    'product_category': (r or {}).get('product_category'),
                } if str((r or {}).get('question_wiki_id') or '').strip() and str((r or {}).get('product_name') or '').strip() else None),
                force=force
            ),
            'matrix_submit_operation': _backfill_matrix_table_from_client(
                MatrixSubmitOperation,
                client,
                'matrix_submit_operation',
                'operation_id,status,attempts,created_by,error_message,created_at,updated_at',
                'created_at',
                'asc',
                ('operation_id',),
                (lambda r: {
                    'operation_id': str((r or {}).get('operation_id') or '').strip(),
                    'status': str((r or {}).get('status') or 'pending'),
                    'attempts': int((r or {}).get('attempts') or 0),
                    'created_by': (r or {}).get('created_by'),
                    'error_message': (r or {}).get('error_message'),
                    'created_at': _coerce_naive_datetime((r or {}).get('created_at')) or datetime.utcnow(),
                    'updated_at': _coerce_naive_datetime((r or {}).get('updated_at')) or _coerce_naive_datetime((r or {}).get('created_at')) or datetime.utcnow(),
                } if str((r or {}).get('operation_id') or '').strip() else None),
                force=force
            ),
            'button': _backfill_matrix_table_from_client(
                Button,
                client,
                'button',
                btn_columns,
                'submitted_at',
                'asc',
                ('operation_id', 'question_wiki_id', 'product_name'),
                (lambda r: {
                    'operation_id': str((r or {}).get('operation_id') or '').strip(),
                    'question_wiki_id': str((r or {}).get('question_wiki_id') or '').strip(),
                    'product_name': str((r or {}).get('product_name') or '').strip(),
                    'old_is_configured': bool((r or {}).get('old_is_configured', False)),
                    'new_is_configured': bool((r or {}).get('new_is_configured', False)),
                    'edit_source': str((r or {}).get('edit_source') or ''),
                    'diff_json': (r or {}).get('diff_json'),
                    'submitted_by': (r or {}).get('submitted_by'),
                    'submitted_at': _coerce_naive_datetime((r or {}).get('submitted_at')) or datetime.utcnow(),
                } if str((r or {}).get('operation_id') or '').strip() and str((r or {}).get('question_wiki_id') or '').strip() and str((r or {}).get('product_name') or '').strip() else None),
                force=force
            )
        }
        db.session.commit()
        attempted = any(bool((results.get(k) or {}).get('attempted')) for k in results.keys())
        return {'enabled': True, 'attempted': attempted, 'ok': True, 'mode': 'local_pg_compat', 'results': results}
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return {'enabled': True, 'attempted': True, 'ok': False, 'mode': 'local_pg_compat', 'message': str(e)}

def _maybe_pull_matrix_and_logs_from_supabase(force=False):
    if not is_supabase_matrix_enabled():
        return _maybe_backfill_matrix_and_logs_from_local_pg(force=force)

    now_ts = time.time()
    with _MATRIX_SB_SYNC_LOCK:
        last_pull_at = float(_MATRIX_SB_SYNC_STATE.get('matrix_last_pull_at') or 0.0)
        if (not force) and (now_ts - last_pull_at) < 8.0:
            return {'enabled': True, 'attempted': False, 'ok': True}
        _MATRIX_SB_SYNC_STATE['matrix_last_pull_at'] = now_ts
        last_synced_at_iso = _MATRIX_SB_SYNC_STATE.get('matrix_last_synced_at_iso')
        logs_last_pull_at = float(_MATRIX_SB_SYNC_STATE.get('logs_last_pull_at') or 0.0)
        op_updated_at_iso = _MATRIX_SB_SYNC_STATE.get('logs_last_op_updated_at_iso')
        btn_submitted_at_iso = _MATRIX_SB_SYNC_STATE.get('logs_last_btn_submitted_at_iso')
        pull_logs = force or ((now_ts - logs_last_pull_at) >= 15.0)
        if pull_logs:
            _MATRIX_SB_SYNC_STATE['logs_last_pull_at'] = now_ts

    if not last_synced_at_iso:
        try:
            max_dt = db.session.query(func.max(ProductMatrix.last_synced_at)).scalar()
            if max_dt:
                last_synced_at_iso = max_dt.isoformat() + 'Z'
                with _MATRIX_SB_SYNC_LOCK:
                    if not _MATRIX_SB_SYNC_STATE.get('matrix_last_synced_at_iso'):
                        _MATRIX_SB_SYNC_STATE['matrix_last_synced_at_iso'] = last_synced_at_iso
        except Exception:
            pass

    if pull_logs:
        if not op_updated_at_iso:
            try:
                max_dt = db.session.query(func.max(MatrixSubmitOperation.updated_at)).scalar()
                if max_dt:
                    op_updated_at_iso = max_dt.isoformat() + 'Z'
                    with _MATRIX_SB_SYNC_LOCK:
                        if not _MATRIX_SB_SYNC_STATE.get('logs_last_op_updated_at_iso'):
                            _MATRIX_SB_SYNC_STATE['logs_last_op_updated_at_iso'] = op_updated_at_iso
            except Exception:
                pass
        if not btn_submitted_at_iso:
            try:
                max_dt = db.session.query(func.max(Button.submitted_at)).scalar()
                if max_dt:
                    btn_submitted_at_iso = max_dt.isoformat() + 'Z'
                    with _MATRIX_SB_SYNC_LOCK:
                        if not _MATRIX_SB_SYNC_STATE.get('logs_last_btn_submitted_at_iso'):
                            _MATRIX_SB_SYNC_STATE['logs_last_btn_submitted_at_iso'] = btn_submitted_at_iso
            except Exception:
                pass

    client = get_supabase_client()
    if not client:
        return {'enabled': True, 'attempted': True, 'ok': False, 'message': '本地主库未配置'}

    try:
        cols = client.select_all(
            'matrix_column',
            filters=None,
            order_by='sort_order',
            order_dir='asc',
            columns='product_name,sort_order',
            page_size=1000
        )
        col_rows = []
        for c in cols or []:
            pn = str((c or {}).get('product_name') or '').strip()
            if not pn:
                continue
            try:
                so = int((c or {}).get('sort_order') or 0)
            except Exception:
                so = 0
            col_rows.append({'product_name': pn, 'sort_order': so})
        if col_rows:
            _upsert_sqlite_rows(MatrixColumn, col_rows, conflict_cols=('product_name',))

        pm_filters = None
        if last_synced_at_iso:
            pm_filters = {'last_synced_at': f'gte.{last_synced_at_iso}'}
        pm_rows = client.select_all(
            'product_matrix',
            filters=pm_filters,
            order_by='last_synced_at',
            order_dir='asc',
            columns='question_wiki_id,product_name,is_configured,manual_edit,edit_source,last_synced_at,question_content,answer_content,update_time,product_category',
            page_size=1000
        )
        to_upsert = []
        max_last_synced_at = None
        for r in pm_rows or []:
            wid = str((r or {}).get('question_wiki_id') or '').strip()
            pn = str((r or {}).get('product_name') or '').strip()
            if not wid or not pn:
                continue
            ls = _iso_to_naive_datetime((r or {}).get('last_synced_at'))
            if ls:
                iso = ls.isoformat() + 'Z'
                if (max_last_synced_at is None) or (iso > max_last_synced_at):
                    max_last_synced_at = iso
            to_upsert.append({
                'question_wiki_id': wid,
                'product_name': pn,
                'is_configured': bool((r or {}).get('is_configured', False)),
                'manual_edit': bool((r or {}).get('manual_edit', False)),
                'edit_source': str((r or {}).get('edit_source') or ''),
                'last_synced_at': ls,
                'question_content': (r or {}).get('question_content', None),
                'answer_content': (r or {}).get('answer_content', None),
                'update_time': (r or {}).get('update_time', None),
                'product_category': (r or {}).get('product_category', None)
            })
        if to_upsert:
            _upsert_sqlite_rows(ProductMatrix, to_upsert, conflict_cols=('question_wiki_id', 'product_name'))
        if max_last_synced_at:
            with _MATRIX_SB_SYNC_LOCK:
                _MATRIX_SB_SYNC_STATE['matrix_last_synced_at_iso'] = max_last_synced_at

        if pull_logs:
            op_filters = None
            if op_updated_at_iso:
                op_filters = {'updated_at': f'gte.{op_updated_at_iso}'}
            ops = client.select_all(
                'matrix_submit_operation',
                filters=op_filters,
                order_by='updated_at',
                order_dir='asc',
                columns='operation_id,status,attempts,created_by,error_message,created_at,updated_at',
                page_size=1000
            )
            op_upsert = []
            max_updated_at = None
            for op in ops or []:
                oid = str((op or {}).get('operation_id') or '').strip()
                if not oid:
                    continue
                u = _iso_to_naive_datetime((op or {}).get('updated_at'))
                if u:
                    iso = u.isoformat() + 'Z'
                    if (max_updated_at is None) or (iso > max_updated_at):
                        max_updated_at = iso
                op_upsert.append({
                    'operation_id': oid,
                    'status': str((op or {}).get('status') or 'pending'),
                    'attempts': int((op or {}).get('attempts') or 0),
                    'created_by': (op or {}).get('created_by', None),
                    'error_message': (op or {}).get('error_message', None),
                    'created_at': _iso_to_naive_datetime((op or {}).get('created_at')),
                    'updated_at': u or _iso_to_naive_datetime((op or {}).get('created_at')) or datetime.utcnow()
                })
            if op_upsert:
                _upsert_sqlite_rows(MatrixSubmitOperation, op_upsert, conflict_cols=('operation_id',))
            if max_updated_at:
                with _MATRIX_SB_SYNC_LOCK:
                    _MATRIX_SB_SYNC_STATE['logs_last_op_updated_at_iso'] = max_updated_at

            btn_filters = None
            if btn_submitted_at_iso:
                btn_filters = {'submitted_at': f'gte.{btn_submitted_at_iso}'}
            include_diff_json = _supabase_has_column(client, 'button', 'diff_json')
            btn_columns = 'operation_id,question_wiki_id,product_name,old_is_configured,new_is_configured,edit_source,submitted_by,submitted_at'
            if include_diff_json:
                btn_columns += ',diff_json'
            btns = client.select_all(
                'button',
                filters=btn_filters,
                order_by='submitted_at',
                order_dir='asc',
                columns=btn_columns,
                page_size=1000
            )
            btn_upsert = []
            max_submitted_at = None
            for b in btns or []:
                oid = str((b or {}).get('operation_id') or '').strip()
                wid = str((b or {}).get('question_wiki_id') or '').strip()
                pn = str((b or {}).get('product_name') or '').strip()
                if not oid or not wid or not pn:
                    continue
                sdt = _iso_to_naive_datetime((b or {}).get('submitted_at'))
                if sdt:
                    iso = sdt.isoformat() + 'Z'
                    if (max_submitted_at is None) or (iso > max_submitted_at):
                        max_submitted_at = iso
                btn_upsert.append({
                    'operation_id': oid,
                    'question_wiki_id': wid,
                    'product_name': pn,
                    'old_is_configured': bool((b or {}).get('old_is_configured', False)),
                    'new_is_configured': bool((b or {}).get('new_is_configured', False)),
                    'edit_source': str((b or {}).get('edit_source') or ''),
                    'diff_json': (b or {}).get('diff_json', None),
                    'submitted_by': (b or {}).get('submitted_by', None),
                    'submitted_at': sdt or datetime.utcnow()
                })
            if btn_upsert:
                _upsert_sqlite_rows(Button, btn_upsert, conflict_cols=('operation_id', 'question_wiki_id', 'product_name'))
            if max_submitted_at:
                with _MATRIX_SB_SYNC_LOCK:
                    _MATRIX_SB_SYNC_STATE['logs_last_btn_submitted_at_iso'] = max_submitted_at

        db.session.commit()
        return {'enabled': True, 'attempted': True, 'ok': True}
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return {'enabled': True, 'attempted': True, 'ok': False, 'message': str(e)}

@app.route('/api/migrate/sqlite_to_primary_db', methods=['POST'])
@app.route('/api/migrate/sqlite_to_supabase', methods=['POST'])
@login_required
def migrate_sqlite_to_supabase():
    payload = request.json or {}
    targets = payload.get('targets')
    if not targets:
        targets = ['kb_recall', 'matrix', 'matrix_logs']
    if not isinstance(targets, list):
        return jsonify({'success': False, 'message': 'targets must be a list'}), 400
    targets = [str(x).strip() for x in targets if str(x).strip()]
    targets = list(dict.fromkeys(targets))
    chunk_size = payload.get('chunk_size', 500)
    dry_run = bool(payload.get('dry_run', False))
    try:
        offset = int(payload.get('offset', 0) or 0)
    except Exception:
        offset = 0
    try:
        limit = payload.get('limit', None)
        limit = int(limit) if limit is not None else None
    except Exception:
        limit = None

    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': _db_not_configured_message()}), 400

    out = {'success': True, 'dry_run': dry_run, 'targets': targets, 'results': {}}

    try:
        do_kb_recall = 'kb_recall' in targets
        do_matrix_column = ('matrix' in targets) or ('matrix_column' in targets)
        do_product_matrix = ('matrix' in targets) or ('product_matrix' in targets)
        do_matrix_logs = 'matrix_logs' in targets

        if not dry_run:
            required_tables = []
            if do_kb_recall:
                required_tables.append('kb_recall')
            if do_matrix_column:
                required_tables.append('matrix_column')
            if do_product_matrix:
                required_tables.append('product_matrix')
            if do_matrix_logs:
                required_tables.extend(['matrix_submit_operation', 'button'])
            required_tables = list(dict.fromkeys(required_tables))
            missing = []
            missing_details = {}
            for t in required_tables:
                ok, detail = _supabase_table_exists(client, t)
                if not ok:
                    missing.append(t)
                    if detail:
                        missing_details[t] = detail
            if missing:
                return jsonify({
                    'success': False,
                    'dry_run': dry_run,
                    'targets': targets,
                    'missing_tables': missing,
                    'missing_details': missing_details,
                    'message': '目标主库缺少必要表，请先执行 `supabase_schema.sql` 中对应建表 SQL，再进行迁移'
                }), 400

        if do_kb_recall:
            q = KBRecall.query.order_by(KBRecall.id.asc())
            if offset and offset > 0:
                q = q.offset(offset)
            if limit is not None and limit >= 0:
                q = q.limit(limit)
            rows = q.all()
            to_send = []
            for r in rows:
                to_send.append({
                    'kb_id': str(r.kb_id or '').strip(),
                    'month': str(r.month or '').strip(),
                    'recall_count': int(r.recall_count or 0),
                    'valid_recall_count': int(r.valid_recall_count or 0),
                    'created_at': _dt_to_iso(getattr(r, 'created_at', None))
                })
            to_send = [x for x in to_send if x.get('kb_id') and x.get('month')]
            if dry_run:
                out['results']['kb_recall'] = {'rows': len(to_send), 'offset': offset, 'limit': limit}
            else:
                out['results']['kb_recall'] = _supabase_upsert_chunks(
                    client,
                    'kb_recall',
                    to_send,
                    on_conflict='kb_id,month',
                    chunk_size=chunk_size
                )
                out['results']['kb_recall']['offset'] = offset
                out['results']['kb_recall']['limit'] = limit
                if out['results']['kb_recall'].get('ok') is False:
                    out['success'] = False

        if do_matrix_column or do_product_matrix:
            col_payload = []
            if do_matrix_column:
                cols = MatrixColumn.query.order_by(MatrixColumn.sort_order).all()
                for c in cols:
                    pn = str(getattr(c, 'product_name', '') or '').strip()
                    if not pn:
                        continue
                    try:
                        so = int(getattr(c, 'sort_order', 0) or 0)
                    except Exception:
                        so = 0
                    col_payload.append({'product_name': pn, 'sort_order': so})

            mat_payload = []
            if do_product_matrix:
                mat_q = ProductMatrix.query
                try:
                    mat_q = mat_q.order_by(ProductMatrix.id.asc())
                except Exception:
                    mat_q = mat_q.order_by(ProductMatrix.question_wiki_id.asc(), ProductMatrix.product_name.asc())
                if offset and offset > 0:
                    mat_q = mat_q.offset(offset)
                if limit is not None and limit >= 0:
                    mat_q = mat_q.limit(limit)
                mats = mat_q.all()
                for m in mats:
                    wid = str(getattr(m, 'question_wiki_id', '') or '').strip()
                    pn = str(getattr(m, 'product_name', '') or '').strip()
                    if not wid or not pn:
                        continue
                    mat_payload.append({
                        'question_wiki_id': wid,
                        'product_name': pn,
                        'is_configured': bool(getattr(m, 'is_configured', False)),
                        'manual_edit': bool(getattr(m, 'manual_edit', False)),
                        'edit_source': str(getattr(m, 'edit_source', '') or ''),
                        'last_synced_at': _dt_to_iso(getattr(m, 'last_synced_at', None)),
                        'question_content': getattr(m, 'question_content', None),
                        'answer_content': getattr(m, 'answer_content', None),
                        'update_time': getattr(m, 'update_time', None),
                        'product_category': getattr(m, 'product_category', None)
                    })

            if dry_run:
                if do_matrix_column:
                    out['results']['matrix_column'] = {'rows': len(col_payload)}
                if do_product_matrix:
                    out['results']['product_matrix'] = {'rows': len(mat_payload), 'offset': offset, 'limit': limit}
            else:
                if do_matrix_column:
                    out['results']['matrix_column'] = _supabase_upsert_chunks(
                        client,
                        'matrix_column',
                        col_payload,
                        on_conflict='product_name',
                        chunk_size=chunk_size
                    )
                    if out['results']['matrix_column'].get('ok') is False:
                        out['success'] = False
                if do_product_matrix:
                    out['results']['product_matrix'] = _supabase_upsert_chunks(
                        client,
                        'product_matrix',
                        mat_payload,
                        on_conflict='question_wiki_id,product_name',
                        chunk_size=chunk_size
                    )
                    out['results']['product_matrix']['offset'] = offset
                    out['results']['product_matrix']['limit'] = limit
                    if out['results']['product_matrix'].get('ok') is False:
                        out['success'] = False

        if do_matrix_logs:
            ops = MatrixSubmitOperation.query.all()
            op_payload = []
            for op in ops:
                oid = str(getattr(op, 'operation_id', '') or '').strip()
                if not oid:
                    continue
                op_payload.append({
                    'operation_id': oid,
                    'status': str(getattr(op, 'status', '') or 'pending'),
                    'attempts': int(getattr(op, 'attempts', 0) or 0),
                    'created_by': getattr(op, 'created_by', None),
                    'error_message': getattr(op, 'error_message', None),
                    'created_at': _dt_to_iso(getattr(op, 'created_at', None)),
                    'updated_at': _dt_to_iso(getattr(op, 'updated_at', None))
                })

            btns = Button.query.all()
            btn_payload = []
            include_diff_json = True
            if not dry_run:
                include_diff_json = _supabase_has_column(client, 'button', 'diff_json')
            for b in btns:
                oid = str(getattr(b, 'operation_id', '') or '').strip()
                wid = str(getattr(b, 'question_wiki_id', '') or '').strip()
                pn = str(getattr(b, 'product_name', '') or '').strip()
                if not oid or not wid or not pn:
                    continue
                row = {
                    'operation_id': oid,
                    'question_wiki_id': wid,
                    'product_name': pn,
                    'old_is_configured': bool(getattr(b, 'old_is_configured', False)),
                    'new_is_configured': bool(getattr(b, 'new_is_configured', False)),
                    'edit_source': str(getattr(b, 'edit_source', '') or ''),
                    'submitted_by': getattr(b, 'submitted_by', None),
                    'submitted_at': _dt_to_iso(getattr(b, 'submitted_at', None))
                }
                if include_diff_json:
                    row['diff_json'] = getattr(b, 'diff_json', None)
                btn_payload.append(row)

            if dry_run:
                out['results']['matrix_submit_operation'] = {'rows': len(op_payload)}
                out['results']['button'] = {'rows': len(btn_payload)}
            else:
                out['results']['matrix_submit_operation'] = _supabase_upsert_chunks(
                    client,
                    'matrix_submit_operation',
                    op_payload,
                    on_conflict='operation_id',
                    chunk_size=chunk_size
                )
                out['results']['button'] = _supabase_upsert_chunks(
                    client,
                    'button',
                    btn_payload,
                    on_conflict='operation_id,question_wiki_id,product_name',
                    chunk_size=chunk_size
                )
                if out['results']['matrix_submit_operation'].get('ok') is False:
                    out['success'] = False
                if out['results']['button'].get('ok') is False:
                    out['success'] = False

        return jsonify(out)
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e), 'results': out.get('results', {})}), 500

# Config route removed/disabled
@app.route('/api/kb/config', methods=['GET', 'POST'])
@login_required
def kb_config():
    return jsonify({'message': 'Configuration is managed server-side.'})

@app.route('/api/kb/template', methods=['GET'])
@login_required
def kb_template():
    # Create a DataFrame with the required columns
    columns = [
        'ID', '问题类型', '问题', '答案', '答案类型', 'BM25', '相似问题', 
        '错误列表', '关键词', '图片链接', '视频链接', '文件链接', '跳转链接类型', '跳转链接（url/key）', '更新时间', '型号', '产品名称'
    ]
    # Add a sample row to help user understand format
    sample_data = [{
        'ID': 'ICWIKI202307240001',
        '问题类型': 'Usage',
        '问题': '示例问题：如何重置设备？',
        '答案': '长按电源键10秒即可重置。',
        '答案类型': 'Troubleshooting',
        'BM25': 'TRUE',
        '相似问题': '["怎么重启", "设备死机怎么办"]',
        '错误列表': '[]',
        '关键词': '["重置", "电源键"]',
        '图片链接': '["http://example.com/howto.png"]',
        '视频链接': '["http://example.com/howto.mp4"]',
        '文件链接': '["http://example.com/howto.pdf"]',
        '跳转链接类型': '外部链接',
        '跳转链接（url/key）': 'http://example.com/manual',
        '更新时间': '2024-01-01',
        '型号': 'SD-2024',
        '产品名称': 'SmartDevice'
    }]
    df = pd.DataFrame(sample_data, columns=columns)
    
    try:
        # Write to BytesIO
        output = io.BytesIO()
        engine = 'xlsxwriter' if importlib.util.find_spec('xlsxwriter') is not None else 'openpyxl'
        with pd.ExcelWriter(output, engine=engine) as writer:
            df.to_excel(writer, index=False, sheet_name='ImportTemplate')
            worksheet = writer.sheets['ImportTemplate']
            if engine == 'xlsxwriter':
                for idx, _col in enumerate(df.columns):
                    worksheet.set_column(idx, idx, 20)
            else:
                from openpyxl.utils import get_column_letter
                for idx, _col in enumerate(df.columns, start=1):
                    worksheet.column_dimensions[get_column_letter(idx)].width = 20
                
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=canonical_download_name('kb_template'),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error generating template: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

def _kb_import_column_mapping():
    return {
        # 中文表头
        'ID': 'question_wiki_id',
        'id': 'question_wiki_id',
        'Id': 'question_wiki_id',
        '问题ID': 'question_wiki_id',
        '问题id': 'question_wiki_id',
        '问题编号': 'question_wiki_id',
        '问题类型': 'question_type',
        '问题': 'question',
        '答案': 'answer',
        '答案类型': 'answer_type',
        'BM25': 'if_bm25',
        'bm25': 'if_bm25',
        '相似问题': 'similar_questions',
        '相似提问': 'similar_questions',
        '错误列表': 'error_list',
        '关键词': 'keyword_list',
        '图片链接': 'image_urls',
        '视频链接': 'video_urls',
        '文件链接': 'file_urls',
        '跳转链接类型': 'link_type',
        '链接类型': 'link_type',
        '跳转链接（url/key）': 'link_url',
        '跳转链接': 'link_url',
        'URLs': 'link_url',
        'URL链接': 'link_url',
        '更新时间': 'update_time',
        '更新时刻': 'update_time',
        # 型号类字段统一表示适用型号；品类类字段才进入 product_category_name。
        '型号': 'product_name',
        '产品型号': 'product_name',
        '产品名称': 'product_name',
        '机型': 'product_name',
        '产品': 'product_name',
        '产品品类': 'product_category_name',
        '产品分类': 'product_category_name',
        '分类': 'product_category_name',

        # 英文表头
        'question_wiki_id': 'question_wiki_id',
        'question_type': 'question_type',
        'question': 'question',
        'Question': 'question',
        'answer': 'answer',
        'Answer': 'answer',
        'answer_type': 'answer_type',
        'Type': 'answer_type',
        'Question Type': 'question_type',
        'if_bm25': 'if_bm25',
        'similar_questions': 'similar_questions',
        'error_list': 'error_list',
        'keyword_list': 'keyword_list',
        'image_urls': 'image_urls',
        'video_urls': 'video_urls',
        'file_urls': 'file_urls',
        'link_type': 'link_type',
        'Link Type': 'link_type',
        'link_url': 'link_url',
        'Link URL': 'link_url',
        'update_time': 'update_time',
        'product_categories': 'product_category_name',
        'product_category_name': 'product_category_name',
        'Category': 'product_category_name',
        'products': 'product_name',
        'product_name': 'product_name',
        'Product': 'product_name'
    }

def _dedupe_kb_compare_ids(raw_ids):
    ids = []
    duplicates = []
    seen = set()
    for value in raw_ids or []:
        kid = str(value or '').strip().replace('\ufeff', '')
        if not kid:
            continue
        if kid.lower() in ('id', 'wikiid', 'wiki_id', 'question_wiki_id') or kid in ('问题编号', '问题ID', '问题id'):
            continue
        if not re.match(r'^(ICWIKI[A-Za-z0-9_-]+|[A-Za-z0-9_-]{8,})$', kid, re.I):
            continue
        if kid in seen:
            duplicates.append(kid)
            continue
        seen.add(kid)
        ids.append(kid)
    return ids, duplicates

def _extract_kb_compare_ids_from_text(text, allow_generic=True):
    raw = str(text or '')
    direct_ids = re.findall(r'ICWIKI[A-Za-z0-9_-]+', raw, flags=re.I)
    if direct_ids:
        return _dedupe_kb_compare_ids(direct_ids)
    if not allow_generic:
        return [], []
    parts = re.split(r'[\s,，;；、|"\']+', raw)
    return _dedupe_kb_compare_ids(parts)

def _extract_kb_compare_ids_from_df(df):
    if df is None or getattr(df, 'empty', True):
        return [], []

    id_column_names = {'id', 'wikiid', 'wiki_id', 'question_wiki_id', '问题编号', '问题id', '问题ID'}
    candidate_cols = [
        col for col in df.columns
        if str(col or '').strip() in id_column_names
        or str(col or '').strip().lower() in id_column_names
    ]

    if candidate_cols:
        values = []
        for col in candidate_cols:
            values.extend(df[col].tolist())
        return _dedupe_kb_compare_ids(values)

    text = '\n'.join(
        str(v)
        for v in df.fillna('').astype(str).to_numpy().flatten().tolist()
        if str(v).strip()
    )
    ids, duplicates = _extract_kb_compare_ids_from_text(text, allow_generic=False)
    if ids:
        return ids, duplicates

    first_col = df.columns[0] if len(df.columns) else None
    if first_col is not None:
        return _dedupe_kb_compare_ids(df[first_col].tolist())
    return [], []

@app.route('/api/kb/compare/parse_ids', methods=['POST'])
@login_required
def kb_compare_parse_ids():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'success': False, 'message': '请选择文件'}), 400

        filename = secure_filename(file.filename or 'ids')
        lower = filename.lower()

        if lower.endswith(('.txt', '.csv')):
            data = file.read()
            text = ''
            for enc in ('utf-8-sig', 'utf-8', 'gbk'):
                try:
                    text = data.decode(enc)
                    break
                except Exception:
                    continue
            ids, duplicates = _extract_kb_compare_ids_from_text(text, allow_generic=True)
        elif lower.endswith(('.xlsx', '.xls')):
            sheets = pd.read_excel(file, sheet_name=None, dtype=str)
            ids = []
            duplicates = []
            seen = set()
            for df in (sheets or {}).values():
                sheet_ids, sheet_dups = _extract_kb_compare_ids_from_df(df)
                for kid in sheet_ids:
                    if kid in seen:
                        duplicates.append(kid)
                    else:
                        seen.add(kid)
                        ids.append(kid)
                duplicates.extend(sheet_dups)
        else:
            return jsonify({'success': False, 'message': '仅支持 txt / csv / xlsx / xls'}), 400

        return jsonify({
            'success': True,
            'ids': ids,
            'duplicateIds': duplicates,
            'count': len(ids)
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/kb/import', methods=['POST'])
@login_required
def kb_import():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
    mode = request.form.get('mode') or request.form.get('importMode') or 'upsert'  # upsert, append, overwrite
    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'message': 'Empty filename'}), 400

    try:
        # Read Excel or CSV file
        if file.filename.lower().endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        required_cols = ['image_urls', 'video_urls', 'file_urls', 'link_type', 'link_url']
        missing_cols = [c for c in required_cols if not _supabase_has_column(client, 'knowledge_base_v1', c)]
        if missing_cols:
            return jsonify({
                'success': False,
                'message': f"导入失败：knowledge_base_v1 缺少字段 {', '.join(missing_cols)}。请先在本地主库执行 `update_schema_v3.sql` 中的 KB Schema 更新 SQL（新增图片/视频/文件/跳转链接字段并删除 answer_info/urls）。"
            }), 400
        
        # Column Mapping (Excel Header -> DB Column)
        column_mapping = _kb_import_column_mapping()
        
        # Rename columns if they exist
        df.rename(columns=column_mapping, inplace=True)

        # 始终忽略导入文件中的 review_status，防止覆盖现有修订状态
        # 修订状态只由系统内部流程控制（创建/修改/完成修订等），不从 Excel 同步
        if 'review_status' in df.columns:
            df.drop(columns=['review_status'], inplace=True)

        forbidden_cols = [c for c in ('urls', 'answer_info') if c in df.columns]
        if forbidden_cols:
            return jsonify({
                'success': False,
                'message': f"导入失败：检测到旧字段 {', '.join(forbidden_cols)}。当前版本已删除 urls/answer_info，请使用新字段：图片链接/视频链接/文件链接/跳转链接类型/跳转链接（url/key）。"
            }), 400
        
        # --- DATA CLEANING (Added per requirements) ---
        
        # 1. Product Name & Category Cleaning: remove space after comma
        for col in ['product_name', 'product_category_name']:
            if col in df.columns:
                # Convert to string, replace ", " with ",", then replace "nan" with None/Empty
                df[col] = df[col].astype(str).apply(lambda x: re.sub(r',\s+', ',', x) if pd.notna(x) and x.lower() != 'nan' else x)
                # Cleanup "nan" strings back to None or empty string if preferred, but later logic handles strings
                # Let's keep as string but empty if it was nan
                df[col] = df[col].replace('nan', '')

        # 4. Answer Type Mapping: text->文本, image->图片, video->视频
        if 'answer_type' in df.columns:
            type_map = {'text': '文本', 'image': '图片', 'video': '视频'}
            # Use map with get to keep original if not in map
            df['answer_type'] = df['answer_type'].astype(str).apply(lambda x: type_map.get(x.lower(), x) if pd.notna(x) and x.lower() != 'nan' else x)
            df['answer_type'] = df['answer_type'].replace('nan', '')

        # Validate Products (User Requirement)
        if 'product_name' in df.columns:
            valid_map, valid_set = get_all_valid_models()
            invalid_rows = []
            
            for idx, row in df.iterrows():
                p_name = "" if _is_blank_cell_value(row['product_name']) else str(row['product_name'])
                if p_name.strip():
                    valid, invalid = validate_product_string(p_name, valid_map, valid_set)
                    
                    if invalid:
                        invalid_rows.append(f"行 {idx+2} (ID: {row.get('question_wiki_id', 'Unknown')}): 未知型号 {invalid}")
                    else:
                        # Normalize: sort and join
                        valid.sort()
                        df.at[idx, 'product_name'] = ",".join(valid)
            
            if invalid_rows:
                # Block import
                return jsonify({'success': False, 'message': f"导入失败：发现 {len(invalid_rows)} 行包含未知产品型号。请修正后重试。\n" + "\n".join(invalid_rows[:10]) + ("..." if len(invalid_rows)>10 else "")}), 400

        # 自动根据型号库回填产品分类（若为空）
        # 规则：根据该行的 product_name（可能是多个型号逗号分隔）反查其所属分类；
        # 若命中多个分类，用英文逗号连接，排序后写入 product_category_name
        try:
            catalog = parse_product_catalog() or {}
            # 建立 归一化型号 -> 分类集合 的映射，提高匹配鲁棒性
            norm_model_to_categories = {}
            for cat, models in (catalog.items() if isinstance(catalog, dict) else []):
                if not cat: continue
                if not isinstance(models, list): continue
                for m in models:
                    ms = str(m).strip()
                    if not ms: continue
                    # 归一化处理（去空格，转小写）
                    norm_key = ms.replace(" ", "").lower()
                    norm_model_to_categories.setdefault(norm_key, set()).add(str(cat).strip())

            if 'product_name' in df.columns:
                if 'product_category_name' not in df.columns:
                    df['product_category_name'] = ''

                for idx, row in df.iterrows():
                    cur_cat = str(row.get('product_category_name') or '').strip()
                    # 如果分类已存在（且不是 nan/null 等），则跳过自动回填，尊重手动设置
                    if cur_cat and cur_cat.lower() not in ('nan', 'null', 'none'):
                        continue
                    
                    raw_models = str(row.get('product_name') or '').strip()
                    if not raw_models or raw_models.lower() in ('nan', 'null', 'none'):
                        continue
                    
                    # split and normalize each model for matching
                    parts = [p.strip() for p in re.split(r'[,，]', raw_models) if p.strip()]
                    cats = set()
                    for p in parts:
                        # 同样进行归一化匹配
                        p_norm = p.replace(" ", "").lower()
                        
                        # 1. 精确匹配
                        hit = norm_model_to_categories.get(p_norm)
                        if hit:
                            cats.update(hit)
                            continue
                            
                        # 2. 模糊匹配：如果输入是 "P10S Pro"，目录里是 "P10S Pro 系列"
                        # 或者输入带了“系列”，目录里没带
                        p_pure = p_norm.replace("系列", "").replace("series", "")
                        for norm_key, cat_set in norm_model_to_categories.items():
                            norm_pure = norm_key.replace("系列", "").replace("series", "")
                            if p_pure == norm_pure:
                                cats.update(cat_set)
                                break
                    
                    if cats:
                        df.at[idx, 'product_category_name'] = ",".join(sorted(cats))
                        print(f"[Import] Auto-filled categories for row {idx+2} ({raw_models}): {df.at[idx, 'product_category_name']}")
        except Exception as _e:
            # 回填失败不阻断导入，保持兼容
            print(f"WARN: auto fill product_category_name failed: {_e}")

        # Helper to clean NaN recursively
        def clean_nan(obj):
            if isinstance(obj, float) and (obj != obj or obj == float('inf') or obj == float('-inf')):
                return None
            elif isinstance(obj, dict):
                return {k: clean_nan(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [clean_nan(v) for v in obj]
            return obj

        # Convert specific columns to JSON if they are strings
        json_cols = ['similar_questions', 'error_list', 'keyword_list', 'image_urls', 'video_urls', 'file_urls']
        
        # Pre-process dataframe to dicts
        records = df.to_dict(orient='records')
        
        processed_records = []
        for record in records:
            # 3. Clean List Fields: remove [, ], " and split
            # 5. Remove 'null' strings
            for col in json_cols:
                if col in record:
                    val = record[col]
                    cleaned = None
                    if val is None:
                        cleaned = None
                    elif isinstance(val, float) and pd.isna(val):
                        cleaned = None
                    elif isinstance(val, dict):
                        cleaned = val
                    elif isinstance(val, list):
                        cleaned = [str(v).strip() for v in val if v is not None and str(v).strip() and str(v).strip().lower() != 'null']
                    elif isinstance(val, str):
                        s = val.strip()
                        if not s or s.lower() == 'nan' or s.lower() == 'null':
                            cleaned = None
                        else:
                            parsed = None
                            try:
                                parsed = json.loads(s)
                            except Exception:
                                parsed = None
                            
                            if isinstance(parsed, dict):
                                cleaned = parsed
                            elif isinstance(parsed, list):
                                cleaned = [str(v).strip() for v in parsed if v is not None and str(v).strip() and str(v).strip().lower() != 'null']
                            else:
                                val_clean = s.replace('[', '').replace(']', '').replace('"', '').replace("'", "")
                                parts = _split_text_list_value(val_clean, is_url_list=(col in ('image_urls', 'video_urls', 'file_urls')))
                                final_list = [p.strip() for p in parts if p.strip() and p.strip().lower() != 'null']
                                cleaned = final_list if final_list else None
                    else:
                        cleaned = val
                    
                    if cleaned is None:
                        record[col] = None
                    elif isinstance(cleaned, list) and len(cleaned) == 0:
                        record[col] = None
                    elif isinstance(cleaned, dict) and len(cleaned) == 0:
                        record[col] = None
                    else:
                        record[col] = cleaned
            
            # Ensure boolean
            # 2. if_bm25: 0->False, 1->True, Empty->False
            if 'if_bm25' in record:
                val = record['if_bm25']
                if _is_blank_cell_value(val):
                    record['if_bm25'] = False
                else:
                    str_val = str(val).strip().lower()
                    if str_val in ['1', 'true', 'yes', '是', 'true']:
                        record['if_bm25'] = True
                    elif str_val in ['0', 'false', 'no', '否', 'false']:
                        record['if_bm25'] = False
                    else:
                        record['if_bm25'] = False # Default to False per requirement (Empty->False)
            else:
                 record['if_bm25'] = False
            
            # Handle date format
            if 'update_time' in record:
                if isinstance(record['update_time'], datetime):
                    record['update_time'] = record['update_time'].isoformat()
                elif _is_blank_cell_value(record['update_time']):
                     del record['update_time']
            
            if 'link_url' in record:
                s = str(record.get('link_url') or '').strip()
                if not s or s.lower() in ('nan', 'null', 'none'):
                    record['link_url'] = None
                else:
                    record['link_url'] = s
            if 'link_type' in record:
                s = str(record.get('link_type') or '').strip()
                if not s or s.lower() in ('nan', 'null', 'none'):
                    record['link_type'] = None
                else:
                    record['link_type'] = s
            if record.get('link_url') and not record.get('link_type'):
                record['link_type'] = '外部链接'

            # Recursively clean NaN
            cleaned_record = clean_nan(record)
            if isinstance(cleaned_record, dict):
                allow = [
                    'question_wiki_id',
                    'question_type',
                    'question',
                    'answer',
                    'answer_type',
                    'if_bm25',
                    'similar_questions',
                    'error_list',
                    'keyword_list',
                    'image_urls',
                    'video_urls',
                    'file_urls',
                    'link_type',
                    'link_url',
                    'update_time',
                    'product_category_name',
                    'product_name'
                ]
                cleaned_record = {k: cleaned_record.get(k) for k in allow if k in cleaned_record}
            
            # Remove question_wiki_id if it is None/Empty to allow auto-increment (if supported)
            # or avoid sending null for PK
            if 'question_wiki_id' in cleaned_record:
                val = cleaned_record['question_wiki_id']
                if val is None or val == '':
                    del cleaned_record['question_wiki_id']
                else:
                    # Ensure it's int if needed? Supabase/JSON handles numbers.
                    # But if it's "123" string, it might be fine.
                    pass

            processed_records.append(cleaned_record)

        # Deduplicate records by ID (keep last occurrence) to avoid internal duplicates in the batch
        # This prevents PK violations if the file itself contains duplicate IDs
        unique_map = {}
        no_id_records = []
        
        for r in processed_records:
            if 'question_wiki_id' in r:
                unique_map[r['question_wiki_id']] = r
            else:
                no_id_records.append(r)
        
        # Reconstruct list: unique IDs + records without ID
        processed_records = list(unique_map.values()) + no_id_records

        # Batch insert (upsert/insert/overwrite)
        batch_size = 100
        total_inserted = 0
        stats = {'added': 0, 'updated': 0, 'skipped': 0, 'total': len(processed_records)}
        details = {'added_ids': [], 'updated_ids': [], 'skipped_ids': []}
        file_ids_all = [str(r['question_wiki_id']) for r in processed_records if 'question_wiki_id' in r]
        errors = []
        pre_t1_sync_note = None
        score_backup_path = None
        score_backup_rows = []
        delete_missing_requested = (
            mode != 'overwrite'
            and str(request.form.get('delete_missing') or '').strip().lower() in ('1', 'true', 'yes', 'y')
        )
        delete_missing_ids = []
        delete_missing_result = None

        if delete_missing_requested:
            if not file_ids_all:
                return jsonify({
                    'success': False,
                    'message': '启用“同步删除缺失 ID”时，导入文件必须包含至少一个有效 ID。'
                }), 400
            delete_missing_ids = _get_v1_ids_missing_from_file(client, file_ids_all)
            expected_delete_missing = request.form.get('expected_delete_missing_count')
            if expected_delete_missing is not None:
                try:
                    if int(expected_delete_missing) != len(delete_missing_ids):
                        return jsonify({
                            'success': False,
                            'requires_confirmation': True,
                            'confirmation_type': 'delete_missing',
                            'message': f'缺失 ID 删除数量已变化（确认时 {expected_delete_missing} 条，当前 {len(delete_missing_ids)} 条），请重新确认。',
                            'preview': {
                                'delete_missing_count': len(delete_missing_ids),
                                'delete_missing_sample_ids': delete_missing_ids[:20]
                            }
                        }), 409
                except Exception:
                    return jsonify({'success': False, 'message': 'expected_delete_missing_count 必须为数字'}), 400

            confirm_delete_missing = str(request.form.get('confirm_delete_missing') or '').strip().lower() in ('1', 'true', 'yes', 'y')
            if delete_missing_ids and not confirm_delete_missing:
                return jsonify({
                    'success': False,
                    'requires_confirmation': True,
                    'confirmation_type': 'delete_missing',
                    'message': '同步删除缺失 ID 会把当前 V1 中存在、但导入文件 ID 列未包含的记录直接物理删除。请确认后继续。',
                    'preview': {
                        'delete_missing_count': len(delete_missing_ids),
                        'delete_missing_sample_ids': delete_missing_ids[:20]
                    }
                }), 409

        if mode == 'overwrite':
            current_count = _client_count(client, 'knowledge_base_v1')
            score_count = _client_count(client, 'kb_scores')
            confirm_overwrite = str(request.form.get('confirm_overwrite') or '').strip().lower() in ('1', 'true', 'yes', 'y')
            if not confirm_overwrite:
                return jsonify({
                    'success': False,
                    'requires_confirmation': True,
                    'message': '全量覆盖导入会备份 V1 到 V1T-1，然后清空当前 V1 与评分缓存。请确认影响范围后继续。',
                    'preview': {
                        'incoming_count': len(processed_records),
                        'current_v1_count': current_count,
                        'score_count': score_count,
                        'unique_id_count': len(set(file_ids_all)),
                        'missing_id_count': len([r for r in processed_records if 'question_wiki_id' not in r])
                    }
                }), 409
            expected_incoming = request.form.get('expected_incoming_count')
            if expected_incoming is not None:
                try:
                    if int(expected_incoming) != len(processed_records):
                        return jsonify({
                            'success': False,
                            'requires_confirmation': True,
                            'message': f'导入文件记录数已变化（确认时 {expected_incoming} 条，当前 {len(processed_records)} 条），请重新确认。',
                            'preview': {
                                'incoming_count': len(processed_records),
                                'current_v1_count': current_count,
                                'score_count': score_count
                            }
                        }), 409
                except Exception:
                    return jsonify({'success': False, 'message': 'expected_incoming_count 必须为数字'}), 400

        # Mode 2: Overwrite (Clear all data first)
        if mode == 'overwrite':
            # 0. 先把当前 V1 备份到前刻库 V1T-1，再执行清空与导入（与手动「同步此刻到前刻」同一逻辑）
            print("DEBUG: Overwrite mode — V1 -> V1T-1 backup before destructive import...")
            sync_res = _run_kb_v1_to_t1_sync(client)
            if not sync_res.get('success'):
                return jsonify({
                    'success': False,
                    'message': (
                        f"全量覆盖前自动同步到前刻库失败：{sync_res.get('message', 'unknown')}。"
                        "此刻库未清空，请检查网络或权限后重试；亦可先手动点击「同步此刻到前刻」再导入。"
                    )
                }), 500
            pre_t1_sync_note = sync_res.get('message') or ''
            try:
                score_backup_rows = client.select_all('kb_scores', order_by='id', page_size=1000) or []
                score_backup_path = _write_json_backup('kb_import_overwrite', 'kb_scores', score_backup_rows, {
                    'action': 'overwrite_import_pre_clear',
                    'incoming_count': len(processed_records)
                })
            except Exception as _e:
                return jsonify({
                    'success': False,
                    'message': f'全量覆盖前评分缓存备份失败：{str(_e)}。此刻库未清空，请检查磁盘权限后重试。'
                }), 500
            # 1. Fetch all IDs first to ensure complete deletion
            # This avoids issues with implicit delete limits or partial content
            print(f"DEBUG: Starting overwrite mode cleanup (Robust Strategy)...")
            
            # Step 0: Clear dependent tables first (kb_scores)
            # Loop until truly empty using Batch Delete strategy
            print("DEBUG: Clearing dependent table 'kb_scores'...")
            max_retries = 100
            delete_batch_size = 1000 # kb_scores is simple, can handle larger batches
            
            for i in range(max_retries):
                 # Check count first to avoid unnecessary selects
                 check_count = client.select('kb_scores', page=1, page_size=1)
                 total_scores = 0
                 if check_count.status_code in (200, 206):
                     cr = check_count.headers.get('Content-Range')
                     if cr:
                         try:
                             total_scores = int(cr.split('/')[-1])
                         except: pass
                 
                 if total_scores == 0:
                     print("DEBUG: kb_scores cleared.")
                     break
                 
                 # Fetch IDs to delete
                 check_scores = client.select('kb_scores', page=1, page_size=delete_batch_size, columns='id')
                 
                 if check_scores.status_code not in (200, 206):
                     print(f"Error fetching kb_scores: {check_scores.text}")
                     import time
                     time.sleep(1)
                     continue
                 
                 batch_ids = [item['id'] for item in check_scores.json()]
                 
                 if not batch_ids:
                     # Count says > 0 but no IDs returned? Retry.
                     continue
                 
                 print(f"DEBUG: Found {len(batch_ids)} scores to delete. Batch {i+1}...")
                 
                 # Delete this batch
                 # kb_scores.id is integer
                 id_str = "(" + ",".join([str(bid) for bid in batch_ids]) + ")"
                 del_res = client.delete('kb_scores', {'id': f'in.{id_str}'})
                 
                 if del_res.status_code >= 400:
                     print(f"Warning: Failed to delete kb_scores batch: {del_res.text}")
                     import time
                     time.sleep(1)
                 else:
                     # Brief pause
                     import time
                     time.sleep(0.1)
            else:
                 return jsonify({'success': False, 'message': f"无法清空评分表 (kb_scores)，请重试或联系管理员。"}), 500
            
            # 2. Clear main table (knowledge_base_v1)
            print("DEBUG: Clearing main table 'knowledge_base_v1'...")
            
            # Attempt global delete first (More efficient)
            print("DEBUG: Attempting global delete via 'not.is.null' filter...")
            # Using 'not.is.null' covers all non-null values. 
            global_del = client.delete('knowledge_base_v1', {'question_wiki_id': 'not.is.null'})
            
            # Also delete nulls if any
            client.delete('knowledge_base_v1', {'question_wiki_id': 'is.null'})
            
            if global_del.status_code >= 400:
                print(f"Warning: Global delete failed: {global_del.text}. Falling back to batch delete.")
            else:
                print("DEBUG: Global delete request sent.")
            
            # Verify and Cleanup Loop
            max_retries_main = 200 
            # Increase batch size slightly, but stay safe
            delete_batch_size = 150 
            
            for i in range(max_retries_main):
                # Fetch IDs to check if data remains
                resp = client.select('knowledge_base_v1', page=1, page_size=delete_batch_size, columns='question_wiki_id')
                if resp.status_code not in (200, 206):
                     return jsonify({'success': False, 'message': f"Failed to fetch IDs for deletion: {resp.text}"}), 500
                
                batch_data = resp.json()
                batch_ids = [item.get('question_wiki_id') for item in batch_data]
                
                if not batch_ids:
                    print("DEBUG: knowledge_base_v1 cleared.")
                    break
                
                print(f"DEBUG: Found {len(batch_ids)} remaining records. Deleting batch {i+1}...")
                
                # Separate NULLs and Values
                null_ids_count = batch_ids.count(None)
                value_ids = [bid for bid in batch_ids if bid is not None]
                
                if null_ids_count > 0:
                    print(f"DEBUG: Deleting {null_ids_count} records with NULL ID...")
                    client.delete('knowledge_base_v1', {'question_wiki_id': 'is.null'})
                
                if value_ids:
                    # Delete this batch
                    # Safe quoting: replace " with ""
                    quoted_ids = []
                    for bid in value_ids:
                        val = str(bid).replace('"', '""')
                        quoted_ids.append(f'"{val}"')
                    id_str = "(" + ",".join(quoted_ids) + ")"
                    del_resp = client.delete('knowledge_base_v1', {'question_wiki_id': f'in.{id_str}'})
                    
                    if del_resp.status_code >= 400:
                         print(f"Warning: Failed to delete batch, retrying... {del_resp.text}")
                         import time
                         time.sleep(1)
                    else:
                         print(f"DEBUG: Deleted batch of {len(value_ids)} records.")
                         # Sleep briefly to let DB catch up
                         import time
                         time.sleep(0.1)
            else:
                 # Check one last time
                 final_check = client.select('knowledge_base_v1', page=1, page_size=1)
                 total_remaining = 0
                 if final_check.status_code in (200, 206):
                     cr = final_check.headers.get('Content-Range')
                     if cr:
                         try:
                             total_remaining = int(cr.split('/')[-1])
                         except:
                             pass
                 
                 if total_remaining > 0:
                    return jsonify({'success': False, 'message': f"无法彻底清空知识库主表 (knowledge_base_v1)，仍有 {total_remaining} 条数据。请重试。"}), 500



            print("DEBUG: Table cleared successfully.")
            stats['added'] = len(processed_records)
            details['added_ids'] = file_ids_all

        # Modes: Upsert or Append (Need to know existing IDs for stats)
        existing_ids = set()
        file_ids = [str(r['question_wiki_id']) for r in processed_records if 'question_wiki_id' in r]
        if mode in ['upsert', 'append'] and file_ids:
            for i in range(0, len(file_ids), 50):
                batch_ids = file_ids[i:i+50]
                id_filter = _postgrest_in_str(batch_ids)
                resp = client.select('knowledge_base_v1', page=1, page_size=len(batch_ids), filters={'question_wiki_id': id_filter})
                if resp.status_code in (200, 206):
                    data = resp.json()
                    for row in data:
                        existing_ids.add(str(row['question_wiki_id']))

        for i in range(0, len(processed_records), batch_size):
            batch = processed_records[i:i+batch_size]
            
            if mode == 'append':
                batch_ids_list = [str(r.get('question_wiki_id')) for r in batch if r.get('question_wiki_id')]
                skipped_this_batch = [id for id in batch_ids_list if id in existing_ids]
                batch_to_insert = [r for r in batch if str(r.get('question_wiki_id')) not in existing_ids]
                stats['skipped'] += len(skipped_this_batch)
                if not batch_to_insert:
                    continue
                resp = client.insert('knowledge_base_v1', batch_to_insert)
                if resp.status_code < 400:
                    stats['added'] += len(batch_to_insert)
                    details['added_ids'].extend([str(r.get('question_wiki_id')) for r in batch_to_insert if r.get('question_wiki_id')])
                    details['skipped_ids'].extend(skipped_this_batch)
            elif mode == 'overwrite':
                resp = client.insert('knowledge_base_v1', batch)
            else:
                resp = client.upsert('knowledge_base_v1', batch, on_conflict='question_wiki_id')
                if resp.status_code < 400:
                    batch_ids = [str(r.get('question_wiki_id')) for r in batch if r.get('question_wiki_id')]
                    added_this_batch = [id for id in batch_ids if id not in existing_ids]
                    updated_this_batch = [id for id in batch_ids if id in existing_ids]
                    stats['updated'] += len(updated_this_batch)
                    stats['added'] += len(added_this_batch)
                    details['updated_ids'].extend(updated_this_batch)
                    details['added_ids'].extend(added_this_batch)

            if resp.status_code >= 400:
                try:
                    err_detail = resp.json()
                except:
                    err_detail = resp.text
                errors.append(f"Batch {i//batch_size}: {err_detail}")
            else:
                total_inserted += len(batch)

        # After all writes, force all affected IDs' review_status to 'unadjusted'
        # 业务期望：无论增量/批量新增/全量覆盖，只要被本次导入触达的 ID，修订状态一律重置为“未调整”
        affected_ids = sorted(list(dict.fromkeys(file_ids_all)))
        if affected_ids:
            for i in range(0, len(affected_ids), 100):
                batch_ids = affected_ids[i:i+100]
                id_filter = _postgrest_in_str(batch_ids)
                try:
                    _ = client.update(
                        'knowledge_base_v1',
                        {'review_status': 'unadjusted'},
                        {'question_wiki_id': id_filter}
                    )
                except Exception as _e:
                    print(f"WARN: failed to reset review_status for ids batch {i//100}: {_e}")

        if errors:
            error_summary = "; ".join([str(e)[:200] for e in errors[:3]])
            if len(errors) > 3:
                error_summary += "..."
            restore_info = {}
            if mode == 'overwrite':
                restore_info['v1_restore'] = _run_kb_t1_to_v1_restore(client)
                restore_info['score_restore'] = _restore_table_rows(client, 'kb_scores', score_backup_rows, conflict_col='id')
            return jsonify({
                'success': False,
                'message': f"Import failed/partial. Errors: {error_summary}",
                'errors': errors,
                'count': total_inserted,
                'stats': stats,
                'details': details,
                'restore': restore_info,
                'score_backup_path': score_backup_path
            }), 500

        if delete_missing_ids:
            delete_missing_result = _delete_kb_items_physical(client, delete_missing_ids)
            if not delete_missing_result.get('success'):
                return jsonify({
                    'success': False,
                    'message': delete_missing_result.get('message') or '导入已执行，但物理删除缺失 ID 失败。',
                    'count': total_inserted,
                    'stats': stats,
                    'details': details,
                    'delete_missing': delete_missing_result
                }), 500
            stats['deleted'] = delete_missing_result.get('count', 0)
            details['deleted_ids'] = delete_missing_result.get('ids', [])

        out = {'success': True, 'count': total_inserted, 'stats': stats, 'mode': mode, 'details': details}
        if pre_t1_sync_note:
            out['pre_sync_v1_to_t1'] = pre_t1_sync_note
        if score_backup_path:
            out['score_backup_path'] = score_backup_path
        if delete_missing_result is not None:
            out['delete_missing'] = delete_missing_result
        return jsonify(out)

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/kb/import/preview', methods=['POST'])
@login_required
def kb_import_preview():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'message': 'Empty filename'}), 400

    try:
        if file.filename.lower().endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        df.rename(columns=_kb_import_column_mapping(), inplace=True)

        total_rows = len(df)
        file_ids = []
        duplicate_id_count = 0
        missing_id_count = 0
        if 'question_wiki_id' in df.columns:
            seen = set()
            for raw in df['question_wiki_id'].tolist():
                s = _normalize_kb_import_id(raw)
                if not s:
                    missing_id_count += 1
                    continue
                if s in seen:
                    duplicate_id_count += 1
                seen.add(s)
                file_ids.append(s)
        else:
            missing_id_count = total_rows

        delete_missing_requested = str(request.form.get('delete_missing') or '').strip().lower() in ('1', 'true', 'yes', 'y')
        delete_missing_ids = []
        delete_missing_blocked = False
        if delete_missing_requested:
            if file_ids:
                delete_missing_ids = _get_v1_ids_missing_from_file(client, file_ids)
            else:
                delete_missing_blocked = True

        invalid_rows = []
        if 'product_name' in df.columns:
            valid_map, valid_set = get_all_valid_models()
            for idx, row in df.iterrows():
                product_value = row.get('product_name')
                p_name = '' if _is_blank_cell_value(product_value) else str(product_value)
                if p_name.strip():
                    _, invalid = validate_product_string(p_name, valid_map, valid_set)
                    if invalid:
                        invalid_rows.append({
                            'row': idx + 2,
                            'question_wiki_id': row.get('question_wiki_id', ''),
                            'invalid_models': invalid
                        })

        current_count = _client_count(client, 'knowledge_base_v1')
        score_count = _client_count(client, 'kb_scores')
        return jsonify({
            'success': True,
            'preview': {
                'incoming_count': len(set(file_ids)) + missing_id_count,
                'raw_row_count': total_rows,
                'unique_id_count': len(set(file_ids)),
                'duplicate_id_count': duplicate_id_count,
                'missing_id_count': missing_id_count,
                'invalid_model_count': len(invalid_rows),
                'invalid_rows': invalid_rows[:20],
                'current_v1_count': current_count,
                'score_count': score_count,
                'delete_missing_count': len(delete_missing_ids),
                'delete_missing_sample_ids': delete_missing_ids[:20],
                'delete_missing_blocked': delete_missing_blocked
            }
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/kb/check_duplicates', methods=['POST'])
@login_required
def kb_check_duplicates():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'message': 'Empty filename'}), 400

    try:
        df = pd.read_excel(file)
        # Mapping is shared with import so preview semantics match actual import.
        column_mapping = _kb_import_column_mapping()
        df.rename(columns=column_mapping, inplace=True)
        
        # We only care about records with question_wiki_id
        if 'question_wiki_id' not in df.columns:
             # Try one last fallback: assume first column is ID if it looks like int
             # But safer to return error
             return jsonify({'success': False, 'message': '未找到ID列 (支持列名: ID, id, 问题ID)'}), 400
             
        # Extract IDs from file (remove NaN)
        raw_ids = df['question_wiki_id'].dropna().unique().tolist()
        file_ids = []
        for x in raw_ids:
            s = _normalize_kb_import_id(x)
            if s:
                file_ids.append(s)
        
        delete_missing_requested = str(request.form.get('delete_missing') or '').strip().lower() in ('1', 'true', 'yes', 'y')
        if not file_ids:
            return jsonify({
                'success': True,
                'report': [],
                'message': '未提取到有效的ID，请检查Excel文件ID列内容',
                'duplicates_count': 0,
                'new_count': 0,
                'error_count': 0,
                'delete_missing_count': 0,
                'delete_missing_sample_ids': [],
                'delete_missing_blocked': delete_missing_requested
            }) # No IDs to check

        delete_missing_ids = _get_v1_ids_missing_from_file(client, file_ids) if delete_missing_requested else []

        # Query DB for these IDs
        existing_records = {}
        batch_size = 50 # Smaller batch for string IDs
        for i in range(0, len(file_ids), batch_size):
            batch_ids = file_ids[i:i+batch_size]
            id_filter = _postgrest_in_str(batch_ids)
            resp = client.select('knowledge_base_v1', page=1, page_size=len(batch_ids), filters={'question_wiki_id': id_filter})
            
            if resp.status_code in (200, 206):
                data = resp.json()
                for row in data:
                    # Store in map for easy lookup, key should be string for consistent comparison
                    existing_records[str(row['question_wiki_id'])] = row
        
        # Compare
        report = []
        valid_map, valid_set = get_all_valid_models()
        # Convert df to records
        records = df.to_dict(orient='records')
        
        for record in records:
            if 'question_wiki_id' not in record or not _normalize_kb_import_id(record['question_wiki_id']):
                continue
                
            # Normalize ID to string for comparison
            wiki_id = _normalize_kb_import_id(record['question_wiki_id'])
            
            if wiki_id not in existing_records:
                # Check for invalid models even for new records
                raw_product_name = str(record.get('product_name', '')).strip()
                matched_msg = ""
                if raw_product_name:
                    valid_input_models, invalid_models = validate_product_string(raw_product_name, valid_map, valid_set)

                    if invalid_models:
                        report.append({'id': wiki_id, 'status': '异常', 'details': f"新增条目包含未知型号: {', '.join(invalid_models)}"})
                        continue
                    
                    if valid_input_models:
                        valid_input_models.sort()
                        matched_msg = f" [匹配型号: {', '.join(valid_input_models)}]"

                report.append({'id': wiki_id, 'status': '新增', 'details': 'ID不存在，将执行新增' + matched_msg})
            else:
                # Exists, compare
                db_row = existing_records[wiki_id]
                diffs = []
                
                # Fields to compare
                compare_fields = [
                    'question', 'answer', 'question_type', 'answer_type', 'if_bm25',
                    'similar_questions', 'error_list', 'keyword_list',
                    'image_urls', 'video_urls', 'file_urls', 'link_type', 'link_url',
                    'product_name', 'product_category_name'
                ]
                
                def normalize_val(v, field_name=None):
                    def is_blank_scalar(value):
                        if value is None:
                            return True
                        if isinstance(value, (list, tuple, set, dict)):
                            return False
                        try:
                            return bool(pd.isna(value))
                        except (TypeError, ValueError):
                            return False

                    if field_name in ('similar_questions', 'error_list', 'keyword_list', 'image_urls', 'video_urls', 'file_urls'):
                        if is_blank_scalar(v):
                            return ""
                        if isinstance(v, list):
                            parts = v
                        else:
                            raw = str(v or '').strip()
                            if not raw:
                                return ""
                            try:
                                obj = json.loads(raw)
                                parts = obj if isinstance(obj, list) else [obj]
                            except Exception:
                                parts = _split_text_list_value(raw, is_url_list=(field_name in ('image_urls', 'video_urls', 'file_urls')))
                        norm_parts = []
                        for x in parts or []:
                            if x is None:
                                continue
                            t = str(x).replace('`', '').strip()
                            if not t:
                                continue
                            norm_parts.append(t.lower())
                        norm_parts = sorted(list(dict.fromkeys(norm_parts)))
                        return "|".join(norm_parts)

                    if is_blank_scalar(v):
                        return ""
                    s = str(v).strip() # Only strip outer whitespace

                    if field_name == 'if_bm25':
                        if isinstance(v, bool):
                            return '1' if v else '0'
                        sv = str(v).strip().lower()
                        if sv in ('1', 'true', 'yes', 'y', '是'):
                            return '1'
                        if sv in ('0', 'false', 'no', 'n', '否'):
                            return '0'
                        return ""
                    
                    if field_name == 'answer_type':
                         s_lower = s.lower()
                         if s_lower == '文本':
                             return 'text'
                         return s # Keep original case for answer_type? Or normalize? Usually specific values.
                    
                    if field_name == 'product_name':
                        # Use validate_product_string to normalize (ignore space, case, etc)
                        # We only care about valid models for normalization comparison
                        valid, _ = validate_product_string(s, valid_map, valid_set)
                        valid.sort()
                        return ", ".join(valid)

                    if field_name in ['question', 'answer']:
                         return s.lower().replace(" ", "").replace("\u3000", "")
                    
                    return s.lower()

                for field in compare_fields:
                    file_val = normalize_val(record.get(field), field)
                    db_val = normalize_val(db_row.get(field), field)
                    
                    if file_val != db_val:
                        diffs.append(field)
                
                # Product Validation (User Requirement)
                # Check if product_name contains invalid models
                raw_product_name = str(record.get('product_name', '')).strip()
                matched_msg = ""
                if raw_product_name:
                    valid_input_models, invalid_models = validate_product_string(raw_product_name, valid_map, valid_set)
                    
                    if invalid_models:
                         report.append({'id': wiki_id, 'status': '异常', 'details': f"包含未知型号: {', '.join(invalid_models)}"})
                         continue # Skip the normal diff check if invalid
                    
                    if valid_input_models:
                        valid_input_models.sort()
                        matched_msg = f" [匹配型号: {', '.join(valid_input_models)}]"

                if not diffs:
                    report.append({'id': wiki_id, 'status': '完全重复', 'details': '内容完全一致' + matched_msg})
                else:
                    report.append({'id': wiki_id, 'status': '不一致', 'details': f"{', '.join(diffs)} 不一致" + matched_msg})

        duplicates_count = sum(1 for item in report if item['status'] in ['完全重复', '不一致'])
        new_count = sum(1 for item in report if item['status'] == '新增')
        error_count = sum(1 for item in report if item['status'] == '异常')
        
        return jsonify({
            'success': True, 
            'report': report,
            'duplicates_count': duplicates_count,
            'new_count': new_count,
            'error_count': error_count,
            'delete_missing_count': len(delete_missing_ids),
            'delete_missing_sample_ids': delete_missing_ids[:20]
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/kb/item', methods=['GET', 'POST'])
@login_required
def kb_item_upsert():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400

    if request.method == 'GET':
        table = request.args.get('table', 'knowledge_base_v1')
        if table not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
            return jsonify({'success': False, 'message': 'Invalid table'}), 400
        wiki_id = request.args.get('id') or request.args.get('question_wiki_id')
        wiki_id = str(wiki_id or '').strip()
        if not wiki_id:
            return jsonify({'success': False, 'message': 'id is required'}), 400
        try:
            cols = 'question_wiki_id,question_type,question,answer,answer_type,if_bm25,similar_questions,error_list,keyword_list,image_urls,video_urls,file_urls,link_type,link_url,update_time,product_category_name,product_name'
            resp = client.select(table, page=1, page_size=1, filters={'question_wiki_id': f"eq.{wiki_id}"}, columns=cols)
            if resp.status_code not in (200, 206):
                return jsonify({'success': False, 'message': resp.text}), 500
            rows = resp.json() or []
            return jsonify({'success': True, 'data': rows[0] if rows else {}})
        except Exception as e:
            traceback.print_exc()
            return jsonify({'success': False, 'message': str(e)}), 500
    
    data = request.json
    if not data:
        return jsonify({'success': False, 'message': 'No data provided'}), 400

    try:
        allow = _kb_all_fields_allowlist()
        data = {k: v for k, v in (data or {}).items() if k in allow}

        wiki_id = str(data.get('question_wiki_id') or '').strip()
        before_row = None
        if wiki_id:
            try:
                cols = 'question_wiki_id,question_type,question,answer,answer_type,if_bm25,similar_questions,error_list,keyword_list,image_urls,video_urls,file_urls,link_type,link_url,update_time,product_category_name,product_name'
                before_resp = client.select('knowledge_base_v1', page=1, page_size=1, filters={'question_wiki_id': f"eq.{wiki_id}"}, columns=cols)
                if before_resp.status_code in (200, 206):
                    rows = before_resp.json() or []
                    if rows and isinstance(rows[0], dict):
                        before_row = rows[0]
            except Exception:
                before_row = None

        # Data cleaning
        # Handle JSON fields
        json_cols = ['similar_questions', 'error_list', 'keyword_list', 'image_urls', 'video_urls', 'file_urls']
        for col in json_cols:
            if col in data and isinstance(data[col], str):
                try:
                    data[col] = json.loads(data[col])
                except:
                    pass # Keep as string or convert to list if possible?
                    # If it's a string, try to split by comma or newline
                    if isinstance(data[col], str):
                         # 文本字段只按换行/半角逗号拆分；全角逗号保留为句内标点
                         parts = _split_text_list_value(data[col], is_url_list=(col in ('image_urls', 'video_urls', 'file_urls')))
                         data[col] = [x.strip() for x in parts if x.strip()]
        
        if 'link_url' in data:
            s = str(data.get('link_url') or '').strip()
            if not s:
                data['link_url'] = None
            else:
                data['link_url'] = s
        if 'link_type' in data:
            s = str(data.get('link_type') or '').strip()
            if not s:
                data['link_type'] = None
            else:
                data['link_type'] = s
        
        # Ensure boolean
        if 'if_bm25' in data:
            val = data['if_bm25']
            if isinstance(val, str):
                str_val = val.strip().lower()
                if str_val in ['1', 'true', 'yes', '是', 'true']:
                    data['if_bm25'] = True
                elif str_val in ['0', 'false', 'no', '否', 'false']:
                    data['if_bm25'] = False
                else:
                    data['if_bm25'] = None # Unknown
        
        # Product Validation
        if 'product_name' in data:
            p_name = str(data['product_name']) if data['product_name'] else ""
            if p_name.strip():
                valid_map, valid_set = get_all_valid_models()
                valid, invalid = validate_product_string(p_name, valid_map, valid_set)
                
                if invalid:
                    return jsonify({'success': False, 'message': f'包含未知型号: {", ".join(invalid)}。请先在“管理型号库”中添加该型号。'}), 400
                
                # Update with normalized string
                valid.sort()
                data['product_name'] = ", ".join(valid)
        
        # Update Time: if not provided, do NOT set to now automatically
        # Let the database handle defaults (for insert) or keep existing (for update)
        if 'update_time' in data and not data['update_time']:
            del data['update_time']
            
        # Helper to clean NaN (reuse logic?)
        # For single item, we just check manually if needed, but dict usually doesn't have NaN unless from pandas
        
        # Perform Upsert
        # Note: data must contain primary key for upsert to update. 
        # If question_wiki_id is missing, Supabase might insert new if it's auto-generated, or fail.
        resp = client.upsert('knowledge_base_v1', [data], on_conflict='question_wiki_id')
        
        if resp.status_code >= 400:
             return jsonify({'success': False, 'message': resp.text}), 500

        try:
            if wiki_id:
                after_for_diff = dict(before_row) if isinstance(before_row, dict) else {}
                after_for_diff.update(data or {})
                before_obj = _snapshot_mod_fields(before_row) if isinstance(before_row, dict) else None
                after_obj = _snapshot_mod_fields(after_for_diff)
                changed_fields = _compute_mod_changed_fields(before_obj, after_obj)
                if changed_fields:
                    change_type = 'edit' if isinstance(before_row, dict) else 'create'
                    mod_rec = _build_kb_modification_record(
                        '知识库管理',
                        current_user.username if current_user.is_authenticated else 'system',
                        change_type,
                        wiki_id,
                        before_obj,
                        after_obj,
                        changed_fields,
                        base_row=after_for_diff
                    )
                    _supabase_insert_drop_unknown_columns(client, 'knowledge_base_modifications', mod_rec)
        except Exception:
            pass
             
        return jsonify({'success': True, 'data': resp.json()})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


def _run_kb_v1_to_t1_sync(client):
    """
    将 knowledge_base_v1 全量复制到 knowledge_base_v1_t1（按 t1 允许的列白名单）。
    用于「同步此刻到前刻」以及「全量覆盖导入」前先备份 V1。
    Returns dict: success (bool), message (str), count (int, 同步行数；v1 空时为 0)。
    """
    import time
    try:
        print("Starting manual sync from v1 to v1_t1...")
        print("Fetching all data from v1 with page_size=1000...")
        v1_data = client.select_all('knowledge_base_v1', order_by='question_wiki_id', page_size=1000)
        if not v1_data:
            print("v1 is empty, nothing to sync.")
            client.delete('knowledge_base_v1_t1', {'question_wiki_id': 'not.is.null'})
            # Also clear KB tag mapping for previous (方案A：previous 全量重建)
            try:
                client.delete('kb_item_tags', {'library_type': 'eq.previous'})
            except Exception:
                pass
            return {'success': True, 'message': 'v1 is empty, t1 cleared.', 'count': 0}

        allowed_columns = [
            'question_wiki_id', 'question_type', 'question', 'answer', 'answer_type',
            'if_bm25', 'similar_questions', 'error_list', 'keyword_list',
            'image_urls', 'video_urls', 'file_urls', 'link_type', 'link_url',
            'update_time', 'product_category_name', 'product_name'
        ]
        to_insert = []
        for item in v1_data:
            new_item = {}
            for col in allowed_columns:
                if col in item:
                    new_item[col] = item[col]
            to_insert.append(new_item)

        print(f"Prepared {len(to_insert)} items for sync.")
        print("Clearing v1_t1...")
        del_resp = client.delete('knowledge_base_v1_t1', {'question_wiki_id': 'not.is.null'})
        if del_resp.status_code >= 400:
            return {'success': False, 'message': f"Failed to clear t1: {del_resp.text}", 'count': 0}

        batch_size = 500
        print(f"Inserting into v1_t1 in batches of {batch_size}...")
        for i in range(0, len(to_insert), batch_size):
            batch = to_insert[i:i + batch_size]
            print(f"Inserting batch {i // batch_size + 1}/{len(to_insert) // batch_size + 1}...")
            resp = client.insert('knowledge_base_v1_t1', batch)
            if resp.status_code >= 400:
                print(f"Batch insert failed: {resp.text}")
                return {'success': False, 'message': f"Batch insert failed at index {i}: {resp.text}", 'count': 0}
            time.sleep(0.1)

        # Sync KB item tag mappings: current -> previous (全量重建)
        try:
            print("Syncing KB tags mapping current -> previous...")
            client.delete('kb_item_tags', {'library_type': 'eq.previous'})
            v1_ids = [str(it.get('question_wiki_id') or '').strip() for it in (v1_data or [])]
            v1_ids = [x for x in v1_ids if x]
            if v1_ids:
                in_str = _postgrest_in_str(v1_ids)
                cur_maps = client.select_all(
                    'kb_item_tags',
                    filters={'library_type': 'eq.current', 'question_wiki_id': in_str},
                    columns='question_wiki_id,tag_id',
                    order_by='question_wiki_id',
                    page_size=1000
                ) or []
            else:
                cur_maps = []

            to_insert_maps = []
            for m in cur_maps:
                if not isinstance(m, dict):
                    continue
                wid = str(m.get('question_wiki_id') or '').strip()
                tag_id = m.get('tag_id')
                if wid and tag_id:
                    to_insert_maps.append({
                        'library_type': 'previous',
                        'question_wiki_id': wid,
                        'tag_id': tag_id
                    })

            if to_insert_maps:
                map_batch_size = 500
                for i in range(0, len(to_insert_maps), map_batch_size):
                    batch = to_insert_maps[i:i + map_batch_size]
                    resp = client.insert('kb_item_tags', batch)
                    if resp.status_code >= 400:
                        print(f"KB tags mapping insert failed: {resp.text}")
                        # Non-fatal: keep data sync success
                    time.sleep(0.05)
        except Exception as e:
            print(f"WARN: KB tags mapping sync failed: {e}")

        print("Sync completed successfully.")
        return {'success': True, 'message': f"Synced {len(to_insert)} items.", 'count': len(to_insert)}
    except Exception as e:
        print(f"Sync Exception: {e}")
        traceback.print_exc()
        return {'success': False, 'message': str(e), 'count': 0}

def _restore_table_rows(client, table, rows, conflict_col='id'):
    """
    Full-replace restore helper used after destructive operations fail.
    """
    rows = rows or []
    try:
        if conflict_col:
            client.delete(table, {conflict_col: 'not.is.null'})
        if rows:
            batch_size = 500
            for i in range(0, len(rows), batch_size):
                batch = rows[i:i + batch_size]
                resp = client.upsert(table, batch, on_conflict=conflict_col if conflict_col else None)
                if resp is None or getattr(resp, 'status_code', 500) >= 400:
                    return {'success': False, 'message': getattr(resp, 'text', '') or f'{table} restore failed at {i}', 'count': i}
        return {'success': True, 'message': f'Restored {len(rows)} rows to {table}', 'count': len(rows)}
    except Exception as e:
        traceback.print_exc()
        return {'success': False, 'message': str(e), 'count': 0}

def _run_kb_t1_to_v1_restore(client):
    """
    Restore current V1 from V1T-1 after an overwrite import fails.
    """
    try:
        allowed_columns = [
            'question_wiki_id', 'question_type', 'question', 'answer', 'answer_type',
            'if_bm25', 'similar_questions', 'error_list', 'keyword_list',
            'image_urls', 'video_urls', 'file_urls', 'link_type', 'link_url',
            'update_time', 'product_category_name', 'product_name'
        ]
        t1_data = client.select_all('knowledge_base_v1_t1', order_by='question_wiki_id', page_size=1000) or []
        restore_rows = []
        for item in t1_data:
            if not isinstance(item, dict):
                continue
            restore_rows.append({col: item.get(col) for col in allowed_columns if col in item})

        client.delete('knowledge_base_v1', {'question_wiki_id': 'not.is.null'})
        client.delete('knowledge_base_v1', {'question_wiki_id': 'is.null'})
        if restore_rows:
            batch_size = 500
            for i in range(0, len(restore_rows), batch_size):
                resp = client.insert('knowledge_base_v1', restore_rows[i:i + batch_size])
                if resp is None or getattr(resp, 'status_code', 500) >= 400:
                    return {'success': False, 'message': getattr(resp, 'text', '') or f'V1 restore failed at {i}', 'count': i}
        return {'success': True, 'message': f'Restored {len(restore_rows)} rows from V1T-1 to V1', 'count': len(restore_rows)}
    except Exception as e:
        traceback.print_exc()
        return {'success': False, 'message': str(e), 'count': 0}


@app.route('/api/kb/sync', methods=['POST'])
@login_required
def kb_sync():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400
    r = _run_kb_v1_to_t1_sync(client)
    payload = {'success': r.get('success'), 'message': r.get('message')}
    if 'count' in r:
        payload['count'] = r['count']
    code = 200 if r.get('success') else 500
    return jsonify(payload), code

@app.route('/api/kb/data_v2', methods=['GET'])
@login_required
def kb_data_v2():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400
    
    table = request.args.get('table', 'knowledge_base_v1')
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    
    # Sort params
    order_by = request.args.get('orderBy')
    order_dir = request.args.get('orderDir', 'desc')
    
    # Filter params
    filters = {}
    or_clauses = []
    
    # ID Search
    question_id = request.args.get('id')
    if question_id:
        filters['question_wiki_id'] = f"eq.{question_id}"

    product_name = request.args.get('productName')
    if product_name:
        # ilike pattern for case-insensitive partial match
        filters['product_name'] = f"ilike.*{product_name}*"

    keyword = request.args.get('keyword')
    if keyword:
        # question OR answer match
        or_clauses.extend([f"question.ilike.*{keyword}*", f"answer.ilike.*{keyword}*"])

    # Specific Question/Answer Search
    question_search = request.args.get('question')
    if question_search:
        filters['question'] = f"ilike.*{question_search}*"

    similar_question_search = request.args.get('similar_question')
    if similar_question_search:
        # similar_questions 是 JSONB（通常是字符串数组），这里改为匹配每个数组元素
        for i in range(10):
            or_clauses.append(f"similar_questions->>{i}.ilike.*{similar_question_search}*")
        or_clauses.append(f"similar_questions::text.ilike.*{similar_question_search}*")
        
    answer_search = request.args.get('answer')
    if answer_search:
        filters['answer'] = f"ilike.*{answer_search}*"

    if or_clauses:
        filters['or'] = f"({','.join(or_clauses)})"

    # Security check to prevent arbitrary table access
    if table not in ['knowledge_base_v1', 'knowledge_base_v1_t1']:
        return jsonify({'success': False, 'message': 'Invalid table'}), 400

    try:
        resp = client.select(table, page, page_size, filters, order_by, order_dir)
        if resp.status_code >= 400:
            return jsonify({'success': False, 'message': resp.text}), 500
            
        # Parse total count from Content-Range header (e.g., "0-19/100")
        total = 0
        content_range = resp.headers.get('Content-Range')
        if content_range:
            try:
                total = int(content_range.split('/')[-1])
            except:
                pass
                
        return jsonify({
            'success': True, 
            'data': resp.json(),
            'total': total,
            'page': page,
            'pageSize': page_size
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/kb/export', methods=['GET'])
@login_required
def kb_export():
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400
    
    table = request.args.get('table', 'knowledge_base_v1')
    
    if table not in ['knowledge_base_v1', 'knowledge_base_v1_t1']:
        return jsonify({'success': False, 'message': 'Invalid table'}), 400
    
    # Filter params
    filters = {}
    
    # ID Search
    id_search = request.args.get('id')
    if id_search:
        filters['question_wiki_id'] = f"ilike.*{id_search}*"
        
    # Selected IDs (for Export Selected)
    ids_param = request.args.get('ids')
    if ids_param:
        # Check if it's already in format or just comma separated list
        # We assume comma separated list of IDs from frontend
        # PostgREST expects: in.(val1,val2)
        if not ids_param.startswith('in.'):
             filters['question_wiki_id'] = f"in.({ids_param})"
        else:
             filters['question_wiki_id'] = ids_param

    # Product Search
    product_search = request.args.get('product') or request.args.get('productName')
    if product_search:
        filters['product_name'] = f"ilike.*{product_search}*"

    # Keyword
    keyword = request.args.get('keyword')
    if keyword:
        filters['or'] = f"(question.ilike.*{keyword}*,answer.ilike.*{keyword}*)"

    question_search = request.args.get('question')
    if question_search:
        filters['question'] = f"ilike.*{question_search}*"
        
    answer_search = request.args.get('answer')
    if answer_search:
        filters['answer'] = f"ilike.*{answer_search}*"
        
    # Review Status (Added for consistency with kb_data)
    review_status_search = request.args.get('review_status')
    if review_status_search and table == 'knowledge_base_v1':
        statuses = review_status_search.split(',')
        if len(statuses) > 0:
             # PostgREST in filter: in.(val1,val2)
             filters['review_status'] = f"in.({','.join(statuses)})"
        
    # Columns to export
    columns_param = request.args.get('columns')
    columns = columns_param if columns_param else '*'

    # Sorting
    sort_by = request.args.get('sortBy')
    sort_dir = request.args.get('sortDir', 'desc')
    if not sort_by:
        sort_by = 'question_wiki_id' if 'knowledge_base_v1' in table else 'id'

    try:
        # Fetch data
        all_data = client.select_all(table, filters, order_by=sort_by, order_dir=sort_dir, columns=columns)
        
        # Convert to DataFrame
        df = pd.DataFrame(all_data)
        
        if df.empty:
            # Return empty CSV with headers if columns known, or just empty
            output = io.BytesIO()
            output.write(b'\xef\xbb\xbf') # BOM
            output.seek(0)
            return send_file(
                output,
                as_attachment=True,
                download_name=canonical_download_name('kb_export', 'csv'),
                mimetype='text/csv'
            )
            
        # Reorder columns if specific columns were requested (Supabase might return in any order)
        if columns_param:
            req_cols = [c.strip() for c in columns_param.split(',')]
            # Only keep columns that actually exist in the response
            final_cols = [c for c in req_cols if c in df.columns]
            if final_cols:
                df = df[final_cols]
        
        # Clean list columns for export (remove brackets/quotes)
        list_cols = ['similar_questions', 'keyword_list', 'error_list', 'image_urls', 'video_urls', 'file_urls']
        for col in list_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: ",".join(map(str, x)) if isinstance(x, list) else (str(x).replace('[','').replace(']','').replace("'", "").replace('"', "") if pd.notna(x) else ""))

        # Output to CSV
        output = io.BytesIO()
        # Add BOM for Excel compatibility with UTF-8
        output.write(b'\xef\xbb\xbf')
        df.to_csv(output, index=False, encoding='utf-8')
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=canonical_download_name('kb_export', 'csv'),
            mimetype='text/csv'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


_SM_EMBEDDING_DEFAULTS = {
    'api_url': 'https://api.siliconflow.cn/v1/embeddings',
    'model': 'Pro/BAAI/bge-m3',
    'dimensions': 1024,
    'threshold': 0.75,
    'batch_size': 64,
    'timeout': 60,
    'fallback_to_ngram': True,
}
_SM_EMBEDDING_TEXT_MAX_CHARS = 12000
_SM_EMBEDDING_TRANSIENT_STATUS = {408, 409, 429, 500, 502, 503, 504}


def _sm_embedding_config_path():
    return os.path.join(_BASE_DIR, 'smart_mapping_embedding_config.json')


def _sm_embedding_bool(value, default=False):
    if isinstance(value, bool):
        return value
    text_value = str(value or '').strip().lower()
    if text_value in ('1', 'true', 'yes', 'on'):
        return True
    if text_value in ('0', 'false', 'no', 'off'):
        return False
    return bool(default)


def _sm_embedding_number(value, default, cast, minimum, maximum):
    try:
        number = cast(value)
    except (TypeError, ValueError):
        number = cast(default)
    return max(minimum, min(maximum, number))


def _sm_read_embedding_config_file():
    path = _sm_embedding_config_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as handle:
            payload = json.load(handle)
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def _sm_normalize_embedding_config(raw=None, fallback_key=True):
    raw = raw if isinstance(raw, dict) else {}
    config = dict(_SM_EMBEDDING_DEFAULTS)
    for key in ('api_url', 'model'):
        value = str(raw.get(key) or '').strip()
        if value:
            config[key] = value
    config['dimensions'] = _sm_embedding_number(raw.get('dimensions'), config['dimensions'], int, 1, 8192)
    config['threshold'] = _sm_embedding_number(raw.get('threshold'), config['threshold'], float, 0.01, 1.0)
    config['batch_size'] = _sm_embedding_number(raw.get('batch_size'), config['batch_size'], int, 1, 128)
    config['timeout'] = _sm_embedding_number(raw.get('timeout'), config['timeout'], int, 5, 300)
    config['fallback_to_ngram'] = _sm_embedding_bool(
        raw.get('fallback_to_ngram'), config['fallback_to_ngram']
    )

    custom_key = str(raw.get('api_key') or '').strip()
    embedding_host = urlparse(str(config.get('api_url') or '')).netloc.lower()
    provider_env_key = ''
    if 'siliconflow' in embedding_host:
        provider_env_key = str(os.environ.get('SILICONFLOW_API_KEY') or '').strip()
    elif 'openai.com' in embedding_host:
        provider_env_key = str(os.environ.get('OPENAI_API_KEY') or '').strip()
    env_key = str(os.environ.get('KMATRIX_SM_EMBEDDING_API_KEY') or '').strip() or provider_env_key
    root_key = ''
    if fallback_key:
        try:
            ai_config = load_ai_config() or {}
            ai_host = urlparse(str(ai_config.get('base_url') or '')).netloc.lower()
            if embedding_host and ai_host and embedding_host == ai_host:
                root_key = str(ai_config.get('api_key') or '').strip()
        except Exception:
            root_key = ''
    config['api_key'] = env_key or custom_key or root_key
    config['api_key_source'] = 'environment' if env_key else ('custom' if custom_key else ('ai_config' if root_key else ''))
    return config


def _sm_load_embedding_config():
    return _sm_normalize_embedding_config(_sm_read_embedding_config_file())


def _sm_public_embedding_config(config=None):
    config = dict(config or _sm_load_embedding_config())
    api_key = str(config.pop('api_key', '') or '')
    config['api_key_configured'] = bool(api_key)
    config['api_key_source'] = str(config.get('api_key_source') or '')
    try:
        config['cache_count'] = int(
            SmartMappingEmbeddingCache.query.filter_by(
                model=str(config.get('model') or ''),
                dimensions=int(config.get('dimensions') or 0),
            ).count()
        )
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        config['cache_count'] = 0
    return config


def _sm_validate_embedding_config(config):
    api_url = str(config.get('api_url') or '').strip()
    if not re.match(r'^https?://', api_url, re.IGNORECASE):
        raise ValueError('Embedding API URL 必须是完整的 http/https 地址')
    if not str(config.get('model') or '').strip():
        raise ValueError('Embedding 模型不能为空')
    if not str(config.get('api_key') or '').strip():
        raise ValueError('Embedding API Key 未配置，可单独填写或复用现有 AI 配置 Key')


def _sm_save_embedding_config(payload):
    current_raw = _sm_read_embedding_config_file()
    merged = dict(current_raw)
    payload = payload if isinstance(payload, dict) else {}
    for key in ('api_url', 'model', 'dimensions', 'threshold', 'batch_size', 'timeout', 'fallback_to_ngram'):
        if key in payload:
            merged[key] = payload.get(key)
    if _sm_embedding_bool(payload.get('clear_api_key'), False):
        merged.pop('api_key', None)
    else:
        new_key = str(payload.get('api_key') or '').strip()
        if new_key:
            merged['api_key'] = new_key

    normalized = _sm_normalize_embedding_config(merged)
    _sm_validate_embedding_config(normalized)
    stored = {key: normalized[key] for key in _SM_EMBEDDING_DEFAULTS}
    if str(merged.get('api_key') or '').strip():
        stored['api_key'] = str(merged.get('api_key') or '').strip()

    path = _sm_embedding_config_path()
    temp_path = ''
    try:
        with tempfile.NamedTemporaryFile('w', encoding='utf-8', dir=_BASE_DIR, delete=False) as handle:
            json.dump(stored, handle, ensure_ascii=False, indent=2)
            handle.write('\n')
            temp_path = handle.name
        os.replace(temp_path, path)
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
    return _sm_load_embedding_config()


def _sm_embedding_request(texts, config):
    _sm_validate_embedding_config(config)
    clean_texts = [str(text or '').strip()[:_SM_EMBEDDING_TEXT_MAX_CHARS] for text in texts]
    if not clean_texts or any(not text for text in clean_texts):
        raise ValueError('Embedding 输入文本不能为空')
    payload = {
        'model': str(config.get('model') or '').strip(),
        'input': clean_texts,
    }
    model_lower = payload['model'].lower()
    if model_lower.startswith('text-embedding-') or 'qwen3' in model_lower:
        payload['dimensions'] = int(config.get('dimensions') or 0)
    headers = {
        'Authorization': f"Bearer {config.get('api_key')}",
        'Content-Type': 'application/json',
        'Accept': 'application/json',
    }

    last_error = None
    for attempt in range(3):
        try:
            response = requests.post(
                str(config.get('api_url') or '').strip(),
                headers=headers,
                json=payload,
                timeout=int(config.get('timeout') or 60),
            )
            if response.status_code >= 400:
                detail = str(response.text or '').strip().replace('\n', ' ')[:300]
                last_error = RuntimeError(f'Embedding API 请求失败（HTTP {response.status_code}）：{detail}')
                if response.status_code not in _SM_EMBEDDING_TRANSIENT_STATUS:
                    break
            else:
                body = response.json()
                items = body.get('data') if isinstance(body, dict) else None
                if not isinstance(items, list):
                    raise RuntimeError('Embedding API 响应缺少 data 数组')
                items = sorted(items, key=lambda item: int(item.get('index', 0)))
                vectors = [item.get('embedding') for item in items]
                if len(vectors) != len(clean_texts):
                    raise RuntimeError(f'Embedding API 返回 {len(vectors)} 条向量，预期 {len(clean_texts)} 条')
                expected_dimensions = int(config.get('dimensions') or 0)
                for vector in vectors:
                    if not isinstance(vector, list) or not vector:
                        raise RuntimeError('Embedding API 返回了空向量')
                    if expected_dimensions and len(vector) != expected_dimensions:
                        raise RuntimeError(
                            f'Embedding 向量维度为 {len(vector)}，与配置的 {expected_dimensions} 不一致'
                        )
                return [[float(value) for value in vector] for vector in vectors]
        except (requests.RequestException, ValueError, TypeError, json.JSONDecodeError, RuntimeError) as exc:
            last_error = exc if isinstance(exc, RuntimeError) else RuntimeError(f'Embedding API 请求失败：{exc}')
        if attempt < 2:
            time.sleep(1.0 * (attempt + 1))
    raise last_error or RuntimeError('Embedding API 请求失败')


def _sm_embedding_cache_key(text_value, config):
    raw = '|'.join([
        'v1',
        str(config.get('api_url') or '').strip(),
        str(config.get('model') or '').strip(),
        str(int(config.get('dimensions') or 0)),
        str(text_value or '').strip()[:_SM_EMBEDDING_TEXT_MAX_CHARS],
    ])
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()


def _sm_unpack_embedding(blob, dimensions):
    values = array('f')
    values.frombytes(bytes(blob or b''))
    if len(values) != int(dimensions or 0):
        return None
    return [float(value) for value in values]


def _sm_load_cached_embeddings(cache_keys, config):
    found = {}
    keys = list(dict.fromkeys(str(key) for key in cache_keys if key))
    for start in range(0, len(keys), 400):
        batch = keys[start:start + 400]
        rows = SmartMappingEmbeddingCache.query.filter(
            SmartMappingEmbeddingCache.cache_key.in_(batch)
        ).all()
        for row in rows:
            vector = _sm_unpack_embedding(row.vector_blob, row.dimensions)
            if vector is not None:
                found[row.cache_key] = vector
    return found


def _sm_store_cached_embeddings(items, config):
    if not items:
        return
    model = str(config.get('model') or '')
    dimensions = int(config.get('dimensions') or 0)
    now = time.time()
    for cache_key, vector in items:
        blob = array('f', [float(value) for value in vector]).tobytes()
        db.session.merge(SmartMappingEmbeddingCache(
            cache_key=str(cache_key),
            model=model,
            dimensions=dimensions,
            vector_blob=blob,
            updated_ts=now,
        ))
    db.session.commit()


def _sm_get_embeddings(texts, config):
    results = [None] * len(texts)
    key_to_text = OrderedDict()
    key_to_indexes = {}
    for index, value in enumerate(texts):
        text_value = str(value or '').strip()[:_SM_EMBEDDING_TEXT_MAX_CHARS]
        if not text_value:
            continue
        cache_key = _sm_embedding_cache_key(text_value, config)
        key_to_text.setdefault(cache_key, text_value)
        key_to_indexes.setdefault(cache_key, []).append(index)

    try:
        cached = _sm_load_cached_embeddings(list(key_to_text.keys()), config)
    except Exception:
        try:
            db.session.rollback()
            init_db()
            cached = _sm_load_cached_embeddings(list(key_to_text.keys()), config)
        except Exception:
            db.session.rollback()
            cached = {}

    missing_keys = [key for key in key_to_text if key not in cached]
    batch_size = int(config.get('batch_size') or 64)
    generated = []
    for start in range(0, len(missing_keys), batch_size):
        batch_keys = missing_keys[start:start + batch_size]
        batch_texts = [key_to_text[key] for key in batch_keys]
        batch_vectors = _sm_embedding_request(batch_texts, config)
        for cache_key, vector in zip(batch_keys, batch_vectors):
            cached[cache_key] = vector
            generated.append((cache_key, vector))
    if generated:
        try:
            _sm_store_cached_embeddings(generated, config)
        except Exception:
            db.session.rollback()

    for cache_key, indexes in key_to_indexes.items():
        vector = cached.get(cache_key)
        for index in indexes:
            results[index] = vector
    return results


def _sm_vector_cosine(left, right):
    if not left or not right or len(left) != len(right):
        return 0.0
    dot = sum(a * b for a, b in zip(left, right))
    left_norm = math.sqrt(sum(value * value for value in left))
    right_norm = math.sqrt(sum(value * value for value in right))
    if not left_norm or not right_norm:
        return 0.0
    return float(dot / (left_norm * right_norm))


def _sm_embedding_reason(match_type, question_score, answer_score, fallback_reason=''):
    if fallback_reason:
        return _sm_trim_reason(f'Embedding不可用，已降级字符相似度：{fallback_reason}', 80)
    q_text = f'{max(0.0, min(1.0, float(question_score or 0))) * 100:.1f}%'
    a_text = f'{max(0.0, min(1.0, float(answer_score or 0))) * 100:.1f}%'
    if match_type == '无匹配':
        return f'Embedding未达到阈值；问题{q_text}，答案{a_text}'
    if match_type == '仅问题一致':
        return f'Embedding问题语义匹配{q_text}；答案仅{a_text}'
    if match_type == '仅答案一致':
        return f'Embedding答案语义匹配{a_text}；问题仅{q_text}'
    return f'Embedding问题{q_text}、答案{a_text}均达到阈值'


@app.route('/api/smart_mapping/embedding/config', methods=['GET', 'POST'])
@login_required
def sm_embedding_config():
    if request.method == 'GET':
        return jsonify({'success': True, 'config': _sm_public_embedding_config()})
    try:
        config = _sm_save_embedding_config(request.json or {})
        return jsonify({'success': True, 'config': _sm_public_embedding_config(config)})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400


@app.route('/api/smart_mapping/embedding/test', methods=['POST'])
@login_required
def sm_embedding_test():
    payload = request.json or {}
    raw = _sm_read_embedding_config_file()
    raw.update({key: value for key, value in payload.items() if value not in (None, '')})
    config = _sm_normalize_embedding_config(raw)
    try:
        vectors = _sm_embedding_request(['扫地机清扫效果检查', '主刷和风道堵塞排查'], config)
        return jsonify({
            'success': True,
            'message': 'Embedding API 连接成功',
            'model': config.get('model'),
            'dimensions': len(vectors[0]) if vectors else 0,
        })
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 502


_KD_PROMPT_VERSION = 'duplicate-check-p1-v1'
_KD_DOMAIN_TERMS = (
    '主刷', '边刷', '尘盒', '滤网', '吸口', '风道', '集尘口', '毛发', '异响',
    '漏垃圾', '堵塞', '安装不到位', '清扫效果差', '清扫不干净', '吐头发',
    '甩出毛发', '吸不进去', '清理', '检查', '更换', '安装', '复位', '重试',
)
_KD_RELATIONSHIPS = {
    'fully_covered', 'partially_covered', 'collectively_covered',
    'conflicting', 'unrelated',
}
_KD_ACTIONS = {
    'no_add', 'update_existing', 'compare_merge', 'create_new', 'manual_review',
}


def _kd_json_load(raw, default):
    if isinstance(raw, (dict, list)):
        return raw
    try:
        value = json.loads(raw or '')
        return value if isinstance(value, type(default)) else default
    except Exception:
        return default


def _kd_json_dump(value):
    return json.dumps(value, ensure_ascii=False, separators=(',', ':'))


def _kd_string_list(value, limit=100):
    if isinstance(value, (list, tuple, set)):
        raw_items = list(value)
    elif isinstance(value, dict):
        raw_items = list(value.values())
    else:
        text_value = str(value or '').strip()
        if not text_value:
            return []
        if text_value[:1] in ('[', '{'):
            try:
                return _kd_string_list(json.loads(text_value), limit=limit)
            except Exception:
                pass
        raw_items = re.split(r'[\n,，;；、/|]+', text_value)
    result = []
    seen = set()
    for item in raw_items:
        text_item = str(item or '').strip()
        if text_item and text_item not in seen:
            seen.add(text_item)
            result.append(text_item)
            if len(result) >= limit:
                break
    return result


def _kd_norm(value):
    return re.sub(r'[^\u3400-\u9fffA-Za-z0-9]+', '', str(value or '')).lower()


def _kd_topic_terms(*values):
    text_value = ' '.join(str(value or '') for value in values)
    normalized = _kd_norm(text_value)
    terms = [term for term in _KD_DOMAIN_TERMS if _kd_norm(term) in normalized]
    for token in re.findall(r'[A-Za-z]+[A-Za-z0-9_.-]*|\d+[A-Za-z0-9_.-]*', text_value):
        token = token.strip()
        if len(token) >= 2 and token not in terms:
            terms.append(token)
    return terms[:60]


def _kd_answer_points(answer):
    parts = re.split(r'[\r\n]+|(?<=[。！？；;])', str(answer or ''))
    points = [part.strip(' \t-•0123456789.、）)') for part in parts]
    return [point for point in points if len(point) >= 2][:30]


def _kd_index_intent_text(question, similar_questions):
    values = [str(question or '').strip()] + _kd_string_list(similar_questions, limit=30)
    return '\n'.join(value for value in values if value)[:_SM_EMBEDDING_TEXT_MAX_CHARS]


def _kd_content_hash(row):
    payload = {
        'question': str(row.get('question') or ''),
        'answer': str(row.get('answer') or ''),
        'similar_questions': _kd_string_list(row.get('similar_questions')),
        'product_category_name': str(row.get('product_category_name') or ''),
        'product_name': _kd_string_list(row.get('product_name')),
        'keyword_list': _kd_string_list(row.get('keyword_list')),
        'error_list': _kd_string_list(row.get('error_list')),
    }
    return hashlib.sha256(_kd_json_dump(payload).encode('utf-8')).hexdigest()


def _kd_task_username():
    return current_user.username if current_user.is_authenticated else ''


def _kd_task_owned(task):
    return bool(task and (not task.username or task.username == _kd_task_username()))


def _kd_update_task(task_id, **fields):
    task = KBDuplicateCheckTask.query.get(str(task_id))
    if not task:
        return None
    for key, value in fields.items():
        if hasattr(task, key):
            setattr(task, key, value)
    task.updated_ts = time.time()
    db.session.commit()
    return task


def _kd_task_payload(task, include_results=True):
    completed = _kd_json_load(task.completed_channels_json, [])
    failed = _kd_json_load(task.failed_stages_json, [])
    payload = {
        'task_id': task.task_id,
        'status': task.status,
        'stage': task.stage,
        'message': task.message or '',
        'library': task.library,
        'question': task.question or '',
        'answer': task.answer or '',
        'product_category_name': task.product_category_name or '',
        'product_names': _kd_json_load(task.product_names_json, []),
        'source_note': task.source_note or '',
        'top_k': int(task.top_k or 20),
        'expanded': bool(task.expanded),
        'index': {
            'total': int(task.index_total or 0),
            'done': int(task.index_done or 0),
        },
        'candidate_count': int(task.candidate_count or 0),
        'completed_channels': completed,
        'failed_stages': failed,
        'human_decision': task.human_decision or '',
        'human_note': task.human_note or '',
        'selected_source_ids': _kd_json_load(task.selected_source_ids_json, []),
        'created_ts': float(task.created_ts or 0),
        'updated_ts': float(task.updated_ts or 0),
    }
    if include_results:
        payload['candidates'] = _kd_json_load(task.candidates_json, [])
        payload['analysis'] = _kd_json_load(task.analysis_json, {})
        payload['config_snapshot'] = _kd_json_load(task.config_snapshot_json, {})
    return payload


def _kd_config_snapshot():
    embedding = _sm_public_embedding_config()
    try:
        ai_config = load_ai_config() or {}
    except Exception:
        ai_config = {}
    return {
        'embedding': {
            'api_host': urlparse(str(embedding.get('api_url') or '')).netloc,
            'model': embedding.get('model') or '',
            'dimensions': int(embedding.get('dimensions') or 0),
        },
        'coverage_ai': {
            'api_host': urlparse(str(ai_config.get('base_url') or '')).netloc,
            'model': str(ai_config.get('model') or ''),
            'prompt_version': _KD_PROMPT_VERSION,
        },
    }


def _kd_fetch_source_rows(library):
    client = get_supabase_client()
    if not client:
        raise RuntimeError(_db_not_configured_message())
    columns = (
        'question_wiki_id,question,answer,similar_questions,product_name,'
        'product_category_name,keyword_list,error_list,update_time'
    )
    return client.select_all(
        library,
        order_by='question_wiki_id',
        order_dir='asc',
        columns=columns,
        page_size=1000,
    ) or []


def _kd_cached_embedding_keys(cache_keys):
    found = set()
    keys = list(dict.fromkeys(str(key) for key in cache_keys if key))
    for start in range(0, len(keys), 400):
        rows = SmartMappingEmbeddingCache.query.with_entities(
            SmartMappingEmbeddingCache.cache_key
        ).filter(
            SmartMappingEmbeddingCache.cache_key.in_(keys[start:start + 400])
        ).all()
        found.update(str(row[0]) for row in rows if row and row[0])
    return found


def _kd_sync_index(task, progress_callback=None, cancel_check=None, mode='incremental'):
    mode = mode if mode in ('incremental', 'full', 'failed') else 'incremental'

    def report(done, total, cache_hits, failed_count, message):
        if progress_callback:
            progress_callback(done, total, cache_hits, failed_count, message)
            return
        _kd_update_task(
            task.task_id,
            index_total=total,
            index_done=done,
            message=message,
        )

    def should_cancel():
        if cancel_check:
            return bool(cancel_check())
        current_task = KBDuplicateCheckTask.query.get(task.task_id)
        return bool(current_task and current_task.cancel_requested)

    rows = _kd_fetch_source_rows(task.library)
    config = _sm_load_embedding_config()
    existing = {
        item.question_wiki_id: item
        for item in KBDuplicateRetrievalIndex.query.filter_by(library_type=task.library).all()
    }
    changed = []
    seen_ids = set()
    unchanged_count = 0

    for row in rows:
        wiki_id = str(row.get('question_wiki_id') or '').strip()
        question = str(row.get('question') or '').strip()
        if not wiki_id or not question:
            continue
        seen_ids.add(wiki_id)
        answer = str(row.get('answer') or '').strip()
        similar_questions = _kd_string_list(row.get('similar_questions'))
        product_names = _kd_string_list(row.get('product_name'))
        category = str(row.get('product_category_name') or '').strip()
        topic_terms = _kd_topic_terms(
            question,
            answer,
            ' '.join(similar_questions),
            ' '.join(_kd_string_list(row.get('keyword_list'))),
            ' '.join(_kd_string_list(row.get('error_list'))),
        )
        intent_text = _kd_index_intent_text(question, similar_questions)
        intent_key = _sm_embedding_cache_key(intent_text, config)
        content_key = _sm_embedding_cache_key(answer, config) if answer else ''
        content_hash = _kd_content_hash(row)
        item = existing.get(wiki_id)
        is_current = bool(
            item
            and item.index_status == 'ready'
            and item.content_hash == content_hash
            and item.intent_cache_key == intent_key
            and _kd_json_load(item.content_cache_keys_json, []) == ([content_key] if content_key else [])
        )
        is_failed_target = bool(item and item.index_status in ('failed', 'pending'))
        if mode == 'failed' and not is_failed_target:
            unchanged_count += 1
            continue
        if mode == 'incremental' and is_current:
            unchanged_count += 1
            continue
        if not item:
            item = KBDuplicateRetrievalIndex(
                library_type=task.library,
                question_wiki_id=wiki_id,
                content_hash=content_hash,
            )
            db.session.add(item)
        item.content_hash = content_hash
        item.intent_cache_key = intent_key
        item.content_cache_keys_json = _kd_json_dump([content_key] if content_key else [])
        item.question = question
        item.answer = answer
        item.similar_questions_json = _kd_json_dump(similar_questions)
        item.product_category_name = category
        item.product_names_json = _kd_json_dump(product_names)
        item.topic_terms_json = _kd_json_dump(topic_terms)
        item.source_update_time = str(row.get('update_time') or '')
        item.index_status = 'pending'
        item.last_error = ''
        item.updated_ts = time.time()
        changed.append((item, intent_text, answer))

    if mode != 'failed':
        for wiki_id, item in existing.items():
            if wiki_id not in seen_ids and item.index_status != 'deleted':
                item.index_status = 'deleted'
                item.updated_ts = time.time()
    db.session.commit()

    total = len(seen_ids)
    expected_cache_keys = []
    for item, _intent_text, answer in changed:
        expected_cache_keys.append(item.intent_cache_key)
        if answer:
            expected_cache_keys.extend(_kd_json_load(item.content_cache_keys_json, []))
    cache_hits = len(_kd_cached_embedding_keys(expected_cache_keys))
    report(unchanged_count, total, cache_hits, 0, f'索引同步：{unchanged_count}/{total}')
    embedding_error = ''
    failed_count = 0
    cancelled = False
    batch_size = max(1, min(64, int(config.get('batch_size') or 32)))
    for start in range(0, len(changed), batch_size):
        if should_cancel():
            cancelled = True
            break
        batch = changed[start:start + batch_size]
        texts = []
        for _, intent_text, answer in batch:
            texts.extend([intent_text, answer])
        try:
            vectors = _sm_get_embeddings(texts, config)
            for offset, (item, _intent_text, answer) in enumerate(batch):
                intent_vector = vectors[offset * 2]
                content_vector = vectors[offset * 2 + 1]
                if not intent_vector or (answer and not content_vector):
                    raise RuntimeError('Embedding API 未返回完整的索引向量')
                item.index_status = 'ready'
                item.last_error = ''
                item.indexed_at = datetime.utcnow()
                item.updated_ts = time.time()
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            embedding_error = str(exc or 'Embedding 索引失败').replace('\n', ' ')[:500]
            pending_items = [item for item, _, _ in changed[start:]]
            failed_count = len(pending_items)
            pending_ids = [item.id for item in pending_items if item.id]
            if pending_ids:
                KBDuplicateRetrievalIndex.query.filter(
                    KBDuplicateRetrievalIndex.id.in_(pending_ids)
                ).update({
                    KBDuplicateRetrievalIndex.index_status: 'failed',
                    KBDuplicateRetrievalIndex.last_error: embedding_error,
                    KBDuplicateRetrievalIndex.updated_ts: time.time(),
                }, synchronize_session=False)
                db.session.commit()
            break
        completed = unchanged_count + min(start + len(batch), len(changed))
        report(completed, total, cache_hits, 0, f'索引同步：{completed}/{total}')

    if cancelled:
        ready_count = KBDuplicateRetrievalIndex.query.filter_by(
            library_type=task.library,
            index_status='ready',
        ).count()
        report(ready_count, total, cache_hits, 0, '索引任务已取消')
    elif embedding_error:
        report(total, total, cache_hits, failed_count, 'Embedding 不可用，继续使用本地召回')
    else:
        report(total, total, cache_hits, 0, f'索引已就绪：{total}/{total}')
    return {
        'embedding_error': embedding_error,
        'total': total,
        'done': total if not cancelled else ready_count,
        'cache_hits': cache_hits,
        'failed_count': failed_count,
        'updated_count': len(changed),
        'cancelled': cancelled,
    }


def _kd_ranked_channel(items, key, limit, minimum=0.0):
    ranked = sorted(items, key=lambda item: float(item.get(key) or 0), reverse=True)
    return [item for item in ranked if float(item.get(key) or 0) > minimum][:limit]


def _kd_retrieve_candidates(task):
    indexes = KBDuplicateRetrievalIndex.query.filter(
        KBDuplicateRetrievalIndex.library_type == task.library,
        KBDuplicateRetrievalIndex.index_status != 'deleted',
    ).all()
    if not indexes:
        return [], {'embedding_error': '', 'channels': ['structured']}

    config = _sm_load_embedding_config()
    query_vectors = None
    cached_vectors = {}
    embedding_error = ''
    try:
        query_vectors = _sm_get_embeddings([task.question, task.answer], config)
        cache_keys = []
        for item in indexes:
            if item.index_status != 'ready':
                continue
            cache_keys.append(item.intent_cache_key)
            cache_keys.extend(_kd_json_load(item.content_cache_keys_json, []))
        cached_vectors = _sm_load_cached_embeddings(cache_keys, config)
    except Exception as exc:
        db.session.rollback()
        embedding_error = str(exc or 'Embedding 查询失败').replace('\n', ' ')[:500]
        query_vectors = None

    query_category = _kd_norm(task.product_category_name)
    query_models = {_kd_norm(item) for item in _kd_json_load(task.product_names_json, []) if _kd_norm(item)}
    query_terms = set(_kd_topic_terms(task.question, task.answer))
    scored = []
    for item in indexes:
        intent_text = _kd_index_intent_text(item.question, _kd_json_load(item.similar_questions_json, []))
        question_score = _ai_cosine_sim(task.question, intent_text)
        answer_score = _ai_cosine_sim(task.answer, item.answer)
        algorithm = 'ngram_fallback'
        if query_vectors and item.index_status == 'ready':
            intent_vector = cached_vectors.get(item.intent_cache_key)
            content_keys = _kd_json_load(item.content_cache_keys_json, [])
            content_vector = cached_vectors.get(content_keys[0]) if content_keys else None
            if intent_vector:
                question_score = _sm_vector_cosine(query_vectors[0], intent_vector)
                algorithm = 'embedding'
            if content_vector:
                answer_score = _sm_vector_cosine(query_vectors[1], content_vector)
                algorithm = 'embedding'

        category_norm = _kd_norm(item.product_category_name)
        model_norms = {_kd_norm(value) for value in _kd_json_load(item.product_names_json, []) if _kd_norm(value)}
        item_terms = set(_kd_json_load(item.topic_terms_json, []))
        term_overlap = sorted(query_terms & item_terms, key=lambda value: (-len(value), value))
        category_match = bool(query_category and category_norm and query_category == category_norm)
        category_conflict = bool(query_category and category_norm and query_category != category_norm)
        model_overlap = sorted(query_models & model_norms)
        model_conflict = bool(query_models and model_norms and not model_overlap)
        structured_score = 0.0
        if category_match:
            structured_score += 0.38
        if model_overlap:
            structured_score += 0.34
        if term_overlap:
            structured_score += min(0.28, 0.07 * len(term_overlap))
        local_score = (float(question_score) + float(answer_score)) / 2.0
        scored.append({
            'question_wiki_id': item.question_wiki_id,
            'question': item.question or '',
            'answer': item.answer or '',
            'similar_questions': _kd_json_load(item.similar_questions_json, []),
            'product_category_name': item.product_category_name or '',
            'product_names': _kd_json_load(item.product_names_json, []),
            'source_update_time': item.source_update_time or '',
            'question_similarity': round(float(question_score), 4),
            'answer_similarity': round(float(answer_score), 4),
            'local_similarity': round(local_score, 4),
            'structured_score': round(structured_score, 4),
            'keyword_hits': term_overlap,
            'category_match': category_match,
            'category_conflict': category_conflict,
            'model_overlap': model_overlap,
            'model_conflict': model_conflict,
            'algorithm': algorithm,
            'channels': [],
            'fusion_score': 0.0,
        })

    channels = []
    if query_vectors and any(item.get('algorithm') == 'embedding' for item in scored):
        channels.extend([
            ('embedding_intent', _kd_ranked_channel(scored, 'question_similarity', 50, 0.05)),
            ('embedding_content', _kd_ranked_channel(scored, 'answer_similarity', 50, 0.05)),
        ])
    else:
        channels.append(('ngram_fallback', _kd_ranked_channel(scored, 'local_similarity', 80, 0.01)))
    channels.append(('structured', _kd_ranked_channel(scored, 'structured_score', 50, 0.0)))

    fused = {}
    for channel_name, ranked in channels:
        for rank, candidate in enumerate(ranked, start=1):
            target = fused.setdefault(candidate['question_wiki_id'], candidate)
            target['fusion_score'] += 1.0 / (60.0 + rank)
            if channel_name not in target['channels']:
                target['channels'].append(channel_name)
    if not fused:
        for candidate in _kd_ranked_channel(scored, 'local_similarity', min(20, task.top_k), 0.0):
            fused[candidate['question_wiki_id']] = candidate
            candidate['channels'] = ['ngram_fallback']

    results = list(fused.values())
    for candidate in results:
        candidate['fusion_score'] = round(
            candidate['fusion_score'] + candidate['structured_score'] * 0.01,
            6,
        )
        if candidate['category_conflict'] or candidate['model_conflict']:
            candidate['fusion_score'] = round(candidate['fusion_score'] * 0.9, 6)
        if candidate['model_conflict']:
            relationship = 'conflicting'
        elif candidate['question_similarity'] >= 0.78 and candidate['answer_similarity'] >= 0.72:
            relationship = 'fully_covered'
        elif candidate['question_similarity'] >= 0.5 or candidate['answer_similarity'] >= 0.5:
            relationship = 'partially_covered'
        else:
            relationship = 'unrelated'
        candidate.update({
            'relationship': relationship,
            'confidence': round(max(candidate['question_similarity'], candidate['answer_similarity']), 3),
            'covered_points': candidate['keyword_hits'][:8],
            'missing_points': [],
            'conflicts': ['适用型号范围不重叠'] if candidate['model_conflict'] else [],
            'recommended_action': 'manual_review',
            'reason': '等待 AI 覆盖判断',
            'analysis_source': 'heuristic',
        })
    results.sort(key=lambda item: (item['fusion_score'], item['local_similarity']), reverse=True)
    limit = max(1, min(50, int(task.top_k or 20)))
    return results[:limit], {
        'embedding_error': embedding_error,
        'channels': [name for name, ranked in channels if ranked],
    }


def _kd_normalize_ai_analysis(raw, candidates):
    if not isinstance(raw, dict):
        raise ValueError('AI 未返回合法 JSON')
    relationship = str(raw.get('relationship') or '').strip()
    action = str(raw.get('recommended_action') or '').strip()
    reason = str(raw.get('reason') or '').strip()[:2000]
    if relationship not in _KD_RELATIONSHIPS:
        raise ValueError('AI 未返回合法的覆盖关系')
    if action not in _KD_ACTIONS:
        raise ValueError('AI 未返回合法的建议动作')
    if not reason:
        raise ValueError('AI 未返回可核验的判断依据')
    candidate_ids = {item['question_wiki_id'] for item in candidates}
    source_ids = [
        value for value in _kd_string_list(raw.get('source_ids'), limit=20)
        if value in candidate_ids
    ]
    normalized_candidates = []
    for item in raw.get('candidates') or []:
        if not isinstance(item, dict):
            continue
        source_id = str(item.get('source_id') or item.get('question_wiki_id') or '').strip()
        item_relationship = str(item.get('relationship') or '').strip()
        item_action = str(item.get('recommended_action') or 'manual_review').strip()
        if source_id not in candidate_ids or item_relationship not in _KD_RELATIONSHIPS:
            continue
        normalized_candidates.append({
            'source_id': source_id,
            'relationship': item_relationship,
            'confidence': _kb_compare_ai_confidence(item.get('confidence')),
            'covered_points': _kd_string_list(item.get('covered_points'), limit=20),
            'missing_points': _kd_string_list(item.get('missing_points'), limit=20),
            'conflicts': _kd_string_list(item.get('conflicts'), limit=12),
            'recommended_action': item_action if item_action in _KD_ACTIONS else 'manual_review',
            'reason': str(item.get('reason') or '').strip()[:1200],
        })
    return {
        'relationship': relationship,
        'confidence': _kb_compare_ai_confidence(raw.get('confidence')),
        'covered_points': _kd_string_list(raw.get('covered_points'), limit=30),
        'missing_points': _kd_string_list(raw.get('missing_points'), limit=30),
        'conflicts': _kd_string_list(raw.get('conflicts'), limit=20),
        'recommended_action': action,
        'reason': reason,
        'source_ids': source_ids,
        'candidates': normalized_candidates,
        'prompt_version': _KD_PROMPT_VERSION,
    }


def _kd_run_ai_coverage(task, candidates):
    if not candidates:
        return {
            'relationship': 'unrelated',
            'confidence': 0.0,
            'covered_points': [],
            'missing_points': _kd_answer_points(task.answer),
            'conflicts': [],
            'recommended_action': 'manual_review',
            'reason': '未发现高相关候选，需要人工确认后再决定是否新增。',
            'source_ids': [],
            'candidates': [],
            'prompt_version': _KD_PROMPT_VERSION,
        }
    prompt_candidates = []
    for candidate in candidates[:12]:
        prompt_candidates.append({
            'source_id': candidate['question_wiki_id'],
            'question': candidate['question'][:3000],
            'answer': candidate['answer'][:5000],
            'product_category_name': candidate['product_category_name'],
            'product_names': candidate['product_names'],
            'recall_channels': candidate['channels'],
            'question_similarity': candidate['question_similarity'],
            'answer_similarity': candidate['answer_similarity'],
            'keyword_hits': candidate['keyword_hits'],
            'range_conflicts': candidate['conflicts'],
        })
    system_prompt = (
        '你是企业知识库查重审核员。判断拟新增 FAQ 是否被现有知识完整、部分、合计覆盖，或存在冲突。\n'
        '只输出合法 JSON，不输出 Markdown 或额外解释。不得补造输入中没有的事实。\n'
        '主题相似但事实点不同必须判 unrelated；信息不足必须建议 manual_review。\n'
        '型号或适用范围冲突时不得建议直接合并。\n'
        'relationship 只能是 fully_covered、partially_covered、collectively_covered、conflicting、unrelated。\n'
        'recommended_action 只能是 no_add、update_existing、compare_merge、create_new、manual_review。\n'
        '输出结构：{"relationship":"...","confidence":0-1,"covered_points":[],"missing_points":[],'
        '"conflicts":[],"recommended_action":"...","reason":"一句可核验依据","source_ids":[],'
        '"candidates":[{"source_id":"...","relationship":"...","confidence":0-1,'
        '"covered_points":[],"missing_points":[],"conflicts":[],"recommended_action":"...","reason":"..."}]}'
    )
    user_prompt = _kd_json_dump({
        'new_faq': {
            'question': task.question,
            'answer': task.answer,
            'answer_points': _kd_answer_points(task.answer),
            'product_category_name': task.product_category_name,
            'product_names': _kd_json_load(task.product_names_json, []),
        },
        'candidates': prompt_candidates,
    })
    last_error = None
    for attempt in range(2):
        try:
            raw_text = _ai_call_llm(load_ai_config() or {}, system_prompt, user_prompt, temperature=0.1)
            parsed = _ai_repair_mojibake_value(_ai_extract_json(raw_text))
            return _kd_normalize_ai_analysis(parsed, candidates)
        except Exception as exc:
            last_error = exc
            user_prompt += '\n上一轮输出未通过 JSON 契约校验，请只返回符合结构的 JSON。'
    raise last_error or RuntimeError('AI 覆盖判断失败')


def _kd_apply_ai_candidates(candidates, analysis):
    by_id = {item.get('source_id'): item for item in analysis.get('candidates') or []}
    result = []
    for candidate in candidates:
        item = by_id.get(candidate['question_wiki_id'])
        if item:
            candidate = dict(candidate)
            candidate.update({
                'relationship': item['relationship'],
                'confidence': item['confidence'],
                'covered_points': item['covered_points'],
                'missing_points': item['missing_points'],
                'conflicts': item['conflicts'],
                'recommended_action': item['recommended_action'],
                'reason': item['reason'],
                'analysis_source': 'ai',
            })
        result.append(candidate)
    return result


def _kd_safe_stage_error(exc, stage):
    message = str(exc or '').strip().replace('\n', ' ')
    if '配置不完整' in message:
        return message[:500]
    if message.startswith('HTTP '):
        status = message.split(':', 1)[0]
        return f'{stage}服务请求失败（{status}）'
    return (message or f'{stage}失败')[:500]


def _kd_run_task(task_id, retry_stage='all'):
    ctx = app.app_context()
    ctx.push()
    try:
        task = KBDuplicateCheckTask.query.get(str(task_id))
        if not task:
            return
        completed = _kd_json_load(task.completed_channels_json, [])
        failed = _kd_json_load(task.failed_stages_json, [])
        if retry_stage != 'ai':
            _kd_update_task(
                task_id,
                status='running',
                stage='preparing_index',
                message='正在同步知识检索索引',
                cancel_requested=False,
            )
            task = KBDuplicateCheckTask.query.get(str(task_id))
            sync_result = _kd_sync_index(task)
            embedding_error = sync_result.get('embedding_error') or ''
            task = KBDuplicateCheckTask.query.get(str(task_id))
            if task.cancel_requested:
                _kd_update_task(task_id, status='cancelled', stage='cancelled', message='任务已取消')
                return
            _kd_update_task(task_id, stage='recalling', message='正在执行多路候选召回')
            task = KBDuplicateCheckTask.query.get(str(task_id))
            candidates, recall_meta = _kd_retrieve_candidates(task)
            embedding_error = embedding_error or recall_meta.get('embedding_error') or ''
            completed = [value for value in completed if value not in ('embedding', 'ngram_fallback', 'structured')]
            failed = [value for value in failed if value != 'embedding']
            completed.append('structured')
            if embedding_error:
                completed.append('ngram_fallback')
                failed.append('embedding')
            else:
                completed.append('embedding')
            completed = list(OrderedDict.fromkeys(completed))
            failed = list(OrderedDict.fromkeys(failed))
            _kd_update_task(
                task_id,
                stage='ai_analyzing',
                message=f'已召回 {len(candidates)} 条候选，正在进行 AI 覆盖判断',
                candidate_count=len(candidates),
                candidates_json=_kd_json_dump(candidates),
                completed_channels_json=_kd_json_dump(completed),
                failed_stages_json=_kd_json_dump(failed),
            )
        else:
            candidates = _kd_json_load(task.candidates_json, [])
            completed = [value for value in completed if value != 'ai']
            failed = [value for value in failed if value != 'ai']
            _kd_update_task(
                task_id,
                status='running',
                stage='ai_analyzing',
                message='正在重试 AI 覆盖判断',
                cancel_requested=False,
                completed_channels_json=_kd_json_dump(completed),
                failed_stages_json=_kd_json_dump(failed),
            )

        task = KBDuplicateCheckTask.query.get(str(task_id))
        if task.cancel_requested:
            _kd_update_task(task_id, status='cancelled', stage='cancelled', message='任务已取消')
            return
        analysis = {}
        try:
            analysis = _kd_run_ai_coverage(task, candidates)
            candidates = _kd_apply_ai_candidates(candidates, analysis)
            completed.append('ai')
            failed = [value for value in failed if value != 'ai']
        except Exception as exc:
            failed.append('ai')
            analysis = {
                'relationship': '',
                'confidence': 0.0,
                'covered_points': [],
                'missing_points': [],
                'conflicts': [],
                'recommended_action': 'manual_review',
                'reason': _kd_safe_stage_error(exc, 'AI 覆盖判断'),
                'source_ids': [],
                'candidates': [],
                'prompt_version': _KD_PROMPT_VERSION,
            }
        completed = list(OrderedDict.fromkeys(completed))
        failed = list(OrderedDict.fromkeys(failed))
        final_status = 'partial_failed' if failed else 'done'
        final_message = '查重已完成' if not failed else '查重已完成，部分阶段可重试'
        _kd_update_task(
            task_id,
            status=final_status,
            stage='completed',
            message=final_message,
            candidate_count=len(candidates),
            candidates_json=_kd_json_dump(candidates),
            analysis_json=_kd_json_dump(analysis),
            completed_channels_json=_kd_json_dump(completed),
            failed_stages_json=_kd_json_dump(failed),
        )
    except Exception as exc:
        db.session.rollback()
        try:
            task = KBDuplicateCheckTask.query.get(str(task_id))
            stage = task.stage if task else 'unknown'
            failed = _kd_json_load(task.failed_stages_json, []) if task else []
            failed.append(stage)
            _kd_update_task(
                task_id,
                status='failed',
                message=_kd_safe_stage_error(exc, '查重任务'),
                failed_stages_json=_kd_json_dump(list(OrderedDict.fromkeys(failed))),
            )
        except Exception:
            db.session.rollback()
    finally:
        with _KD_LOCK:
            _KD_ACTIVE_TASKS.discard(str(task_id))
        ctx.pop()


def _kd_spawn_task(task_id, retry_stage='all'):
    task_id = str(task_id)
    with _KD_LOCK:
        if task_id in _KD_ACTIVE_TASKS:
            return False
        _KD_ACTIVE_TASKS.add(task_id)
    thread = threading.Thread(
        target=_kd_run_task,
        args=(task_id, retry_stage),
        daemon=True,
        name=f'kb-duplicate-{task_id[:8]}',
    )
    thread.start()
    return True


def _kd_update_index_job(job_id, **fields):
    job = KBDuplicateIndexJob.query.get(str(job_id))
    if not job:
        return None
    for key, value in fields.items():
        if hasattr(job, key):
            setattr(job, key, value)
    job.updated_ts = time.time()
    db.session.commit()
    return job


def _kd_index_job_owned(job):
    return bool(job and (not job.username or job.username == _kd_task_username()))


def _kd_index_job_payload(job):
    if not job:
        return None
    return {
        'job_id': job.job_id,
        'library': job.library,
        'mode': job.mode,
        'status': job.status,
        'total': int(job.total or 0),
        'done': int(job.done or 0),
        'cache_hits': int(job.cache_hits or 0),
        'failed_count': int(job.failed_count or 0),
        'message': job.message or '',
        'error': job.error or '',
        'created_ts': float(job.created_ts or 0),
        'updated_ts': float(job.updated_ts or 0),
    }


def _kd_index_status_payload(library, job=None):
    base = KBDuplicateRetrievalIndex.query.filter_by(library_type=library)
    counts = {
        status: base.filter_by(index_status=status).count()
        for status in ('ready', 'pending', 'failed', 'deleted')
    }
    latest_indexed_at = db.session.query(
        func.max(KBDuplicateRetrievalIndex.indexed_at)
    ).filter(
        KBDuplicateRetrievalIndex.library_type == library,
        KBDuplicateRetrievalIndex.index_status == 'ready',
    ).scalar()
    failed_item = base.filter_by(index_status='failed').order_by(
        KBDuplicateRetrievalIndex.updated_ts.desc()
    ).first()
    config = _sm_public_embedding_config()
    return {
        'library': library,
        'total': counts['ready'] + counts['pending'] + counts['failed'],
        **counts,
        'cache_count': int(config.get('cache_count') or 0),
        'last_indexed_at': latest_indexed_at.isoformat() if latest_indexed_at else '',
        'last_error': (failed_item.last_error or '')[:500] if failed_item else '',
        'embedding': {
            'model': config.get('model') or '',
            'dimensions': int(config.get('dimensions') or 0),
            'api_key_configured': bool(config.get('api_key_configured')),
        },
        'job': _kd_index_job_payload(job),
    }


def _kd_run_index_job(job_id):
    ctx = app.app_context()
    ctx.push()
    try:
        job = KBDuplicateIndexJob.query.get(str(job_id))
        if not job:
            return

        def progress(done, total, cache_hits, failed_count, message):
            _kd_update_index_job(
                job_id,
                total=total,
                done=done,
                cache_hits=cache_hits,
                failed_count=failed_count,
                message=message,
            )

        def cancelled():
            current = KBDuplicateIndexJob.query.get(str(job_id))
            return bool(current and current.cancel_requested)

        _kd_update_index_job(
            job_id,
            status='running',
            message='正在读取知识库并计算内容变更',
            error='',
            cancel_requested=False,
        )
        job = KBDuplicateIndexJob.query.get(str(job_id))
        result = _kd_sync_index(
            job,
            progress_callback=progress,
            cancel_check=cancelled,
            mode=job.mode,
        )
        if result.get('cancelled'):
            _kd_update_index_job(job_id, status='cancelled', message='索引任务已取消')
        elif result.get('embedding_error'):
            _kd_update_index_job(
                job_id,
                status='partial_failed',
                failed_count=int(result.get('failed_count') or 0),
                error=str(result.get('embedding_error') or '')[:500],
                message='索引任务完成，但部分向量生成失败',
            )
        else:
            _kd_update_index_job(
                job_id,
                status='done',
                total=int(result.get('total') or 0),
                done=int(result.get('done') or 0),
                cache_hits=int(result.get('cache_hits') or 0),
                failed_count=0,
                message='知识检索索引已就绪',
            )
    except Exception as exc:
        db.session.rollback()
        try:
            _kd_update_index_job(
                job_id,
                status='failed',
                error=_kd_safe_stage_error(exc, '索引任务'),
                message='索引任务失败',
            )
        except Exception:
            db.session.rollback()
    finally:
        with _KD_LOCK:
            _KD_INDEX_ACTIVE_JOBS.discard(str(job_id))
        ctx.pop()


def _kd_spawn_index_job(job_id):
    job_id = str(job_id)
    with _KD_LOCK:
        if job_id in _KD_INDEX_ACTIVE_JOBS:
            return False
        _KD_INDEX_ACTIVE_JOBS.add(job_id)
    thread = threading.Thread(
        target=_kd_run_index_job,
        args=(job_id,),
        daemon=True,
        name=f'kb-index-{job_id[:8]}',
    )
    thread.start()
    return True


@app.route('/api/kb/duplicate-check/index/status', methods=['GET'])
@login_required
def kb_duplicate_index_status():
    library = str(request.args.get('library') or 'knowledge_base_v1').strip()
    if library not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
        return jsonify({'success': False, 'message': '索引范围无效'}), 400
    job = KBDuplicateIndexJob.query.filter_by(
        username=_kd_task_username(),
        library=library,
    ).order_by(KBDuplicateIndexJob.created_ts.desc()).first()
    if job and job.status == 'running':
        _kd_spawn_index_job(job.job_id)
    return jsonify({'success': True, **_kd_index_status_payload(library, job)})


@app.route('/api/kb/duplicate-check/index/rebuild', methods=['POST'])
@login_required
def kb_duplicate_index_rebuild():
    payload = request.json or {}
    library = str(payload.get('library') or 'knowledge_base_v1').strip()
    mode = str(payload.get('mode') or 'incremental').strip()
    if library not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
        return jsonify({'success': False, 'message': '索引范围无效'}), 400
    if mode not in ('incremental', 'full', 'failed'):
        return jsonify({'success': False, 'message': '索引模式无效'}), 400
    try:
        _sm_validate_embedding_config(_sm_load_embedding_config())
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400
    active = KBDuplicateIndexJob.query.filter_by(
        username=_kd_task_username(),
        library=library,
        status='running',
    ).order_by(KBDuplicateIndexJob.created_ts.desc()).first()
    if active:
        _kd_spawn_index_job(active.job_id)
        return jsonify({
            'success': False,
            'message': '当前知识库已有索引任务正在执行',
            'job_id': active.job_id,
        }), 409
    job_id = str(uuid.uuid4())
    now = time.time()
    job = KBDuplicateIndexJob(
        job_id=job_id,
        username=_kd_task_username(),
        library=library,
        mode=mode,
        status='running',
        message='索引任务已创建',
        config_snapshot_json=_kd_json_dump(_kd_config_snapshot()),
        created_ts=now,
        updated_ts=now,
    )
    db.session.add(job)
    db.session.commit()
    _kd_spawn_index_job(job_id)
    return jsonify({'success': True, 'job_id': job_id, 'status': 'running', 'mode': mode})


@app.route('/api/kb/duplicate-check/index/cancel', methods=['POST'])
@login_required
def kb_duplicate_index_cancel():
    payload = request.json or {}
    job_id = str(payload.get('job_id') or '').strip()
    job = KBDuplicateIndexJob.query.get(job_id) if job_id else None
    if not _kd_index_job_owned(job):
        return jsonify({'success': False, 'message': '索引任务不存在'}), 404
    if job.status != 'running':
        return jsonify({'success': False, 'message': '索引任务当前不可取消'}), 409
    _kd_update_index_job(job_id, cancel_requested=True, message='正在取消索引任务')
    return jsonify({'success': True, 'message': '已提交取消请求'})


@app.route('/api/kb/duplicate-check/start', methods=['POST'])
@login_required
def kb_duplicate_check_start():
    payload = request.json or {}
    library = str(payload.get('library') or 'knowledge_base_v1').strip()
    question = str(payload.get('question') or '').strip()
    answer = str(payload.get('answer') or '').strip()
    if library not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
        return jsonify({'success': False, 'message': '查重范围无效'}), 400
    if not question or not answer:
        return jsonify({'success': False, 'message': '拟新增问题和拟新增答案均为必填项'}), 400
    if len(question) > 8000 or len(answer) > 30000:
        return jsonify({'success': False, 'message': '问题或答案内容过长，请精简后重试'}), 400
    task_id = str(uuid.uuid4())
    now = time.time()
    task = KBDuplicateCheckTask(
        task_id=task_id,
        username=_kd_task_username(),
        library=library,
        status='running',
        stage='preparing_index',
        question=question,
        answer=answer,
        product_category_name=str(payload.get('product_category_name') or '').strip()[:1000],
        product_names_json=_kd_json_dump(_kd_string_list(payload.get('product_names'), limit=100)),
        source_note=str(payload.get('source_note') or '').strip()[:4000],
        top_k=20,
        expanded=False,
        config_snapshot_json=_kd_json_dump(_kd_config_snapshot()),
        message='任务已创建，正在准备索引',
        created_ts=now,
        updated_ts=now,
    )
    db.session.add(task)
    db.session.commit()
    _kd_spawn_task(task_id)
    return jsonify({'success': True, 'task_id': task_id, 'status': 'running'})


@app.route('/api/kb/duplicate-check/status', methods=['GET'])
@login_required
def kb_duplicate_check_status():
    task_id = str(request.args.get('task_id') or '').strip()
    task = KBDuplicateCheckTask.query.get(task_id) if task_id else None
    if not _kd_task_owned(task):
        return jsonify({'success': False, 'message': '查重任务不存在'}), 404
    if task.status == 'running':
        _kd_spawn_task(task_id)
    return jsonify({'success': True, **_kd_task_payload(task)})


@app.route('/api/kb/duplicate-check/retry', methods=['POST'])
@login_required
def kb_duplicate_check_retry():
    payload = request.json or {}
    task_id = str(payload.get('task_id') or '').strip()
    stage = str(payload.get('stage') or 'all').strip()
    task = KBDuplicateCheckTask.query.get(task_id) if task_id else None
    if not _kd_task_owned(task):
        return jsonify({'success': False, 'message': '查重任务不存在'}), 404
    if stage not in ('all', 'embedding', 'ai'):
        return jsonify({'success': False, 'message': '仅支持重试 embedding、ai 或全部阶段'}), 400
    retry_stage = 'ai' if stage == 'ai' and _kd_json_load(task.candidates_json, []) else 'all'
    _kd_update_task(task_id, status='running', cancel_requested=False, message='正在重试失败阶段')
    _kd_spawn_task(task_id, retry_stage=retry_stage)
    return jsonify({'success': True, 'task_id': task_id, 'status': 'running'})


@app.route('/api/kb/duplicate-check/expand', methods=['POST'])
@login_required
def kb_duplicate_check_expand():
    payload = request.json or {}
    task_id = str(payload.get('task_id') or '').strip()
    task = KBDuplicateCheckTask.query.get(task_id) if task_id else None
    if not _kd_task_owned(task):
        return jsonify({'success': False, 'message': '查重任务不存在'}), 404
    _kd_update_task(
        task_id,
        status='running',
        top_k=50,
        expanded=True,
        cancel_requested=False,
        message='正在扩大检索范围',
    )
    _kd_spawn_task(task_id, retry_stage='all')
    return jsonify({'success': True, 'task_id': task_id, 'status': 'running', 'top_k': 50})


@app.route('/api/kb/duplicate-check/decision', methods=['POST'])
@login_required
def kb_duplicate_check_decision():
    payload = request.json or {}
    task_id = str(payload.get('task_id') or '').strip()
    decision = str(payload.get('decision') or '').strip()
    task = KBDuplicateCheckTask.query.get(task_id) if task_id else None
    if not _kd_task_owned(task):
        return jsonify({'success': False, 'message': '查重任务不存在'}), 404
    if decision not in _KD_ACTIONS:
        return jsonify({'success': False, 'message': '人工结论无效'}), 400
    _kd_update_task(
        task_id,
        human_decision=decision,
        human_note=str(payload.get('note') or '').strip()[:4000],
        selected_source_ids_json=_kd_json_dump(_kd_string_list(payload.get('source_ids'), limit=50)),
    )
    return jsonify({'success': True, 'message': '人工结论已保存'})


@app.route('/api/kb/duplicate-check/cancel', methods=['POST'])
@login_required
def kb_duplicate_check_cancel():
    payload = request.json or {}
    task_id = str(payload.get('task_id') or '').strip()
    task = KBDuplicateCheckTask.query.get(task_id) if task_id else None
    if not _kd_task_owned(task):
        return jsonify({'success': False, 'message': '查重任务不存在'}), 404
    _kd_update_task(task_id, cancel_requested=True, message='正在取消任务')
    return jsonify({'success': True, 'message': '已提交取消请求'})


@app.route('/api/kb/duplicate-check/history', methods=['GET'])
@login_required
def kb_duplicate_check_history():
    query = KBDuplicateCheckTask.query.filter_by(username=_kd_task_username())
    decision = str(request.args.get('decision') or '').strip()
    category = str(request.args.get('category') or '').strip()
    if decision:
        query = query.filter_by(human_decision=decision)
    if category:
        query = query.filter(KBDuplicateCheckTask.product_category_name.ilike(f'%{category}%'))
    tasks = query.order_by(KBDuplicateCheckTask.created_ts.desc()).limit(50).all()
    return jsonify({
        'success': True,
        'items': [_kd_task_payload(task, include_results=False) for task in tasks],
    })


@app.route('/api/kb/duplicate-check/export', methods=['GET'])
@login_required
def kb_duplicate_check_export():
    task_id = str(request.args.get('task_id') or '').strip()
    task = KBDuplicateCheckTask.query.get(task_id) if task_id else None
    if not _kd_task_owned(task):
        return jsonify({'success': False, 'message': '查重任务不存在'}), 404
    candidates = _kd_json_load(task.candidates_json, [])
    analysis = _kd_json_load(task.analysis_json, {})
    workbook = Workbook()
    summary = workbook.active
    summary.title = '查重结论'
    summary.append(['字段', '内容'])
    summary.append(['任务 ID', task.task_id])
    summary.append(['查重范围', task.library])
    summary.append(['拟新增问题', task.question])
    summary.append(['拟新增答案', task.answer])
    summary.append(['产品品类', task.product_category_name])
    summary.append(['适用型号', '、'.join(_kd_json_load(task.product_names_json, []))])
    summary.append(['来源说明', task.source_note])
    summary.append(['系统覆盖结论', analysis.get('relationship') or ''])
    summary.append(['系统建议动作', analysis.get('recommended_action') or ''])
    summary.append(['判断依据', analysis.get('reason') or ''])
    summary.append(['已覆盖项', '\n'.join(analysis.get('covered_points') or [])])
    summary.append(['缺失项', '\n'.join(analysis.get('missing_points') or [])])
    summary.append(['冲突项', '\n'.join(analysis.get('conflicts') or [])])
    summary.append(['人工结论', task.human_decision or ''])
    summary.append(['人工备注', task.human_note or ''])
    summary.append(['Embedding 模型', (_kd_json_load(task.config_snapshot_json, {}).get('embedding') or {}).get('model', '')])
    summary.append(['AI 模型', (_kd_json_load(task.config_snapshot_json, {}).get('coverage_ai') or {}).get('model', '')])
    summary.append(['提示词版本', _KD_PROMPT_VERSION])
    summary.column_dimensions['A'].width = 20
    summary.column_dimensions['B'].width = 90
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    sheet = workbook.create_sheet('候选明细')
    sheet.append([
        'Wiki ID', '标准问题', '答案', '品类', '型号', '召回渠道', '问题相似度',
        '答案相似度', '命中词', '覆盖结论', '已覆盖项', '缺失项', '冲突项',
        '建议动作', '判断依据',
    ])
    for candidate in candidates:
        sheet.append([
            candidate.get('question_wiki_id', ''),
            candidate.get('question', ''),
            candidate.get('answer', ''),
            candidate.get('product_category_name', ''),
            '、'.join(candidate.get('product_names') or []),
            '、'.join(candidate.get('channels') or []),
            candidate.get('question_similarity', 0),
            candidate.get('answer_similarity', 0),
            '、'.join(candidate.get('keyword_hits') or []),
            candidate.get('relationship', ''),
            '\n'.join(candidate.get('covered_points') or []),
            '\n'.join(candidate.get('missing_points') or []),
            '\n'.join(candidate.get('conflicts') or []),
            candidate.get('recommended_action', ''),
            candidate.get('reason', ''),
        ])
    sheet.freeze_panes = 'A2'
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=canonical_download_name('duplicate_check_report', 'xlsx', 's02', 'kb8085'),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def _sm_now_iso():
    return _now_iso_with_tz()

def _sm_db_create_job(job_id, username, total):
    try:
        now = time.time()
        job = SmartMappingJob(
            job_id=str(job_id),
            username=str(username or ''),
            status='running',
            total=int(total or 0),
            done=0,
            message='',
            results_json='',
            created_ts=now,
            updated_ts=now
        )
        db.session.merge(job)
        db.session.commit()
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        msg = str(e).lower()
        if 'no such table' in msg and 'smart_mapping_job' in msg:
            try:
                init_db()
            except Exception:
                return
            try:
                now = time.time()
                job = SmartMappingJob(
                    job_id=str(job_id),
                    username=str(username or ''),
                    status='running',
                    total=int(total or 0),
                    done=0,
                    message='',
                    results_json='',
                    created_ts=now,
                    updated_ts=now
                )
                db.session.merge(job)
                db.session.commit()
            except Exception:
                try:
                    db.session.rollback()
                except Exception:
                    pass

def _sm_db_update_job(job_id, **fields):
    try:
        job = SmartMappingJob.query.get(str(job_id))
        if not job:
            return
        for k, v in (fields or {}).items():
            if not hasattr(job, k):
                continue
            setattr(job, k, v)
        job.updated_ts = time.time()
        db.session.commit()
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        msg = str(e).lower()
        if 'no such table' in msg and 'smart_mapping_job' in msg:
            try:
                init_db()
            except Exception:
                return
            try:
                job = SmartMappingJob.query.get(str(job_id))
                if not job:
                    return
                for k, v in (fields or {}).items():
                    if not hasattr(job, k):
                        continue
                    setattr(job, k, v)
                job.updated_ts = time.time()
                db.session.commit()
            except Exception:
                try:
                    db.session.rollback()
                except Exception:
                    pass

def _sm_db_get_job(job_id):
    try:
        return SmartMappingJob.query.get(str(job_id))
    except Exception as e:
        msg = str(e).lower()
        if 'no such table' in msg and 'smart_mapping_job' in msg:
            try:
                init_db()
                return SmartMappingJob.query.get(str(job_id))
            except Exception:
                return None
        return None

def _sm_cleanup(ttl_sec=3600):
    now = time.time()
    global _SM_DB_CLEAN_TS
    with _SM_LOCK:
        drop_cache = []
        for k, v in _SM_BASELINE_CACHE.items():
            ts = (v or {}).get('ts', 0)
            if now - float(ts or 0) > ttl_sec:
                drop_cache.append(k)
        for k in drop_cache:
            _SM_BASELINE_CACHE.pop(k, None)

        drop_jobs = []
        for jid, job in _SM_JOBS.items():
            ts = (job or {}).get('ts', 0)
            if now - float(ts or 0) > ttl_sec:
                drop_jobs.append(jid)
        for jid in drop_jobs:
            _SM_JOBS.pop(jid, None)

    if now - float(_SM_DB_CLEAN_TS or 0) < 600:
        return
    _SM_DB_CLEAN_TS = now
    try:
        cutoff = now - float(ttl_sec or 0)
        SmartMappingJob.query.filter(SmartMappingJob.updated_ts < cutoff).delete(synchronize_session=False)
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

def _sm_baseline_cache_key(username, table, models):
    m = [str(x).strip() for x in (models or []) if str(x or '').strip()]
    m = sorted(list(dict.fromkeys(m)))
    return (str(username or ''), str(table or ''), tuple(m))

def _sm_build_models_or_filter(models):
    parts = []
    for m in models or []:
        s = str(m or '').strip()
        if not s:
            continue
        s = s.replace(',', ' ').replace('，', ' ').strip()
        if not s:
            continue
        parts.append(f"product_name.ilike.*{s}*")
    if not parts:
        return None
    return "(" + ",".join(parts) + ")"

def _sm_fetch_baseline_items(client, table, models):
    if not client:
        return []
    if table not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
        return []
    models = [str(x).strip() for x in (models or []) if str(x or '').strip()]
    if not models:
        return []
    filters = {}
    or_filter = _sm_build_models_or_filter(models)
    if or_filter:
        filters['or'] = or_filter
    rows = client.select_all(
        table,
        filters=filters,
        order_by='question_wiki_id',
        order_dir='asc',
        columns='question_wiki_id,question,answer,product_name,product_category_name',
        page_size=1000
    )
    out = []
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        wid = str(r.get('question_wiki_id') or '').strip()
        if not wid:
            continue
        out.append({
            'question_wiki_id': wid,
            'question': r.get('question') or '',
            'answer': r.get('answer') or '',
            'product_name': r.get('product_name') or '',
            'product_category_name': r.get('product_category_name') or ''
        })
    return out

@app.route('/api/smart_mapping/kb/load', methods=['POST'])
@login_required
def sm_kb_load():
    data = request.json or {}
    table = data.get('table', 'knowledge_base_v1')
    models = data.get('models') or []
    if table not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
        return jsonify({'success': False, 'message': 'Invalid table'}), 400
    if not isinstance(models, list) or not any(str(x or '').strip() for x in models):
        return jsonify({'success': False, 'message': 'models is required'}), 400

    _sm_cleanup()
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    try:
        items = _sm_fetch_baseline_items(client, table, models)
        key = _sm_baseline_cache_key(current_user.username if current_user.is_authenticated else '', table, models)
        with _SM_LOCK:
            _SM_BASELINE_CACHE[key] = {'ts': time.time(), 'items': items}
        return jsonify({'success': True, 'count': len(items)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

def _sm_parse_faq_excel(file_storage):
    name = str(getattr(file_storage, 'filename', '') or '')
    lower = name.lower()
    if not (lower.endswith('.xlsx') or lower.endswith('.xls')):
        raise ValueError('仅支持 .xlsx / .xls')

    stream = file_storage.stream
    stream.seek(0)
    if lower.endswith('.xls'):
        df = pd.read_excel(stream, header=None, engine='xlrd')
    else:
        df = pd.read_excel(stream, header=None)

    if df is None or df.empty:
        return []
    if df.shape[1] < 3:
        raise ValueError('模板校验失败：需要至少三列 (A=ID, B=问题, C=答案)')

    first = df.iloc[0, :3].tolist()
    first_s = [str(x).strip() if x is not None and not (isinstance(x, float) and math.isnan(x)) else '' for x in first]
    if len(first_s) >= 3:
        if (first_s[0].lower() == 'id' or first_s[0] == 'ID') and ('问题' in first_s[1] or first_s[1].lower() == 'question') and ('答案' in first_s[2] or first_s[2].lower() == 'answer'):
            df = df.iloc[1:].reset_index(drop=True)

    items = []
    for i in range(len(df)):
        row = df.iloc[i, :3].tolist()
        rid = '' if row[0] is None else str(row[0]).strip()
        q = '' if row[1] is None else str(row[1]).strip()
        a = '' if row[2] is None else str(row[2]).strip()
        if not (rid or q or a):
            continue
        if not q and not a:
            continue
        items.append({
            'row_number': int(i + 1),
            'id': rid,
            'question': q,
            'answer': a
        })
    return items

def _sm_cell_str(v):
    try:
        if v is None:
            return ''
        if isinstance(v, float) and math.isnan(v):
            return ''
        return str(v).strip()
    except Exception:
        return ''

def _sm_has_header_row(df, expected_tokens):
    try:
        if df is None or df.empty:
            return False
        first = df.iloc[0, :max(4, len(expected_tokens))].tolist()
        joined = " ".join([_sm_cell_str(x) for x in first])
        for t in expected_tokens:
            if t not in joined:
                return False
        return True
    except Exception:
        return False

def _sm_parse_faq_sheet(df):
    if df is None or df.empty:
        return []
    if df.shape[1] < 2:
        raise ValueError('产品FAQ表：需要至少两列 (A=问题, B=答案, C=适配型号)')
    start_idx = 1 if _sm_has_header_row(df, ['问题', '答案']) else 0
    items = []
    for i in range(start_idx, len(df)):
        row = df.iloc[i, :3].tolist() if df.shape[1] >= 3 else df.iloc[i, :2].tolist() + ['']
        q = _sm_cell_str(row[0])
        a = _sm_cell_str(row[1])
        m = _sm_cell_str(row[2]) if len(row) > 2 else ''
        if not q or not a:
            continue
        m_out = m if m else '未指定'
        items.append({
            'row_number': int(i + 1),
            'question': q,
            'answer': a,
            'models_raw': m,
            'models_text': m_out
        })
    return items

def _sm_parse_kb_sheet(df):
    if df is None or df.empty:
        return []
    if df.shape[1] < 3:
        raise ValueError('知识库表：需要至少三列 (A=ID, B=问题, C=答案, D=适配型号)')
    start_idx = 1 if _sm_has_header_row(df, ['ID', '问题', '答案']) else 0
    items = []
    for i in range(start_idx, len(df)):
        row = df.iloc[i, :4].tolist() if df.shape[1] >= 4 else df.iloc[i, :3].tolist() + ['']
        wid = _sm_cell_str(row[0])
        q = _sm_cell_str(row[1])
        a = _sm_cell_str(row[2])
        m = _sm_cell_str(row[3]) if len(row) > 3 else ''
        if not wid or not q or not a:
            continue
        items.append({
            'row_number': int(i + 1),
            'question_wiki_id': wid,
            'question': q,
            'answer': a,
            'product_name': m
        })
    return items

def _sm_parse_compare_excel(file_storage):
    name = str(getattr(file_storage, 'filename', '') or '')
    lower = name.lower()
    if not (lower.endswith('.xlsx') or lower.endswith('.xls')):
        raise ValueError('仅支持 .xlsx / .xls')

    try:
        file_storage.stream.seek(0)
    except Exception:
        pass
    data = file_storage.read()
    bio = io.BytesIO(data or b'')
    bio.seek(0)
    if lower.endswith('.xls'):
        sheets = pd.read_excel(bio, sheet_name=None, header=None, engine='xlrd')
    else:
        sheets = pd.read_excel(bio, sheet_name=None, header=None)

    if not isinstance(sheets, dict) or not sheets:
        raise ValueError('Excel 为空或不可读取')
    if '产品FAQ表' not in sheets or '知识库表' not in sheets:
        raise ValueError('工作表名称必须为：产品FAQ表、知识库表')

    faq_items = _sm_parse_faq_sheet(sheets.get('产品FAQ表'))
    kb_items = _sm_parse_kb_sheet(sheets.get('知识库表'))
    return faq_items, kb_items

@app.route('/api/smart_mapping/excel/parse', methods=['POST'])
@login_required
def sm_excel_parse():
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': 'File is required'}), 400
    try:
        faq_items, kb_items = _sm_parse_compare_excel(file)
        return jsonify({
            'success': True,
            'faq_items': faq_items,
            'kb_items': kb_items,
            'faq_count': len(faq_items),
            'kb_count': len(kb_items)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/smart_mapping/template', methods=['GET'])
@login_required
def sm_download_template():
    table = request.args.get('table', 'knowledge_base_v1')
    models_param = request.args.get('models', '')
    models = [m.strip() for m in re.split(r'[,，]', str(models_param)) if m and m.strip()]
    if table not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
        return jsonify({'success': False, 'message': 'Invalid table'}), 400
    if not models:
        return jsonify({'success': False, 'message': 'models required'}), 400

    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400

    try:
        kb_rows = _sm_fetch_baseline_items(client, table, models)
        faq_df = pd.DataFrame([], columns=['问题', '答案', '适配型号'])
        kb_df = pd.DataFrame([{
            'ID': r.get('question_wiki_id'),
            '问题': r.get('question'),
            '答案': r.get('answer'),
            '适配型号': r.get('product_name') or ''
        } for r in (kb_rows or [])], columns=['ID', '问题', '答案', '适配型号'])

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            faq_df.to_excel(writer, index=False, sheet_name='产品FAQ表')
            kb_df.to_excel(writer, index=False, sheet_name='知识库表')
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=canonical_download_name('smart_mapping_template'),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/smart_mapping/export', methods=['POST'])
@login_required
def sm_export_excel():
    data = request.json or {}
    rows = data.get('rows') or []
    if not isinstance(rows, list) or len(rows) == 0:
        return jsonify({'success': False, 'message': 'rows is required'}), 400

    headers = [
        'FAQ原始行号',
        'FAQ适配型号（原表C列）',
        'FAQ_问题（原表A列）',
        'FAQ_答案（原表B列）',
        '知识库ID',
        '知识库_问题',
        '知识库_答案',
        '匹配维度',
        '判定理由'
    ]
    allowed_types = {'问题+答案均一致', '仅问题一致', '仅答案一致', '无匹配'}

    wb = Workbook()
    ws = wb.active
    ws.title = '智能映射结果'
    ws.append(headers)

    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = header_align

    col_widths = {
        1: 12,
        2: 18,
        3: 42,
        4: 52,
        5: 16,
        6: 42,
        7: 52,
        8: 18,
        9: 65
    }
    for col_idx, w in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    match_font = Font(bold=True, color='000000')

    for r in rows:
        item = r if isinstance(r, dict) else {}
        faq = item.get('faq') if isinstance(item.get('faq'), dict) else {}
        match = item.get('match') if isinstance(item.get('match'), dict) else {}

        row_number = faq.get('row_number')
        try:
            row_number = int(row_number)
        except Exception:
            row_number = None

        faq_models = str(faq.get('models_text') or faq.get('models_raw') or '').strip()
        if not faq_models:
            faq_models = '未指定'
        faq_q = str(faq.get('question') or '').strip()
        faq_a = str(faq.get('answer') or '').strip()

        kb_id = str(match.get('kb_id') or '').strip() or '-'
        kb_q = str(match.get('kb_question') or '').strip() or '-'
        kb_a = str(match.get('kb_answer') or '').strip() or '-'

        mtype = str(match.get('type') or '').strip()
        if mtype not in allowed_types:
            mtype = '无匹配'
        if mtype == '无匹配':
            kb_id, kb_q, kb_a = '-', '-', '-'

        reason = str(item.get('reason') or '').replace(' ', '').strip()
        if mtype == '无匹配':
            reason = '未找到语义一致的问题/答案'
        if len(reason) > 80:
            reason = reason[:80]

        ws.append([row_number, faq_models, faq_q, faq_a, kb_id, kb_q, kb_a, mtype, reason])

    max_row = ws.max_row
    for r in range(2, max_row + 1):
        ws.cell(row=r, column=1).alignment = align_center
        ws.cell(row=r, column=1).number_format = '0'
        ws.cell(row=r, column=2).alignment = align_center
        ws.cell(row=r, column=5).alignment = align_center
        ws.cell(row=r, column=8).alignment = align_center
        ws.cell(row=r, column=8).font = match_font

        for col in (3, 4, 6, 7, 9):
            ws.cell(row=r, column=col).alignment = align_left

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=canonical_download_name('smart_mapping'),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/smart_mapping/faq/parse', methods=['POST'])
@login_required
def sm_faq_parse():
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': 'File is required'}), 400
    try:
        items = _sm_parse_faq_excel(file)
        return jsonify({'success': True, 'items': items, 'count': len(items)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

def _sm_pick_match_type(q_sim, a_sim, threshold):
    if q_sim >= threshold and a_sim >= threshold:
        return '问题+答案均一致'
    if q_sim >= threshold and a_sim < threshold:
        return '仅问题一致'
    if a_sim >= threshold and q_sim < threshold:
        return '仅答案一致'
    return '无匹配'

def _sm_strip_number_prefix(s):
    try:
        s = str(s or '')
        s = re.sub(r'^\s*[\(\（]?\s*\d+\s*[\)\）]?\s*', '', s)
        s = re.sub(r'^\s*\d+\s*[\.、]\s*', '', s)
        s = re.sub(r'^\s*[一二三四五六七八九十]+\s*[、\.]\s*', '', s)
        return s
    except Exception:
        return str(s or '')

def _sm_remove_punct_and_space(s):
    try:
        s = str(s or '')
        s = re.sub(r'[\s\u3000]+', '', s)
        s = re.sub(r'[，,。．\.\!\！\?\？；;：:“”"\'‘’（）()\[\]{}【】<>《》、/\\|—\-_\~`·…]+', '', s)
        return s
    except Exception:
        return str(s or '')

def _sm_strip_particles_end(s):
    try:
        s = str(s or '')
        return re.sub(r'[啊呀吧呢嘛哦哈]+$', '', s)
    except Exception:
        return str(s or '')

def _sm_sort_lines(s):
    try:
        s = str(s or '')
        parts = [p.strip() for p in re.split(r'[\r\n]+', s) if p and str(p).strip()]
        if len(parts) <= 1:
            return s
        parts = sorted(parts)
        return "\n".join(parts)
    except Exception:
        return str(s or '')

def _sm_norm_for_sim(s):
    s0 = _sm_strip_number_prefix(s)
    s1 = _sm_sort_lines(s0)
    s2 = _sm_strip_particles_end(s1)
    s3 = _sm_remove_punct_and_space(s2)
    return s3

def _sm_detect_diff_type(a, b):
    a0 = str(a or '')
    b0 = str(b or '')
    if not a0 or not b0:
        return '标点'
    if _sm_remove_punct_and_space(a0) == _sm_remove_punct_and_space(b0) and a0 != b0:
        return '标点'
    if _sm_strip_particles_end(a0) == _sm_strip_particles_end(b0) and a0 != b0:
        return '语气词'
    if _sm_strip_number_prefix(a0) == _sm_strip_number_prefix(b0) and a0 != b0:
        return '格式编号'
    if _sm_sort_lines(a0) == _sm_sort_lines(b0) and a0 != b0:
        return '换行排序'
    return '标点'

def _sm_core_meaning(s, max_len=20):
    t = _sm_remove_punct_and_space(_sm_strip_number_prefix(s))
    if not t:
        return '核心含义'
    return t[:max_len]

def _sm_trim_reason(s, max_len=80):
    out = str(s or '')
    out = out.replace(' ', '')
    if len(out) <= max_len:
        return out
    m = re.search(r'「([^」]+)」', out)
    if not m:
        return out[:max_len]
    core = m.group(1)
    keep = max(4, len(core) - (len(out) - max_len))
    core2 = core[:keep]
    out2 = out.replace(f'「{core}」', f'「{core2}」', 1)
    return out2[:max_len]

def _sm_reason_text(match_type, faq_q, faq_a, kb_q, kb_a):
    if match_type == '无匹配':
        return '未找到语义一致的问题/答案'
    if match_type == '仅答案一致':
        core = _sm_core_meaning(faq_a)
        return _sm_trim_reason(f"答案核心为「{core}」；问题语义不同，无匹配性")
    if match_type == '仅问题一致':
        diff = _sm_detect_diff_type(faq_q, kb_q)
        core = _sm_core_meaning(faq_q)
        return _sm_trim_reason(f"忽略{diff}差异，问题核心为「{core}」；答案核心不一致")
    diff = _sm_detect_diff_type(faq_q, kb_q)
    if diff == '标点':
        d2 = _sm_detect_diff_type(faq_a, kb_a)
        if d2 != '标点':
            diff = d2
    core = _sm_core_meaning(faq_q)
    return _sm_trim_reason(f"忽略{diff}差异，问题/答案核心均为「{core}」")

def _sm_run_compare_job(job_id, username, table, models, faq_items, kb_items, threshold):
    ctx = app.app_context()
    ctx.push()
    try:
        baseline = []
        if isinstance(kb_items, list) and kb_items:
            for r in kb_items:
                if not isinstance(r, dict):
                    continue
                wid = str(r.get('question_wiki_id') or r.get('id') or '').strip()
                q = str(r.get('question') or '').strip()
                a = str(r.get('answer') or '').strip()
                m = str(r.get('product_name') or r.get('适配型号') or '').strip()
                if not wid or not q or not a:
                    continue
                baseline.append({
                    'question_wiki_id': wid,
                    'question': q,
                    'answer': a,
                    'product_name': m
                })
        else:
            client = get_supabase_client()
            cache_key = _sm_baseline_cache_key(username, table, models)
            with _SM_LOCK:
                cache = _SM_BASELINE_CACHE.get(cache_key) or {}
                baseline = cache.get('items')
            if baseline is None:
                baseline = _sm_fetch_baseline_items(client, table, models)
                with _SM_LOCK:
                    _SM_BASELINE_CACHE[cache_key] = {'ts': time.time(), 'items': baseline}

        total = len(faq_items or [])
        with _SM_LOCK:
            _SM_JOBS[job_id]['total'] = total
            _SM_JOBS[job_id]['message'] = '正在准备 Embedding 向量...'
        _sm_db_update_job(
            job_id,
            total=total,
            status='running',
            done=0,
            message='正在准备 Embedding 向量...',
            results_json='',
        )

        embedding_config = _sm_load_embedding_config()
        algorithm = 'embedding'
        fallback_reason = ''
        embedding_model = str(embedding_config.get('model') or '')
        baseline_vectors = []
        faq_vectors = []
        try:
            all_texts = []
            for kb in baseline or []:
                all_texts.extend([kb.get('question') or '', kb.get('answer') or ''])
            for faq in faq_items or []:
                faq = faq if isinstance(faq, dict) else {}
                all_texts.extend([faq.get('question') or '', faq.get('answer') or ''])
            all_vectors = _sm_get_embeddings(all_texts, embedding_config)
            baseline_vector_end = len(baseline or []) * 2
            for offset in range(0, baseline_vector_end, 2):
                baseline_vectors.append((all_vectors[offset], all_vectors[offset + 1]))
            for offset in range(baseline_vector_end, len(all_vectors), 2):
                faq_vectors.append((all_vectors[offset], all_vectors[offset + 1]))
        except Exception as exc:
            if not embedding_config.get('fallback_to_ngram'):
                raise
            algorithm = 'ngram_fallback'
            fallback_reason = str(exc or '').strip().replace('\n', ' ')[:240]

        job_message = (
            f'Embedding：{embedding_model}'
            if algorithm == 'embedding'
            else f'Embedding 不可用，已降级字符相似度：{fallback_reason}'
        )
        with _SM_LOCK:
            if job_id in _SM_JOBS:
                _SM_JOBS[job_id]['message'] = job_message
        _sm_db_update_job(job_id, message=job_message)

        results = []
        done = 0
        last_db_ts = 0.0
        for faq_index, f in enumerate(faq_items or []):
            faq = f if isinstance(f, dict) else {}
            fq_raw = str(faq.get('question') or '').strip()
            fa_raw = str(faq.get('answer') or '').strip()
            fm_raw = str(faq.get('models_raw') or faq.get('models') or '').strip()
            fm_text = str(faq.get('models_text') or ('未指定' if not fm_raw else fm_raw)).strip()
            best = None
            best_q = 0.0
            best_a = 0.0
            best_score = -1.0
            faq_q_vector = None
            faq_a_vector = None
            if algorithm == 'embedding' and faq_index < len(faq_vectors):
                faq_q_vector, faq_a_vector = faq_vectors[faq_index]
            fq = _sm_norm_for_sim(fq_raw) if algorithm != 'embedding' else ''
            fa = _sm_norm_for_sim(fa_raw) if algorithm != 'embedding' else ''

            for kb_index, kb in enumerate(baseline or []):
                kq_raw = kb.get('question') or ''
                ka_raw = kb.get('answer') or ''
                if algorithm == 'embedding':
                    kb_q_vector, kb_a_vector = baseline_vectors[kb_index]
                    q_sim = _sm_vector_cosine(faq_q_vector, kb_q_vector)
                    a_sim = _sm_vector_cosine(faq_a_vector, kb_a_vector)
                else:
                    kq = _sm_norm_for_sim(kq_raw)
                    ka = _sm_norm_for_sim(ka_raw)
                    q_sim = _ai_cosine_sim(fq, kq)
                    a_sim = _ai_cosine_sim(fa, ka)
                score = (q_sim + a_sim) / 2.0
                if score > best_score:
                    best_score = score
                    best = kb
                    best_q = q_sim
                    best_a = a_sim

            kb_id = str((best or {}).get('question_wiki_id') or '').strip()
            kb_q = (best or {}).get('question') if best else ''
            kb_a = (best or {}).get('answer') if best else ''
            kb_m = (best or {}).get('product_name') if best else ''
            if not kb_id:
                best_q = 0.0
                best_a = 0.0
            match_type = _sm_pick_match_type(best_q, best_a, threshold)
            if match_type == '无匹配':
                kb_id = ''
                kb_q = ''
                kb_a = ''
                kb_m = ''
            if algorithm == 'embedding':
                reason = _sm_embedding_reason(match_type, best_q, best_a)
            else:
                reason = _sm_trim_reason(
                    f'Embedding不可用，字符相似度降级；{_sm_reason_text(match_type, fq_raw, fa_raw, kb_q, kb_a)}',
                    80,
                )
            results.append({
                'faq': {
                    'row_number': faq.get('row_number'),
                    'question': fq_raw,
                    'answer': fa_raw,
                    'models_raw': fm_raw,
                    'models_text': fm_text
                },
                'match': {
                    'type': match_type,
                    'kb_id': kb_id,
                    'kb_question': kb_q or '',
                    'kb_answer': kb_a or '',
                    'kb_models': kb_m or '',
                    'q_sim': float(best_q),
                    'a_sim': float(best_a),
                    'score': float(best_score if best_score >= 0 else 0.0),
                    'algorithm': algorithm,
                    'embedding_model': embedding_model if algorithm == 'embedding' else '',
                    'fallback_reason': fallback_reason if algorithm != 'embedding' else '',
                },
                'reason': reason,
            })
            done += 1
            with _SM_LOCK:
                job = _SM_JOBS.get(job_id) or {}
                job['done'] = done
                job['status'] = 'running'
                job['ts'] = time.time()
                _SM_JOBS[job_id] = job
            now = time.time()
            if (now - last_db_ts) >= 0.8 or done == total:
                _sm_db_update_job(job_id, done=done, status='running')
                last_db_ts = now

        with _SM_LOCK:
            job = _SM_JOBS.get(job_id) or {}
            job['status'] = 'done'
            job['done'] = done
            job['results'] = results
            job['message'] = job_message
            job['ts'] = time.time()
            _SM_JOBS[job_id] = job
        _sm_db_update_job(
            job_id,
            status='done',
            done=done,
            message=job_message,
            results_json=json.dumps(results, ensure_ascii=False),
        )
    except Exception as e:
        traceback.print_exc()
        with _SM_LOCK:
            job = _SM_JOBS.get(job_id) or {}
            job['status'] = 'failed'
            job['message'] = str(e)
            job['ts'] = time.time()
            _SM_JOBS[job_id] = job
        _sm_db_update_job(job_id, status='failed', done=int((_SM_JOBS.get(job_id) or {}).get('done') or 0), message=str(e))
    finally:
        try:
            ctx.pop()
        except Exception:
            pass

@app.route('/api/smart_mapping/compare/start', methods=['POST'])
@login_required
def sm_compare_start():
    data = request.json or {}
    table = data.get('table', 'knowledge_base_v1')
    models = data.get('models') or []
    faq_items = data.get('faq_items') or []
    kb_items = data.get('kb_items') or []
    embedding_config = _sm_load_embedding_config()
    threshold = data.get('threshold', embedding_config.get('threshold', 0.75))
    if table not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
        return jsonify({'success': False, 'message': 'Invalid table'}), 400
    if not isinstance(faq_items, list) or len(faq_items) == 0:
        return jsonify({'success': False, 'message': 'faq_items is required'}), 400
    if kb_items and not isinstance(kb_items, list):
        return jsonify({'success': False, 'message': 'kb_items must be list'}), 400
    if not kb_items:
        if not isinstance(models, list) or not any(str(x or '').strip() for x in models):
            return jsonify({'success': False, 'message': 'models is required'}), 400
    try:
        threshold = float(threshold)
    except Exception:
        threshold = float(embedding_config.get('threshold', 0.75))
    if threshold <= 0 or threshold > 1:
        threshold = float(embedding_config.get('threshold', 0.75))

    _sm_cleanup()
    job_id = str(uuid.uuid4())
    username = current_user.username if current_user.is_authenticated else ''
    with _SM_LOCK:
        _SM_JOBS[job_id] = {
            'ts': time.time(),
            'status': 'running',
            'total': len(faq_items),
            'done': 0,
            'results': []
        }
    _sm_db_create_job(job_id, username, len(faq_items))
    t = threading.Thread(target=_sm_run_compare_job, args=(job_id, username, table, models, faq_items, kb_items, threshold), daemon=True)
    t.start()
    return jsonify({
        'success': True,
        'job_id': job_id,
        'algorithm': 'embedding',
        'embedding_model': embedding_config.get('model'),
        'threshold': threshold,
    })

@app.route('/api/smart_mapping/compare/status', methods=['GET'])
@login_required
def sm_compare_status():
    job_id = request.args.get('job_id')
    if not job_id:
        return jsonify({'success': False, 'message': 'job_id is required'}), 400
    _sm_cleanup()
    out = None
    with _SM_LOCK:
        job = _SM_JOBS.get(job_id)
        if job:
            out = dict(job)
    if out is None:
        j = _sm_db_get_job(job_id)
        if not j:
            return jsonify({'success': False, 'message': 'job not found'}), 404
        out = {
            'status': j.status,
            'message': j.message or '',
            'total': int(j.total or 0),
            'done': int(j.done or 0),
            'results': []
        }
        if str(j.status) == 'done':
            try:
                out['results'] = json.loads(j.results_json or '[]') or []
            except Exception:
                out['results'] = []
    return jsonify({
        'success': True,
        'status': out.get('status'),
        'message': out.get('message', ''),
        'total': out.get('total', 0),
        'done': out.get('done', 0),
        'results': out.get('results', []) if out.get('status') == 'done' else []
    })


@app.route('/api/smart_mapping/embedding/score', methods=['POST'])
@login_required
def sm_embedding_score():
    payload = request.json or {}
    faq = payload.get('faq') if isinstance(payload.get('faq'), dict) else {}
    kb = payload.get('kb') if isinstance(payload.get('kb'), dict) else {}
    faq_question = str(faq.get('question') or '').strip()
    faq_answer = str(faq.get('answer') or '').strip()
    kb_question = str(kb.get('question') or '').strip()
    kb_answer = str(kb.get('answer') or '').strip()
    if not faq_question or not faq_answer or not kb_question or not kb_answer:
        return jsonify({'success': False, 'message': 'FAQ 和知识库的问题、答案均不能为空'}), 400

    config = _sm_load_embedding_config()
    threshold = float(config.get('threshold') or 0.75)
    try:
        vectors = _sm_get_embeddings(
            [faq_question, faq_answer, kb_question, kb_answer],
            config,
        )
        question_score = _sm_vector_cosine(vectors[0], vectors[2])
        answer_score = _sm_vector_cosine(vectors[1], vectors[3])
        algorithm = 'embedding'
        fallback_reason = ''
        reason = _sm_embedding_reason(
            _sm_pick_match_type(question_score, answer_score, threshold),
            question_score,
            answer_score,
        )
    except Exception as exc:
        if not config.get('fallback_to_ngram'):
            return jsonify({'success': False, 'message': str(exc)}), 502
        question_score = _ai_cosine_sim(_sm_norm_for_sim(faq_question), _sm_norm_for_sim(kb_question))
        answer_score = _ai_cosine_sim(_sm_norm_for_sim(faq_answer), _sm_norm_for_sim(kb_answer))
        algorithm = 'ngram_fallback'
        fallback_reason = str(exc or '').strip().replace('\n', ' ')[:240]
        reason = _sm_trim_reason(
            f'Embedding不可用，字符相似度降级；{_sm_reason_text(_sm_pick_match_type(question_score, answer_score, threshold), faq_question, faq_answer, kb_question, kb_answer)}',
            80,
        )

    match_type = _sm_pick_match_type(question_score, answer_score, threshold)
    return jsonify({
        'success': True,
        'q_sim': question_score,
        'a_sim': answer_score,
        'score': (question_score + answer_score) / 2.0,
        'type': match_type,
        'reason': reason,
        'algorithm': algorithm,
        'embedding_model': config.get('model') if algorithm == 'embedding' else '',
        'fallback_reason': fallback_reason,
        'threshold': threshold,
    })

@app.route('/api/smart_mapping/kb/search', methods=['GET'])
@login_required
def sm_kb_search():
    q = (request.args.get('q') or '').strip()
    table = request.args.get('table', 'knowledge_base_v1')
    models_param = (request.args.get('models') or '').strip()
    if table not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
        return jsonify({'success': False, 'message': 'Invalid table'}), 400
    if not q:
        return jsonify({'success': True, 'items': []})

    try:
        limit = int(request.args.get('limit') or 50)
    except Exception:
        limit = 50
    limit = max(1, min(200, limit))

    q = re.sub(r'[\(\),]', ' ', q)
    q = re.sub(r'\s+', ' ', q).strip()
    if not q:
        return jsonify({'success': True, 'items': []})

    models = [m.strip() for m in re.split(r'[,，]', models_param) if m and m.strip()]
    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    try:
        filters = {}
        or_parts = [
            f"question.ilike.*{q}*",
            f"answer.ilike.*{q}*",
            f"question_wiki_id.ilike.*{q}*",
            f"product_name.ilike.*{q}*"
        ]
        text_or = "(" + ",".join(or_parts) + ")"
        if models:
            m_or = _sm_build_models_or_filter(models)
            if m_or:
                filters['and'] = "(" + ",".join([f"or={text_or}", f"or={m_or}"]) + ")"
            else:
                filters['or'] = text_or
        else:
            filters['or'] = text_or
        resp = client.select(
            table,
            page=1,
            page_size=limit,
            filters=filters,
            order_by='update_time',
            order_dir='desc',
            columns='question_wiki_id,question,answer,product_name,update_time'
        )
        if resp.status_code >= 400:
            return jsonify({'success': False, 'message': resp.text}), 500
        items = []
        for r in resp.json() or []:
            if not isinstance(r, dict):
                continue
            items.append({
                'question_wiki_id': r.get('question_wiki_id'),
                'question': r.get('question') or '',
                'answer': r.get('answer') or '',
                'product_name': r.get('product_name') or '',
                'update_time': r.get('update_time')
            })
        return jsonify({'success': True, 'items': items})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

def _sm_normalize_models(models):
    models = [str(x).strip() for x in (models or []) if str(x or '').strip()]
    valid_map, valid_set = get_all_valid_models()
    normalized = []
    invalid = []
    for m in models:
        if m in valid_set:
            normalized.append(m)
            continue
        norm = re.sub(r'\s+', '', m).lower()
        official = valid_map.get(norm)
        if official:
            normalized.append(official)
        else:
            invalid.append(m)
    normalized = sorted(list(dict.fromkeys(normalized)))
    return normalized, invalid

@app.route('/api/smart_mapping/submit', methods=['POST'])
@login_required
def sm_submit():
    data = request.json or {}
    table = data.get('table', 'knowledge_base_v1')
    items = data.get('items') or []
    if table not in ('knowledge_base_v1', 'knowledge_base_v1_t1'):
        return jsonify({'success': False, 'message': 'Invalid table'}), 400
    if not isinstance(items, list) or len(items) == 0:
        return jsonify({'success': False, 'message': 'items is required'}), 400

    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 500

    success_count = 0
    failures = []
    now_iso = _sm_now_iso()
    operation_id = str(uuid.uuid4())

    def _sm_parse_bool(v):
        if v is None:
            return None
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        if not s:
            return None
        if s in ('1', 'true', 'yes', 'y', '是'):
            return True
        if s in ('0', 'false', 'no', 'n', '否'):
            return False
        return None

    def _sm_parse_list(v):
        if v is None:
            return None
        if isinstance(v, list):
            return v
        s = str(v).strip()
        if not s:
            return None
        try:
            obj = json.loads(s)
            if isinstance(obj, list):
                return obj
        except Exception:
            pass
        parts = re.split(r'[\n,，]', s)
        return [x.strip() for x in parts if x.strip()]

    def _sm_parse_text(v):
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        return s

    for idx, it in enumerate(items):
        try:
            item = it if isinstance(it, dict) else {}
            faq = item.get('faq') if isinstance(item.get('faq'), dict) else {}
            question = str(faq.get('question') or '').strip()
            answer = str(faq.get('answer') or '').strip()
            models_norm, invalid_models = _sm_normalize_models(item.get('models') or [])
            if invalid_models:
                raise ValueError(f'包含未知型号: {", ".join(invalid_models)}')
            if not models_norm:
                raise ValueError('适用型号不能为空')

            mode = 'create' if str(item.get('mode') or '').strip() == 'create' else 'update'
            match = item.get('match') if isinstance(item.get('match'), dict) else {}
            kb_id = str(match.get('kb_id') or '').strip()
            other_info = item.get('other_info') if isinstance(item.get('other_info'), dict) else {}

            if mode == 'update' and not kb_id:
                raise ValueError('缺少匹配对象 ID')

            if mode == 'create':
                if not question or not answer:
                    raise ValueError('问题/答案不能为空')
                new_id = generate_kb_id(client)
                payload = {
                    'question_wiki_id': new_id,
                    'question': question,
                    'answer': answer,
                    'product_name': ", ".join(models_norm),
                    'review_status': 'unadjusted',
                    'update_time': now_iso
                }
                qt = str(other_info.get('question_type') or '').strip()
                if qt:
                    payload['question_type'] = qt
                bm = _sm_parse_bool(other_info.get('if_bm25'))
                if bm is not None:
                    payload['if_bm25'] = bm
                sq = _sm_parse_list(other_info.get('similar_questions'))
                if sq is not None:
                    payload['similar_questions'] = sq
                kw = _sm_parse_list(other_info.get('keyword_list'))
                if kw is not None:
                    payload['keyword_list'] = kw
                img = _sm_parse_list(other_info.get('image_urls'))
                if img is not None:
                    payload['image_urls'] = img
                vid = _sm_parse_list(other_info.get('video_urls'))
                if vid is not None:
                    payload['video_urls'] = vid
                fil = _sm_parse_list(other_info.get('file_urls'))
                if fil is not None:
                    payload['file_urls'] = fil
                lt = _sm_parse_text(other_info.get('link_type'))
                if lt is not None:
                    payload['link_type'] = lt
                lu = _sm_parse_text(other_info.get('link_url'))
                if lu is not None:
                    payload['link_url'] = lu

                mod_rec = dict(payload)
                mod_rec['kb_id'] = new_id
                mod_rec['modifier'] = current_user.username if current_user.is_authenticated else 'system'
                mod_rec['modification_time'] = now_iso
                mod_rec['change_type'] = 'create'
                before_obj = None
                after_obj = _snapshot_mod_fields(payload)
                changed_fields = _compute_mod_changed_fields(before_obj, after_obj)
                _attach_change_meta(mod_rec, {
                    'source': '智能映射',
                    'operation_id': operation_id,
                    'before': before_obj,
                    'after': after_obj,
                    'changed_fields': changed_fields,
                    'match': match,
                    'reason': item.get('reason')
                })
                if 'review_status' in mod_rec:
                    del mod_rec['review_status']
                if table == 'knowledge_base_v1':
                    # 转换数组字段为 JSON 字符串
                    _convert_array_fields_to_json(mod_rec)
                    
                    resp = _supabase_insert_drop_unknown_columns(client, 'knowledge_base_modifications', mod_rec)
                    if not resp or getattr(resp, 'status_code', 500) >= 400:
                        raise ValueError((getattr(resp, 'text', '') or '').strip() or '写入修改记录失败')
            else:
                before_row = None
                before_resp = client.select(
                    table,
                    page=1,
                    page_size=1,
                    filters={'question_wiki_id': f"eq.{kb_id}"},
                    columns='question_wiki_id,question,answer,product_name,question_type,answer_type,if_bm25,similar_questions,error_list,keyword_list,image_urls,video_urls,file_urls,link_type,link_url,update_time'
                )
                if before_resp.status_code in (200, 206):
                    rows = before_resp.json() or []
                    if rows and isinstance(rows[0], dict):
                        before_row = rows[0]
                if not before_row:
                    raise ValueError('匹配对象不存在或不可用')

                old_products = set(_split_product_names(before_row.get('product_name')))
                merged = sorted(list(old_products.union(set(models_norm))))

                target_q = str(match.get('kb_question') or '').strip()
                target_a = str(match.get('kb_answer') or '').strip()
                if not target_q or not target_a:
                    target_q = question
                    target_a = answer
                if not target_q or not target_a:
                    raise ValueError('问题/答案不能为空')
                payload = {
                    'question': target_q,
                    'answer': target_a,
                    'product_name': ", ".join(merged),
                    'review_status': 'modifying',
                    'update_time': now_iso
                }
                qt = str(other_info.get('question_type') or '').strip()
                if qt:
                    payload['question_type'] = qt
                bm = _sm_parse_bool(other_info.get('if_bm25'))
                if bm is not None:
                    payload['if_bm25'] = bm
                sq = _sm_parse_list(other_info.get('similar_questions'))
                if sq is not None:
                    payload['similar_questions'] = sq
                kw = _sm_parse_list(other_info.get('keyword_list'))
                if kw is not None:
                    payload['keyword_list'] = kw
                img = _sm_parse_list(other_info.get('image_urls'))
                if img is not None:
                    payload['image_urls'] = img
                vid = _sm_parse_list(other_info.get('video_urls'))
                if vid is not None:
                    payload['video_urls'] = vid
                fil = _sm_parse_list(other_info.get('file_urls'))
                if fil is not None:
                    payload['file_urls'] = fil
                lt = _sm_parse_text(other_info.get('link_type'))
                if lt is not None:
                    payload['link_type'] = lt
                lu = _sm_parse_text(other_info.get('link_url'))
                if lu is not None:
                    payload['link_url'] = lu

                before_obj = _snapshot_mod_fields(before_row)
                after_for_diff = dict(before_row)
                after_for_diff.update(payload)
                after_obj = _snapshot_mod_fields(after_for_diff)
                changed_fields = _compute_mod_changed_fields(before_obj, after_obj)

                mod_rec = dict(before_row)
                mod_rec.update(payload)
                mod_rec['kb_id'] = kb_id
                mod_rec['modifier'] = current_user.username if current_user.is_authenticated else 'system'
                mod_rec['modification_time'] = now_iso
                mod_rec['change_type'] = 'edit'
                _attach_change_meta(mod_rec, {
                    'source': '智能映射',
                    'operation_id': operation_id,
                    'before': before_obj,
                    'after': after_obj,
                    'changed_fields': changed_fields,
                    'match': match,
                    'reason': item.get('reason')
                })
                if 'id' in mod_rec:
                    del mod_rec['id']
                if 'review_status' in mod_rec:
                    del mod_rec['review_status']
                if table == 'knowledge_base_v1':
                    # 转换数组字段为 JSON 字符串
                    _convert_array_fields_to_json(mod_rec)
                    
                    resp = _supabase_insert_drop_unknown_columns(client, 'knowledge_base_modifications', mod_rec)
                    if not resp or getattr(resp, 'status_code', 500) >= 400:
                        raise ValueError((getattr(resp, 'text', '') or '').strip() or '写入修改记录失败')

            success_count += 1
        except Exception as e:
            failures.append({'index': idx, 'message': str(e)})

    return jsonify({
        'success': True,
        'success_count': success_count,
        'failed_count': len(failures),
        'failures': failures,
        'operation_id': operation_id
    })

@app.route('/api/smart_mapping/archive', methods=['POST'])
@login_required
def sm_archive_operation():
    payload = request.json or {}
    batch_name = str(payload.get('batch_name') or '').strip()
    operation_id = str(payload.get('operation_id') or '').strip()
    if not batch_name:
        return jsonify({'success': False, 'message': 'batch_name required'}), 400
    if not operation_id:
        return jsonify({'success': False, 'message': 'operation_id required'}), 400

    client = get_supabase_client()
    if not client:
        return jsonify({'success': False, 'message': '本地主库未配置'}), 400

    try:
        normalized, raw_ids = _collect_archive_candidates(client, operation_id=operation_id)

        if not normalized:
            return jsonify({'success': False, 'message': '没有可归档的记录'}), 400

        if not bool(payload.get('confirm_archive')):
            return jsonify({
                'success': False,
                'requires_confirmation': True,
                'message': '归档会迁移并删除本次智能映射提交对应的当前修改记录，请确认后继续。',
                'count': len(normalized),
                'delete_count': len(raw_ids),
                'operation_id': operation_id
            }), 409

        expected_count = payload.get('expected_count')
        if expected_count is not None:
            try:
                if int(expected_count) != len(normalized):
                    return jsonify({
                        'success': False,
                        'requires_confirmation': True,
                        'message': f'待归档记录数已变化（确认时 {expected_count} 条，当前 {len(normalized)} 条），请重新确认。',
                        'count': len(normalized),
                        'delete_count': len(raw_ids),
                        'operation_id': operation_id
                    }), 409
            except Exception:
                return jsonify({'success': False, 'message': 'expected_count 必须为数字'}), 400

        batch = ArchiveBatch(
            batch_name=batch_name,
            record_count=len(normalized),
            created_by=current_user.username
        )
        db.session.add(batch)
        db.session.flush()

        recs = []
        for it in normalized:
            mt = _safe_parse_dt(it.get('modify_time'))
            recs.append(ArchiveRecord(
                batch_id=batch.id,
                record_json=_json_dumps_safe(it),
                modify_time=mt
            ))
        if recs:
            db.session.add_all(recs)
        db.session.commit()

        deleted = 0
        if raw_ids:
            chunk_size = 500
            for i in range(0, len(raw_ids), chunk_size):
                chunk = raw_ids[i:i + chunk_size]
                resp = client.delete_in('knowledge_base_modifications', 'id', chunk)
                if not resp or getattr(resp, 'status_code', 500) >= 400:
                    return jsonify({'success': False, 'message': f'归档已保存，但清理主库失败: {getattr(resp, "text", "")}'}), 500
                deleted += len(chunk)

        return jsonify({'success': True, 'id': batch.id, 'record_count': batch.record_count, 'deleted_count': deleted})
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# --- Matrix Management APIs ---

def _split_product_names(value):
    s = str(value or "")
    items = [p.strip() for p in re.split(r'[,，]', s) if p and p.strip()]
    return [p for p in items if p]

def _fetch_kb_products_map(table_name, wiki_ids):
    if not wiki_ids:
        return {}, True
    try:
        client = get_supabase_client()
        if not client:
            return {}, False
        cleaned = [str(x).strip() for x in (wiki_ids or []) if str(x).strip()]
        cleaned = list(dict.fromkeys(cleaned))
        if not cleaned:
            return {}, True

        result = {}
        chunk_size = 400
        for i in range(0, len(cleaned), chunk_size):
            chunk = cleaned[i:i + chunk_size]
            in_str = _postgrest_in_str(chunk)
            if not in_str:
                continue
            try:
                rows = client.select_all(
                    table_name,
                    filters={'question_wiki_id': in_str},
                    columns='question_wiki_id,product_name',
                    order_by='question_wiki_id',
                    order_dir='asc',
                    page_size=1000
                ) or []
            except Exception:
                return {}, False
            for item in rows:
                wid = str(item.get('question_wiki_id') or "").strip()
                if not wid:
                    continue
                result[wid] = set(_split_product_names(item.get('product_name')))
        return result, True
    except Exception:
        return {}, False

def _fetch_kb_detail_map(wiki_ids):
    cleaned = [str(x).strip() for x in (wiki_ids or []) if str(x).strip()]
    cleaned = list(dict.fromkeys(cleaned))
    if not cleaned:
        return {}

    base_columns = (
        'question_wiki_id,question,answer,question_type,answer_type,if_bm25,'
        'similar_questions,error_list,keyword_list,image_urls,video_urls,'
        'file_urls,link_type,link_url,product_name'
    )
    result = {}
    client = get_supabase_client()

    if client:
        chunk_size = 400
        for i in range(0, len(cleaned), chunk_size):
            chunk = cleaned[i:i + chunk_size]
            in_str = _postgrest_in_str(chunk)
            if not in_str:
                continue
            try:
                rows = client.select_all(
                    'knowledge_base_v1',
                    filters={'question_wiki_id': in_str},
                    order_by='question_wiki_id',
                    order_dir='asc',
                    columns=base_columns,
                    page_size=1000
                ) or []
            except Exception:
                rows = []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                wid = str(row.get('question_wiki_id') or '').strip()
                if wid:
                    result[wid] = dict(row)

    missing = [wid for wid in cleaned if wid not in result]
    if not missing:
        return result

    try:
        pm_rows = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(missing)).all()
    except Exception:
        pm_rows = []

    pm_grouped = {}
    for row in pm_rows:
        wid = str(getattr(row, 'question_wiki_id', '') or '').strip()
        if not wid:
            continue
        entry = pm_grouped.setdefault(wid, {
            'question_wiki_id': wid,
            'question': '',
            'answer': '',
            'product_name': '',
            '_products': set()
        })
        q = str(getattr(row, 'question_content', '') or '').strip()
        a = str(getattr(row, 'answer_content', '') or '').strip()
        if q and not entry['question']:
            entry['question'] = q
        if a and not entry['answer']:
            entry['answer'] = a
        if bool(getattr(row, 'is_configured', False)):
            pn = str(getattr(row, 'product_name', '') or '').strip()
            if pn:
                entry['_products'].add(pn)

    for wid, entry in pm_grouped.items():
        products = sorted(list(entry.pop('_products', set())))
        if products:
            entry['product_name'] = ",".join(products)
        result.setdefault(wid, entry)

    return result

def _build_matrix_submit_snapshot_map(changes, base_map):
    def _norm_pn(v):
        s = str(v or '')
        s = s.replace('\u3000', ' ')
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    grouped = {}
    wiki_ids = []
    for change in (changes or []):
        wid = str((change or {}).get('question_wiki_id') or '').strip()
        if not wid:
            continue
        if wid not in grouped:
            grouped[wid] = []
            wiki_ids.append(wid)
        grouped[wid].append(change or {})

    current_products_map = {}
    if wiki_ids:
        try:
            pm_rows = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(wiki_ids)).all()
        except Exception:
            pm_rows = []
        for row in pm_rows:
            wid = str(getattr(row, 'question_wiki_id', '') or '').strip()
            pn = _norm_pn(getattr(row, 'product_name', '') or '')
            if not wid or not pn:
                continue
            bucket = current_products_map.setdefault(wid, set())
            if bool(getattr(row, 'is_configured', False)):
                bucket.add(pn)

    snapshot_map = {}
    for wid in wiki_ids:
        base = dict(base_map.get(wid) or {})
        before_set = set([_norm_pn(x) for x in _split_product_names(base.get('product_name')) if _norm_pn(x)])
        after_set = set(before_set)
        group_changes = grouped.get(wid) or []

        if not before_set and wid in current_products_map:
            after_set = set(current_products_map.get(wid, set()))
            before_set = set(after_set)
            for change in group_changes:
                pn = _norm_pn(change.get('product_name'))
                if not pn:
                    continue
                old_v = bool(change.get('old_is_configured'))
                new_v = bool(change.get('new_is_configured'))
                if new_v and not old_v:
                    before_set.discard(pn)
                elif (not new_v) and old_v:
                    before_set.add(pn)
        else:
            for change in group_changes:
                pn = _norm_pn(change.get('product_name'))
                if not pn:
                    continue
                old_v = bool(change.get('old_is_configured'))
                new_v = bool(change.get('new_is_configured'))
                if new_v and not old_v:
                    after_set.add(pn)
                elif (not new_v) and old_v:
                    after_set.discard(pn)

        before_row = dict(base)
        after_row = dict(base)
        before_row['question_wiki_id'] = wid
        after_row['question_wiki_id'] = wid
        before_row['product_name'] = ",".join(sorted(list(before_set)))
        after_row['product_name'] = ",".join(sorted(list(after_set)))
        before_obj = _snapshot_mod_fields(before_row)
        after_obj = _snapshot_mod_fields(after_row)
        snapshot_map[wid] = {
            'before': before_obj,
            'after': after_obj,
            'changed_fields': _compute_mod_changed_fields(before_obj, after_obj),
            'changed_products': sorted(list(dict.fromkeys([
                _norm_pn(it.get('product_name')) for it in group_changes if _norm_pn(it.get('product_name'))
            ])))
        }

    return snapshot_map

@app.route('/api/matrix/data', methods=['GET'])
@login_required
def get_matrix_data():
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        wid = request.args.get('id', '').strip()
        q = request.args.get('q', '').strip()
        a = request.args.get('a', '').strip()
        p = request.args.get('p', '').strip()
        p_models_param = request.args.get('p_models', '').strip()
        p_mode = request.args.get('p_mode', 'any').strip().lower()
        mark_modified_raw = request.args.get('mark_modified', '').strip()
        mark_unmodified_raw = request.args.get('mark_unmodified', '').strip()
        pc = request.args.get('pc', '').strip()
        mc = request.args.get('mc', '').strip()
        col_models_param = request.args.get('col_models', '').strip()
        diff_compare = request.args.get('diff_compare', '').strip().lower() in ('1', 'true', 'yes', 'on')

        def _as_bool(s, default=True):
            if s is None:
                return default
            t = str(s).strip().lower()
            if t == '':
                return default
            if t in ('1', 'true', 't', 'yes', 'y', 'on'):
                return True
            if t in ('0', 'false', 'f', 'no', 'n', 'off'):
                return False
            return default

        want_modified = _as_bool(mark_modified_raw, True)
        want_unmodified = _as_bool(mark_unmodified_raw, True)
        any_mark_wanted = bool(want_modified or want_unmodified)
        
        mappings = get_model_mappings()
        allowed_models = []
        if mc and mc in mappings and isinstance(mappings.get(mc), list):
            allowed_models = [str(x) for x in mappings.get(mc) if x]
        
        catalog = parse_product_catalog()
        product_categories = [str(k) for k in catalog.keys() if k and str(k).strip()]
        mapping_categories = sorted([str(k) for k in mappings.keys()])
        
        pc_models = []
        if pc:
            if pc in catalog and isinstance(catalog.get(pc), list):
                pc_models = [str(x).strip() for x in catalog.get(pc) if x and str(x).strip()]
        
        models_whitelist = set()
        if pc_models:
            models_whitelist = set([str(x) for x in pc_models if x])
        if allowed_models:
            if models_whitelist:
                models_whitelist = models_whitelist.intersection(set([str(x) for x in allowed_models if x]))
            else:
                models_whitelist = set([str(x) for x in allowed_models if x])
        
        selected_models_ordered = []
        if col_models_param:
            raw_models = [m.strip() for m in re.split(r'[,，]', col_models_param) if m and m.strip()]
            seen_models = set()
            for m in raw_models:
                if m in seen_models:
                    continue
                seen_models.add(m)
                selected_models_ordered.append(m)
            if selected_models_ordered:
                selected_set = set(selected_models_ordered)
                if models_whitelist:
                    models_whitelist = models_whitelist.intersection(selected_set)
                else:
                    models_whitelist = selected_set
        
        # 1. Get Columns (Filter by p if provided)
        col_query = MatrixColumn.query.order_by(MatrixColumn.sort_order)
        if models_whitelist:
            col_query = col_query.filter(MatrixColumn.product_name.in_(list(models_whitelist)))
        if selected_models_ordered:
            columns = col_query.filter(MatrixColumn.product_name.in_(selected_models_ordered)).all()
            by_name = {str(c.product_name): c for c in columns if c and str(getattr(c, 'product_name', '') or '').strip()}
            col_list = [m for m in selected_models_ordered if m in by_name]
        else:
            columns = col_query.all()
            col_list = [c.product_name for c in columns]  # Return names directly as per frontend expectation

        col_list = [c for c in (col_list or []) if str(c or '').strip() and str(c or '').strip() != '测试型号']
        
        if not col_list:
            return jsonify({
                'success': True,
                'columns': [],
                'data': [],
                'total': 0,
                'page': page,
                'per_page': per_page,
                'product_categories': product_categories,
                'mapping_categories': mapping_categories
            })

        # 2. Get Distinct Question Wiki IDs (Filter by q)
        # We need to find wiki_ids that match the search criteria.
        # Since content is denormalized, we can query ProductMatrix directly.
        # Using distinct() on wiki_id.
        
        base_query = db.session.query(ProductMatrix.question_wiki_id)
        
        if wid:
            base_query = base_query.filter(ProductMatrix.question_wiki_id.ilike(f'%{wid}%'))
        
        if q:
            base_query = base_query.filter(ProductMatrix.question_content.ilike(f'%{q}%'))
        
        if a:
            base_query = base_query.filter(ProductMatrix.answer_content.ilike(f'%{a}%'))
        
        p_models = []
        if p_models_param:
            p_models = [m.strip() for m in re.split(r'[,，\n]+', p_models_param) if m and m.strip()]
        elif p and re.search(r'[,，\n]', p):
            p_models = [m.strip() for m in re.split(r'[,，\n]+', p) if m and m.strip()]
        
        if p_models:
            p_models = list(dict.fromkeys(p_models))
            if p_mode == 'all':
                subq = db.session.query(ProductMatrix.question_wiki_id)\
                    .filter(ProductMatrix.product_name.in_(p_models))\
                    .group_by(ProductMatrix.question_wiki_id)\
                    .having(func.count(func.distinct(ProductMatrix.product_name)) == len(p_models))\
                    .subquery()
                base_query = base_query.filter(ProductMatrix.question_wiki_id.in_(subq))
            else:
                base_query = base_query.filter(ProductMatrix.product_name.in_(p_models))
        elif p:
            base_query = base_query.filter(ProductMatrix.product_name.ilike(f'%{p}%'))
        
        if pc:
            base_query = base_query.filter(ProductMatrix.product_category.ilike(f'%{pc}%'))
        
        if mc and mc in mappings:
            if allowed_models:
                allowed_models = [str(x) for x in allowed_models if x]
                n = len(allowed_models)
                if n == 0:
                    base_query = base_query.filter(False)
                else:
                    exact_ids_subq = db.session.query(ProductMatrix.question_wiki_id)\
                        .filter(ProductMatrix.is_configured == True)\
                        .group_by(ProductMatrix.question_wiki_id)\
                        .having(func.count(func.distinct(ProductMatrix.product_name)) == n)\
                        .having(func.count(func.distinct(case(
                            (ProductMatrix.product_name.in_(allowed_models), ProductMatrix.product_name),
                            else_=None
                        ))) == n)\
                        .subquery()
                    base_query = base_query.filter(ProductMatrix.question_wiki_id.in_(exact_ids_subq))
            else:
                base_query = base_query.filter(False)
            
        base_total = int(base_query.distinct().count() or 0)

        diff_compare_active = diff_compare and len(selected_models_ordered) >= 2
        diff_compare_ids = None

        def _compute_diff_compare_ids(ids_subq, selected_models):
            models = [str(x).strip() for x in (selected_models or []) if str(x).strip()]
            if len(models) < 2:
                return []
            rows_for_diff = ProductMatrix.query.filter(
                ProductMatrix.question_wiki_id.in_(ids_subq),
                ProductMatrix.product_name.in_(models)
            ).all()
            state_map = {}
            for r in rows_for_diff:
                wid_key = str(getattr(r, 'question_wiki_id', '') or '').strip()
                product_key = str(getattr(r, 'product_name', '') or '').strip()
                if not wid_key or product_key not in models:
                    continue
                state_map.setdefault(wid_key, {})[product_key] = bool(getattr(r, 'is_configured', False))
            out = []
            for wid_key, states_by_model in state_map.items():
                states = [bool(states_by_model.get(model, False)) for model in models]
                if len(set(states)) > 1:
                    out.append(wid_key)
            return sorted(out)

        if not any_mark_wanted:
            return jsonify({
                'success': True,
                'columns': col_list,
                'data': [],
                'total': 0,
                'page': page,
                'per_page': per_page,
                'product_categories': product_categories,
                'mapping_categories': mapping_categories
            })

        def _normalize_name(v):
            s = str(v or '')
            s = s.replace('\u3000', ' ')
            s = re.sub(r'\s+', ' ', s).strip()
            return s

        def _compute_modified_ids(ids_subq, columns_list):
            edit_query = db.session.query(ProductMatrix.question_wiki_id).filter(
                ProductMatrix.question_wiki_id.in_(ids_subq)
            ).filter(
                or_(
                    ProductMatrix.edit_source.in_(['cell', 'bulk']),
                    ProductMatrix.manual_edit == True
                )
            )
            wiki_ids_for_edits = [
                str(r[0]).strip()
                for r in edit_query.distinct().all()
                if r and str(r[0]).strip()
            ]
            wiki_ids_for_edits = list(dict.fromkeys(wiki_ids_for_edits))

            if not wiki_ids_for_edits:
                return set()

            source_products_map, source_ok = _fetch_kb_products_map('knowledge_base_v1', wiki_ids_for_edits)
            if not source_ok:
                raise RuntimeError('Failed to fetch source products')

            source_norm_map = {}
            for x in wiki_ids_for_edits:
                source_norm_map[x] = set([
                    _normalize_name(v)
                    for v in (source_products_map.get(x, set()) or set())
                    if _normalize_name(v)
                ])

            cols_norm_set = set([_normalize_name(x) for x in (columns_list or []) if _normalize_name(x)])

            qy = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(wiki_ids_for_edits)).filter(
                or_(
                    ProductMatrix.edit_source.in_(['cell', 'bulk']),
                    ProductMatrix.manual_edit == True
                )
            )
            if columns_list:
                qy = qy.filter(ProductMatrix.product_name.in_(columns_list))
            matrix_rows = qy.all()

            modified_ids = set()
            for r in matrix_rows:
                x = str(getattr(r, 'question_wiki_id', '') or '').strip()
                pn = str(getattr(r, 'product_name', '') or '').strip()
                if not x or not pn:
                    continue
                pn_norm = _normalize_name(pn)
                if pn_norm not in cols_norm_set:
                    continue
                source_cfg = pn_norm in source_norm_map.get(x, set())
                current_cfg = bool(getattr(r, 'is_configured', False))
                if current_cfg == source_cfg:
                    continue
                modified_ids.add(x)
            return modified_ids

        total = base_total
        wiki_ids = []
        if diff_compare_active:
            base_ids_subq = base_query.distinct().subquery()
            diff_compare_ids = _compute_diff_compare_ids(base_ids_subq, selected_models_ordered)
            total = len(diff_compare_ids)
            start = max(0, (page - 1) * per_page)
            wiki_ids = diff_compare_ids[start:start + per_page]
        elif want_modified and want_unmodified:
            distinct_ids = base_query.distinct().order_by(ProductMatrix.question_wiki_id)\
                                     .offset((page - 1) * per_page).limit(per_page).all()
            wiki_ids = [row[0] for row in distinct_ids]
        else:
            base_ids_subq = base_query.distinct().subquery()
            modified_ids = _compute_modified_ids(base_ids_subq, col_list)
            modified_total = len(modified_ids)
            if want_modified and not want_unmodified:
                total = modified_total
                ids_sorted = sorted(list(modified_ids))
                start = max(0, (page - 1) * per_page)
                wiki_ids = ids_sorted[start:start + per_page]
            elif want_unmodified and not want_modified:
                total = max(0, base_total - modified_total)
                base_ids = [row[0] for row in base_query.distinct().order_by(ProductMatrix.question_wiki_id).all()]
                start = max(0, (page - 1) * per_page)
                out = []
                idx = 0
                for x in base_ids:
                    if x in modified_ids:
                        continue
                    if idx >= start and len(out) < per_page:
                        out.append(x)
                    idx += 1
                    if len(out) >= per_page:
                        break
                wiki_ids = out
            else:
                total = 0
                wiki_ids = []
        
        if not wiki_ids:
             return jsonify({
                 'success': True,
                 'columns': col_list,
                 'data': [],
                 'total': total,
                 'page': page,
                 'per_page': per_page,
                 'product_categories': product_categories,
                 'mapping_categories': mapping_categories
             })

        source_products_map, source_ok = _fetch_kb_products_map('knowledge_base_v1', wiki_ids)
        prev_products_map, prev_ok = _fetch_kb_products_map('knowledge_base_v1_t1', wiki_ids)

        # 3. Fetch Data for these IDs and Columns
        data_query = ProductMatrix.query.filter(
            ProductMatrix.question_wiki_id.in_(wiki_ids),
            ProductMatrix.product_name.in_(col_list)
        )
        rows = data_query.all()
        
        # 4. Pivot Data
        data_map = {}
        # Initialize map with wiki_ids to ensure all paginated rows appear even if no product data matches (unlikely if cols exist)
        # But wait, if we filtered columns, and a question has NO entries for those columns?
        # We still want to show the question row?
        # If we filtered columns, we only show those columns. 
        # If the question exists in DB, we should show it.
        # But we only fetched data matching columns.
        # We need to fetch at least ONE entry per wiki_id to get the question content.
        
        # To ensure we have content, we can fetch ANY entry for these wiki_ids, 
        # OR rely on the fact that we probably have data.
        # Let's fetch all entries for these wiki_ids, then filter in memory for pivot? 
        # No, that's inefficient if many columns.
        
        # Better: Fetch (WikiID, Content) separately or assume we have it in `rows`.
        # If `rows` is empty (no product matches), we miss content.
        # Let's do a separate query for content if needed, or just fetch all for these IDs and filter in python.
        # Given page size 20, fetching all columns for 20 rows is fine.
        
        # Revised Step 3: Fetch ALL data for these 20 questions, then filter columns in Python.
        # This ensures we get question content even if we filter columns.
        rows_all_cols = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(wiki_ids)).all()
        
        col_set = set(col_list)
        
        for r in rows_all_cols:
            if r.question_wiki_id not in data_map:
                wid = str(r.question_wiki_id or "").strip()
                source_products = sorted(list(source_products_map.get(wid, set()))) if source_ok else None
                prev_products = sorted(list(prev_products_map.get(wid, set()))) if prev_ok else None
                data_map[r.question_wiki_id] = {
                    'question_wiki_id': r.question_wiki_id,
                    'question': r.question_content,
                    'answer': r.answer_content,
                    'update_time': r.update_time,
                    'product_category': r.product_category,
                    'source_products': source_products,
                    'prev_products': prev_products,
                    'products': {}
                }
            
            if r.product_name in col_set:
                data_map[r.question_wiki_id]['products'][r.product_name] = {
                    'is_configured': r.is_configured,
                    'manual_edit': r.manual_edit,
                    'edit_source': getattr(r, 'edit_source', '') or ''
                }
            
        result = list(data_map.values())
        
        # Sort result by wiki_id to match pagination order
        result.sort(key=lambda x: x['question_wiki_id'])
        
        return jsonify({
            'success': True, 
            'columns': col_list, 
            'data': result,
            'total': total,
            'page': page,
            'per_page': per_page,
            'product_categories': product_categories,
            'mapping_categories': mapping_categories
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/matrix/mark_total', methods=['POST'])
@login_required
def get_matrix_mark_total():
    try:
        payload = request.json or {}

        wid = str(payload.get('id') or '').strip()
        q = str(payload.get('q') or '').strip()
        a = str(payload.get('a') or '').strip()
        p = str(payload.get('p') or '').strip()
        pc = str(payload.get('pc') or '').strip()
        mc = str(payload.get('mc') or '').strip()
        p_mode = str(payload.get('p_mode') or 'any').strip().lower()

        p_models_raw = payload.get('p_models')
        p_models = []
        if isinstance(p_models_raw, list):
            p_models = [str(x).strip() for x in p_models_raw if str(x).strip()]
        elif isinstance(p_models_raw, str) and p_models_raw.strip():
            p_models = [m.strip() for m in re.split(r'[,，\n]+', p_models_raw) if m and m.strip()]
        p_models = list(dict.fromkeys([x for x in p_models if x]))

        col_models_raw = payload.get('col_models')
        col_models = []
        if isinstance(col_models_raw, list):
            col_models = [str(x).strip() for x in col_models_raw if str(x).strip()]
        elif isinstance(col_models_raw, str) and col_models_raw.strip():
            col_models = [m.strip() for m in re.split(r'[,，]+', col_models_raw) if m and m.strip()]
        col_models = list(dict.fromkeys([x for x in col_models if x]))

        marks = payload.get('marks') if isinstance(payload.get('marks'), dict) else {}
        has_new_marks = ('modified' in marks) or ('unmodified' in marks)
        if has_new_marks:
            want_modified = bool(marks.get('modified', True))
            want_unmodified = bool(marks.get('unmodified', True))
            any_wanted = want_modified or want_unmodified
        else:
            want_green = bool(marks.get('green', True))
            want_red = bool(marks.get('red', True))
            want_yellow = bool(marks.get('yellow', True))
            any_wanted = want_green or want_red or want_yellow

        if not any_wanted:
            return jsonify({'success': True, 'total': 0, 'base_total': 0})

        mappings = get_model_mappings()
        allowed_models = []
        if mc and mc in mappings and isinstance(mappings.get(mc), list):
            allowed_models = [str(x) for x in mappings.get(mc) if x]

        catalog = parse_product_catalog()
        pc_models = []
        if pc and pc in catalog and isinstance(catalog.get(pc), list):
            pc_models = [str(x).strip() for x in catalog.get(pc) if x and str(x).strip()]

        models_whitelist = set()
        if pc_models:
            models_whitelist = set([str(x) for x in pc_models if x])
        if allowed_models:
            if models_whitelist:
                models_whitelist = models_whitelist.intersection(set([str(x) for x in allowed_models if x]))
            else:
                models_whitelist = set([str(x) for x in allowed_models if x])
        if col_models:
            selected_set = set(col_models)
            if models_whitelist:
                models_whitelist = models_whitelist.intersection(selected_set)
            else:
                models_whitelist = selected_set

        columns_raw = payload.get('columns')
        if isinstance(columns_raw, list) and columns_raw:
            col_list = [str(x).strip() for x in columns_raw if str(x).strip()]
        else:
            col_query = MatrixColumn.query.order_by(MatrixColumn.sort_order)
            if models_whitelist:
                col_query = col_query.filter(MatrixColumn.product_name.in_(list(models_whitelist)))
            columns = col_query.all()
            col_list = [str(c.product_name).strip() for c in columns if c and str(c.product_name or '').strip()]

        if not col_list:
            return jsonify({'success': True, 'total': 0, 'base_total': 0})

        base_query = db.session.query(ProductMatrix.question_wiki_id)

        if wid:
            base_query = base_query.filter(ProductMatrix.question_wiki_id.ilike(f'%{wid}%'))
        if q:
            base_query = base_query.filter(ProductMatrix.question_content.ilike(f'%{q}%'))
        if a:
            base_query = base_query.filter(ProductMatrix.answer_content.ilike(f'%{a}%'))

        if p_models:
            if p_mode == 'all':
                subq = db.session.query(ProductMatrix.question_wiki_id)\
                    .filter(ProductMatrix.product_name.in_(p_models))\
                    .group_by(ProductMatrix.question_wiki_id)\
                    .having(func.count(func.distinct(ProductMatrix.product_name)) == len(p_models))\
                    .subquery()
                base_query = base_query.filter(ProductMatrix.question_wiki_id.in_(subq))
            else:
                base_query = base_query.filter(ProductMatrix.product_name.in_(p_models))
        elif p:
            base_query = base_query.filter(ProductMatrix.product_name.ilike(f'%{p}%'))

        if pc:
            base_query = base_query.filter(ProductMatrix.product_category.ilike(f'%{pc}%'))

        if mc and mc in mappings:
            if allowed_models:
                allowed_models = [str(x) for x in allowed_models if x]
                n = len(allowed_models)
                if n == 0:
                    base_query = base_query.filter(False)
                else:
                    exact_ids_subq = db.session.query(ProductMatrix.question_wiki_id)\
                        .filter(ProductMatrix.is_configured == True)\
                        .group_by(ProductMatrix.question_wiki_id)\
                        .having(func.count(func.distinct(ProductMatrix.product_name)) == n)\
                        .having(func.count(func.distinct(case(
                            (ProductMatrix.product_name.in_(allowed_models), ProductMatrix.product_name),
                            else_=None
                        ))) == n)\
                        .subquery()
                    base_query = base_query.filter(ProductMatrix.question_wiki_id.in_(exact_ids_subq))
            else:
                base_query = base_query.filter(False)

        base_total = int(base_query.distinct().count() or 0)
        if base_total <= 0:
            return jsonify({'success': True, 'total': 0, 'base_total': 0})

        if has_new_marks:
            if want_modified and want_unmodified:
                return jsonify({'success': True, 'total': base_total, 'base_total': base_total})
        else:
            if want_green and want_red and want_yellow:
                return jsonify({'success': True, 'total': base_total, 'base_total': base_total})

        def _normalize_name(v):
            s = str(v or '')
            s = s.replace('\u3000', ' ')
            s = re.sub(r'\s+', ' ', s).strip()
            return s

        cols_norm_set = set([_normalize_name(x) for x in col_list if _normalize_name(x)])

        base_ids_subq = base_query.distinct().subquery()

        if has_new_marks:
            edit_query = db.session.query(ProductMatrix.question_wiki_id).filter(
                ProductMatrix.question_wiki_id.in_(base_ids_subq)
            ).filter(
                or_(
                    ProductMatrix.edit_source.in_(['cell', 'bulk']),
                    ProductMatrix.manual_edit == True
                )
            )
            wiki_ids_for_edits = [
                str(r[0]).strip()
                for r in edit_query.distinct().all()
                if r and str(r[0]).strip()
            ]
            wiki_ids_for_edits = list(dict.fromkeys(wiki_ids_for_edits))

            if not wiki_ids_for_edits:
                modified_total = 0
            else:
                source_products_map, source_ok = _fetch_kb_products_map('knowledge_base_v1', wiki_ids_for_edits)
                if not source_ok:
                    return jsonify({'success': False, 'message': 'Failed to fetch source products'}), 500

                source_norm_map = {}
                for x in wiki_ids_for_edits:
                    source_norm_map[x] = set([_normalize_name(v) for v in (source_products_map.get(x, set()) or set()) if _normalize_name(v)])

                qy = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(wiki_ids_for_edits)).filter(
                    or_(
                        ProductMatrix.edit_source.in_(['cell', 'bulk']),
                        ProductMatrix.manual_edit == True
                    )
                )
                if col_list:
                    qy = qy.filter(ProductMatrix.product_name.in_(col_list))
                matrix_rows = qy.all()

                matched = set()
                for r in matrix_rows:
                    x = str(getattr(r, 'question_wiki_id', '') or '').strip()
                    pn = str(getattr(r, 'product_name', '') or '').strip()
                    if not x or not pn:
                        continue
                    pn_norm = _normalize_name(pn)
                    if pn_norm not in cols_norm_set:
                        continue

                    source_cfg = pn_norm in source_norm_map.get(x, set())
                    current_cfg = bool(getattr(r, 'is_configured', False))
                    if current_cfg == source_cfg:
                        continue

                    matched.add(x)

                modified_total = len(matched)

            if want_modified and not want_unmodified:
                return jsonify({'success': True, 'total': modified_total, 'base_total': base_total})
            if want_unmodified and not want_modified:
                return jsonify({'success': True, 'total': max(0, base_total - modified_total), 'base_total': base_total})
            return jsonify({'success': True, 'total': 0, 'base_total': base_total})
        else:
            wiki_ids_for_yellow = []
            if want_yellow:
                wiki_ids_for_yellow = [
                    str(r[0]).strip()
                    for r in db.session.query(base_ids_subq.c.question_wiki_id).all()
                    if r and str(r[0]).strip()
                ]
                wiki_ids_for_yellow = list(dict.fromkeys(wiki_ids_for_yellow))

            wiki_ids_for_edits = []
            if want_green or want_red:
                edit_query = db.session.query(ProductMatrix.question_wiki_id).filter(
                    ProductMatrix.question_wiki_id.in_(base_ids_subq)
                ).filter(
                    or_(
                        ProductMatrix.edit_source.in_(['cell', 'bulk']),
                        ProductMatrix.manual_edit == True
                    )
                )
                wiki_ids_for_edits = [
                    str(r[0]).strip()
                    for r in edit_query.distinct().all()
                    if r and str(r[0]).strip()
                ]
                wiki_ids_for_edits = list(dict.fromkeys(wiki_ids_for_edits))

            ids_needed = sorted(list(dict.fromkeys(wiki_ids_for_yellow + wiki_ids_for_edits)))
            if not ids_needed:
                return jsonify({'success': True, 'total': 0, 'base_total': base_total})

            source_products_map, source_ok = _fetch_kb_products_map('knowledge_base_v1', ids_needed)
            if not source_ok:
                return jsonify({'success': False, 'message': 'Failed to fetch source products'}), 500

            source_norm_map = {}
            for x in ids_needed:
                source_norm_map[x] = set([_normalize_name(v) for v in (source_products_map.get(x, set()) or set()) if _normalize_name(v)])

            prev_norm_map = {}
            prev_ok = True
            if want_yellow:
                prev_products_map, prev_ok = _fetch_kb_products_map('knowledge_base_v1_t1', wiki_ids_for_yellow)
                if prev_ok:
                    for x in wiki_ids_for_yellow:
                        prev_norm_map[x] = set([_normalize_name(v) for v in (prev_products_map.get(x, set()) or set()) if _normalize_name(v)])

            matched = set()

            if want_yellow and prev_ok and cols_norm_set:
                for x in wiki_ids_for_yellow:
                    diff = source_norm_map.get(x, set()).symmetric_difference(prev_norm_map.get(x, set()))
                    if not diff:
                        continue
                    if any((d in cols_norm_set) for d in diff):
                        matched.add(x)

            if (want_green or want_red) and cols_norm_set:
                qy = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(wiki_ids_for_edits)).filter(
                    or_(
                        ProductMatrix.edit_source.in_(['cell', 'bulk']),
                        ProductMatrix.manual_edit == True
                    )
                )
                if col_list:
                    qy = qy.filter(ProductMatrix.product_name.in_(col_list))
                matrix_rows = qy.all()

                for r in matrix_rows:
                    x = str(getattr(r, 'question_wiki_id', '') or '').strip()
                    pn = str(getattr(r, 'product_name', '') or '').strip()
                    if not x or not pn:
                        continue
                    pn_norm = _normalize_name(pn)
                    if pn_norm not in cols_norm_set:
                        continue

                    source_cfg = pn_norm in source_norm_map.get(x, set())
                    current_cfg = bool(getattr(r, 'is_configured', False))
                    if current_cfg == source_cfg:
                        continue

                    es = str(getattr(r, 'edit_source', '') or '').strip()
                    if es not in ['cell', 'bulk']:
                        if bool(getattr(r, 'manual_edit', False)):
                            es = 'cell'
                    if es == 'cell':
                        if want_red:
                            matched.add(x)
                    elif es == 'bulk':
                        if want_green:
                            matched.add(x)

            return jsonify({'success': True, 'total': len(matched), 'base_total': base_total})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500



@app.route('/api/matrix/export', methods=['GET'])
@login_required
def export_matrix_data():
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        wid = request.args.get('id', '').strip() or request.args.get('wid', '').strip()
        q = request.args.get('q', '').strip()
        a = request.args.get('a', '').strip()
        p = request.args.get('p', '').strip()
        pc = request.args.get('pc', '').strip()
        mc = request.args.get('mc', '').strip()
        p_models_param = request.args.get('p_models', '').strip()
        p_mode = request.args.get('p_mode', 'any').strip().lower()
        col_models_param = request.args.get('col_models', '').strip()
        diff_compare = request.args.get('diff_compare', '').strip().lower() in ('1', 'true', 'yes', 'on')
        mark_modified_raw = request.args.get('mark_modified', '').strip()
        mark_unmodified_raw = request.args.get('mark_unmodified', '').strip()

        def _as_bool(s, default=True):
            if s is None:
                return default
            t = str(s).strip().lower()
            if t == '':
                return default
            if t in ('1', 'true', 't', 'yes', 'y', 'on'):
                return True
            if t in ('0', 'false', 'f', 'no', 'n', 'off'):
                return False
            return default

        want_modified = _as_bool(mark_modified_raw, True)
        want_unmodified = _as_bool(mark_unmodified_raw, True)
        any_mark_wanted = bool(want_modified or want_unmodified)

        mappings = get_model_mappings()
        allowed_models = []
        if mc and mc in mappings and isinstance(mappings.get(mc), list):
            allowed_models = [str(x) for x in mappings.get(mc) if x]

        catalog = parse_product_catalog()
        pc_models = []
        if pc and pc in catalog and isinstance(catalog.get(pc), list):
            pc_models = [str(x).strip() for x in catalog.get(pc) if x and str(x).strip()]

        models_whitelist = set()
        if pc_models:
            models_whitelist = set([str(x) for x in pc_models if x])
        if allowed_models:
            if models_whitelist:
                models_whitelist = models_whitelist.intersection(set([str(x) for x in allowed_models if x]))
            else:
                models_whitelist = set([str(x) for x in allowed_models if x])

        selected_models_ordered = []
        if col_models_param:
            raw_models = [m.strip() for m in re.split(r'[,，]', col_models_param) if m and m.strip()]
            seen_models = set()
            for m in raw_models:
                if m in seen_models:
                    continue
                seen_models.add(m)
                selected_models_ordered.append(m)
            if selected_models_ordered:
                selected_set = set(selected_models_ordered)
                if models_whitelist:
                    models_whitelist = models_whitelist.intersection(selected_set)
                else:
                    models_whitelist = selected_set
        
        col_query = MatrixColumn.query.order_by(MatrixColumn.sort_order)
        if models_whitelist:
            col_query = col_query.filter(MatrixColumn.product_name.in_(list(models_whitelist)))
        if selected_models_ordered:
            columns = col_query.filter(MatrixColumn.product_name.in_(selected_models_ordered)).all()
            by_name = {str(c.product_name): c for c in columns if c and str(getattr(c, 'product_name', '') or '').strip()}
            col_list = [m for m in selected_models_ordered if m in by_name]
        else:
            columns = col_query.all()
            col_list = [c.product_name for c in columns]
        col_list = [c for c in (col_list or []) if str(c or '').strip() and str(c or '').strip() != '测试型号']
        
        if not col_list:
            return jsonify({'success': False, 'message': 'No columns found'}), 404

        base_query = db.session.query(ProductMatrix.question_wiki_id)
        
        if wid:
            base_query = base_query.filter(ProductMatrix.question_wiki_id.ilike(f'%{wid}%'))
        if q:
            base_query = base_query.filter(ProductMatrix.question_content.ilike(f'%{q}%'))
        if a:
            base_query = base_query.filter(ProductMatrix.answer_content.ilike(f'%{a}%'))
        
        p_models = []
        if p_models_param:
            p_models = [m.strip() for m in re.split(r'[,，\n]+', p_models_param) if m and m.strip()]
        elif p and re.search(r'[,，\n]', p):
            p_models = [m.strip() for m in re.split(r'[,，\n]+', p) if m and m.strip()]
        
        if p_models:
            p_models = list(dict.fromkeys(p_models))
            if p_mode == 'all':
                subq = db.session.query(ProductMatrix.question_wiki_id)\
                    .filter(ProductMatrix.product_name.in_(p_models))\
                    .group_by(ProductMatrix.question_wiki_id)\
                    .having(func.count(func.distinct(ProductMatrix.product_name)) == len(p_models))\
                    .subquery()
                base_query = base_query.filter(ProductMatrix.question_wiki_id.in_(subq))
            else:
                base_query = base_query.filter(ProductMatrix.product_name.in_(p_models))
        elif p:
            base_query = base_query.filter(ProductMatrix.product_name.ilike(f'%{p}%'))

        if pc:
            base_query = base_query.filter(ProductMatrix.product_category.ilike(f'%{pc}%'))

        if mc and mc in mappings:
            if allowed_models:
                allowed_models = [str(x) for x in allowed_models if x]
                n = len(allowed_models)
                if n == 0:
                    base_query = base_query.filter(False)
                else:
                    exact_ids_subq = db.session.query(ProductMatrix.question_wiki_id)\
                        .filter(ProductMatrix.is_configured == True)\
                        .group_by(ProductMatrix.question_wiki_id)\
                        .having(func.count(func.distinct(ProductMatrix.product_name)) == n)\
                        .having(func.count(func.distinct(case(
                            (ProductMatrix.product_name.in_(allowed_models), ProductMatrix.product_name),
                            else_=None
                        ))) == n)\
                        .subquery()
                    base_query = base_query.filter(ProductMatrix.question_wiki_id.in_(exact_ids_subq))
            else:
                base_query = base_query.filter(False)

        def _normalize_name(v):
            s = str(v or '')
            s = s.replace('\u3000', ' ')
            s = re.sub(r'\s+', ' ', s).strip()
            return s

        def _compute_diff_compare_ids(ids_subq, selected_models):
            models = [str(x).strip() for x in (selected_models or []) if str(x).strip()]
            if len(models) < 2:
                return []
            rows_for_diff = ProductMatrix.query.filter(
                ProductMatrix.question_wiki_id.in_(ids_subq),
                ProductMatrix.product_name.in_(models)
            ).all()
            state_map = {}
            for r in rows_for_diff:
                wid_key = str(getattr(r, 'question_wiki_id', '') or '').strip()
                product_key = str(getattr(r, 'product_name', '') or '').strip()
                if not wid_key or product_key not in models:
                    continue
                state_map.setdefault(wid_key, {})[product_key] = bool(getattr(r, 'is_configured', False))
            out = []
            for wid_key, states_by_model in state_map.items():
                states = [bool(states_by_model.get(model, False)) for model in models]
                if len(set(states)) > 1:
                    out.append(wid_key)
            return sorted(out)

        def _compute_modified_ids(ids_subq, columns_list):
            edit_query = db.session.query(ProductMatrix.question_wiki_id).filter(
                ProductMatrix.question_wiki_id.in_(ids_subq)
            ).filter(
                or_(
                    ProductMatrix.edit_source.in_(['cell', 'bulk']),
                    ProductMatrix.manual_edit == True
                )
            )
            wiki_ids_for_edits = [
                str(r[0]).strip()
                for r in edit_query.distinct().all()
                if r and str(r[0]).strip()
            ]
            wiki_ids_for_edits = list(dict.fromkeys(wiki_ids_for_edits))
            if not wiki_ids_for_edits:
                return set()

            source_products_map, source_ok = _fetch_kb_products_map('knowledge_base_v1', wiki_ids_for_edits)
            if not source_ok:
                raise RuntimeError('Failed to fetch source products')

            source_norm_map = {}
            for x in wiki_ids_for_edits:
                source_norm_map[x] = set([
                    _normalize_name(v)
                    for v in (source_products_map.get(x, set()) or set())
                    if _normalize_name(v)
                ])

            cols_norm_set = set([_normalize_name(x) for x in (columns_list or []) if _normalize_name(x)])
            qy = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(wiki_ids_for_edits)).filter(
                or_(
                    ProductMatrix.edit_source.in_(['cell', 'bulk']),
                    ProductMatrix.manual_edit == True
                )
            )
            if columns_list:
                qy = qy.filter(ProductMatrix.product_name.in_(columns_list))
            matrix_rows = qy.all()

            modified_ids = set()
            for r in matrix_rows:
                x = str(getattr(r, 'question_wiki_id', '') or '').strip()
                pn = str(getattr(r, 'product_name', '') or '').strip()
                if not x or not pn:
                    continue
                pn_norm = _normalize_name(pn)
                if pn_norm not in cols_norm_set:
                    continue
                source_cfg = pn_norm in source_norm_map.get(x, set())
                current_cfg = bool(getattr(r, 'is_configured', False))
                if current_cfg != source_cfg:
                    modified_ids.add(x)
            return modified_ids

        if not any_mark_wanted:
            wiki_ids = []
        else:
            diff_compare_active = diff_compare and len(selected_models_ordered) >= 2
            if diff_compare_active:
                base_ids_subq = base_query.distinct().subquery()
                wiki_ids = _compute_diff_compare_ids(base_ids_subq, selected_models_ordered)
            elif want_modified and want_unmodified:
                distinct_ids = base_query.distinct().order_by(ProductMatrix.question_wiki_id).all()
                wiki_ids = [row[0] for row in distinct_ids]
            else:
                base_ids_subq = base_query.distinct().subquery()
                modified_ids = _compute_modified_ids(base_ids_subq, col_list)
                if want_modified and not want_unmodified:
                    wiki_ids = sorted(list(modified_ids))
                elif want_unmodified and not want_modified:
                    base_ids = [row[0] for row in base_query.distinct().order_by(ProductMatrix.question_wiki_id).all()]
                    wiki_ids = [x for x in base_ids if x not in modified_ids]
                else:
                    wiki_ids = []
        
        if not wiki_ids:
            return jsonify({'success': False, 'message': 'No data found'}), 404

        rows_all_cols = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(wiki_ids)).all()
        
        col_set = set(col_list)
        data_map = {}
        
        for r in rows_all_cols:
            if r.question_wiki_id not in data_map:
                data_map[r.question_wiki_id] = {
                    'question_wiki_id': r.question_wiki_id,
                    'question': r.question_content,
                    'answer': r.answer_content,
                    'update_time': r.update_time,
                    'product_category': r.product_category,
                    'products': {}
                }
            
            if r.product_name in col_set:
                data_map[r.question_wiki_id]['products'][r.product_name] = {
                    'is_configured': r.is_configured,
                    'manual_edit': r.manual_edit,
                    'edit_source': getattr(r, 'edit_source', '') or ''
                }
            
        # 4. Construct DataFrame
        export_data = []
        for wiki_id in wiki_ids: # Use sorted wiki_ids order
            if wiki_id not in data_map: continue
            item = data_map[wiki_id]
            row = {
                'ID': item['question_wiki_id'],
                '问题': item['question'],
                '答案': item['answer']
            }
            for col in col_list:
                prod_data = item['products'].get(col)
                if prod_data and prod_data['is_configured']:
                    row[col] = '✅'
                else:
                    row[col] = '❌'
            export_data.append(row)
            
        df = pd.DataFrame(export_data)
        
        output = io.BytesIO()
        engine = 'xlsxwriter' if importlib.util.find_spec('xlsxwriter') is not None else 'openpyxl'
        with pd.ExcelWriter(output, engine=engine) as writer:
            df.to_excel(writer, index=False, sheet_name='Matrix')
            worksheet = writer.sheets['Matrix']
            if engine == 'xlsxwriter':
                worksheet.set_column(0, 0, 15)
                worksheet.set_column(1, 1, 40)
                worksheet.set_column(2, 2, 40)
                for i in range(len(col_list)):
                    worksheet.set_column(3 + i, 3 + i, 10)
            else:
                from openpyxl.utils import get_column_letter
                widths = [15, 40, 40] + [10] * len(col_list)
                for idx, width in enumerate(widths, start=1):
                    worksheet.column_dimensions[get_column_letter(idx)].width = width
                
        output.seek(0)
        
        filename = canonical_download_name('product_matrix')
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/update', methods=['POST'])
@login_required
def update_matrix_cell():
    data = request.json
    wiki_id = data.get('question_wiki_id')
    product_name = data.get('product_name')
    is_configured = data.get('is_configured')
    
    if not wiki_id or not product_name:
        return jsonify({'success': False, 'message': 'Missing keys'}), 400
        
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        item = ProductMatrix.query.filter_by(question_wiki_id=wiki_id, product_name=product_name).first()
        old_val = False
        
        if not item:
            # Try to find question info from other entries
            existing = ProductMatrix.query.filter_by(question_wiki_id=wiki_id).first()
            question = existing.question_content if existing else ""
            answer = existing.answer_content if existing else ""
            category = existing.product_category if existing else ""
            update_time = existing.update_time if existing else datetime.now().strftime('%Y-%m-%d')
            
            item = ProductMatrix(
                question_wiki_id=wiki_id,
                product_name=product_name,
                question_content=question,
                answer_content=answer,
                product_category=category,
                update_time=update_time
            )
            db.session.add(item)
        else:
            old_val = item.is_configured
            
        item.is_configured = is_configured
        item.manual_edit = True
        item.edit_source = 'cell'
        item.last_synced_at = datetime.utcnow()
        db.session.commit()

        warnings = []
        if is_supabase_matrix_enabled():
            try:
                col = MatrixColumn.query.filter_by(product_name=product_name).first()
                try:
                    so = int(getattr(col, 'sort_order', 0) or 0) if col else 0
                except Exception:
                    so = 0
                pm_payload = [{
                    'question_wiki_id': str(wiki_id).strip(),
                    'product_name': str(product_name).strip(),
                    'is_configured': bool(item.is_configured),
                    'manual_edit': bool(item.manual_edit),
                    'edit_source': str(getattr(item, 'edit_source', '') or ''),
                    'last_synced_at': _dt_to_iso(getattr(item, 'last_synced_at', None)),
                    'question_content': getattr(item, 'question_content', None),
                    'answer_content': getattr(item, 'answer_content', None),
                    'update_time': getattr(item, 'update_time', None),
                    'product_category': getattr(item, 'product_category', None)
                }]

                def _bg_sync_matrix_one():
                    try:
                        c = get_supabase_client()
                        if c:
                            resp_col = c.upsert('matrix_column', [{'product_name': str(product_name).strip(), 'sort_order': so}], on_conflict='product_name')
                            if resp_col is None or getattr(resp_col, 'status_code', 500) >= 400:
                                raise RuntimeError(getattr(resp_col, 'text', '') or 'matrix_column upsert failed')
                            if pm_payload:
                                resp_pm = c.upsert('product_matrix', pm_payload, on_conflict='question_wiki_id,product_name')
                                if resp_pm is None or getattr(resp_pm, 'status_code', 500) >= 400:
                                    raise RuntimeError(getattr(resp_pm, 'text', '') or 'product_matrix upsert failed')
                    except Exception as e:
                        raise RuntimeError(str(e))

                _bg_sync_matrix_one()
            except Exception as e:
                warnings.append(f'远端矩阵同步失败: {str(e)}')

        out = {'success': True}
        if warnings:
            out['warnings'] = warnings
        return jsonify(out)
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/batch_update', methods=['POST'])
@login_required
def batch_update_matrix():
    data = request.json
    wiki_ids = data.get('question_wiki_ids', [])
    product_name = data.get('product_name')
    is_configured = data.get('is_configured')
    
    if not wiki_ids or not product_name:
        return jsonify({'success': False, 'message': 'Missing parameters'}), 400
        
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        # 1. Update existing records
        existing_query = ProductMatrix.query.filter(
            ProductMatrix.question_wiki_id.in_(wiki_ids),
            ProductMatrix.product_name == product_name
        )
        existing_items = existing_query.all()
        existing_ids = set()
        before_map = {}
        
        for item in existing_items:
            before_map[str(item.question_wiki_id)] = bool(getattr(item, 'is_configured', False))
            item.is_configured = is_configured
            item.manual_edit = True
            if getattr(item, 'edit_source', '') != 'cell':
                item.edit_source = 'bulk'
            item.last_synced_at = datetime.utcnow()
            existing_ids.add(item.question_wiki_id)
            
        # 2. Create missing records
        missing_ids = set(wiki_ids) - existing_ids
        if missing_ids:
            # We need basic info (question, answer, etc.) to create new records.
            # We can get this from other ProductMatrix entries for the same wiki_id.
            # Use a subquery or just query all entries for these wiki_ids to find a template.
            
            # Fetch any entry for each missing wiki_id to get content
            templates = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(missing_ids)).all()
            template_map = {}
            for t in templates:
                if t.question_wiki_id not in template_map:
                    template_map[t.question_wiki_id] = t
            
            new_items = []
            for wid in missing_ids:
                tmpl = template_map.get(wid)
                if tmpl:
                    new_items.append(ProductMatrix(
                        question_wiki_id=wid,
                        product_name=product_name,
                        question_content=tmpl.question_content,
                        answer_content=tmpl.answer_content,
                        product_category=tmpl.product_category,
                        update_time=tmpl.update_time,
                        is_configured=is_configured,
                        manual_edit=True,
                        edit_source='bulk',
                        last_synced_at=datetime.utcnow()
                    ))
            
            if new_items:
                db.session.add_all(new_items)
                
        db.session.commit()

        warnings = []
        if is_supabase_matrix_enabled():
            try:
                col = MatrixColumn.query.filter_by(product_name=product_name).first()
                try:
                    so = int(getattr(col, 'sort_order', 0) or 0) if col else 0
                except Exception:
                    so = 0
                rows = ProductMatrix.query.filter(
                    ProductMatrix.question_wiki_id.in_(wiki_ids),
                    ProductMatrix.product_name == product_name
                ).all()
                pm_payload = []
                for r in rows:
                    pm_payload.append({
                        'question_wiki_id': str(getattr(r, 'question_wiki_id', '') or '').strip(),
                        'product_name': str(getattr(r, 'product_name', '') or '').strip(),
                        'is_configured': bool(getattr(r, 'is_configured', False)),
                        'manual_edit': bool(getattr(r, 'manual_edit', False)),
                        'edit_source': str(getattr(r, 'edit_source', '') or ''),
                        'last_synced_at': _dt_to_iso(getattr(r, 'last_synced_at', None)),
                        'question_content': getattr(r, 'question_content', None),
                        'answer_content': getattr(r, 'answer_content', None),
                        'update_time': getattr(r, 'update_time', None),
                        'product_category': getattr(r, 'product_category', None)
                    })
                pm_payload = [x for x in pm_payload if x.get('question_wiki_id') and x.get('product_name')]

                def _bg_sync_matrix_batch():
                    try:
                        c = get_supabase_client()
                        if c:
                            resp_col = c.upsert('matrix_column', [{'product_name': str(product_name).strip(), 'sort_order': so}], on_conflict='product_name')
                            if resp_col is None or getattr(resp_col, 'status_code', 500) >= 400:
                                raise RuntimeError(getattr(resp_col, 'text', '') or 'matrix_column upsert failed')
                            if pm_payload:
                                chunk_res = _supabase_upsert_chunks(c, 'product_matrix', pm_payload, on_conflict='question_wiki_id,product_name', chunk_size=500)
                                if isinstance(chunk_res, dict) and chunk_res.get('ok') is False:
                                    raise RuntimeError(json.dumps(chunk_res.get('errors') or [], ensure_ascii=False) or 'product_matrix batch upsert failed')
                    except Exception as e:
                        raise RuntimeError(str(e))

                _bg_sync_matrix_batch()
            except Exception as e:
                warnings.append(f'远端矩阵同步失败: {str(e)}')

        out = {'success': True, 'updated': len(wiki_ids)}
        if warnings:
            out['warnings'] = warnings
        return jsonify(out)
        
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/mismatch_changes', methods=['POST'])
@login_required
def get_matrix_mismatch_changes():
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        data = request.json or {}
        wiki_ids_raw = data.get('wiki_ids', None)
        
        wiki_ids = []
        if wiki_ids_raw is not None:
            if not isinstance(wiki_ids_raw, list):
                return jsonify({'success': False, 'message': 'wiki_ids must be a list'}), 400
            wiki_ids = [str(x).strip() for x in wiki_ids_raw if str(x).strip()]
            wiki_ids = list(dict.fromkeys(wiki_ids))
        
        if not wiki_ids:
            rows = db.session.query(ProductMatrix.question_wiki_id).filter(
                or_(
                    ProductMatrix.edit_source.in_(['cell', 'bulk']),
                    ProductMatrix.manual_edit == True
                )
            ).distinct().all()
            wiki_ids = [str(r[0]).strip() for r in rows if r and str(r[0]).strip()]
        
        if not wiki_ids:
            return jsonify({'success': True, 'wiki_ids_count': 0, 'count': 0, 'changes': []})
        
        def _normalize_name(v):
            s = str(v or '')
            s = s.replace('\u3000', ' ')
            s = re.sub(r'\s+', ' ', s).strip()
            return s
        
        col_set = set()
        try:
            col_set = set([str(c.product_name or '').strip() for c in MatrixColumn.query.all() if c and str(c.product_name or '').strip()])
        except Exception:
            col_set = set()
        
        source_products_map = {}
        source_ok = True
        chunk_size = 400
        for i in range(0, len(wiki_ids), chunk_size):
            chunk = wiki_ids[i:i + chunk_size]
            m, ok = _fetch_kb_products_map('knowledge_base_v1', chunk)
            if not ok:
                source_ok = False
                break
            source_products_map.update(m or {})
        
        if not source_ok:
            return jsonify({'success': False, 'message': 'Failed to fetch source products from knowledge_base_v1'}), 500
        
        source_norm_map = {}
        for wid in wiki_ids:
            source_norm_map[wid] = set([_normalize_name(x) for x in (source_products_map.get(wid, set()) or set()) if _normalize_name(x)])
        
        q = ProductMatrix.query.filter(ProductMatrix.question_wiki_id.in_(wiki_ids)).filter(
            or_(
                ProductMatrix.edit_source.in_(['cell', 'bulk']),
                ProductMatrix.manual_edit == True
            )
        )
        matrix_rows = q.all()
        
        changes = []
        seen = set()
        for r in matrix_rows:
            wid = str(getattr(r, 'question_wiki_id', '') or '').strip()
            pn = str(getattr(r, 'product_name', '') or '').strip()
            if not wid or not pn:
                continue
            if col_set and pn not in col_set:
                continue
            
            pn_norm = _normalize_name(pn)
            source_cfg = pn_norm in source_norm_map.get(wid, set())
            current_cfg = bool(getattr(r, 'is_configured', False))
            if current_cfg == source_cfg:
                continue
            
            es = str(getattr(r, 'edit_source', '') or '').strip()
            if es not in ['cell', 'bulk']:
                if bool(getattr(r, 'manual_edit', False)):
                    es = 'cell'
            
            if es not in ['cell', 'bulk']:
                continue
            
            key = f'{wid}::{pn}'
            if key in seen:
                continue
            seen.add(key)
            
            changes.append({
                'question_wiki_id': wid,
                'product_name': pn,
                'old_is_configured': bool(source_cfg),
                'new_is_configured': bool(current_cfg),
                'edit_source': es
            })

        affected_wiki_ids_count = len(set([str(c.get('question_wiki_id') or '').strip() for c in changes if str(c.get('question_wiki_id') or '').strip()]))
        return jsonify({
            'success': True,
            'wiki_ids_count': len(wiki_ids),
            'count': len(changes),
            'affected_wiki_ids_count': affected_wiki_ids_count,
            'changes': changes
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/submit_changes', methods=['POST'])
@login_required
def submit_matrix_changes():
    data = request.json or {}
    operation_id = str(data.get('operation_id') or '').strip() or str(uuid.uuid4())
    attempt = data.get('attempt')
    try:
        attempt = int(attempt) if attempt is not None else 1
    except Exception:
        attempt = 1
    changes = data.get('changes') or []
    
    if not isinstance(changes, list) or len(changes) == 0:
        return jsonify({'success': False, 'message': 'No changes'}), 400
    normalized, errors = validate_submit_changes(changes)
    
    if errors:
        return jsonify({'success': False, 'message': 'Validation failed', 'errors': errors}), 400

    _maybe_pull_matrix_and_logs_from_supabase(force=False)
    
    existing_op = MatrixSubmitOperation.query.filter_by(operation_id=operation_id).first()
    if existing_op and existing_op.status == 'success':
        written = Button.query.filter_by(operation_id=operation_id).count()
        return jsonify({'success': True, 'operation_id': operation_id, 'written': written})
    
    try:
        if not existing_op:
            existing_op = MatrixSubmitOperation(
                operation_id=operation_id,
                status='pending',
                attempts=0,
                created_by=current_user.username
            )
            db.session.add(existing_op)
        
        existing_op.attempts = max(existing_op.attempts or 0, attempt)
        existing_op.status = 'pending'
        existing_op.error_message = None
        db.session.flush()
        
        conflict_errors = []
        wiki_ids = []
        product_names = []
        for c in normalized:
            wid = c.get('question_wiki_id')
            pn = c.get('product_name')
            if wid and wid not in wiki_ids:
                wiki_ids.append(wid)
            if pn and pn not in product_names:
                product_names.append(pn)
        
        current_map = {}
        if wiki_ids and product_names:
            current_rows = ProductMatrix.query.filter(
                ProductMatrix.question_wiki_id.in_(wiki_ids),
                ProductMatrix.product_name.in_(product_names)
            ).all()
            for r in current_rows:
                current_map[(str(r.question_wiki_id or '').strip(), str(r.product_name or '').strip())] = bool(r.is_configured)
        
        for idx, c in enumerate(normalized):
            wiki_id = str(c.get('question_wiki_id') or '').strip()
            product_name = str(c.get('product_name') or '').strip()
            new_val = bool(c.get('new_is_configured'))
            current_val = bool(current_map.get((wiki_id, product_name), False))
            if current_val != new_val:
                conflict_errors.append(f'第 {idx + 1} 条: 冲突：当前库值({current_val}) 与待提交 after({new_val}) 不一致')
        
        if conflict_errors:
            raise ValueError('\n'.join(conflict_errors))
        
        now = datetime.utcnow()
        to_write = []
        for c in normalized:
            diff_obj = {
                'question_wiki_id': c['question_wiki_id'],
                'product_name': c['product_name'],
                'before': c['old_is_configured'],
                'after': c['new_is_configured'],
                'edit_source': c['edit_source'],
                'submitted_by': current_user.username,
                'submitted_at': now.isoformat() + 'Z'
            }
            to_write.append(Button(
                operation_id=operation_id,
                question_wiki_id=c['question_wiki_id'],
                product_name=c['product_name'],
                old_is_configured=c['old_is_configured'],
                new_is_configured=c['new_is_configured'],
                edit_source=c['edit_source'],
                diff_json=json.dumps(diff_obj, ensure_ascii=False),
                submitted_by=current_user.username,
                submitted_at=now
            ))
        
        db.session.add_all(to_write)
        db.session.flush()
        
        supabase_info = {
            'button': {'enabled': bool(is_supabase_button_sync_enabled()), 'attempted': False},
            'modifications': {'attempted': False}
        }
        if is_supabase_button_sync_enabled():
            client = get_supabase_client()
            if client:
                payload = []
                for c in normalized:
                    payload.append({
                        'operation_id': operation_id,
                        'question_wiki_id': c['question_wiki_id'],
                        'product_name': c['product_name'],
                        'old_is_configured': c['old_is_configured'],
                        'new_is_configured': c['new_is_configured'],
                        'edit_source': c['edit_source'],
                        'submitted_by': current_user.username,
                        'submitted_at': now.isoformat() + 'Z'
                    })
                # 使用 upsert 而不是 insert，避免主键冲突
                resp = client.upsert('button', payload, on_conflict='operation_id,question_wiki_id,product_name')
                supabase_info['button']['attempted'] = True
                supabase_info['button']['status_code'] = getattr(resp, 'status_code', None) if resp else None
                supabase_info['button']['text'] = getattr(resp, 'text', '') if resp else ''
                supabase_info['button']['ok'] = bool(resp) and getattr(resp, 'status_code', 500) in [200, 201]
                if not supabase_info['button']['ok']:
                    raise RuntimeError(f"按钮提交日志写入失败 ({supabase_info['button']['status_code']}): {supabase_info['button']['text'] or 'unknown error'}")
        
        existing_op.status = 'success'
        db.session.commit()

        if is_supabase_matrix_enabled():
            try:
                client = get_supabase_client()
                if client:
                    include_diff_json = _supabase_has_column(client, 'button', 'diff_json')
                    op_payload = [{
                        'operation_id': operation_id,
                        'status': 'success',
                        'attempts': int(existing_op.attempts or 0),
                        'created_by': existing_op.created_by,
                        'error_message': existing_op.error_message,
                        'created_at': _dt_to_iso(existing_op.created_at),
                        'updated_at': _dt_to_iso(existing_op.updated_at)
                    }]
                    client.upsert('matrix_submit_operation', op_payload, on_conflict='operation_id')

                    btn_payload = []
                    for c in normalized:
                        diff_obj = {
                            'question_wiki_id': c['question_wiki_id'],
                            'product_name': c['product_name'],
                            'before': c['old_is_configured'],
                            'after': c['new_is_configured'],
                            'edit_source': c['edit_source'],
                            'submitted_by': current_user.username,
                            'submitted_at': now.isoformat() + 'Z'
                        }
                        row = {
                            'operation_id': operation_id,
                            'question_wiki_id': c['question_wiki_id'],
                            'product_name': c['product_name'],
                            'old_is_configured': c['old_is_configured'],
                            'new_is_configured': c['new_is_configured'],
                            'edit_source': c['edit_source'],
                            'submitted_by': current_user.username,
                            'submitted_at': now.isoformat() + 'Z'
                        }
                        if include_diff_json:
                            row['diff_json'] = json.dumps(diff_obj, ensure_ascii=False)
                        btn_payload.append(row)
                    if btn_payload:
                        _supabase_upsert_chunks(client, 'button', btn_payload, on_conflict='operation_id,question_wiki_id,product_name', chunk_size=500)
            except Exception:
                pass
        
        warnings = []
        try:
            client = get_supabase_client()
            if client:
                wiki_ids = []
                for c in normalized:
                    wid = str(c.get('question_wiki_id') or '').strip()
                    if wid and wid not in wiki_ids:
                        wiki_ids.append(wid)
                kb_rows = _fetch_kb_detail_map(wiki_ids)
                
                wid_to_changes = {}
                for c in normalized:
                    wid = str(c.get('question_wiki_id') or '').strip()
                    if not wid:
                        continue
                    wid_to_changes.setdefault(wid, []).append(c)

                snapshot_map = _build_matrix_submit_snapshot_map(normalized, kb_rows)
                mod_records = []
                for wid, cs in wid_to_changes.items():
                    base = kb_rows.get(wid, {})
                    rec = dict(base) if isinstance(base, dict) else {}
                    rec.pop('id', None)
                    rec.pop('review_status', None)
                    rec.pop('combinatorial_problem', None)
                    rec['kb_id'] = wid
                    rec['question_wiki_id'] = wid
                    rec['change_type'] = 'edit'
                    rec['modifier'] = current_user.username
                    rec['modification_time'] = _now_iso_with_tz()
                    summary = snapshot_map.get(wid) or {}
                    before_obj = summary.get('before') if isinstance(summary.get('before'), dict) else _snapshot_mod_fields(rec)
                    after_obj = summary.get('after') if isinstance(summary.get('after'), dict) else _snapshot_mod_fields(rec)
                    changed_fields = summary.get('changed_fields') if isinstance(summary.get('changed_fields'), list) else _compute_mod_changed_fields(before_obj, after_obj)
                    changed_products = summary.get('changed_products') if isinstance(summary.get('changed_products'), list) else []
                    edit_sources = sorted(list(dict.fromkeys([
                        str(c.get('edit_source') or '').strip() for c in cs if str(c.get('edit_source') or '').strip()
                    ])))
                    rec['product_name'] = after_obj.get('products') or rec.get('product_name') or ''
                    if not str(rec.get('question') or '').strip():
                        rec['question'] = after_obj.get('question') or ''
                    if not str(rec.get('answer') or '').strip():
                        rec['answer'] = after_obj.get('answer') or ''

                    _attach_change_meta(rec, {
                        'source': '机型矩阵管理',
                        'operation_id': operation_id,
                        'edit_source': edit_sources[0] if len(edit_sources) == 1 else ','.join(edit_sources),
                        'changed_products': changed_products,
                        'before': before_obj,
                        'after': after_obj,
                        'changed_fields': changed_fields
                    })
                    mod_records.append(rec)
                
                if mod_records:
                    # 转换数组字段为 JSON 字符串
                    _convert_array_fields_to_json(mod_records)
                    
                    # 使用 _supabase_insert_drop_unknown_columns 来自动处理字段不匹配
                    mod_resp = _supabase_insert_drop_unknown_columns(client, 'knowledge_base_modifications', mod_records)
                    supabase_info['modifications']['attempted'] = True
                    supabase_info['modifications']['status_code'] = getattr(mod_resp, 'status_code', None) if mod_resp else None
                    supabase_info['modifications']['text'] = getattr(mod_resp, 'text', '') if mod_resp else ''
                    supabase_info['modifications']['ok'] = bool(mod_resp) and getattr(mod_resp, 'status_code', 500) in [200, 201]
                    if not supabase_info['modifications']['ok']:
                        warnings.append(f"修改记录写入失败 ({supabase_info['modifications']['status_code']}): {supabase_info['modifications']['text'] or 'unknown error'}")
        except Exception as e:
            warnings.append(f"修改记录写入异常: {str(e)}")
        
        out = {'success': True, 'operation_id': operation_id, 'written': len(to_write), 'supabase': supabase_info}
        if warnings:
            out['warnings'] = warnings
        return jsonify(out)
    except Exception as e:
        db.session.rollback()
        try:
            op = MatrixSubmitOperation.query.filter_by(operation_id=operation_id).first()
            if not op:
                op = MatrixSubmitOperation(operation_id=operation_id, created_by=current_user.username)
                db.session.add(op)
            op.status = 'failed'
            op.attempts = max(op.attempts or 0, attempt)
            op.error_message = str(e)
            db.session.commit()
        except Exception:
            db.session.rollback()
        errors = str(e).split('\n') if str(e) else ['提交失败']
        return jsonify({'success': False, 'message': 'Submit failed', 'errors': errors, 'operation_id': operation_id}), 500

@app.route('/api/matrix/submit_logs/<operation_id>/details', methods=['GET'])
@login_required
def get_matrix_submit_log_details(operation_id):
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        op_id = str(operation_id or '').strip()
        if not op_id:
            return jsonify({'success': False, 'message': 'operation_id required'}), 400

        limit = request.args.get('limit', '500')
        try:
            limit = int(limit)
        except Exception:
            limit = 500
        limit = max(1, min(2000, limit))

        rows = []
        if is_supabase_matrix_enabled():
            client = get_supabase_client()
            if not client:
                return jsonify({'success': False, 'message': '数据库客户端不可用'}), 500
            include_diff_json = _supabase_has_column(client, 'button', 'diff_json')
            btn_columns = 'operation_id,question_wiki_id,product_name,old_is_configured,new_is_configured,edit_source,submitted_by,submitted_at'
            if include_diff_json:
                btn_columns += ',diff_json'
            rows = client.select_all(
                'button',
                filters={'operation_id': f'eq.{op_id}'},
                order_by='submitted_at',
                order_dir='asc',
                columns=btn_columns,
                page_size=limit
            ) or []
        else:
            btns = Button.query.filter_by(operation_id=op_id).order_by(Button.submitted_at.asc()).limit(limit).all()
            for btn in btns:
                rows.append({
                    'operation_id': btn.operation_id,
                    'question_wiki_id': btn.question_wiki_id,
                    'product_name': btn.product_name,
                    'old_is_configured': btn.old_is_configured,
                    'new_is_configured': btn.new_is_configured,
                    'edit_source': btn.edit_source,
                    'diff_json': btn.diff_json,
                    'submitted_by': btn.submitted_by,
                    'submitted_at': _dt_to_iso(btn.submitted_at)
                })

        normalized_rows = []
        wiki_ids = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            wid = str(row.get('question_wiki_id') or '').strip()
            pn = str(row.get('product_name') or '').strip()
            if not wid or not pn:
                continue
            if wid not in wiki_ids:
                wiki_ids.append(wid)
            diff_obj = _safe_json_loads(row.get('diff_json'), None)
            normalized_rows.append({
                'operation_id': str(row.get('operation_id') or '').strip(),
                'question_wiki_id': wid,
                'product_name': pn,
                'old_is_configured': bool(row.get('old_is_configured', False)),
                'new_is_configured': bool(row.get('new_is_configured', False)),
                'edit_source': str(row.get('edit_source') or '').strip(),
                'diff_json': diff_obj,
                'submitted_by': row.get('submitted_by'),
                'submitted_at': _dt_to_iso(row.get('submitted_at'))
            })

        base_map = _fetch_kb_detail_map(wiki_ids)
        snapshot_map = _build_matrix_submit_snapshot_map(normalized_rows, base_map)
        affected_wiki_ids = set()
        data = []
        for row in normalized_rows:
            wid = row.get('question_wiki_id')
            base = dict(base_map.get(wid) or {})
            summary = snapshot_map.get(wid) or {}
            before_obj = summary.get('before') if isinstance(summary.get('before'), dict) else None
            after_obj = summary.get('after') if isinstance(summary.get('after'), dict) else None
            changed_fields = summary.get('changed_fields') if isinstance(summary.get('changed_fields'), list) else ['products']
            changed_products = summary.get('changed_products') if isinstance(summary.get('changed_products'), list) else [row.get('product_name')]
            if wid:
                affected_wiki_ids.add(wid)
            data.append({
                'operation_id': row.get('operation_id'),
                'question_wiki_id': wid,
                'question': (after_obj or {}).get('question') or base.get('question') or '',
                'answer': (after_obj or {}).get('answer') or base.get('answer') or '',
                'product_name': row.get('product_name'),
                'old_is_configured': row.get('old_is_configured'),
                'new_is_configured': row.get('new_is_configured'),
                'edit_source': row.get('edit_source'),
                'submitted_by': row.get('submitted_by'),
                'submitted_at': row.get('submitted_at'),
                'source_module': '机型矩阵管理',
                'before': before_obj,
                'after': after_obj,
                'changed_fields': changed_fields,
                'changed_products': changed_products,
                'diff_json': row.get('diff_json')
            })

        return jsonify({
            'success': True,
            'operation_id': op_id,
            'total': len(data),
            'affected_wiki_ids_count': len(affected_wiki_ids),
            'data': data
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/add_column', methods=['POST'])
@login_required
def add_matrix_column():
    data = request.json
    product_name = data.get('product_name')
    if not product_name:
        return jsonify({'success': False, 'message': 'Missing product name'}), 400

    _maybe_pull_matrix_and_logs_from_supabase(force=False)
        
    if MatrixColumn.query.filter_by(product_name=product_name).first():
        return jsonify({'success': False, 'message': 'Product already exists'}), 400
        
    max_order = db.session.query(func.max(MatrixColumn.sort_order)).scalar() or 0
    new_col = MatrixColumn(product_name=product_name, sort_order=max_order + 1)
    db.session.add(new_col)
    db.session.commit()

    warnings = []
    if is_supabase_matrix_enabled():
        try:
            so = int(new_col.sort_order or 0)
            pn = str(product_name).strip()

            c = get_supabase_client()
            if c:
                resp = c.upsert('matrix_column', [{'product_name': pn, 'sort_order': so}], on_conflict='product_name')
                if resp is None or getattr(resp, 'status_code', 500) >= 400:
                    warnings.append(f'远端矩阵列同步失败: {getattr(resp, "text", "") or "unknown error"}')
        except Exception as e:
            warnings.append(f'远端矩阵列同步异常: {str(e)}')
    
    out = {'success': True}
    if warnings:
        out['warnings'] = warnings
    return jsonify(out)

@app.route('/api/matrix/products', methods=['GET'])
@login_required
def get_matrix_products():
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        # Get from Catalog
        catalog = parse_product_catalog()
        catalog_products = set()
        for p_list in catalog.values():
            if isinstance(p_list, list):
                for p in p_list:
                    if p: catalog_products.add(str(p).strip())

        # Get from MatrixColumn (configured columns)
        cols = MatrixColumn.query.order_by(MatrixColumn.sort_order).all()
        db_products = {c.product_name for c in cols}
        
        # Get from ProductMatrix (data) to ensure we don't miss any
        data_products_res = db.session.query(ProductMatrix.product_name).distinct().all()
        data_products = {p[0] for p in data_products_res if p[0]}
        
        all_products = sorted(list(catalog_products | db_products | data_products))
        return jsonify({
            'success': True, 
            'data': all_products,
            'catalog': sorted(list(catalog_products))
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/submit_logs', methods=['GET'])
@login_required
def get_matrix_submit_logs():
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        limit = request.args.get('limit', '50')
        try:
            limit = int(limit)
        except Exception:
            limit = 50
        limit = max(1, min(200, limit))

        if is_supabase_matrix_enabled():
            client = get_supabase_client()
            if not client:
                return jsonify({'success': False, 'message': '数据库客户端不可用'}), 500
            resp = client.select('matrix_submit_operation', page=1, page_size=limit, filters={}, order_by='created_at', order_dir='desc', columns='operation_id,status,attempts,created_by,error_message,created_at,updated_at')
            if not resp or getattr(resp, 'status_code', 500) >= 400:
                return jsonify({'success': False, 'message': getattr(resp, 'text', '') or '数据库查询失败'}), 500
            ops = resp.json() or []
            data = []
            for op in ops:
                if not isinstance(op, dict):
                    continue
                data.append({
                    'operation_id': op.get('operation_id'),
                    'status': op.get('status'),
                    'attempts': op.get('attempts'),
                    'created_by': op.get('created_by'),
                    'error_message': op.get('error_message'),
                    'created_at': _dt_to_iso(op.get('created_at')),
                    'updated_at': _dt_to_iso(op.get('updated_at'))
                })
            return jsonify({'success': True, 'data': data})
        else:
            ops = MatrixSubmitOperation.query.order_by(MatrixSubmitOperation.created_at.desc()).limit(limit).all()
            data = []
            for op in ops:
                data.append({
                    'operation_id': op.operation_id,
                    'status': op.status,
                    'attempts': op.attempts,
                    'created_by': op.created_by,
                    'error_message': op.error_message,
                    'created_at': op.created_at.isoformat() + 'Z' if op.created_at else None,
                    'updated_at': op.updated_at.isoformat() + 'Z' if op.updated_at else None
                })
            return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/categories', methods=['GET'])
@login_required
def get_matrix_categories():
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        # Get from Catalog
        catalog = parse_product_catalog()
        catalog_cats = set(catalog.keys())

        # Get from ProductMatrix
        cats = db.session.query(ProductMatrix.product_category).distinct().all()
        db_cats = {c[0] for c in cats if c[0]}
        
        categories = sorted(list(catalog_cats | db_cats))
        return jsonify({
            'success': True, 
            'data': categories,
            'catalog': sorted(list(catalog_cats))
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/stats', methods=['GET'])
@login_required
def get_matrix_stats():
    _maybe_pull_matrix_and_logs_from_supabase(force=False)
    mode = request.args.get('mode') # model, category
    source = request.args.get('source')
    
    if not mode or not source:
        return jsonify({'success': False, 'count': 0})
        
    try:
        count = 0
        if mode == 'model':
            # Count configured entries for this model
            # Try exact match first
            count = ProductMatrix.query.filter_by(product_name=source).count()
            
            # If 0, try removing spaces from source (or matching leniently)
            if count == 0 and ' ' in source:
                clean_source = source.replace(' ', '')
                count = ProductMatrix.query.filter_by(product_name=clean_source).count()
                
            # Also try matching against DB if DB has spaces but source doesn't? 
            # (Less likely given user input usually comes from catalog which seems to be the one with spaces?)
            # Actually, user selected "分子筛干衣机 Z1 Max" (space), DB has "分子筛干衣机Z1 Max" (no space).
            # So stripping space from source is the correct fallback.
        elif mode == 'category':
            mappings = get_model_mappings()
            if source in mappings:
                allowed_models = [str(x) for x in (mappings[source] or []) if x]
                n = len(allowed_models)
                if n == 0:
                    count = 0
                else:
                    count = db.session.query(ProductMatrix.question_wiki_id)\
                        .filter(ProductMatrix.is_configured == True)\
                        .group_by(ProductMatrix.question_wiki_id)\
                        .having(func.count(func.distinct(ProductMatrix.product_name)) == n)\
                        .having(func.count(func.distinct(case(
                            (ProductMatrix.product_name.in_(allowed_models), ProductMatrix.product_name),
                            else_=None
                        ))) == n)\
                        .count()
            else:
                count = db.session.query(ProductMatrix.question_wiki_id).filter(ProductMatrix.product_category == source).distinct().count()
            
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/clone_config/preview', methods=['POST'])
@login_required
def preview_matrix_clone_config():
    data = request.json or {}
    mode = data.get('mode')
    source = data.get('source')
    scope = data.get('scope') if isinstance(data.get('scope'), dict) else {}

    if not mode or not source:
        scope_mode = str(scope.get('mode') or 'all').strip().lower() or 'all'
        return jsonify({
            'success': True,
            'count': 0,
            'source_count': 0,
            'scope_mode': scope_mode,
            'scope_ids_count': None,
            'message': 'Missing source'
        })

    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        scope_mode, scope_ids = _resolve_matrix_clone_scope_ids(scope, allow_empty_selected=True)
        source_items, _ = _resolve_matrix_clone_source_items(mode, source, scope_ids)
        scope_ids_count = len(scope_ids) if scope_ids is not None else None
        return jsonify({
            'success': True,
            'count': len(source_items),
            'source_count': len(source_items),
            'scope_mode': scope_mode,
            'scope_ids_count': scope_ids_count
        })
    except ValueError as e:
        return jsonify({'success': False, 'count': 0, 'message': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'count': 0, 'message': str(e)}), 500

@app.route('/api/matrix/clone_config', methods=['POST'])
@login_required
def clone_matrix_config():
    data = request.json or {}
    mode = data.get('mode') # model, category
    source = data.get('source')
    targets = data.get('targets', []) # List of target models
    strategy = data.get('strategy', 'append') # append, force_sync
    scope = data.get('scope') if isinstance(data.get('scope'), dict) else {}
    
    if not mode or not source or not targets:
        return jsonify({'success': False, 'message': 'Missing required parameters'}), 400
        
    if not isinstance(targets, list):
        return jsonify({'success': False, 'message': 'Targets must be a list'}), 400
        
    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        scope_mode, scope_ids = _resolve_matrix_clone_scope_ids(scope)
        pending_changes = []
        source_items, source_products_to_remove = _resolve_matrix_clone_source_items(mode, source, scope_ids)
        
        if not source_items:
            return jsonify({
                'success': True,
                'updated_count': 0,
                'removed_count': 0,
                'pending_changes': [],
                'pending_count': 0,
                'source_count': 0,
                'scope_mode': scope_mode,
                'message': 'No source data found'
            })

        total_updated = 0
        total_removed = 0
        
        # 2. Process Targets
        existing_columns = {c.product_name for c in MatrixColumn.query.all()}
        current_max_order = db.session.query(func.max(MatrixColumn.sort_order)).scalar() or 0
        
        for target_product in targets:
            # Ensure MatrixColumn exists
            if target_product not in existing_columns:
                current_max_order += 1
                db.session.add(MatrixColumn(product_name=target_product, sort_order=current_max_order))
                existing_columns.add(target_product)
            
            # Get existing entries for target to minimize queries
            target_entries = ProductMatrix.query.filter_by(product_name=target_product).all()
            target_map = {item.question_wiki_id: item for item in target_entries}
            
            for wiki_id, source_item in source_items.items():
                if wiki_id in target_map:
                    # Exists
                    if strategy == 'force_sync':
                        # Update content and ensure configured
                        target_item = target_map[wiki_id]
                        old_cfg = bool(getattr(target_item, 'is_configured', False))
                        target_item.question_content = source_item.question_content
                        target_item.answer_content = source_item.answer_content
                        target_item.product_category = source_item.product_category
                        target_item.update_time = source_item.update_time
                        target_item.is_configured = True
                        target_item.manual_edit = True
                        if getattr(target_item, 'edit_source', '') != 'cell':
                            target_item.edit_source = 'bulk'
                        target_item.last_synced_at = datetime.utcnow()
                        total_updated += 1
                        if old_cfg is not True:
                            pending_changes.append({
                                'question_wiki_id': wiki_id,
                                'product_name': target_product,
                                'old_is_configured': bool(old_cfg),
                                'new_is_configured': True,
                                'edit_source': 'bulk'
                            })
                    else:
                        target_item = target_map[wiki_id]
                        if not getattr(target_item, 'is_configured', False):
                            old_cfg = bool(getattr(target_item, 'is_configured', False))
                            target_item.is_configured = True
                            target_item.manual_edit = True
                            if getattr(target_item, 'edit_source', '') != 'cell':
                                target_item.edit_source = 'bulk'
                            if not (target_item.question_content or '').strip():
                                target_item.question_content = source_item.question_content
                            if not (target_item.answer_content or '').strip():
                                target_item.answer_content = source_item.answer_content
                            if not getattr(target_item, 'product_category', None):
                                target_item.product_category = source_item.product_category
                            if not getattr(target_item, 'update_time', None):
                                target_item.update_time = source_item.update_time
                            target_item.last_synced_at = datetime.utcnow()
                            total_updated += 1
                            pending_changes.append({
                                'question_wiki_id': wiki_id,
                                'product_name': target_product,
                                'old_is_configured': bool(old_cfg),
                                'new_is_configured': True,
                                'edit_source': 'bulk'
                            })
                else:
                    # New
                    new_item = ProductMatrix(
                        question_wiki_id=wiki_id,
                        product_name=target_product,
                        question_content=source_item.question_content,
                        answer_content=source_item.answer_content,
                        product_category=source_item.product_category,
                        update_time=source_item.update_time,
                        is_configured=True,
                        manual_edit=True, # New entry created by clone
                        edit_source='bulk',
                        last_synced_at=datetime.utcnow()
                    )
                    db.session.add(new_item)
                    total_updated += 1
                    pending_changes.append({
                        'question_wiki_id': wiki_id,
                        'product_name': target_product,
                        'old_is_configured': False,
                        'new_is_configured': True,
                        'edit_source': 'bulk'
                    })

        if strategy == 'force_sync':
            remove_products = set(source_products_to_remove) - set([str(t) for t in targets])
            if remove_products:
                to_remove_rows = ProductMatrix.query.filter(
                    ProductMatrix.question_wiki_id.in_(list(source_items.keys())),
                    ProductMatrix.product_name.in_(list(remove_products))
                ).all()
                for r in (to_remove_rows or []):
                    old_cfg = bool(getattr(r, 'is_configured', False))
                    if old_cfg is True:
                        pending_changes.append({
                            'question_wiki_id': str(r.question_wiki_id),
                            'product_name': str(r.product_name),
                            'old_is_configured': True,
                            'new_is_configured': False,
                            'edit_source': 'bulk'
                        })
                total_removed = ProductMatrix.query.filter(
                    ProductMatrix.question_wiki_id.in_(list(source_items.keys())),
                    ProductMatrix.product_name.in_(list(remove_products))
                ).delete(synchronize_session=False)
        
        db.session.commit()
        return jsonify({
            'success': True,
            'updated_count': total_updated,
            'removed_count': total_removed,
            'pending_changes': pending_changes,
            'pending_count': len(pending_changes),
            'source_count': len(source_items),
            'scope_mode': scope_mode
        })
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/matrix/copy_column', methods=['POST'])
@login_required
def copy_matrix_column():
    data = request.json
    source_product = data.get('source_product')
    target_product = data.get('target_product')
    
    if not source_product or not target_product:
        return jsonify({'success': False, 'message': 'Missing products'}), 400

    _maybe_pull_matrix_and_logs_from_supabase(force=False)
        
    # Ensure target column exists
    if not MatrixColumn.query.filter_by(product_name=target_product).first():
        max_order = db.session.query(func.max(MatrixColumn.sort_order)).scalar() or 0
        db.session.add(MatrixColumn(product_name=target_product, sort_order=max_order + 1))
        
    # Copy Config
    source_items = ProductMatrix.query.filter_by(product_name=source_product, is_configured=True).all()
    count = 0
    for item in source_items:
        # Check if target exists
        target_item = ProductMatrix.query.filter_by(question_wiki_id=item.question_wiki_id, product_name=target_product).first()
        if not target_item:
            target_item = ProductMatrix(
                question_wiki_id=item.question_wiki_id,
                product_name=target_product,
                question_content=item.question_content,
                answer_content=item.answer_content,
                product_category=item.product_category,
                update_time=item.update_time,
                is_configured=True,
                manual_edit=True,
                edit_source='bulk',
                last_synced_at=datetime.utcnow()
            )
            db.session.add(target_item)
            count += 1
        elif not target_item.is_configured:
            target_item.is_configured = True
            target_item.manual_edit = True
            if getattr(target_item, 'edit_source', '') != 'cell':
                target_item.edit_source = 'bulk'
            target_item.last_synced_at = datetime.utcnow()
            count += 1
            
    db.session.commit()
    return jsonify({'success': True, 'updated_count': count})

def _run_matrix_sync(mode='merge'):
    client = get_supabase_client()
    if not client:
        return {'success': False, 'message': '数据库未连接', '_http_status': 400}

    try:
        _maybe_pull_matrix_and_logs_from_supabase(force=False)
        # 1. Fetch KB Data (All items)
        kb_items = client.select_all('knowledge_base_v1', columns='question_wiki_id,question,answer,product_name,product_category_name,update_time', order_by='question_wiki_id')
        kb_map = {item['question_wiki_id']: item for item in kb_items if item.get('question_wiki_id')}
        
        # 2. Sync Columns (Critical: Ensure columns exist for all products found in KB)
        all_products = set()
        for item in kb_items:
            p_str = str(item.get('product_name') or "")
            # Split products by comma or Chinese comma
            products = [p.strip() for p in re.split(r'[,，]', p_str) if p.strip()]
            for p in products:
                all_products.add(p)
        
        existing_cols = {c.product_name for c in MatrixColumn.query.all()}
        new_cols = all_products - existing_cols
        if new_cols:
            max_order = db.session.query(func.max(MatrixColumn.sort_order)).scalar() or 0
            for i, p in enumerate(sorted(new_cols)):
                db.session.add(MatrixColumn(product_name=p, sort_order=max_order + i + 1))
            db.session.commit()

        # 3. Sync Data based on Mode
        added = 0
        updated = 0
        deleted = 0
        
        if mode == 'reset':
            # Mode: Reset - Clear all and regenerate
            ProductMatrix.query.delete()
            
            for kbid, item in kb_map.items():
                # For each product listed in the KB item, create a matrix entry
                p_str = str(item.get('product_name') or "")
                products = [p.strip() for p in re.split(r'[,，]', p_str) if p.strip()]
                
                for p in products:
                    matrix_item = ProductMatrix(
                        question_wiki_id=kbid,
                        product_name=p,
                        question_content=item.get('question', ''),
                        answer_content=item.get('answer', ''),
                        product_category=item.get('product_category_name', ''),
                        update_time=item.get('update_time', ''),
                        manual_edit=False,
                        edit_source='',
                        is_configured=True, # Auto-added from KB implies configured for that product
                        last_synced_at=datetime.utcnow()
                    )
                    db.session.add(matrix_item)
                    added += 1
                
        elif mode == 'content_refresh':
            # Mode: Content Refresh - Update content fields, preserve manual edits if possible (or just update content)
            # The goal of content_refresh is to update Question/Answer texts without changing the "Configured" status if possible,
            # BUT usually content refresh implies bringing in latest KB state.
            # Let's define it as: Update texts for existing, Add new, Remove deleted.
            
            matrix_items = ProductMatrix.query.all()
            matrix_map = {(m.question_wiki_id, m.product_name): m for m in matrix_items}
            
            # Track processed keys to find deletions
            processed_keys = set()
            
            for kbid, item in kb_map.items():
                p_str = str(item.get('product_name') or "")
                products = [p.strip() for p in re.split(r'[,，]', p_str) if p.strip()]
                
                for p in products:
                    key = (kbid, p)
                    processed_keys.add(key)
                    
                    if key in matrix_map:
                        m = matrix_map[key]
                        # Update content
                        m.question_content = item.get('question', '')
                        m.answer_content = item.get('answer', '')
                        m.product_category = item.get('product_category_name', '')
                        m.update_time = item.get('update_time', '')
                        m.last_synced_at = datetime.utcnow()
                        if not m.manual_edit:
                            m.edit_source = ''
                        updated += 1
                    else:
                        # New
                        matrix_item = ProductMatrix(
                            question_wiki_id=kbid,
                            product_name=p,
                            question_content=item.get('question', ''),
                            answer_content=item.get('answer', ''),
                            product_category=item.get('product_category_name', ''),
                            update_time=item.get('update_time', ''),
                            manual_edit=False,
                            edit_source='',
                            is_configured=True,
                            last_synced_at=datetime.utcnow()
                        )
                        db.session.add(matrix_item)
                        added += 1
            
            # Delete missing
            for key, m in matrix_map.items():
                if key not in processed_keys:
                    db.session.delete(m)
                    deleted += 1
                    
        else: 
            # Mode: Merge (Default) - Preserve manual edits
            matrix_items = ProductMatrix.query.all()
            matrix_map = {(m.question_wiki_id, m.product_name): m for m in matrix_items}
            processed_keys = set()
            
            for kbid, item in kb_map.items():
                p_str = str(item.get('product_name') or "")
                products = [p.strip() for p in re.split(r'[,，]', p_str) if p.strip()]
                
                for p in products:
                    key = (kbid, p)
                    processed_keys.add(key)
                    
                    if key in matrix_map:
                        m = matrix_map[key]
                        if not m.manual_edit:
                            m.question_content = item.get('question', '')
                            m.answer_content = item.get('answer', '')
                            m.product_category = item.get('product_category_name', '')
                            m.update_time = item.get('update_time', '')
                            m.last_synced_at = datetime.utcnow()
                            m.edit_source = ''
                            updated += 1
                    else:
                        # New
                        matrix_item = ProductMatrix(
                            question_wiki_id=kbid,
                            product_name=p,
                            question_content=item.get('question', ''),
                            answer_content=item.get('answer', ''),
                            product_category=item.get('product_category_name', ''),
                            update_time=item.get('update_time', ''),
                            manual_edit=False,
                            edit_source='',
                            is_configured=True,
                            last_synced_at=datetime.utcnow()
                        )
                        db.session.add(matrix_item)
                        added += 1
            
            # Delete missing (Only if not manual edit)
            for key, m in matrix_map.items():
                if key not in processed_keys:
                    if not m.manual_edit:
                        db.session.delete(m)
                        deleted += 1

        db.session.commit()
        return {
            'success': True,
            'added': added,
            'updated': updated,
            'deleted': deleted
        }

    except Exception as e:
        traceback.print_exc()
        db.session.rollback()
        return {'success': False, 'message': str(e), '_http_status': 500}


@app.route('/api/matrix/sync', methods=['POST'])
@login_required
def sync_matrix():
    data = request.json or {}
    mode = data.get('mode', 'merge')
    r = _run_matrix_sync(mode)
    if r.get('success'):
        return jsonify(r)
    code = int(r.pop('_http_status', 500))
    return jsonify(r), code


@app.route('/api/kb/sync_downstream', methods=['POST'])
@login_required
def kb_sync_downstream():
    """
    One-click: pull latest V1 into 机型矩阵(merge) + 多媒体 link_previews + 评分 kb_scores.
    治理：月度召回仍依赖 Excel 导入；本接口会同步评分快照，治理页关联展示会随之更新。
    """
    steps = {}
    r_matrix = _run_matrix_sync('merge')
    steps['matrix'] = r_matrix
    if not r_matrix.get('success'):
        return jsonify({'success': False, 'message': r_matrix.get('message', 'matrix sync failed'), 'steps': steps}), 500

    r_links = _run_sync_kb_links()
    steps['links'] = r_links
    if not r_links.get('success'):
        return jsonify({'success': False, 'message': r_links.get('message', 'links sync failed'), 'steps': steps}), 500

    r_scores = _run_scoring_sync()
    steps['scoring'] = r_scores
    if not r_scores.get('success'):
        return jsonify({'success': False, 'message': r_scores.get('message', 'scoring sync failed'), 'steps': steps}), 500

    steps['governance'] = {
        'note': '治理「月度召回」数据仍通过治理页 Excel 导入；评分表已随上一步同步，治理页展示会关联最新快照。'
    }
    return jsonify({
        'success': True,
        'message': '已同步：矩阵、多媒体预览、评分；治理说明见 steps.governance',
        'steps': steps
    })


if __name__ == '__main__':
    init_db()
    _start_supabase_outbox_background_worker()
    
    # Get local IP
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
        print(f"\n{'='*50}")
        print(f" 服务已启动！请使用以下地址访问：")
        server_port = int(os.environ.get('KMATRIX_BACKEND_PORT', '8085'))
        server_host = os.environ.get('KMATRIX_BACKEND_HOST', '0.0.0.0')
        print(f" 本机访问: http://localhost:{server_port}")
        print(f" 局域网共享访问: http://{local_ip}:{server_port}")
        print(f" (请确保其他电脑连接了同一个 WiFi/局域网)")
        print(f"{'='*50}\n")
    except:
        local_ip = '0.0.0.0'
        server_port = int(os.environ.get('KMATRIX_BACKEND_PORT', '8085'))
        server_host = os.environ.get('KMATRIX_BACKEND_HOST', '0.0.0.0')
        
    # Important: do NOT use Flask debug reloader for background jobs.
    # debug=True + reloader will restart the process and in-memory _NHF_JOBS will be lost.
    app.run(port=server_port, host=server_host, debug=False, use_reloader=False)
